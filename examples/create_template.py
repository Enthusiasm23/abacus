"""数据映射模板生成器

使用方法:
    python create_template.py                    # 生成4个源表的基线模板
    python create_template.py -n 6               # 生成6个源表的模板
    python create_template.py -o my_template.xlsx  # 指定输出路径
    python create_template.py -n 8 -o custom.xlsx # 自定义源表数和路径
    python create_template.py -h                 # 显示帮助信息
"""

import sys
sys.path.insert(0, 'src')

from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_template(source_count=4, output=None, quiet=False):
    """
    创建数据映射模板
    
    Args:
        source_count: 源表数量（默认4个）
        output: 输出文件路径（默认: mapping_template_{timestamp}.xlsx）
        quiet: 静默模式，不输出日志
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "{目标表中文名}"
    
    # ==================== 样式定义 ====================
    title_font = Font(bold=True, size=12, name='微软雅黑', color='FFFFFF')
    label_font = Font(bold=True, size=12, name='微软雅黑')
    placeholder_font = Font(size=12, name='微软雅黑', color='FF0000')
    data_font = Font(size=12, name='微软雅黑')
    
    target_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    source_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    field_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    data_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def set_cell(cell, value, font=None, fill=None, align=None):
        cell.value = value
        if font: cell.font = font
        if fill: cell.fill = fill
        if align: cell.alignment = align
        cell.border = thin_border
    
    # ==================== 动态计算 ====================
    base_rows = 5
    extra_rows = max(0, source_count - 4)
    data_last_row = base_rows + extra_rows
    field_row = data_last_row + 1
    
    # ==================== 标题行 ====================
    ws.merge_cells(f'A1:A{data_last_row}')
    set_cell(ws['A1'], '目标表', title_font, target_fill, center_align)
    
    set_cell(ws['B1'], '目标表英文名', label_font, data_fill, center_align)
    
    ws.merge_cells('C1:F1')
    set_cell(ws['C1'], '{目标表英文名}', placeholder_font, data_fill, center_align)
    
    ws.merge_cells(f'G1:G{data_last_row}')
    set_cell(ws['G1'], '源表', title_font, source_fill, center_align)
    
    for col, name in enumerate(['源表中文名', '源表英文名', '关联关系', '过滤条件', '源表别名', 'SCHEMA', '源表备注'], 8):
        set_cell(ws.cell(row=1, column=col), name, label_font, data_fill, center_align)
    
    # ==================== 目标表属性 ====================
    set_cell(ws['B2'], '目标表中文名', label_font, data_fill, center_align)
    ws.merge_cells('C2:F2')
    set_cell(ws['C2'], '{目标表中文名}', placeholder_font, data_fill, center_align)
    
    set_cell(ws['B3'], '主键/分布键/粒度', label_font, data_fill, center_align)
    ws.merge_cells('C3:F3')
    set_cell(ws['C3'], '{主键}', placeholder_font, data_fill, center_align)
    
    set_cell(ws['B4'], '频率/调度/时间', label_font, data_fill, center_align)
    ws.merge_cells('C4:F4')
    set_cell(ws['C4'], '{调度策略}', placeholder_font, data_fill, center_align)
    
    set_cell(ws['B5'], '目标表业务定义', label_font, data_fill, center_align)
    if data_last_row > 5:
        ws.merge_cells(f'B5:B{data_last_row}')
        ws.merge_cells(f'C5:F{data_last_row}')
    else:
        ws.merge_cells('C5:F5')
    set_cell(ws['C5'], '{业务逻辑实体定义}', placeholder_font, data_fill, center_align)
    
    # ==================== 源表区域 ====================
    for row in range(2, data_last_row + 1):
        if row - 1 <= source_count:
            suffix = str(row - 1) if row > 2 else ''
            set_cell(ws[f'H{row}'], f'{{源表中文名{suffix}}}', placeholder_font, data_fill, center_align)
            set_cell(ws[f'I{row}'], f'{{源表英文名{suffix}}}', placeholder_font, data_fill, center_align)
            set_cell(ws[f'J{row}'], '{关联关系}', placeholder_font, data_fill, center_align)
            set_cell(ws[f'K{row}'], '{过滤条件}', placeholder_font, data_fill, center_align)
            set_cell(ws[f'L{row}'], f'{{别名{suffix}}}', placeholder_font, data_fill, center_align)
            set_cell(ws[f'M{row}'], f'{{SCHEMA{suffix}}}', placeholder_font, data_fill, center_align)
            set_cell(ws[f'N{row}'], '{源表备注}', placeholder_font, data_fill, center_align)
        else:
            for col in range(8, 15):
                cell = ws.cell(row=row, column=col)
                cell.fill = data_fill
                cell.border = thin_border
                cell.alignment = center_align
    
    for row in range(1, data_last_row + 1):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            if not cell.fill or cell.fill.patternType is None:
                cell.fill = data_fill
            if not cell.border or cell.border.left.style is None:
                cell.border = thin_border
            if not cell.alignment or cell.alignment.horizontal is None:
                cell.alignment = center_align
    
    # ==================== 字段映射表 ====================
    headers = ['字段序号', '字段英文名', '字段中文名', '字段属性说明', '类型及长度(位数)',
               '是否可为空', '来源别名', '来源表名', '来源字段', '加工逻辑', '加工说明',
               '基线版本', '版本日期', '目标表备注']
    for col, h in enumerate(headers, 1):
        set_cell(ws.cell(row=field_row, column=col), h, title_font, field_fill, center_align)
    
    set_cell(ws[f'A{field_row+1}'], 1, data_font, data_fill, center_align)
    set_cell(ws[f'G{field_row+1}'], '{别名}', placeholder_font, data_fill, center_align)
    set_cell(ws[f'A{field_row+2}'], 2, data_font, data_fill, center_align)
    set_cell(ws[f'G{field_row+2}'], '{别名}', placeholder_font, data_fill, center_align)
    
    for row in [field_row+1, field_row+2]:
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            if not cell.fill or cell.fill.patternType is None:
                cell.fill = data_fill
            if not cell.border or cell.border.left.style is None:
                cell.border = thin_border
            if not cell.alignment or cell.alignment.horizontal is None:
                cell.alignment = center_align
    
    # ==================== 样式设置 ====================
    for col, w in {'A':9.75, 'B':18.625, 'C':11.875, 'D':14.125, 'E':16.375, 'F':11.875, 'G':9.75,
                   'H':14.5, 'I':14.5, 'J':11.125, 'K':11.125, 'L':9.75, 'M':12.875, 'N':11.875}.items():
        ws.column_dimensions[col].width = w
    
    for row in range(1, 5):
        ws.row_dimensions[row].height = 18.0
    for row in range(5, data_last_row + 1):
        ws.row_dimensions[row].height = 17.25
    ws.row_dimensions[field_row].height = 18.0
    for row in [field_row+1, field_row+2]:
        ws.row_dimensions[row].height = 17.25
    
    # ==================== 填写说明 ====================
    ws2 = wb.create_sheet(title="填写说明")
    
    title_font_doc = Font(bold=True, size=16, name='微软雅黑', color='4472C4')
    h1_font = Font(bold=True, size=14, name='微软雅黑', color='333333')
    h2_font = Font(bold=True, size=12, name='微软雅黑', color='4472C4')
    h3_font = Font(bold=True, size=11, name='微软雅黑', color='333333')
    normal_font = Font(size=11, name='微软雅黑')
    
    docs = [
        ("A1", "数据映射模板 - 填写说明", title_font_doc),
        ("A2", "本文档详细说明如何填写数据映射模板，适用于开发人员和AI系统。", normal_font),
        
        ("A4", "一、模板概述", h1_font),
        ("A5", "本模板用于定义数据仓库中【目标表】与【源表】之间的映射关系。", normal_font),
        ("A6", "模板包含表级映射信息和字段级映射信息。", normal_font),
        ("A7", "所有需要填写的地方都用 {占位符} 格式标记，红色字体显示。", normal_font),
        
        ("A9", "二、表格结构详解", h1_font),
        ("A11", "【Row 1: 标题行】", h2_font),
        ("A12", "  A1 (蓝色背景): '目标表' - 左侧为目标表信息区域", normal_font),
        ("A13", "  B1: '目标表英文名' - 属性名称", normal_font),
        ("A14", "  C1:F1 (合并): 填写目标表的英文名称", normal_font),
        ("A15", "  G1 (绿色背景): '源表' - 右侧为源表信息区域", normal_font),
        ("A16", "  H1-N1: 源表属性名（中文名、英文名、关联关系、过滤条件、别名、SCHEMA、备注）", normal_font),
        
        ("A18", "【Row 2-5: 数据区域】", h2_font),
        ("A19", "  Row 2: 目标表中文名 + 源表1的信息", normal_font),
        ("A20", "  Row 3: 主键/分布键/粒度 + 源表2的信息", normal_font),
        ("A21", "  Row 4: 频率/调度/时间 + 源表3的信息", normal_font),
        ("A22", "  Row 5+: 目标表业务定义 + 源表4+的信息", normal_font),
        
        ("A24", "【Row 6+: 字段映射表头 (橙色背景)】", h2_font),
        ("A25", "  定义字段级别的映射关系", normal_font),
        ("A26", "  A: 字段序号 | B: 字段英文名 | C: 字段中文名", normal_font),
        ("A27", "  D: 字段属性说明 | E: 类型及长度 | F: 是否可为空", normal_font),
        ("A28", "  G: 来源别名 | H: 来源表名 | I: 来源字段", normal_font),
        ("A29", "  J: 加工逻辑 | K: 加工说明 | L: 基线版本", normal_font),
        ("A30", "  M: 版本日期 | N: 目标表备注", normal_font),
        
        ("A32", "【示例数据行】", h2_font),
        ("A33", "  提供字段映射的示例，可删除或修改", normal_font),
        
        ("A35", "三、左侧目标表属性详解", h1_font),
        ("A37", "【B2-C2: 目标表中文名】", h3_font),
        ("A38", "  含义: 目标表的中文名称，替换后会重命名Sheet", normal_font),
        ("A39", "  填写: 如 '用户信息表'", normal_font),
        
        ("A41", "【B3-C3: 主键/分布键/粒度】", h3_font),
        ("A42", "  含义: 表的主键字段", normal_font),
        ("A43", "  填写: 如 'user_id' 或 'user_id, order_id'", normal_font),
        
        ("A45", "【B4-C4: 频率/调度/时间】", h3_font),
        ("A46", "  含义: 数据更新频率", normal_font),
        ("A47", "  填写格式: '频率 | 时间 | 间隔'", normal_font),
        ("A48", "  示例: 'T+1 | 0点，12点，17点 | 半小时'", normal_font),
        
        ("A50", "【B5-C5: 目标表业务定义】", h3_font),
        ("A51", "  含义: 目标表的业务含义描述", normal_font),
        ("A52", "  示例: '用户基本信息实体，包含用户的注册、登录、权限等核心数据'", normal_font),
        
        ("A54", "四、右侧源表属性详解", h1_font),
        ("A56", "【H列: 源表中文名】", h3_font),
        ("A57", "  含义: 源表的中文名称", normal_font),
        ("A59", "【I列: 源表英文名】", h3_font),
        ("A60", "  含义: 源表的英文名称，如 'ods_user'", normal_font),
        
        ("A62", "【J列: 关联关系】", h3_font),
        ("A63", "  含义: 多表关联时的JOIN条件和类型", normal_font),
        ("A64", "  格式: 'JOIN_TYPE ON 条件' 或 'form 表别名'", normal_font),
        ("A66", "  JOIN类型说明:", h3_font),
        ("A67", "    - form t1: 单表查询（无JOIN）", normal_font),
        ("A68", "    - left join t2 ON ...: 左连接（保留左表全部）", normal_font),
        ("A69", "    - right join t2 ON ...: 右连接（保留右表全部）", normal_font),
        ("A70", "    - inner join t2 ON ...: 内连接（只保留匹配行）", normal_font),
        ("A71", "    - full join t2 ON ...: 全连接（保留两表全部）", normal_font),
        ("A73", "  示例:", h3_font),
        ("A74", "    form t1", normal_font),
        ("A75", "    left join t2 ON t1.order_id = t2.order_id", normal_font),
        ("A76", "    inner join t3 ON t2.product_id = t3.product_id", normal_font),
        
        ("A66", "【K列: 过滤条件】", h3_font),
        ("A67", "  含义: 数据筛选的WHERE条件", normal_font),
        ("A68", "  填写: 如 \"where t1.del_flag = 'N'\"", normal_font),
        
        ("A70", "【L列: 源表别名】", h3_font),
        ("A71", "  含义: SQL中使用的表别名，如 't1', 't2'", normal_font),
        ("A72", "  规则: 每个源表的别名必须唯一", normal_font),
        
        ("A74", "【M列: SCHEMA】", h3_font),
        ("A75", "  含义: 数据库Schema名称，如 'sdiipd'", normal_font),
        
        ("A77", "【N列: 源表备注】", h3_font),
        ("A78", "  含义: 源表的补充说明（可选）", normal_font),
        
        ("A80", "五、占位符系统", h1_font),
        ("A81", "所有需要填写的地方都用 {xxx} 格式的占位符标记。", normal_font),
        ("A83", "【表级占位符】", h2_font),
        ("A84", "  {目标表中文名} - 替换后会重命名Sheet", normal_font),
        ("A85", "  {目标表英文名} - 目标表英文名", normal_font),
        ("A86", "  {主键} - 主键字段名", normal_font),
        ("A87", "  {调度策略} - 更新频率描述", normal_font),
        ("A88", "  {业务逻辑实体定义} - 业务含义描述", normal_font),
        
        ("A90", "【源表占位符】", h2_font),
        ("A91", "  第1个源表: {源表中文名}, {源表英文名}, {关联关系}, {过滤条件}, {别名}, {SCHEMA}, {源表备注}", normal_font),
        ("A92", "  第2+个源表: 后缀递增，如 {源表中文名2}, {别名3} 等", normal_font),
        
        ("A94", "六、颜色编码说明", h1_font),
        ("A95", "  蓝色背景 (A1): 目标表标题区域", normal_font),
        ("A96", "  绿色背景 (G1): 源表标题区域", normal_font),
        ("A97", "  橙色背景 (字段表头行): 字段映射表头", normal_font),
        ("A98", "  浅蓝背景: 数据区域", normal_font),
        ("A99", "  红色字体: 占位符，需要替换", normal_font),
        
        ("A101", "七、源表扩展规则", h1_font),
        ("A102", "当源表数量超过4个时，模板会自动扩展：", normal_font),
        ("A103", "  - 目标表标题(A列)向下合并", normal_font),
        ("A104", "  - 目标表业务定义(B5)向下合并", normal_font),
        ("A105", "  - 业务逻辑实体定义(C5:F5)向下合并", normal_font),
        ("A106", "  - 源表标题(G列)向下合并", normal_font),
        
        ("A108", "八、填写示例", h1_font),
        ("A109", "【目标表信息】", h2_font),
        ("A110", "  英文名: dwd_user_info_t | 中文名: 用户信息表", normal_font),
        ("A111", "  主键: user_id | 调度: T+1 | 0点 | 半小时", normal_font),
        ("A112", "  业务定义: 用户基本信息实体", normal_font),
        ("A114", "【源表信息】", h2_font),
        ("A115", "  源表1: ods_user (用户原始数据)", normal_font),
        ("A116", "  别名: t1 | Schema: sdiipd | 过滤: where t1.del_flag = 'N'", normal_font),
        
        ("A118", "九、注意事项", h1_font),
        ("A119", "1. Sheet名称最长31个字符，不能包含 : \\ / ? * [ ] 等特殊字符", normal_font),
        ("A120", "2. 未匹配的占位符会保留原样", normal_font),
        ("A121", "3. 每个源表的别名必须唯一", normal_font),
        ("A122", "4. 版本日期格式: YYYY-MM-DD", normal_font),
    ]
    
    for ref, text, font in docs:
        ws2[ref] = text
        ws2[ref].font = font
    
    for row in range(1, 124):
        if ws2[f'A{row}'].value:
            ws2.row_dimensions[row].height = 20
    ws2.row_dimensions[1].height = 30
    ws2.row_dimensions[2].height = 25
    ws2.column_dimensions['A'].width = 80
    
    # ==================== 示例Sheet ====================
    ws3 = wb.create_sheet(title="填写示例")
    
    no_wrap_align = Alignment(horizontal='left', vertical='center')
    no_wrap_center = Alignment(horizontal='center', vertical='center')
    
    # Row 1: 标题行
    ws3.merge_cells('A1:A5')
    set_cell(ws3['A1'], '目标表', title_font, target_fill, center_align)
    set_cell(ws3['B1'], '目标表英文名', label_font, data_fill, center_align)
    ws3.merge_cells('C1:F1')
    set_cell(ws3['C1'], 'dwd_user_order_detail', data_font, data_fill, center_align)
    ws3.merge_cells('G1:G5')
    set_cell(ws3['G1'], '源表', title_font, source_fill, center_align)
    for col, name in enumerate(['源表中文名', '源表英文名', '关联关系', '过滤条件', '源表别名', 'SCHEMA', '源表备注'], 8):
        set_cell(ws3.cell(row=1, column=col), name, label_font, data_fill, no_wrap_center)
    
    # Row 2-5: 数据区域（先设置样式，再合并）
    # Row 2
    set_cell(ws3['B2'], '目标表中文名', label_font, data_fill, no_wrap_center)
    set_cell(ws3['C2'], '用户订单明细表', data_font, data_fill, no_wrap_center)
    for col in ['D', 'E', 'F']:
        set_cell(ws3[f'{col}2'], None, data_font, data_fill, no_wrap_center)
    ws3.merge_cells('C2:F2')
    
    # Row 3
    set_cell(ws3['B3'], '主键/分布键/粒度', label_font, data_fill, no_wrap_center)
    set_cell(ws3['C3'], 'order_detail_id', data_font, data_fill, no_wrap_center)
    for col in ['D', 'E', 'F']:
        set_cell(ws3[f'{col}3'], None, data_font, data_fill, no_wrap_center)
    ws3.merge_cells('C3:F3')
    
    # Row 4
    set_cell(ws3['B4'], '频率/调度/时间', label_font, data_fill, no_wrap_center)
    set_cell(ws3['C4'], 'T+1 | 0点 | 每日', data_font, data_fill, no_wrap_center)
    for col in ['D', 'E', 'F']:
        set_cell(ws3[f'{col}4'], None, data_font, data_fill, no_wrap_center)
    ws3.merge_cells('C4:F4')
    
    # Row 5
    set_cell(ws3['B5'], '目标表业务定义', label_font, data_fill, no_wrap_center)
    set_cell(ws3['C5'], '用户订单明细宽表，整合订单、商品、用户等多维度信息', data_font, data_fill, no_wrap_center)
    for col in ['D', 'E', 'F']:
        set_cell(ws3[f'{col}5'], None, data_font, data_fill, no_wrap_center)
    ws3.merge_cells('C5:F5')
    
    # 源表1
    set_cell(ws3['H2'], '订单主表', data_font, data_fill, no_wrap_align)
    set_cell(ws3['I2'], 'ods_order_master', data_font, data_fill, no_wrap_align)
    set_cell(ws3['J2'], 'form t1', data_font, data_fill, no_wrap_align)
    set_cell(ws3['K2'], "where t1.del_flag = 'N'", data_font, data_fill, no_wrap_align)
    set_cell(ws3['L2'], 't1', data_font, data_fill, no_wrap_center)
    set_cell(ws3['M2'], 'sdiipd', data_font, data_fill, no_wrap_center)
    set_cell(ws3['N2'], '订单主表', data_font, data_fill, no_wrap_align)
    # 源表2
    set_cell(ws3['H3'], '订单明细表', data_font, data_fill, no_wrap_align)
    set_cell(ws3['I3'], 'ods_order_detail', data_font, data_fill, no_wrap_align)
    set_cell(ws3['J3'], 'left join t2 ON t1.order_id = t2.order_id', data_font, data_fill, no_wrap_align)
    set_cell(ws3['K3'], "where t2.del_flag = 'N'", data_font, data_fill, no_wrap_align)
    set_cell(ws3['L3'], 't2', data_font, data_fill, no_wrap_center)
    set_cell(ws3['M3'], 'sdiipd', data_font, data_fill, no_wrap_center)
    set_cell(ws3['N3'], '订单明细', data_font, data_fill, no_wrap_align)
    # 源表3
    set_cell(ws3['H4'], '商品信息表', data_font, data_fill, no_wrap_align)
    set_cell(ws3['I4'], 'ods_product', data_font, data_fill, no_wrap_align)
    set_cell(ws3['J4'], 'inner join t3 ON t2.product_id = t3.product_id', data_font, data_fill, no_wrap_align)
    set_cell(ws3['K4'], "where t3.del_flag = 'N'", data_font, data_fill, no_wrap_align)
    set_cell(ws3['L4'], 't3', data_font, data_fill, no_wrap_center)
    set_cell(ws3['M4'], 'sdiipd', data_font, data_fill, no_wrap_center)
    set_cell(ws3['N4'], '商品主数据', data_font, data_fill, no_wrap_align)
    # 源表4
    set_cell(ws3['H5'], '用户信息表', data_font, data_fill, no_wrap_align)
    set_cell(ws3['I5'], 'ods_user', data_font, data_fill, no_wrap_align)
    set_cell(ws3['J5'], 'left join t4 ON t1.user_id = t4.user_id', data_font, data_fill, no_wrap_align)
    set_cell(ws3['K5'], "where t4.del_flag = 'N'", data_font, data_fill, no_wrap_align)
    set_cell(ws3['L5'], 't4', data_font, data_fill, no_wrap_center)
    set_cell(ws3['M5'], 'sdiipd', data_font, data_fill, no_wrap_center)
    set_cell(ws3['N5'], '用户主数据', data_font, data_fill, no_wrap_align)
    
    # Row 6: 字段映射表头
    for col, h in enumerate(['字段序号', '字段英文名', '字段中文名', '字段属性说明', '类型及长度(位数)',
                             '是否可为空', '来源别名', '来源表名', '来源字段', '加工逻辑', '加工说明',
                             '基线版本', '版本日期', '目标表备注'], 1):
        set_cell(ws3.cell(row=6, column=col), h, title_font, field_fill, center_align)
    
    # Row 7+: 字段映射数据
    example_data = [
        [1, 'order_detail_id', '订单明细ID', '订单明细唯一标识', 'varchar(64)', 'NO', 't2', 'ods_order_detail', 'detail_id', 't2.detail_id', '直接映射', 'v1.0', '2024-01-15', '主键'],
        [2, 'order_id', '订单号', '订单唯一标识', 'varchar(64)', 'NO', 't1', 'ods_order_master', 'order_id', 't1.order_id', '直接映射', 'v1.0', '2024-01-15', '关联键'],
        [3, 'user_id', '用户ID', '用户唯一标识', 'varchar(64)', 'NO', 't1', 'ods_order_master', 'user_id', 't1.user_id', '直接映射', 'v1.0', '2024-01-15', '关联键'],
        [4, 'product_id', '商品ID', '商品唯一标识', 'varchar(64)', 'NO', 't2', 'ods_order_detail', 'product_id', 't2.product_id', '直接映射', 'v1.0', '2024-01-15', '关联键'],
        [5, 'product_name', '商品名称', '商品中文名称', 'varchar(200)', 'NO', 't3', 'ods_product', 'product_name', 't3.product_name', '直接映射', 'v1.0', '2024-01-15', ''],
        [6, 'product_category', '商品分类', '商品所属分类', 'varchar(50)', 'YES', 't3', 'ods_product', 'category', 't3.category', '直接映射', 'v1.0', '2024-01-15', ''],
        [7, 'quantity', '购买数量', '用户购买的商品数量', 'int', 'NO', 't2', 'ods_order_detail', 'quantity', 't2.quantity', '直接映射', 'v1.0', '2024-01-15', ''],
        [8, 'unit_price', '商品单价', '商品销售单价（元）', 'decimal(10,2)', 'NO', 't2', 'ods_order_detail', 'price', 't2.price', '直接映射', 'v1.0', '2024-01-15', ''],
        [9, 'total_amount', '订单金额', '订单总金额', 'decimal(12,2)', 'NO', 't1', 'ods_order_master', 'total_amount', 't1.total_amount', '直接映射', 'v1.0', '2024-01-15', ''],
        [10, 'discount_amount', '优惠金额', '订单优惠减免金额', 'decimal(10,2)', 'YES', 't1', 'ods_order_master', 'discount_amount', 't1.discount_amount', '直接映射', 'v1.0', '2024-01-15', ''],
        [11, 'pay_amount', '实付金额', '用户实际支付金额', 'decimal(12,2)', 'NO', 't1', 'ods_order_master', 'pay_amount', 't1.total_amount - IFNULL(t1.discount_amount, 0)', '订单金额减去优惠金额', 'v1.0', '2024-01-15', ''],
        [12, 'order_status', '订单状态', '订单当前状态', 'varchar(20)', 'NO', 't1', 'ods_order_master', 'status', "CASE t1.status WHEN 0 THEN '待付款' WHEN 1 THEN '已付款' WHEN 2 THEN '已发货' WHEN 3 THEN '已完成' WHEN 4 THEN '已取消' ELSE '未知' END", '状态码转中文', 'v1.0', '2024-01-15', ''],
        [13, 'payment_method', '支付方式', '用户支付方式', 'varchar(20)', 'YES', 't1', 'ods_order_master', 'pay_method', "CASE t1.pay_method WHEN 1 THEN '微信支付' WHEN 2 THEN '支付宝' WHEN 3 THEN '银行卡' ELSE '其他' END", '支付方式码转中文', 'v1.0', '2024-01-15', ''],
        [14, 'user_name', '用户姓名', '用户真实姓名', 'varchar(50)', 'YES', 't4', 'ods_user', 'user_name', 't4.user_name', '直接映射', 'v1.0', '2024-01-15', ''],
        [15, 'user_phone', '用户手机', '用户手机号码', 'varchar(20)', 'YES', 't4', 'ods_user', 'phone', 'CONCAT(LEFT(t4.phone, 3), \'****\', RIGHT(t4.phone, 4))', '手机脱敏处理', 'v1.0', '2024-01-15', '隐私字段'],
        [16, 'user_level', '用户等级', '用户会员等级', 'varchar(20)', 'YES', 't4', 'ods_user', 'level', "CASE t4.level WHEN 1 THEN '普通会员' WHEN 2 THEN '银卡会员' WHEN 3 THEN '金卡会员' WHEN 4 THEN '钻石会员' ELSE '未知' END", '等级码转中文', 'v1.0', '2024-01-15', ''],
        [17, 'order_time', '下单时间', '用户下单时间', 'datetime', 'NO', 't1', 'ods_order_master', 'create_time', 't1.create_time', '直接映射', 'v1.0', '2024-01-15', ''],
        [18, 'pay_time', '支付时间', '用户支付时间', 'datetime', 'YES', 't1', 'ods_order_master', 'pay_time', 't1.pay_time', '直接映射', 'v1.0', '2024-01-15', ''],
        [19, 'ship_time', '发货时间', '商品发货时间', 'datetime', 'YES', 't1', 'ods_order_master', 'ship_time', 't1.ship_time', '直接映射', 'v1.0', '2024-01-15', ''],
        [20, 'receive_time', '收货时间', '用户确认收货时间', 'datetime', 'YES', 't1', 'ods_order_master', 'receive_time', 't1.receive_time', '直接映射', 'v1.0', '2024-01-15', ''],
        [21, 'province', '省份', '收货地址-省份', 'varchar(50)', 'YES', 't1', 'ods_order_master', 'province', 't1.province', '直接映射', 'v1.0', '2024-01-15', ''],
        [22, 'city', '城市', '收货地址-城市', 'varchar(50)', 'YES', 't1', 'ods_order_master', 'city', 't1.city', '直接映射', 'v1.0', '2024-01-15', ''],
        [23, 'district', '区县', '收货地址-区县', 'varchar(50)', 'YES', 't1', 'ods_order_master', 'district', 't1.district', '直接映射', 'v1.0', '2024-01-15', ''],
        [24, 'remark', '订单备注', '用户订单备注信息', 'varchar(500)', 'YES', 't1', 'ods_order_master', 'remark', 't1.remark', '直接映射', 'v1.0', '2024-01-15', ''],
        [25, 'etl_time', 'ETL时间', '数据抽取时间', 'datetime', 'NO', '-', '-', '-', 'CURRENT_TIMESTAMP', '系统生成', 'v1.0', '2024-01-15', '技术字段'],
    ]
    
    for row_idx, field in enumerate(example_data, 7):
        for col_idx, value in enumerate(field, 1):
            set_cell(ws3.cell(row=row_idx, column=col_idx), value, data_font, data_fill, 
                     no_wrap_center if col_idx in [1, 6, 7, 12, 13] else no_wrap_align)
    
    # 示例Sheet列宽和行高
    for col, w in {'A':9.75, 'B':18.625, 'C':11.875, 'D':14.125, 'E':16.375, 'F':11.875, 'G':9.75,
                   'H':14.5, 'I':14.5, 'J':11.125, 'K':11.125, 'L':9.75, 'M':12.875, 'N':11.875}.items():
        ws3.column_dimensions[col].width = w
    
    for row in range(1, 6):
        ws3.row_dimensions[row].height = 18.0
    ws3.row_dimensions[6].height = 18.0
    for row in range(7, 32):
        ws3.row_dimensions[row].height = 17.25
    
    # ==================== 保存 ====================
    if output is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        script_dir = Path(__file__).parent
        output = script_dir / f"mapping_template_{timestamp}.xlsx"
    else:
        output = Path(output)
    
    # 确保目录存在
    output.parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(output)
    
    if not quiet:
        print(f"✅ {output}")
    
    return output


def show_help():
    """显示帮助信息"""
    print("""
╔══════════════════════════════════════════════════════════════╗
║           数据映射模板生成器                                  ║
╚══════════════════════════════════════════════════════════════╝

【功能】
    生成数据仓库的数据映射模板Excel文件

【用法】
    python create_template.py [选项]

【选项】
    -n, --sources    源表数量（默认: 4）
    -o, --output     输出文件路径（默认: mapping_template_{时间戳}.xlsx）
    -q, --quiet      静默模式，不输出日志
    -h, --help       显示帮助信息

【示例】
    python create_template.py                  生成4个源表的模板
    python create_template.py -n 6             生成6个源表的模板
    python create_template.py -o my.xlsx       指定输出路径
    python create_template.py -n 8 -q          静默生成8个源表的模板
""")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description='数据映射模板生成器',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('-n', '--sources', type=int, default=4, 
                        help='源表数量（默认: 4）')
    parser.add_argument('-o', '--output', type=str, default=None,
                        help='输出文件路径（默认: mapping_template_{时间戳}.xlsx）')
    parser.add_argument('-q', '--quiet', action='store_true',
                        help='静默模式，不输出日志')
    
    args = parser.parse_args()
    
    create_template(source_count=args.sources, output=args.output, quiet=args.quiet)
