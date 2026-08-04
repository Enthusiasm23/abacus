"""测试 auto_sum.py 的完整覆盖"""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from abacus.core.dimension.auto_sum import AutoSumCapability
from abacus.core.exceptions import DataError, FileNotFoundError, ValidationError


@pytest.fixture
def auto_sum_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "auto_sum_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # 添加表头
    ws.append(["Product", "Q1", "Q2", "Q3", "Q4"])
    
    # 添加数据
    ws.append(["A", 100, 150, 200, 250])
    ws.append(["B", 110, 160, 210, 260])
    ws.append(["C", 120, 170, 220, 270])
    
    wb.save(file_path)
    wb.close()
    return file_path


class TestAutoSumCoverage:
    """测试 auto_sum.py 的完整覆盖"""
    
    def test_missing_file_raises_error(self):
        """测试缺少文件参数"""
        cap = AutoSumCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 file"):
            cap.execute(None, sheet="Data", range="B2:E4")
    
    def test_missing_sheet_raises_error(self, auto_sum_excel):
        """测试缺少工作表参数"""
        cap = AutoSumCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 sheet"):
            cap.execute(None, file=str(auto_sum_excel), range="B2:E4")
    
    def test_missing_range_raises_error(self, auto_sum_excel):
        """测试缺少范围参数"""
        cap = AutoSumCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 range"):
            cap.execute(None, file=str(auto_sum_excel), sheet="Data")
    
    def test_invalid_direction_raises_error(self, auto_sum_excel):
        """测试无效方向"""
        cap = AutoSumCapability()
        with pytest.raises(ValidationError, match="direction 必须为"):
            cap.execute(None, file=str(auto_sum_excel), sheet="Data",
                       range="B2:E4", direction="invalid")
    
    def test_file_not_found_raises_error(self, tmp_path):
        """测试文件不存在"""
        cap = AutoSumCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "nonexistent.xlsx"),
                       sheet="Data", range="B2:E4")
    
    def test_sheet_not_found_raises_error(self, auto_sum_excel):
        """测试工作表不存在"""
        cap = AutoSumCapability()
        with pytest.raises(DataError, match="不存在"):
            cap.execute(None, file=str(auto_sum_excel),
                       sheet="NoSuchSheet", range="B2:E4")
    
    def test_auto_sum_down(self, auto_sum_excel):
        """测试向下求和"""
        cap = AutoSumCapability()
        result = cap.execute(None, file=str(auto_sum_excel), sheet="Data",
                           range="B2:E4", direction="down")
        
        assert result["success"] is True
        assert result["direction"] == "down"
        assert result["formulas_set"] == 3
        
        # 验证公式
        wb = load_workbook(auto_sum_excel)
        ws = wb["Data"]
        assert ws["B6"].value == "=SUM(B2:B5)"
        assert ws["C6"].value == "=SUM(C2:C5)"
        assert ws["D6"].value == "=SUM(D2:D5)"
        wb.close()
    
    def test_auto_sum_right(self, auto_sum_excel):
        """测试向右求和"""
        cap = AutoSumCapability()
        result = cap.execute(None, file=str(auto_sum_excel), sheet="Data",
                           range="B2:E4", direction="right")
        
        assert result["success"] is True
        assert result["direction"] == "right"
        assert result["formulas_set"] == 4
        
        # 验证公式
        wb = load_workbook(auto_sum_excel)
        ws = wb["Data"]
        assert ws["E2"].value == "=SUM(B2:D2)"
        assert ws["E3"].value == "=SUM(B3:D3)"
        assert ws["E4"].value == "=SUM(B4:D4)"
        assert ws["E5"].value == "=SUM(B5:D5)"
        wb.close()
    
    def test_auto_sum_default_direction(self, auto_sum_excel):
        """测试默认方向（down）"""
        cap = AutoSumCapability()
        result = cap.execute(None, file=str(auto_sum_excel), sheet="Data",
                           range="B2:E4")
        
        assert result["success"] is True
        assert result["direction"] == "down"
    
    def test_schema_has_required_params(self):
        """测试 schema 包含必要参数"""
        cap = AutoSumCapability()
        names = [s.name for s in cap.schema]
        
        assert "file" in names
        assert "sheet" in names
        assert "range" in names
        assert "direction" in names
    
    def test_capability_properties(self):
        """测试能力属性"""
        cap = AutoSumCapability()
        assert cap.name == "auto_sum"
        assert cap.chapter == "equation"
        assert "求和" in cap.description