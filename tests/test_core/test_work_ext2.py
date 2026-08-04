"""测试商功章扩展能力（批注、冻结、筛选、隐藏、分组、保护、缩放、打包、打印区域）"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.work import (
    CommentCapability,
    FreezePaneCapability,
    AutoFilterCapability,
    HideShowCapability,
    GroupRowsCapability,
    ProtectWorkbookCapability,
    ProtectSheetCapability,
    UnprotectSheetCapability,
    ZoomCapability,
    PrintAreaCapability,
    PackFileCapability,
    UnpackFileCapability,
)


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["A3"] = "Bob"
    ws["B3"] = 200
    wb.save(file_path)
    wb.close()
    return file_path


class TestComment:
    def test_add_comment(self, sample_excel):
        """添加批注"""
        cap = CommentCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           action="add", cell="A1", text="Test comment")
        assert result["action"] == "add"
        assert result["cell"] == "A1"
        assert result["text"] == "Test comment"

    def test_list_comments(self, sample_excel):
        """列出批注"""
        cap = CommentCapability()
        cap.execute(None, file=str(sample_excel), sheet="Data",
                   action="add", cell="A1", text="Test")

        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           action="list")
        assert result["action"] == "list"
        assert len(result["comments"]) >= 1

    def test_get_comment(self, sample_excel):
        """获取批注"""
        cap = CommentCapability()
        cap.execute(None, file=str(sample_excel), sheet="Data",
                   action="add", cell="A1", text="Hello", author="Tester")

        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           action="get", cell="A1")
        assert result["action"] == "get"
        assert result["text"] == "Hello"
        assert result["author"] == "Tester"

    def test_delete_comment(self, sample_excel):
        """删除批注"""
        cap = CommentCapability()
        cap.execute(None, file=str(sample_excel), sheet="Data",
                   action="add", cell="A1", text="To delete")

        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           action="delete", cell="A1")
        assert result["action"] == "delete"


class TestFreezePanes:
    def test_freeze_cell(self, sample_excel):
        """冻结窗格"""
        cap = FreezePaneCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data", cell="B2")
        assert result["action"] == "freeze"
        assert result["cell"] == "B2"

    def test_freeze_rows_columns(self, sample_excel):
        """按行列数冻结"""
        cap = FreezePaneCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           rows=1, columns=1)
        assert result["action"] == "freeze"

    def test_unfreeze(self, sample_excel):
        """解除冻结"""
        cap = FreezePaneCapability()
        cap.execute(None, file=str(sample_excel), sheet="Data", cell="B2")

        result = cap.execute(None, file=str(sample_excel), sheet="Data")
        assert result["action"] == "unfreeze"


class TestAutoFilter:
    def test_set_filter(self, sample_excel):
        """设置筛选"""
        cap = AutoFilterCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           action="set", range="A1:B3")
        assert result["action"] == "set"
        assert result["range"] == "A1:B3"

    def test_get_filter(self, sample_excel):
        """获取筛选"""
        cap = AutoFilterCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           action="get")
        assert result["action"] == "get"
        assert "has_filter" in result

    def test_remove_filter(self, sample_excel):
        """删除筛选"""
        cap = AutoFilterCapability()
        cap.execute(None, file=str(sample_excel), sheet="Data",
                   action="set", range="A1:B3")

        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           action="remove")
        assert result["action"] == "remove"


class TestHideShow:
    def test_hide_row(self, sample_excel):
        """隐藏行"""
        cap = HideShowCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           action="hide", dimension="row", index=2)
        assert result["action"] == "hide"
        assert result["applied"] is True

    def test_show_row(self, sample_excel):
        """显示行"""
        cap = HideShowCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           action="show", dimension="row", index=2)
        assert result["action"] == "show"
        assert result["applied"] is True

    def test_hide_column(self, sample_excel):
        """隐藏列"""
        cap = HideShowCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           action="hide", dimension="column", index=1)
        assert result["action"] == "hide"
        assert result["applied"] is True


class TestGroupRows:
    def test_group_rows(self, sample_excel):
        """分组行"""
        cap = GroupRowsCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           start_row=2, end_row=3, level=1)
        assert result["success"] is True
        assert result["rows_grouped"] == 2


class TestProtection:
    def test_protect_workbook(self, sample_excel):
        """保护工作簿"""
        cap = ProtectWorkbookCapability()
        result = cap.execute(None, file=str(sample_excel), password="test")
        assert result["success"] is True
        assert result["action"] == "protect_workbook"

    def test_protect_sheet(self, sample_excel):
        """保护工作表"""
        cap = ProtectSheetCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data", password="test")
        assert result["success"] is True
        assert result["action"] == "protect_sheet"

    def test_unprotect_sheet(self, sample_excel):
        """解除保护"""
        cap_protect = ProtectSheetCapability()
        cap_protect.execute(None, file=str(sample_excel), sheet="Data", password="test")

        cap_unprotect = UnprotectSheetCapability()
        result = cap_unprotect.execute(None, file=str(sample_excel), sheet="Data")
        assert result["success"] is True
        assert result["action"] == "unprotect_sheet"


class TestZoom:
    def test_set_zoom(self, sample_excel):
        """设置缩放"""
        cap = ZoomCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data", zoom=150)
        assert result["set"] is True
        assert result["zoom"] == 150


class TestPrintArea:
    def test_set_print_area(self, sample_excel):
        """设置打印区域"""
        cap = PrintAreaCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data", range="A1:B3")
        assert result["set"] is True
        assert result["print_area"] == "A1:B3"


class TestPackUnpack:
    def test_pack_file(self, sample_excel, tmp_path):
        """打包文件"""
        cap = PackFileCapability()
        output = tmp_path / "packed.zip"
        result = cap.execute(None, file=str(sample_excel), output=str(output))
        assert result["packed"] is True
        assert output.exists()

    def test_unpack_file(self, sample_excel, tmp_path):
        """解包文件"""
        pack_cap = PackFileCapability()
        zip_path = tmp_path / "packed.zip"
        pack_cap.execute(None, file=str(sample_excel), output=str(zip_path))

        unpack_cap = UnpackFileCapability()
        output_dir = tmp_path / "unpacked"
        result = unpack_cap.execute(None, file=str(zip_path), output=str(output_dir))
        assert result["unpacked"] is True
        assert output_dir.exists()
