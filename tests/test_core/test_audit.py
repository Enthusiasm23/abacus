"""测试审计工具能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.balance import ExcelLintCapability, FileAnalyzeCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["C1"] = "=SUM(B:B)"  # 整列引用
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def sample_python_code():
    """创建测试用 Python 代码"""
    return """
from openpyxl import load_workbook

wb = load_workbook("test.xlsx", data_only=True)
ws = wb.active
ws["A1"] = "=SUM(B1:B10)"
wb.save("test.xlsx")
"""


class TestExcelLint:
    def test_lint_data_only_save(self, sample_python_code):
        """检查 data_only=True 后 save()"""
        cap = ExcelLintCapability()
        result = cap.execute(None, code=sample_python_code)
        assert result["total_issues"] > 0
        assert any(i["rule"] == "XL001" for i in result["issues"])
    
    def test_lint_pattern_fill(self):
        """检查 PatternFill 缺 fill_type"""
        code = """
from openpyxl.styles import PatternFill
cell.fill = PatternFill(fgColor="FF0000")
"""
        cap = ExcelLintCapability()
        result = cap.execute(None, code=code)
        assert any(i["rule"] == "XL006" for i in result["issues"])


class TestFileAnalyze:
    def test_analyze_formula_issues(self, sample_excel):
        """检查公式问题"""
        cap = FileAnalyzeCapability()
        result = cap.execute(None, file=str(sample_excel))
        assert result["total_issues"] > 0
        # 应该有整列引用的警告
        assert any(i["rule"] == "XA009" for i in result["issues"])
