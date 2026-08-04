"""测试 validate_range.py 的完整覆盖"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.balance.validate_range import ValidateRangeCapability
from abacus.core.exceptions import DataError, FileNotFoundError


@pytest.fixture
def validate_range_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "validate_range_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # 添加表头
    ws.append(["ID", "Name", "Amount", "Status"])
    
    # 添加数据
    ws.append([1, "Alice", 100, "active"])
    ws.append([2, "Bob", 200, "inactive"])
    ws.append([3, "Charlie", 150, "active"])
    
    wb.save(file_path)
    wb.close()
    return file_path


class TestValidateRangeCoverage:
    """测试 validate_range.py 的完整覆盖"""
    
    def test_missing_file_raises_error(self):
        """测试缺少文件参数"""
        cap = ValidateRangeCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Data", range="A1:D4")
    
    def test_file_not_found_raises_error(self, tmp_path):
        """测试文件不存在"""
        cap = ValidateRangeCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "nonexistent.xlsx"), 
                       sheet="Data", range="A1:D4")
    
    def test_sheet_not_found_raises_error(self, validate_range_excel):
        """测试工作表不存在"""
        cap = ValidateRangeCapability()
        with pytest.raises(DataError, match="not found"):
            cap.execute(None, file=str(validate_range_excel), 
                       sheet="NoSuchSheet", range="A1:D4")
    
    def test_valid_range_no_issues(self, validate_range_excel):
        """测试有效范围无问题"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="A1:D4")
        
        assert result["valid"] is True
        assert result["issues_count"] == 0
        assert result["total_cells"] == 16  # 4x4 = 16 个单元格
    
    def test_check_empty_cells(self, validate_range_excel):
        """测试检查空单元格"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="A1:D4",
                           rules={"no_empty": True})
        
        assert result["valid"] is True
        assert result["empty_cells"] == 0
    
    def test_check_empty_cells_with_empty(self, tmp_path):
        """测试检查空单元格 - 有空值"""
        file_path = tmp_path / "empty_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, None])
        ws.append([2, 3])
        wb.save(file_path)
        wb.close()
        
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(file_path), 
                           sheet="Sheet", range="A1:B3",
                           rules={"no_empty": True})
        
        assert result["valid"] is False
        assert result["empty_cells"] == 1
        assert len(result["issues"]) == 1
        assert result["issues"][0]["type"] == "empty"
    
    def test_check_type_number(self, validate_range_excel):
        """测试检查数字类型"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="A1:A4",
                           rules={"expected_type": "number"})
        
        # A1 是表头 "ID"，不是数字，所以应该有类型不匹配
        assert result["valid"] is False
        assert any(i["type"] == "type_mismatch" for i in result["issues"])
    
    def test_check_type_number_mismatch(self, validate_range_excel):
        """测试检查数字类型 - 类型不匹配"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="B1:B4",
                           rules={"expected_type": "number"})
        
        assert result["valid"] is False
        assert any(i["type"] == "type_mismatch" for i in result["issues"])
    
    def test_check_type_text(self, validate_range_excel):
        """测试检查文本类型"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="B1:B4",
                           rules={"expected_type": "text"})
        
        assert result["valid"] is True
    
    def test_check_type_text_mismatch(self, validate_range_excel):
        """测试检查文本类型 - 类型不匹配"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="A1:A4",
                           rules={"expected_type": "text"})
        
        assert result["valid"] is False
        assert any(i["type"] == "type_mismatch" for i in result["issues"])
    
    def test_check_min_value(self, validate_range_excel):
        """测试检查最小值"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="C1:C4",
                           rules={"min_value": 100})
        
        assert result["valid"] is True
    
    def test_check_min_value_below(self, tmp_path):
        """测试检查最小值 - 低于最小值"""
        file_path = tmp_path / "min_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Value"])
        ws.append([50])
        ws.append([150])
        wb.save(file_path)
        wb.close()
        
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(file_path), 
                           sheet="Sheet", range="A1:A3",
                           rules={"min_value": 100})
        
        assert result["valid"] is False
        assert any(i["type"] == "below_min" for i in result["issues"])
    
    def test_check_max_value(self, validate_range_excel):
        """测试检查最大值"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="C1:C4",
                           rules={"max_value": 300})
        
        assert result["valid"] is True
    
    def test_check_max_value_above(self, tmp_path):
        """测试检查最大值 - 超过最大值"""
        file_path = tmp_path / "max_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Value"])
        ws.append([50])
        ws.append([150])
        wb.save(file_path)
        wb.close()
        
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(file_path), 
                           sheet="Sheet", range="A1:A3",
                           rules={"max_value": 100})
        
        assert result["valid"] is False
        assert any(i["type"] == "above_max" for i in result["issues"])
    
    def test_range_without_end_row(self, validate_range_excel):
        """测试范围不指定结束行"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="A1:D")
        
        assert result["valid"] is True
        assert result["total_cells"] == 16  # 4x4 = 16 个单元格
    
    def test_range_without_end_col(self, validate_range_excel):
        """测试范围不指定结束列"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="A1:4")
        
        assert result["valid"] is True
    
    def test_multiple_rules(self, validate_range_excel):
        """测试多个规则"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(validate_range_excel), 
                           sheet="Data", range="A1:D4",
                           rules={
                               "no_empty": True,
                               "expected_type": "number",
                               "min_value": 0,
                               "max_value": 1000
                           })
        
        # 表头行不是数字，所以会有类型不匹配
        assert result["valid"] is False
        assert any(i["type"] == "type_mismatch" for i in result["issues"])
    
    def test_schema_has_required_params(self):
        """测试 schema 包含必要参数"""
        cap = ValidateRangeCapability()
        names = [s.name for s in cap.schema]
        
        assert "file" in names
        assert "sheet" in names
        assert "range" in names
        assert "rules" in names
    
    def test_capability_properties(self):
        """测试能力属性"""
        cap = ValidateRangeCapability()
        assert cap.name == "validate_range"
        assert cap.chapter == "balance"
        assert "验证" in cap.description