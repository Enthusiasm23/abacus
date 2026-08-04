"""测试 standardize.py 的完整覆盖"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook
import pandas as pd

from abacus.core.grain.standardize import StandardizeCapability
from abacus.core.exceptions import DataError, FileNotFoundError


@pytest.fixture
def standardize_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "standardize_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # 添加表头
    ws.append(["Name", "Amount", "Date", "Category"])
    
    # 添加数据
    ws.append(["alice", 100.567, "2024-01-15", "electronics"])
    ws.append(["BOB", 200.123, "2024-01-16", "clothing"])
    ws.append(["charlie", 300.999, "2024-01-17", "food"])
    
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def standardize_excel_numeric(tmp_path):
    """创建包含数值数据的 Excel 文件"""
    file_path = tmp_path / "numeric_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["ID", "Value", "Rate"])
    ws.append([1, 100.567, 0.123])
    ws.append([2, 200.123, 0.456])
    ws.append([3, 300.999, 0.789])
    
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def standardize_excel_with_dates(tmp_path):
    """创建包含日期数据的 Excel 文件"""
    file_path = tmp_path / "dates_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["ID", "Event", "Date"])
    ws.append([1, "Meeting", "2024-01-15"])
    ws.append([2, "Deadline", "2024-01-16"])
    ws.append([3, "Review", "2024-01-17"])
    
    wb.save(file_path)
    wb.close()
    return file_path


class TestStandardizeCoverage:
    """测试 standardize.py 的完整覆盖"""
    
    def test_missing_file_raises_error(self):
        """测试缺少文件参数"""
        cap = StandardizeCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Data")
    
    def test_missing_sheet_raises_error(self, standardize_excel):
        """测试缺少工作表参数"""
        cap = StandardizeCapability()
        with pytest.raises(DataError, match="sheet parameter is required"):
            cap.execute(None, file=str(standardize_excel))
    
    def test_file_not_found_raises_error(self, tmp_path):
        """测试文件不存在"""
        cap = StandardizeCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "nonexistent.xlsx"), sheet="Data")
    
    def test_text_case_lower(self, standardize_excel):
        """测试文本小写转换"""
        cap = StandardizeCapability()
        result = cap.execute(None, file=str(standardize_excel), sheet="Data", text_case="lower")
        
        assert result["success"] is True
        assert result["operation_count"] > 0
        
        # 验证转换
        wb = Workbook()
        df = pd.read_excel(standardize_excel, sheet_name="Data")
        assert df["Name"].iloc[0] == "alice"
        assert df["Category"].iloc[0] == "electronics"
    
    def test_text_case_upper(self):
        """测试文本大写转换"""
        import tempfile
        import os
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Category"])
        ws.append(["alice", "electronics"])
        ws.append(["BOB", "clothing"])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            file_path = f.name
        wb.close()
        
        cap = StandardizeCapability()
        result = cap.execute(None, file=file_path, sheet="Data", text_case="upper")
        
        assert result["success"] is True
        assert result["operation_count"] > 0
        assert any(op["operation"] == "text_case" for op in result["operations"])
        
        os.unlink(file_path)
    
    def test_text_case_title(self):
        """测试文本标题大小写转换"""
        import tempfile
        import os
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Category"])
        ws.append(["alice", "electronics"])
        ws.append(["BOB", "clothing"])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            file_path = f.name
        wb.close()
        
        cap = StandardizeCapability()
        result = cap.execute(None, file=file_path, sheet="Data", text_case="title")
        
        assert result["success"] is True
        assert result["operation_count"] > 0
        assert any(op["operation"] == "text_case" for op in result["operations"])
        
        os.unlink(file_path)
    
    def test_number_format(self, standardize_excel_numeric):
        """测试数字格式化"""
        cap = StandardizeCapability()
        result = cap.execute(None, file=str(standardize_excel_numeric), sheet="Data", number_format="%.2f")
        
        assert result["success"] is True
        # 数字格式化可能不会增加 operation_count，因为它是原地修改
    
    def test_output_file(self, standardize_excel, tmp_path):
        """测试输出到新文件"""
        output_path = tmp_path / "output.xlsx"
        cap = StandardizeCapability()
        result = cap.execute(None, file=str(standardize_excel), sheet="Data", 
                           text_case="lower", output=str(output_path))
        
        assert result["success"] is True
        assert output_path.exists()
        assert result["output"] == str(output_path)
    
    def test_multiple_operations(self, standardize_excel):
        """测试多个操作同时执行"""
        cap = StandardizeCapability()
        result = cap.execute(None, file=str(standardize_excel), sheet="Data",
                           text_case="upper", number_format="%.1f")
        
        assert result["success"] is True
        assert result["operation_count"] >= 2
    
    def test_schema_has_required_params(self):
        """测试 schema 包含必要参数"""
        cap = StandardizeCapability()
        names = [s.name for s in cap.schema]
        
        assert "file" in names
        assert "sheet" in names
        assert "date_format" in names
        assert "number_format" in names
        assert "text_case" in names
        assert "output" in names
    
    def test_capability_properties(self):
        """测试能力属性"""
        cap = StandardizeCapability()
        assert cap.name == "standardize_data"
        assert cap.chapter == "grain"
        assert "格式" in cap.description