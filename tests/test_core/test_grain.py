"""测试粟米章（格式转换）能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.grain import ConvertFormatCapability, ConvertUnitCapability, ConvertTypeCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["C1"] = "Date"
    ws["A2"] = "Alice"
    ws["B2"] = 100.5
    ws["C2"] = "2024-01-15"
    ws["A3"] = "Bob"
    ws["B3"] = 200
    ws["C3"] = "2024-02-20"
    wb.save(file_path)
    wb.close()
    return file_path


class TestConvertFormat:
    def test_convert_to_number(self, sample_excel):
        """转换为数字格式"""
        cap = ConvertFormatCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B2:B3", format_type="number")
        assert result["cells_formatted"] == 2
    
    def test_convert_to_date(self, sample_excel):
        """转换为日期格式"""
        cap = ConvertFormatCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="C2:C3", format_type="date")
        assert result["cells_formatted"] == 2
    
    def test_convert_to_percentage(self, sample_excel):
        """转换为百分比格式"""
        cap = ConvertFormatCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B2:B3", format_type="percentage")
        assert result["cells_formatted"] == 2


class TestConvertUnit:
    def test_km_to_m(self, sample_excel):
        """千米转米"""
        cap = ConvertUnitCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B2:B3", from_unit="km", to_unit="m")
        assert result["cells_converted"] == 2
    
    def test_kg_to_g(self, sample_excel):
        """千克转克"""
        cap = ConvertUnitCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B2:B3", from_unit="kg", to_unit="g")
        assert result["cells_converted"] == 2


class TestConvertType:
    def test_convert_to_text(self, sample_excel):
        """转换为文本"""
        cap = ConvertTypeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B2:B3", target_type="text")
        assert result["cells_converted"] == 2
