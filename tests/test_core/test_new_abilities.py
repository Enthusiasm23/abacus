"""测试新能力：transform_pipeline, advanced_filter, create_mapping_template"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.grain.transform_pipeline import TransformPipelineCapability
from abacus.core.work.advanced_filter import AdvancedFilterCapability
from abacus.core.work.mapping_template import CreateMappingTemplateCapability
from abacus.core.exceptions import DataError, FileNotFoundError


# ── Fixtures ──────────────────────────────────────────────────────────

@pytest.fixture
def pipeline_excel(tmp_path):
    """创建测试用 Excel 文件（用于管道测试）"""
    file_path = tmp_path / "pipeline_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["C1"] = "Status"
    ws["A2"] = "Alice"
    ws["B2"] = "100"
    ws["C2"] = "active"
    ws["A3"] = "Bob"
    ws["B3"] = "200"
    ws["C3"] = "inactive"
    ws["A4"] = "Charlie"
    ws["B4"] = None
    ws["C4"] = "active"
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def filter_excel(tmp_path):
    """创建测试用 Excel 文件（用于筛选测试）"""
    file_path = tmp_path / "filter_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Name"
    ws["B1"] = "Region"
    ws["C1"] = "Amount"
    ws["D1"] = "Status"
    ws["A2"] = "Alice"
    ws["B2"] = "East"
    ws["C2"] = 1000
    ws["D2"] = "active"
    ws["A3"] = "Bob"
    ws["B3"] = "West"
    ws["C3"] = 2500
    ws["D3"] = "active"
    ws["A4"] = "Charlie"
    ws["B4"] = "East"
    ws["C4"] = 800
    ws["D4"] = "inactive"
    ws["A5"] = "Diana"
    ws["B5"] = "West"
    ws["C5"] = 3000
    ws["D5"] = "active"
    ws["A6"] = "Eve"
    ws["B6"] = "East"
    ws["C6"] = 1500
    ws["D6"] = "active"
    wb.save(file_path)
    wb.close()
    return file_path


# ── TransformPipeline Tests ───────────────────────────────────────────

class TestTransformPipeline:
    def test_capability_properties(self):
        cap = TransformPipelineCapability()
        assert cap.name == "transform_pipeline"
        assert cap.chapter == "grain"
        assert "转换管道" in cap.description

    def test_schema_has_required_params(self):
        cap = TransformPipelineCapability()
        names = [s.name for s in cap.schema]
        assert "file" in names
        assert "steps" in names

    def test_missing_file_raises_error(self):
        cap = TransformPipelineCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, steps=[{"type": "fill_value"}])

    def test_missing_steps_raises_error(self):
        cap = TransformPipelineCapability()
        with pytest.raises(DataError, match="steps parameter is required"):
            cap.execute(None, file="dummy.xlsx")

    def test_file_not_found_raises_error(self, tmp_path):
        cap = TransformPipelineCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "nonexistent.xlsx"),
                       steps=[{"type": "fill_value"}])

    def test_single_fill_value_step(self, pipeline_excel):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_excel), sheet="Data",
                           steps=[{"type": "fill_value", "range": "B4:B4", "value": 0}])
        assert result["steps_executed"] == 1
        assert result["steps_succeeded"] == 1

        from openpyxl import load_workbook
        wb = load_workbook(pipeline_excel)
        assert wb["Data"]["B4"].value == 0
        wb.close()

    def test_single_replace_value_step(self, pipeline_excel):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_excel), sheet="Data",
                           steps=[{"type": "replace_value", "range": "C2:C4",
                                   "old_value": "inactive", "new_value": "retired"}])
        assert result["steps_succeeded"] == 1

        from openpyxl import load_workbook
        wb = load_workbook(pipeline_excel)
        assert wb["Data"]["C3"].value == "retired"
        assert wb["Data"]["C2"].value == "active"
        wb.close()

    def test_convert_type_step(self, pipeline_excel):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_excel), sheet="Data",
                           steps=[{"type": "convert_type", "range": "B2:B3",
                                   "target_type": "int"}])
        assert result["steps_succeeded"] == 1

        from openpyxl import load_workbook
        wb = load_workbook(pipeline_excel)
        assert wb["Data"]["B2"].value == 100
        assert isinstance(wb["Data"]["B2"].value, int)
        wb.close()

    def test_convert_format_step(self, pipeline_excel):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_excel), sheet="Data",
                           steps=[{"type": "convert_format", "range": "B2:B3",
                                   "format_type": "number"}])
        assert result["steps_succeeded"] == 1

        from openpyxl import load_workbook
        wb = load_workbook(pipeline_excel)
        assert wb["Data"]["B2"].number_format == "#,##0.00"
        wb.close()

    def test_chained_steps(self, pipeline_excel):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_excel), sheet="Data", steps=[
            {"type": "fill_value", "range": "B4:B4", "value": 300},
            {"type": "convert_type", "range": "B2:B4", "target_type": "float"},
            {"type": "replace_value", "range": "C2:C4",
             "old_value": "inactive", "new_value": "retired"},
        ])
        assert result["steps_executed"] == 3
        assert result["steps_succeeded"] == 3

    def test_stop_on_error_true(self, pipeline_excel):
        cap = TransformPipelineCapability()
        with pytest.raises(DataError, match="Pipeline failed at step 2"):
            cap.execute(None, file=str(pipeline_excel), sheet="Data",
                       stop_on_error=True, steps=[
                {"type": "fill_value", "range": "B4:B4", "value": 0},
                {"type": "fill_value"},  # missing range -> error
            ])

    def test_stop_on_error_false(self, pipeline_excel):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_excel), sheet="Data",
                           stop_on_error=False, steps=[
            {"type": "fill_value", "range": "B4:B4", "value": 0},
            {"type": "fill_value"},  # missing range -> error
        ])
        assert result["steps_executed"] == 2
        assert result["results"][0]["status"] == "success"
        assert result["results"][1]["status"] == "failed"

    def test_unknown_step_type_raises_error(self, pipeline_excel):
        cap = TransformPipelineCapability()
        with pytest.raises(DataError, match="Unknown step type"):
            cap.execute(None, file=str(pipeline_excel), sheet="Data",
                       steps=[{"type": "nonexistent"}])

    def test_fill_value_skips_non_empty(self, pipeline_excel):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_excel), sheet="Data",
                           steps=[{"type": "fill_value", "range": "B2:B4", "value": 999}])
        filled = result["results"][0]["result"]["filled"]
        assert filled == 1  # only B4 was empty

    def test_replace_value_no_match(self, pipeline_excel):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_excel), sheet="Data",
                           steps=[{"type": "replace_value", "range": "C2:C4",
                                   "old_value": "nonexistent", "new_value": "x"}])
        replaced = result["results"][0]["result"]["replaced"]
        assert replaced == 0


# ── AdvancedFilter Tests ──────────────────────────────────────────────

class TestAdvancedFilter:
    def test_capability_properties(self):
        cap = AdvancedFilterCapability()
        assert cap.name == "advanced_filter"
        assert cap.chapter == "work"
        assert "筛选" in cap.description

    def test_missing_file_raises_error(self):
        cap = AdvancedFilterCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Sheet1", conditions={"type": "condition"})

    def test_missing_sheet_raises_error(self):
        cap = AdvancedFilterCapability()
        with pytest.raises(DataError, match="sheet parameter is required"):
            cap.execute(None, file="dummy.xlsx", conditions={"type": "condition"})

    def test_missing_conditions_raises_error(self):
        cap = AdvancedFilterCapability()
        with pytest.raises(DataError, match="conditions parameter is required"):
            cap.execute(None, file="dummy.xlsx", sheet="Sheet1")

    def test_file_not_found_raises_error(self, tmp_path):
        cap = AdvancedFilterCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "non.xlsx"), sheet="Sheet1",
                       conditions={"type": "condition"})

    def test_sheet_not_found_raises_error(self, filter_excel):
        cap = AdvancedFilterCapability()
        with pytest.raises(DataError, match="not found"):
            cap.execute(None, file=str(filter_excel), sheet="NoSuchSheet",
                       conditions={"type": "condition", "field": "Name",
                                   "operator": "==", "value": "Alice"})

    # ── 比较运算符 ──

    def test_equal_operator(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Name",
                                       "operator": "==", "value": "Bob"})
        assert result["total_matched"] == 1
        assert result["rows"][0]["Name"] == "Bob"

    def test_not_equal_operator(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Region",
                                       "operator": "!=", "value": "East"})
        assert result["total_matched"] == 2

    def test_greater_than_operator(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Amount",
                                       "operator": ">", "value": 2000})
        assert result["total_matched"] == 2

    def test_less_than_operator(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Amount",
                                       "operator": "<", "value": 1000})
        assert result["total_matched"] == 1

    def test_gte_operator(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Amount",
                                       "operator": ">=", "value": 3000})
        assert result["total_matched"] == 1

    def test_lte_operator(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Amount",
                                       "operator": "<=", "value": 800})
        assert result["total_matched"] == 1

    def test_between_operator(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Amount",
                                       "operator": "between", "value": [1000, 2000]})
        assert result["total_matched"] == 2

    # ── 文本筛选 ──

    def test_contains_operator(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Name",
                                       "operator": "contains", "value": "li"})
        assert result["total_matched"] == 2  # Alice + Charlie both contain "li"

    def test_starts_with_operator(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Name",
                                       "operator": "starts_with", "value": "D"})
        assert result["total_matched"] == 1
        assert result["rows"][0]["Name"] == "Diana"

    def test_ends_with_operator(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Name",
                                       "operator": "ends_with", "value": "a"})
        assert result["total_matched"] == 1  # only Diana ends with "a"

    # ── 逻辑组合 ──

    def test_and_logic(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "group", "logic": "AND", "conditions": [
                               {"type": "condition", "field": "Region", "operator": "==", "value": "East"},
                               {"type": "condition", "field": "Amount", "operator": ">", "value": 1000},
                           ]})
        assert result["total_matched"] == 1
        assert result["rows"][0]["Name"] == "Eve"

    def test_or_logic(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "group", "logic": "OR", "conditions": [
                               {"type": "condition", "field": "Amount", "operator": ">", "value": 2500},
                               {"type": "condition", "field": "Name", "operator": "==", "value": "Alice"},
                           ]})
        assert result["total_matched"] == 2  # Alice(1000? no), Bob(2500? no), Diana(3000 yes), Eve(1500 no)
        # Actually: Diana(3000>2500), Alice(name=Alice) => 2

    def test_not_logic(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "group", "logic": "NOT", "conditions": [
                               {"type": "condition", "field": "Region", "operator": "==", "value": "East"},
                           ]})
        assert result["total_matched"] == 2  # Bob, Diana (West)

    def test_nested_and_or(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "group", "logic": "AND", "conditions": [
                               {"type": "condition", "field": "Region", "operator": "==", "value": "East"},
                               {"type": "group", "logic": "OR", "conditions": [
                                   {"type": "condition", "field": "Amount", "operator": ">", "value": 1200},
                                   {"type": "condition", "field": "Name", "operator": "==", "value": "Alice"},
                               ]},
                           ]})
        assert result["total_matched"] == 2  # Alice + Eve

    # ── 返回行号 ──

    def test_return_type_rows(self, filter_excel):
        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(filter_excel), sheet="Sales",
                           range="A1:D6",
                           conditions={"type": "condition", "field": "Amount",
                                       "operator": ">", "value": 2000},
                           return_type="rows")
        assert "row_numbers" in result
        assert len(result["row_numbers"]) == 2

    def test_return_type_invalid(self, filter_excel):
        cap = AdvancedFilterCapability()
        with pytest.raises(DataError, match="return_type"):
            cap.execute(None, file=str(filter_excel), sheet="Sales",
                       conditions={"type": "condition", "field": "Amount",
                                   "operator": ">", "value": 0},
                       return_type="invalid")

    # ── 空值处理 ──

    def test_null_value_equal(self, tmp_path):
        file_path = tmp_path / "null_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "ID"
        ws["B1"] = "Value"
        ws["A2"] = 1
        ws["B2"] = None
        ws["A3"] = 2
        ws["B3"] = "hello"
        wb.save(file_path)
        wb.close()

        cap = AdvancedFilterCapability()
        result = cap.execute(None, file=str(file_path), sheet="Sheet",
                           conditions={"type": "condition", "field": "Value",
                                       "operator": "==", "value": None})
        assert result["total_matched"] == 1

    def test_field_not_found_raises_error(self, filter_excel):
        cap = AdvancedFilterCapability()
        with pytest.raises(DataError, match="not found in data"):
            cap.execute(None, file=str(filter_excel), sheet="Sales",
                       conditions={"type": "condition", "field": "NoSuchField",
                                   "operator": "==", "value": "x"})


# ── CreateMappingTemplate Tests ───────────────────────────────────────

class TestCreateMappingTemplate:
    def test_capability_properties(self):
        cap = CreateMappingTemplateCapability()
        assert cap.name == "create_mapping_template"
        assert cap.chapter == "work"
        assert "映射模板" in cap.description

    def test_create_default_template(self, tmp_path):
        cap = CreateMappingTemplateCapability()
        output = tmp_path / "template.xlsx"
        result = cap.execute(None, output=str(output))
        assert output.exists()
        assert "output" in result
        assert result["source_count"] == 4
        assert len(result["sheets"]) == 3

    def test_create_with_custom_source_count(self, tmp_path):
        cap = CreateMappingTemplateCapability()
        output = tmp_path / "template_2src.xlsx"
        result = cap.execute(None, output=str(output), source_count=2)
        assert output.exists()
        assert result["source_count"] == 2

    def test_create_with_many_sources(self, tmp_path):
        cap = CreateMappingTemplateCapability()
        output = tmp_path / "template_8src.xlsx"
        result = cap.execute(None, output=str(output), source_count=8)
        assert output.exists()
        assert result["source_count"] == 8

    def test_sheets_created(self, tmp_path):
        cap = CreateMappingTemplateCapability()
        output = tmp_path / "template_sheets.xlsx"
        result = cap.execute(None, output=str(output))
        assert "目标表" in result["sheets"] or len(result["sheets"]) >= 2

    def test_quiet_mode(self, tmp_path, caplog):
        cap = CreateMappingTemplateCapability()
        output = tmp_path / "quiet.xlsx"
        result = cap.execute(None, output=str(output), quiet=True)
        assert output.exists()
        assert "Created template" not in caplog.text

    def test_output_auto_generated_when_none(self, tmp_path, monkeypatch):
        cap = CreateMappingTemplateCapability()
        monkeypatch.chdir(tmp_path)
        result = cap.execute(None)
        assert Path(result["output"]).exists()

    def test_template_contains_placeholder(self, tmp_path):
        cap = CreateMappingTemplateCapability()
        output = tmp_path / "check.xlsx"
        cap.execute(None, output=str(output))

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb.active
        # Check that placeholders exist
        found_placeholder = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "{" in cell.value:
                    found_placeholder = True
                    break
            if found_placeholder:
                break
        assert found_placeholder, "Template should contain {placeholder} values"
        wb.close()

    def test_instructions_sheet_created(self, tmp_path):
        cap = CreateMappingTemplateCapability()
        output = tmp_path / "instructions.xlsx"
        cap.execute(None, output=str(output))

        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert "填写说明" in wb.sheetnames
        wb.close()

    def test_example_sheet_created(self, tmp_path):
        cap = CreateMappingTemplateCapability()
        output = tmp_path / "example.xlsx"
        cap.execute(None, output=str(output))

        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert "填写示例" in wb.sheetnames
        wb.close()
