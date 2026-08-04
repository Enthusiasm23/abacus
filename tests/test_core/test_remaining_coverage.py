"""测试更多未覆盖的核心模块"""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
import tempfile
import os

from abacus.core.grain.convert_format import ConvertFormatCapability
from abacus.core.grain.convert_type import ConvertTypeCapability
from abacus.core.grain.text_to_columns import TextToColumnsCapability
from abacus.core.grain.fuzzy_match import FuzzyMatchCapability
from abacus.core.grain.transpose import TransposeCapability
from abacus.core.work.comment import CommentCapability
from abacus.core.work.batch_execute import BatchExecuteCapability
from abacus.core.work.batch_transform import BatchTransformCapability
from abacus.core.named_range.named_range import NamedRangeCapability
from abacus.core.triangle.analyze_stats import AnalyzeStatsCapability
from abacus.core.balance.analyze import FileAnalyzeCapability
from abacus.core.exceptions import DataError, FileNotFoundError, ValidationError


@pytest.fixture
def data_excel(tmp_path):
    file_path = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Name", "Sales", "Date"])
    ws.append([1, "Alice", 100, "2024-01-15"])
    ws.append([2, "Bob", 200, "2024-01-16"])
    ws.append([3, "Charlie", 150, "2024-01-17"])
    wb.save(file_path)
    wb.close()
    return file_path


class TestConvertFormat:
    def test_missing_file(self):
        cap = ConvertFormatCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Data", range="A1:A3", format_type="number")

    def test_file_not_found(self, tmp_path):
        cap = ConvertFormatCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"), sheet="Data",
                       range="A1:A3", format_type="number")

    def test_convert_number_format(self, data_excel):
        cap = ConvertFormatCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           range="C2:C4", format_type="number")
        assert result["cells_formatted"] == 3

    def test_convert_date_format(self, data_excel):
        cap = ConvertFormatCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           range="D2:D4", format_type="date")
        assert result["cells_formatted"] == 3

    def test_convert_currency_format(self, data_excel):
        cap = ConvertFormatCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           range="C2:C4", format_type="currency")
        assert result["cells_formatted"] == 3

    def test_convert_percentage_format(self, data_excel):
        cap = ConvertFormatCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           range="C2:C4", format_type="percentage")
        assert result["cells_formatted"] == 3

    def test_convert_text_format(self, data_excel):
        cap = ConvertFormatCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           range="C2:C4", format_type="text")
        assert result["cells_formatted"] == 3


class TestConvertType:
    def test_missing_file(self):
        cap = ConvertTypeCapability()
        with pytest.raises(DataError, match="参数 file 是必需的"):
            cap.execute(None, sheet="Data", range="A1:A3", target_type="float")

    def test_convert_to_float(self, data_excel):
        cap = ConvertTypeCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           range="C2:C4", target_type="float")
        assert "cells_converted" in result

    def test_convert_to_str(self, data_excel):
        cap = ConvertTypeCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           range="C2:C4", target_type="str")
        assert "cells_converted" in result

    def test_convert_to_int(self, data_excel):
        cap = ConvertTypeCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           range="C2:C4", target_type="int")
        assert "cells_converted" in result

    def test_convert_text_to_float(self, data_excel):
        cap = ConvertTypeCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           range="B2:B4", target_type="float")
        assert "cells_converted" in result


class TestTextToColumns:
    def test_missing_file(self):
        cap = TextToColumnsCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Data", column="A")

    def test_file_not_found(self, tmp_path):
        cap = TextToColumnsCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"), sheet="Data", column="A")

    def test_text_to_columns_comma(self, tmp_path):
        file_path = tmp_path / "split.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Name,Value,Region"
        ws["A2"] = "Alice,100,North"
        wb.save(file_path)
        wb.close()

        cap = TextToColumnsCapability()
        result = cap.execute(None, file=str(file_path), sheet="Sheet",
                           column="A", delimiter=",")
        assert result["success"] is True


class TestFuzzyMatch:
    def test_missing_file(self):
        cap = FuzzyMatchCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Data", target_columns=["Name"])

    def test_fuzzy_match(self, data_excel):
        cap = FuzzyMatchCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           target_columns=["Name", "Sales"])
        assert result["success"] is True


class TestTranspose:
    def test_missing_file(self):
        cap = TransposeCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Data", range="A1:D4")

    def test_transpose(self, data_excel):
        cap = TransposeCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data", range="A1:D4")
        assert result["success"] is True


class TestComment:
    def test_missing_file(self):
        cap = CommentCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Data", action="add", cell="A1", text="test")

    def test_add_comment(self, data_excel):
        cap = CommentCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           action="add", cell="A1", text="Test comment")
        assert result["action"] == "add"

    def test_get_comment(self, data_excel):
        cap = CommentCapability()
        cap.execute(None, file=str(data_excel), sheet="Data",
                   action="add", cell="A1", text="Test")
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           action="get", cell="A1")
        assert result["action"] == "get"

    def test_list_comments(self, data_excel):
        cap = CommentCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data", action="list")
        assert result["action"] == "list"

    def test_delete_comment(self, data_excel):
        cap = CommentCapability()
        cap.execute(None, file=str(data_excel), sheet="Data",
                   action="add", cell="A1", text="Test")
        result = cap.execute(None, file=str(data_excel), sheet="Data",
                           action="delete", cell="A1")
        assert result["action"] == "delete"


class TestBatchExecute:
    def test_missing_file(self):
        cap = BatchExecuteCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, operations=[])

    def test_batch_execute(self, data_excel):
        cap = BatchExecuteCapability()
        result = cap.execute(None, file=str(data_excel), operations=[
            {"type": "write", "sheet": "Data", "cell": "E1", "value": "Test"}
        ])
        assert result["executed"] == 1

    def test_batch_execute_empty(self, data_excel):
        cap = BatchExecuteCapability()
        result = cap.execute(None, file=str(data_excel), operations=[])
        assert result["total"] == 0


class TestBatchTransform:
    def test_missing_file(self):
        cap = BatchTransformCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, operations=[])

    def test_batch_transform(self, data_excel):
        cap = BatchTransformCapability()
        result = cap.execute(None, file=str(data_excel), operations=[
            {"type": "convert_type", "range": "C2:C4", "target_type": "float"}
        ])
        assert "operations" in result


class TestNamedRange:
    def test_missing_file(self):
        cap = NamedRangeCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, action="list")

    def test_list_named_ranges(self, data_excel):
        cap = NamedRangeCapability()
        result = cap.execute(None, file=str(data_excel), action="list")
        assert "named_ranges" in result

    def test_create_named_range(self, data_excel):
        cap = NamedRangeCapability()
        result = cap.execute(None, file=str(data_excel), action="create",
                           name="TestData", refers_to="Data!$A$1:$D$4")
        assert result["action"] == "create"

    def test_read_named_range(self, data_excel):
        cap = NamedRangeCapability()
        cap.execute(None, file=str(data_excel), action="create",
                   name="TestData", refers_to="Data!$A$1:$D$4")
        result = cap.execute(None, file=str(data_excel), action="read", name="TestData")
        assert result["action"] == "read"

    def test_delete_named_range(self, data_excel):
        cap = NamedRangeCapability()
        cap.execute(None, file=str(data_excel), action="create",
                   name="ToDelete", refers_to="Data!$A$1:$D$4")
        result = cap.execute(None, file=str(data_excel), action="delete", name="ToDelete")
        assert result["action"] == "delete"


class TestAnalyzeStats:
    def test_missing_file(self):
        cap = AnalyzeStatsCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 file"):
            cap.execute(None, sheet="Data", range="C2:C4")

    def test_analyze_stats(self, data_excel):
        cap = AnalyzeStatsCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data", range="C2:C4")
        assert "statistics" in result


class TestFileAnalyze:
    def test_missing_file(self):
        cap = FileAnalyzeCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 file"):
            cap.execute(None, sheet="Data", range="C2:C4")

    def test_file_analyze(self, data_excel):
        cap = FileAnalyzeCapability()
        result = cap.execute(None, file=str(data_excel), sheet="Data", range="C2:C4")
        assert "total_issues" in result
