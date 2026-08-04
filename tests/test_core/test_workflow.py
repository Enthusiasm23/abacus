"""测试工作流能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.workflow import SpreadsheetWorkflowCapability, FormattingWorkflowCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    wb.save(file_path)
    wb.close()
    return file_path


class TestSpreadsheetWorkflow:
    def test_read_workflow(self, sample_excel):
        """读取工作流"""
        cap = SpreadsheetWorkflowCapability()
        result = cap.execute(None, action="read", file=str(sample_excel), sheet="Data")
        assert result["action"] == "read"
        assert result["rows"] >= 2
        assert result["columns"] >= 2

    def test_create_workflow(self, tmp_path):
        """创建工作流"""
        cap = SpreadsheetWorkflowCapability()
        output = tmp_path / "created.xlsx"
        result = cap.execute(None, action="create", output=str(output),
                           sheet="TestSheet", data={"rows": [["A", 1], ["B", 2]]})
        assert result["action"] == "create"
        assert result["rows"] == 2
        assert output.exists()

    def test_edit_workflow(self, sample_excel):
        """编辑工作流"""
        cap = SpreadsheetWorkflowCapability()
        result = cap.execute(None, action="edit", file=str(sample_excel), sheet="Data",
                           data={"rows": [["Charlie", 300]]})
        assert result["action"] == "edit"
        assert result["rows_written"] == 1

    def test_format_workflow(self, sample_excel):
        """格式化工作流"""
        cap = SpreadsheetWorkflowCapability()
        result = cap.execute(None, action="format", file=str(sample_excel), sheet="Data",
                           range="A1:B1", bold=True)
        assert result["action"] == "format"
        assert result["range"] == "A1:B1"


class TestFormattingWorkflow:
    def test_format_cells(self, sample_excel):
        """格式化单元格"""
        cap = FormattingWorkflowCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B1", font={"bold": True})
        assert result["cells_formatted"] == 2

    def test_format_with_fill(self, sample_excel):
        """填充格式化"""
        cap = FormattingWorkflowCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B1", fill={"color": "4472C4"})
        assert result["cells_formatted"] == 2
