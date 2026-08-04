"""测试数据可视化能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.triangle.visualize import VisualizeCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["A3"] = "Bob"
    ws["B3"] = 200
    ws["A4"] = "Charlie"
    ws["B4"] = 150
    wb.save(file_path)
    wb.close()
    return file_path


class TestVisualize:
    def test_create_bar_chart(self, sample_excel, tmp_path):
        """创建柱状图"""
        cap = VisualizeCapability()
        output = tmp_path / "bar.png"
        result = cap.execute(None, file=str(sample_excel), chart_type="bar",
                           output=str(output), title="Sales")
        assert result["created"] is True
        assert output.exists()

    def test_create_line_chart(self, sample_excel, tmp_path):
        """创建折线图"""
        cap = VisualizeCapability()
        output = tmp_path / "line.png"
        result = cap.execute(None, file=str(sample_excel), chart_type="line",
                           output=str(output), x_column="Name", y_column="Value")
        assert result["created"] is True
        assert output.exists()

    def test_create_pie_chart(self, sample_excel, tmp_path):
        """创建饼图"""
        cap = VisualizeCapability()
        output = tmp_path / "pie.png"
        result = cap.execute(None, file=str(sample_excel), chart_type="pie",
                           output=str(output), x_column="Name", y_column="Value")
        assert result["created"] is True
        assert output.exists()

    def test_create_scatter_chart(self, sample_excel, tmp_path):
        """创建散点图"""
        cap = VisualizeCapability()
        output = tmp_path / "scatter.png"
        result = cap.execute(None, file=str(sample_excel), chart_type="scatter",
                           output=str(output), x_column="Name", y_column="Value")
        assert result["created"] is True
        assert output.exists()

    def test_create_heatmap(self, sample_excel, tmp_path):
        """创建热力图"""
        cap = VisualizeCapability()
        output = tmp_path / "heatmap.png"
        result = cap.execute(None, file=str(sample_excel), chart_type="heatmap",
                           output=str(output))
        assert result["created"] is True
        assert output.exists()
