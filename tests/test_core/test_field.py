import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.field.measure_range import MeasureRangeCapability
from abacus.core.field.measure_cells import MeasureCellsCapability
from abacus.core.field.measure_structure import MeasureStructureCapability
from abacus.core.field.list_sheets import ListSheetsCapability
from abacus.core.field.peek_preview import PeekPreviewCapability
from abacus.core.field.detect_columns import DetectColumnsCapability
from abacus.core.field.search_content import SearchContentCapability
from abacus.core.field.get_summary import GetSummaryCapability
from abacus.core.field.get_sample_data import GetSampleDataCapability
from abacus.core.exceptions import DataError, SheetNotFoundError, FileNotFoundError


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["A2"] = "Alice"
    ws["B2"] = 30
    ws["A3"] = "Bob"
    ws["B3"] = 25
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def multi_sheet_excel(tmp_path):
    """创建包含多个工作表的 Excel 文件"""
    file_path = tmp_path / "multi_sheet.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1["A1"] = "Product"
    ws1["B1"] = "Amount"
    ws1["A2"] = "Widget"
    ws1["B2"] = 100
    ws2 = wb.create_sheet("Inventory")
    ws2["A1"] = "Item"
    ws2["B1"] = "Quantity"
    ws2["A2"] = "Gadget"
    ws2["B2"] = 50
    ws3 = wb.create_sheet("Config")
    ws3["A1"] = "Setting"
    ws3["B1"] = "Value"
    ws3["A2"] = "Version"
    ws3["B2"] = "1.0"
    wb.save(file_path)
    wb.close()
    return file_path


class TestMeasureRange:
    def test_capability_properties(self):
        cap = MeasureRangeCapability()
        assert cap.name == "measure_range"
        assert cap.chapter == "field"
    
    def test_read_range(self, sample_excel):
        cap = MeasureRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1", range="A1:B3")
        
        assert result["sheet"] == "Sheet1"
        assert result["rows"] == 3
        assert result["data"][0] == ["Name", "Age"]
        assert result["data"][1] == ["Alice", 30]
    
    def test_missing_file(self):
        cap = MeasureRangeCapability()
        with pytest.raises(DataError):
            cap.execute(None, file="nonexistent.xlsx", sheet="Sheet1", range="A1:B3")
    
    def test_missing_sheet(self, sample_excel):
        cap = MeasureRangeCapability()
        with pytest.raises(SheetNotFoundError):
            cap.execute(None, file=str(sample_excel), sheet="NonExistent", range="A1:B3")


class TestMeasureCells:
    def test_capability_properties(self):
        cap = MeasureCellsCapability()
        assert cap.name == "measure_cells"
        assert cap.chapter == "field"
    
    def test_read_cells(self, sample_excel):
        cap = MeasureCellsCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1", range="A1:B2")
        
        assert result["count"] == 4
        assert result["cells"][0]["address"] == "A1"
        assert result["cells"][0]["value"] == "Name"


class TestMeasureStructure:
    def test_capability_properties(self):
        cap = MeasureStructureCapability()
        assert cap.name == "measure_structure"
        assert cap.chapter == "field"
    
    def test_read_structure(self, sample_excel):
        cap = MeasureStructureCapability()
        result = cap.execute(None, file=str(sample_excel))
        
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Sheet1"
        assert result["sheets"][0]["max_row"] == 3


class TestListSheets:
    def test_capability_properties(self):
        cap = ListSheetsCapability()
        assert cap.name == "list_sheets"
        assert cap.chapter == "field"
    
    def test_list_sheets_single(self, sample_excel):
        cap = ListSheetsCapability()
        result = cap.execute(None, file=str(sample_excel))
        
        assert result["file"] == str(sample_excel)
        assert result["sheets"] == ["Sheet1"]
        assert result["count"] == 1
    
    def test_list_sheets_multiple(self, multi_sheet_excel):
        cap = ListSheetsCapability()
        result = cap.execute(None, file=str(multi_sheet_excel))
        
        assert result["count"] == 3
        assert "Sales" in result["sheets"]
        assert "Inventory" in result["sheets"]
        assert "Config" in result["sheets"]
    
    def test_missing_file(self):
        cap = ListSheetsCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx")


class TestPeekPreview:
    def test_capability_properties(self):
        cap = PeekPreviewCapability()
        assert cap.name == "peek_preview"
        assert cap.chapter == "field"
    
    def test_peek_preview_single_sheet(self, sample_excel):
        cap = PeekPreviewCapability()
        result = cap.execute(None, file=str(sample_excel), rows=3)
        
        assert len(result["preview"]) == 1
        assert result["preview"][0]["sheet"] == "Sheet1"
        assert result["preview"][0]["rows_previewed"] == 3
    
    def test_peek_preview_multiple_sheets(self, multi_sheet_excel):
        cap = PeekPreviewCapability()
        result = cap.execute(None, file=str(multi_sheet_excel), rows=2)
        
        assert len(result["preview"]) == 3
    
    def test_missing_file(self):
        cap = PeekPreviewCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx")


class TestDetectColumns:
    def test_capability_properties(self):
        cap = DetectColumnsCapability()
        assert cap.name == "detect_columns"
        assert cap.chapter == "field"
    
    def test_detect_columns(self, sample_excel):
        cap = DetectColumnsCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1")
        
        assert result["columns"] == ["Name", "Age"]
        assert result["column_count"] == 2
        assert result["column_details"]["Name"]["type"] == "string"
        assert result["column_details"]["Age"]["type"] == "number"
    
    def test_missing_sheet(self, sample_excel):
        cap = DetectColumnsCapability()
        with pytest.raises(DataError):
            cap.execute(None, file=str(sample_excel), sheet="NonExistent")
    
    def test_missing_file(self):
        cap = DetectColumnsCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx", sheet="Sheet1")


class TestSearchContent:
    def test_capability_properties(self):
        cap = SearchContentCapability()
        assert cap.name == "search_content"
        assert cap.chapter == "field"
    
    def test_search_content(self, sample_excel):
        cap = SearchContentCapability()
        result = cap.execute(None, file=str(sample_excel), keyword="Alice")
        
        assert result["total_found"] == 1
        assert result["results"][0]["value"] == "Alice"
    
    def test_search_content_not_found(self, sample_excel):
        cap = SearchContentCapability()
        result = cap.execute(None, file=str(sample_excel), keyword="NotExists")
        
        assert result["total_found"] == 0
    
    def test_missing_keyword(self, sample_excel):
        cap = SearchContentCapability()
        with pytest.raises(DataError):
            cap.execute(None, file=str(sample_excel), keyword="")
    
    def test_missing_file(self):
        cap = SearchContentCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx", keyword="test")


class TestGetSummary:
    def test_capability_properties(self):
        cap = GetSummaryCapability()
        assert cap.name == "get_summary"
        assert cap.chapter == "field"
    
    def test_get_summary(self, sample_excel):
        cap = GetSummaryCapability()
        result = cap.execute(None, file=str(sample_excel))
        
        assert result["sheet_count"] == 1
        assert result["total_rows"] == 3
        assert result["sheets"][0]["name"] == "Sheet1"
    
    def test_get_summary_multiple_sheets(self, multi_sheet_excel):
        cap = GetSummaryCapability()
        result = cap.execute(None, file=str(multi_sheet_excel))
        
        assert result["sheet_count"] == 3
    
    def test_missing_file(self):
        cap = GetSummaryCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx")


class TestGetSampleData:
    def test_capability_properties(self):
        cap = GetSampleDataCapability()
        assert cap.name == "get_sample_data"
        assert cap.chapter == "field"
    
    def test_get_sample_data(self, sample_excel):
        cap = GetSampleDataCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1", rows=2)
        
        assert result["columns"] == ["Name", "Age"]
        assert result["rows_returned"] == 2
        assert result["data"][0]["Name"] == "Alice"
        assert result["data"][0]["Age"] == 30
    
    def test_missing_sheet(self, sample_excel):
        cap = GetSampleDataCapability()
        with pytest.raises(DataError):
            cap.execute(None, file=str(sample_excel), sheet="NonExistent")
    
    def test_missing_file(self):
        cap = GetSampleDataCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx", sheet="Sheet1")
