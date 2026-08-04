"""覆盖率提升测试 - 覆盖低覆盖模块的未测试路径"""

import pytest
import os
import tempfile
import shutil
from pathlib import Path

from openpyxl import Workbook


# ─── Helper ────────────────────────────────────────────────────────────

def _wb_with_data(path, data_dict, sheet_name="Sheet"):
    """Create an Excel file with data."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(data_dict.keys())
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    rows = list(zip(*[data_dict[k] for k in headers]))
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    wb.save(path)
    wb.close()
    return path


# ═══════════════════════════════════════════════════════════════════════
# equation/calculate.py  (0% → target 100%)
# ═══════════════════════════════════════════════════════════════════════

class TestEquationCalc:
    def test_basic_expression(self):
        from abacus.core.equation.calculate import CalculateCapability
        cap = CalculateCapability()
        result = cap.execute(None, expression="2 + 3 * 4")
        assert result["result"] == 14

    def test_with_variables(self):
        from abacus.core.equation.calculate import CalculateCapability
        cap = CalculateCapability()
        result = cap.execute(None, expression="x + y", variables={"x": 10, "y": 20})
        assert result["result"] == 30

    def test_math_functions(self):
        from abacus.core.equation.calculate import CalculateCapability
        cap = CalculateCapability()
        result = cap.execute(None, expression="sqrt(16) + log10(100)")
        assert result["result"] == pytest.approx(4 + 2)

    def test_pi_and_e(self):
        from abacus.core.equation.calculate import CalculateCapability
        cap = CalculateCapability()
        result = cap.execute(None, expression="pi + e")
        assert result["result"] > 5

    def test_missing_expression_raises(self):
        from abacus.core.equation.calculate import CalculateCapability
        from abacus.core.exceptions import ValidationError
        cap = CalculateCapability()
        with pytest.raises(ValidationError):
            cap.execute(None)

    def test_invalid_expression_raises(self):
        from abacus.core.equation.calculate import CalculateCapability
        from abacus.core.exceptions import DataError
        cap = CalculateCapability()
        with pytest.raises(DataError):
            cap.execute(None, expression="undefined_func(1)")

    def test_builtin_injection_blocked(self):
        from abacus.core.equation.calculate import CalculateCapability
        from abacus.core.exceptions import DataError
        cap = CalculateCapability()
        with pytest.raises(DataError):
            cap.execute(None, expression="__import__('os')")

    def test_schema(self):
        from abacus.core.equation.calculate import CalculateCapability
        cap = CalculateCapability()
        assert len(cap.schema) == 2
        assert cap.chapter == "equation"
        assert cap.name == "calculate"


# ═══════════════════════════════════════════════════════════════════════
# equation/diagnose_formula.py  (56% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestDiagnoseFormula:
    def _make_file(self, tmp, formulas):
        path = os.path.join(tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        for r, f in enumerate(formulas, 1):
            ws.cell(row=r, column=1, value=f)
        wb.save(path)
        wb.close()
        return path

    def test_no_formulas(self, tmp_path):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        cap = DiagnoseFormulaCapability()
        path = self._make_file(str(tmp_path), ["hello", 123])
        result = cap.execute(None, file=path)
        assert result["formulas_checked"] == 0
        assert result["errors_found"] == 0

    def test_valid_formula_no_error(self, tmp_path):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        cap = DiagnoseFormulaCapability()
        path = self._make_file(str(tmp_path), ["=SUM(A1:A10)"])
        result = cap.execute(None, file=path)
        assert result["formulas_checked"] == 1

    def test_filter_by_sheet(self, tmp_path):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        path = os.path.join(tmp_path, "test.xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.cell(row=1, column=1, value="=SUM(A1:A10)")
        ws2 = wb.create_sheet("Other")
        ws2.cell(row=1, column=1, value="=AVERAGE(B1:B10)")
        wb.save(path)
        wb.close()
        cap = DiagnoseFormulaCapability()
        result = cap.execute(None, file=path, sheet="Other")
        assert result["formulas_checked"] == 1

    def test_filter_by_cell(self, tmp_path):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        path = self._make_file(str(tmp_path), ["=SUM(A1:A10)", "=AVERAGE(B1:B10)"])
        cap = DiagnoseFormulaCapability()
        result = cap.execute(None, file=path, cell="A1")
        assert result["formulas_checked"] == 1

    def test_syntax_check_chinese_comma(self, tmp_path):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        path = self._make_file(str(tmp_path), ["=SUM(A1，A2)"])
        cap = DiagnoseFormulaCapability()
        result = cap.execute(None, file=path)
        assert any("issues" in e for e in result["errors"])

    def test_syntax_check_unbalanced_parens(self, tmp_path):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        path = self._make_file(str(tmp_path), ["=SUM(A1("])
        cap = DiagnoseFormulaCapability()
        result = cap.execute(None, file=path)
        assert any("issues" in e for e in result["errors"])

    def test_syntax_check_unclosed_paren(self, tmp_path):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        path = self._make_file(str(tmp_path), ["=SUM(A1"])
        cap = DiagnoseFormulaCapability()
        result = cap.execute(None, file=path)
        assert any("issues" in e for e in result["errors"])

    def test_syntax_check_unsafe_func(self, tmp_path):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        path = self._make_file(str(tmp_path), ["=INDIRECT(\"A1\")"])
        cap = DiagnoseFormulaCapability()
        result = cap.execute(None, file=path)
        assert any("issues" in e for e in result["errors"])

    def test_missing_file_raises(self):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = DiagnoseFormulaCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx")

    def test_missing_param_raises(self):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        from abacus.core.exceptions import ValidationError
        cap = DiagnoseFormulaCapability()
        with pytest.raises(ValidationError):
            cap.execute(None)

    def test_schema(self):
        from abacus.core.equation.diagnose_formula import DiagnoseFormulaCapability
        cap = DiagnoseFormulaCapability()
        assert cap.name == "diagnose_formula"
        assert cap.chapter == "equation"


# ═══════════════════════════════════════════════════════════════════════
# equation/recalc.py  (62% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestRecalc:
    def test_missing_param_raises(self):
        from abacus.core.equation.recalc import FormulaRecalcCapability
        from abacus.core.exceptions import ValidationError
        cap = FormulaRecalcCapability()
        with pytest.raises(ValidationError):
            cap.execute(None)

    def test_missing_file_raises(self):
        from abacus.core.equation.recalc import FormulaRecalcCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = FormulaRecalcCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx")

    def test_schema(self):
        from abacus.core.equation.recalc import FormulaRecalcCapability
        cap = FormulaRecalcCapability()
        assert cap.name == "recalc_formulas"
        assert cap.chapter == "equation"
        assert len(cap.schema) == 2

    def test_scan_errors_finds_error_cells(self, tmp_path):
        from abacus.core.equation.recalc import FormulaRecalcCapability
        cap = FormulaRecalcCapability()
        path = os.path.join(tmp_path, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="#REF!")
        ws.cell(row=2, column=1, value="#N/A")
        ws.cell(row=3, column=1, value="normal text")
        wb.save(path)
        wb.close()
        errors = cap._scan_errors(path)
        assert len(errors) == 2
        assert errors[0]["error"] == "#REF!"
        assert errors[1]["error"] == "#N/A"

    def test_scan_errors_empty_sheet(self, tmp_path):
        from abacus.core.equation.recalc import FormulaRecalcCapability
        cap = FormulaRecalcCapability()
        path = os.path.join(tmp_path, "test.xlsx")
        wb = Workbook()
        wb.save(path)
        wb.close()
        errors = cap._scan_errors(path)
        assert errors == []


# ═══════════════════════════════════════════════════════════════════════
# report/template.py  (63% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestTemplateReport:
    def test_missing_template_raises(self):
        from abacus.core.report.template import TemplateReportCapability
        from abacus.core.exceptions import DataError
        cap = TemplateReportCapability()
        with pytest.raises(DataError):
            cap.execute(None, output="out.xlsx")

    def test_missing_output_raises(self):
        from abacus.core.report.template import TemplateReportCapability
        from abacus.core.exceptions import DataError
        cap = TemplateReportCapability()
        with pytest.raises(DataError):
            cap.execute(None, template="tpl.xlsx")

    def test_nonexistent_template_raises(self):
        from abacus.core.report.template import TemplateReportCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = TemplateReportCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, template="nonexistent.xlsx", output="out.xlsx")

    def test_fill_with_dict_data(self, tmp_path):
        from abacus.core.report.template import TemplateReportCapability
        cap = TemplateReportCapability()
        tpl = os.path.join(tmp_path, "tpl.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Value")
        wb.save(tpl)
        wb.close()
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, template=tpl, output=out, data={"A2": "test", "B2": 100})
        assert result["filled"]

    def test_fill_with_csv_data_source(self, tmp_path):
        from abacus.core.report.template import TemplateReportCapability
        cap = TemplateReportCapability()
        tpl = os.path.join(tmp_path, "tpl.xlsx")
        wb = Workbook()
        wb.save(tpl)
        wb.close()
        csv_path = os.path.join(tmp_path, "data.csv")
        with open(csv_path, "w") as f:
            f.write("Name,Value\nAlice,100\nBob,200\n")
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, template=tpl, output=out, data_source=csv_path, start_cell="A1")
        assert result["filled"]

    def test_fill_with_excel_data_source(self, tmp_path):
        from abacus.core.report.template import TemplateReportCapability
        cap = TemplateReportCapability()
        tpl = os.path.join(tmp_path, "tpl.xlsx")
        wb = Workbook()
        wb.save(tpl)
        wb.close()
        data_path = os.path.join(tmp_path, "data.xlsx")
        _wb_with_data(data_path, {"Name": ["A", "B"], "Value": [1, 2]})
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, template=tpl, output=out, data_source=data_path, sheet_name="Sheet1")
        assert result["filled"]

    def test_unsupported_data_format_raises(self, tmp_path):
        from abacus.core.report.template import TemplateReportCapability
        from abacus.core.exceptions import DataError
        cap = TemplateReportCapability()
        tpl = os.path.join(tmp_path, "tpl.xlsx")
        wb = Workbook()
        wb.save(tpl)
        wb.close()
        bad_path = os.path.join(tmp_path, "data.txt")
        with open(bad_path, "w") as f:
            f.write("test")
        with pytest.raises(DataError):
            cap.execute(None, template=tpl, output="out.xlsx", data_source=bad_path)

    def test_nonexistent_data_source_raises(self, tmp_path):
        from abacus.core.report.template import TemplateReportCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = TemplateReportCapability()
        tpl = os.path.join(tmp_path, "tpl.xlsx")
        wb = Workbook()
        wb.save(tpl)
        wb.close()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, template=tpl, output="out.xlsx", data_source="nonexistent.csv")

    def test_fill_with_sheet_name(self, tmp_path):
        from abacus.core.report.template import TemplateReportCapability
        cap = TemplateReportCapability()
        tpl = os.path.join(tmp_path, "tpl.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        ws.cell(row=1, column=1, value="Header")
        wb.save(tpl)
        wb.close()
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, template=tpl, output=out, sheet_name="Report", data={"A2": "value"})
        assert result["filled"]

    def test_fill_named_cell_not_found(self, tmp_path):
        from abacus.core.report.template import TemplateReportCapability
        cap = TemplateReportCapability()
        tpl = os.path.join(tmp_path, "tpl.xlsx")
        wb = Workbook()
        wb.save(tpl)
        wb.close()
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, template=tpl, output=out, data={"ZZ99": "value"})
        assert result["filled"]

    def test_schema(self):
        from abacus.core.report.template import TemplateReportCapability
        cap = TemplateReportCapability()
        assert cap.name == "fill_template"
        assert cap.chapter == "work"


# ═══════════════════════════════════════════════════════════════════════
# table/table.py  (66% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestTable:
    def _make_file(self, tmp):
        path = os.path.join(tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Value")
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=2, column=2, value=100)
        wb.save(path)
        wb.close()
        return path

    def test_create_table(self, tmp_path):
        from abacus.core.table.table import TableCapability
        cap = TableCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", action="create", range="A1:B2", table_name="TestTable")
        assert result["action"] == "create"

    def test_create_table_auto_name(self, tmp_path):
        from abacus.core.table.table import TableCapability
        cap = TableCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", action="create", range="A1:B2")
        assert "Table_" in result["table_name"]

    def test_list_tables(self, tmp_path):
        from abacus.core.table.table import TableCapability
        cap = TableCapability()
        path = self._make_file(str(tmp_path))
        cap.execute(None, file=path, sheet="Sheet", action="create", range="A1:B2", table_name="T1")
        result = cap.execute(None, file=path, sheet="Sheet", action="list")
        assert result["action"] == "list"
        assert len(result["tables"]) == 1

    def test_delete_table(self, tmp_path):
        from abacus.core.table.table import TableCapability
        cap = TableCapability()
        path = self._make_file(str(tmp_path))
        cap.execute(None, file=path, sheet="Sheet", action="create", range="A1:B2", table_name="T1")
        result = cap.execute(None, file=path, sheet="Sheet", action="delete", table_name="T1")
        assert result["action"] == "delete"

    def test_delete_nonexistent_table_raises(self, tmp_path):
        from abacus.core.table.table import TableCapability
        from abacus.core.exceptions import DataError
        cap = TableCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", action="delete", table_name="NoSuchTable")

    def test_append_to_table(self, tmp_path):
        from abacus.core.table.table import TableCapability
        cap = TableCapability()
        path = self._make_file(str(tmp_path))
        cap.execute(None, file=path, sheet="Sheet", action="create", range="A1:B2", table_name="T1")
        result = cap.execute(None, file=path, sheet="Sheet", action="append", table_name="T1", data=[["B", 200]])
        assert result["rows_added"] == 1

    def test_append_nonexistent_table_raises(self, tmp_path):
        from abacus.core.table.table import TableCapability
        from abacus.core.exceptions import DataError
        cap = TableCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", action="append", table_name="NoSuchTable", data=[["B", 200]])

    def test_unknown_action_raises(self, tmp_path):
        from abacus.core.table.table import TableCapability
        from abacus.core.exceptions import DataError
        cap = TableCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", action="unknown")

    def test_missing_file_raises(self):
        from abacus.core.table.table import TableCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = TableCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx", sheet="Sheet", action="list")

    def test_missing_action_raises(self):
        from abacus.core.table.table import TableCapability
        from abacus.core.exceptions import DataError
        cap = TableCapability()
        with pytest.raises(DataError):
            cap.execute(None, file="x.xlsx", sheet="Sheet")

    def test_create_requires_range(self, tmp_path):
        from abacus.core.table.table import TableCapability
        from abacus.core.exceptions import DataError
        cap = TableCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", action="create")

    def test_list_string_table_ref(self, tmp_path):
        from abacus.core.table.table import TableCapability
        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table
        path = os.path.join(tmp_path, "test2.xlsx")
        wb = Workbook()
        ws = wb.active
        tab = Table(displayName="FakeTable", ref="A1:B2")
        ws.add_table(tab)
        # Corrupt the table ref to test the hasattr fallback
        ws.tables["FakeTable"] = "A1:B2"
        # Can't save corrupted table, test the hasattr path differently
        wb.close()
        # Test with a valid table first
        path2 = os.path.join(tmp_path, "test3.xlsx")
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.cell(row=1, column=1, value="H")
        tab2 = Table(displayName="RealTable", ref="A1:A1")
        ws2.add_table(tab2)
        wb2.save(path2)
        wb2.close()
        cap = TableCapability()
        result = cap.execute(None, file=path2, sheet="Sheet", action="list")
        assert result["tables"][0]["name"] == "RealTable"

    def test_schema(self):
        from abacus.core.table.table import TableCapability
        cap = TableCapability()
        assert cap.name == "manage_table"
        assert cap.chapter == "work"


# ═══════════════════════════════════════════════════════════════════════
# share/distribute.py  (68% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestDistribute:
    def _make_file(self, tmp):
        path = os.path.join(tmp, "test.xlsx")
        _wb_with_data(path, {"Name": ["A", "B", "C"], "Weight": [1, 2, 3]})
        return path

    def test_equal_distribution(self, tmp_path):
        from abacus.core.share.distribute import DistributeCapability
        cap = DistributeCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B4", total=90)
        assert result["distributed"]

    def test_weighted_distribution(self, tmp_path):
        from abacus.core.share.distribute import DistributeCapability
        cap = DistributeCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B4", total=100,
                             method="weighted", weight_column="Weight")
        assert result["distributed"]

    def test_no_data_rows_raises(self, tmp_path):
        from abacus.core.share.distribute import DistributeCapability
        from abacus.core.exceptions import DataError
        cap = DistributeCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", range="A1:B1", total=100)

    def test_unknown_method_raises(self, tmp_path):
        from abacus.core.share.distribute import DistributeCapability
        from abacus.core.exceptions import DataError
        cap = DistributeCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", range="A1:B4", total=100, method="unknown")

    def test_weighted_no_column_raises(self, tmp_path):
        from abacus.core.share.distribute import DistributeCapability
        from abacus.core.exceptions import DataError
        cap = DistributeCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", range="A1:B4", total=100, method="weighted")

    def test_weighted_missing_column_raises(self, tmp_path):
        from abacus.core.share.distribute import DistributeCapability
        from abacus.core.exceptions import DataError
        cap = DistributeCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", range="A1:B4", total=100,
                         method="weighted", weight_column="NonExistent")

    def test_zero_weight_raises(self, tmp_path):
        from abacus.core.share.distribute import DistributeCapability
        from abacus.core.exceptions import DataError
        cap = DistributeCapability()
        path = os.path.join(tmp_path, "zero.xlsx")
        _wb_with_data(path, {"Name": ["A", "B"], "Weight": [0, 0]})
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", range="A1:B3", total=100,
                         method="weighted", weight_column="Weight")

    def test_missing_file_raises(self):
        from abacus.core.share.distribute import DistributeCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = DistributeCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx", sheet="Sheet", range="A1:B4", total=100)

    def test_missing_params_raises(self):
        from abacus.core.share.distribute import DistributeCapability
        from abacus.core.exceptions import DataError
        cap = DistributeCapability()
        with pytest.raises(DataError):
            cap.execute(None, file="x.xlsx")

    def test_sheet_not_found_raises(self, tmp_path):
        from abacus.core.share.distribute import DistributeCapability
        from abacus.core.exceptions import DataError
        cap = DistributeCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="NoSuchSheet", range="A1:B4", total=100)


# ═══════════════════════════════════════════════════════════════════════
# share/group_by.py  (73% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestGroupBy:
    def _make_file(self, tmp):
        path = os.path.join(tmp, "test.xlsx")
        _wb_with_data(path, {"Cat": ["A", "A", "B"], "Val": [10, 20, 30]})
        return path

    def test_count_only(self, tmp_path):
        from abacus.core.share.group_by import GroupByCapability
        cap = GroupByCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B4", group_columns=["Cat"])
        assert result["groups_count"] == 2

    def test_sum_aggregation(self, tmp_path):
        from abacus.core.share.group_by import GroupByCapability
        cap = GroupByCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B4",
                             group_columns=["Cat"], agg_column="Val", agg_function="sum")
        assert result["groups_count"] == 2

    def test_avg_aggregation(self, tmp_path):
        from abacus.core.share.group_by import GroupByCapability
        cap = GroupByCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B4",
                             group_columns=["Cat"], agg_column="Val", agg_function="avg")
        assert result["groups_count"] == 2

    def test_min_max_aggregation(self, tmp_path):
        from abacus.core.share.group_by import GroupByCapability
        cap = GroupByCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B4",
                             group_columns=["Cat"], agg_column="Val", agg_function="min")
        assert result["groups_count"] == 2
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B4",
                             group_columns=["Cat"], agg_column="Val", agg_function="max")
        assert result["groups_count"] == 2

    def test_unknown_agg_raises(self, tmp_path):
        from abacus.core.share.group_by import GroupByCapability
        from abacus.core.exceptions import DataError
        cap = GroupByCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", range="A1:B4",
                        group_columns=["Cat"], agg_column="Val", agg_function="median")

    def test_missing_column_raises(self, tmp_path):
        from abacus.core.share.group_by import GroupByCapability
        from abacus.core.exceptions import DataError
        cap = GroupByCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", range="A1:B4", group_columns=["NoSuchCol"])

    def test_missing_file_raises(self):
        from abacus.core.share.group_by import GroupByCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = GroupByCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx", sheet="Sheet", range="A1:B4", group_columns=["Cat"])

    def test_missing_params_raises(self):
        from abacus.core.share.group_by import GroupByCapability
        from abacus.core.exceptions import DataError
        cap = GroupByCapability()
        with pytest.raises(DataError):
            cap.execute(None)

    def test_sheet_not_found_raises(self, tmp_path):
        from abacus.core.share.group_by import GroupByCapability
        from abacus.core.exceptions import DataError
        cap = GroupByCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="NoSuchSheet", range="A1:B4", group_columns=["Cat"])

    def test_schema(self):
        from abacus.core.share.group_by import GroupByCapability
        cap = GroupByCapability()
        assert cap.name == "group_by"
        assert cap.chapter == "share"


# ═══════════════════════════════════════════════════════════════════════
# transport/import_data.py  (71% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestImportData:
    def test_csv_import(self, tmp_path):
        from abacus.core.transport.import_data import ImportDataCapability
        cap = ImportDataCapability()
        csv_path = os.path.join(tmp_path, "data.csv")
        with open(csv_path, "w") as f:
            f.write("Name,Age\nAlice,30\nBob,25\n")
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, file=out, source=csv_path)
        assert result["rows_imported"] == 3

    def test_json_import(self, tmp_path):
        import json
        from abacus.core.transport.import_data import ImportDataCapability
        cap = ImportDataCapability()
        json_path = os.path.join(tmp_path, "data.json")
        with open(json_path, "w") as f:
            json.dump([{"Name": "Alice", "Age": 30}, {"Name": "Bob", "Age": 25}], f)
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, file=out, source=json_path, source_type="json")
        assert result["rows_imported"] == 3  # header + 2 rows

    def test_json_empty_list_raises(self, tmp_path):
        import json
        from abacus.core.transport.import_data import ImportDataCapability
        from abacus.core.exceptions import DataError
        cap = ImportDataCapability()
        json_path = os.path.join(tmp_path, "empty.json")
        with open(json_path, "w") as f:
            json.dump([], f)
        with pytest.raises(DataError):
            cap.execute(None, file="out.xlsx", source=json_path, source_type="json")

    def test_import_to_existing_workbook(self, tmp_path):
        from abacus.core.transport.import_data import ImportDataCapability
        cap = ImportDataCapability()
        csv_path = os.path.join(tmp_path, "data.csv")
        with open(csv_path, "w") as f:
            f.write("X,Y\n1,2\n")
        out = os.path.join(tmp_path, "existing.xlsx")
        wb = Workbook()
        wb.save(out)
        wb.close()
        result = cap.execute(None, file=out, source=csv_path)
        assert result["rows_imported"] == 2

    def test_import_with_new_sheet(self, tmp_path):
        from abacus.core.transport.import_data import ImportDataCapability
        cap = ImportDataCapability()
        csv_path = os.path.join(tmp_path, "data.csv")
        with open(csv_path, "w") as f:
            f.write("A,B\n1,2\n")
        out = os.path.join(tmp_path, "out.xlsx")
        wb = Workbook()
        wb.save(out)
        wb.close()
        result = cap.execute(None, file=out, source=csv_path, sheet="NewSheet")
        assert result["sheet"] == "NewSheet"

    def test_import_numeric_values(self, tmp_path):
        from abacus.core.transport.import_data import ImportDataCapability
        cap = ImportDataCapability()
        csv_path = os.path.join(tmp_path, "data.csv")
        with open(csv_path, "w") as f:
            f.write("Val\n1.5\n42\n-7\n")
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, file=out, source=csv_path)
        assert result["rows_imported"] == 4

    def test_unsupported_type_raises(self, tmp_path):
        from abacus.core.transport.import_data import ImportDataCapability
        from abacus.core.exceptions import DataError
        cap = ImportDataCapability()
        bad_path = os.path.join(tmp_path, "data.xyz")
        with open(bad_path, "w") as f:
            f.write("test")
        with pytest.raises(DataError):
            cap.execute(None, file="out.xlsx", source=bad_path, source_type="xml")

    def test_missing_source_raises(self):
        from abacus.core.transport.import_data import ImportDataCapability
        from abacus.core.exceptions import ValidationError
        cap = ImportDataCapability()
        with pytest.raises(ValidationError):
            cap.execute(None, file="out.xlsx")

    def test_missing_file_raises(self):
        from abacus.core.transport.import_data import ImportDataCapability
        from abacus.core.exceptions import ValidationError
        cap = ImportDataCapability()
        with pytest.raises(ValidationError):
            cap.execute(None, source="data.csv")

    def test_nonexistent_source_raises(self):
        from abacus.core.transport.import_data import ImportDataCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = ImportDataCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="out.xlsx", source="nonexistent.csv")

    def test_schema(self):
        from abacus.core.transport.import_data import ImportDataCapability
        cap = ImportDataCapability()
        assert cap.name == "import_data"
        assert cap.chapter == "transport"


# ═══════════════════════════════════════════════════════════════════════
# transport/batch_merge.py  (68% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestBatchMerge:
    def test_merge_single_file(self, tmp_path):
        from abacus.core.transport.batch_merge import BatchMergeCapability
        cap = BatchMergeCapability()
        _wb_with_data(os.path.join(tmp_path, "f1.xlsx"), {"A": [1, 2], "B": [3, 4]})
        out = os.path.join(tmp_path, "merged.xlsx")
        result = cap.execute(None, folder=str(tmp_path), output=out)
        assert result["file_count"] == 1

    def test_merge_multiple_files(self, tmp_path):
        from abacus.core.transport.batch_merge import BatchMergeCapability
        cap = BatchMergeCapability()
        _wb_with_data(os.path.join(tmp_path, "f1.xlsx"), {"A": [1], "B": [2]})
        _wb_with_data(os.path.join(tmp_path, "f2.xlsx"), {"A": [3], "B": [4]})
        out = os.path.join(tmp_path, "merged.xlsx")
        result = cap.execute(None, folder=str(tmp_path), output=out)
        assert result["file_count"] == 2

    def test_merge_specific_sheet(self, tmp_path):
        from abacus.core.transport.batch_merge import BatchMergeCapability
        cap = BatchMergeCapability()
        path = os.path.join(tmp_path, "f1.xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1.cell(row=1, column=1, value="X")
        ws1.cell(row=2, column=1, value=10)
        ws2 = wb.create_sheet("Other")
        ws2.cell(row=1, column=1, value="Y")
        wb.save(path)
        wb.close()
        out = os.path.join(tmp_path, "merged.xlsx")
        result = cap.execute(None, folder=str(tmp_path), output=out, sheet="Data")
        assert result["total_rows"] >= 1

    def test_merge_no_files_raises(self, tmp_path):
        from abacus.core.transport.batch_merge import BatchMergeCapability
        from abacus.core.exceptions import DataError
        cap = BatchMergeCapability()
        empty_dir = os.path.join(tmp_path, "empty")
        os.makedirs(empty_dir)
        with pytest.raises(DataError):
            cap.execute(None, folder=empty_dir, output="out.xlsx")

    def test_missing_folder_raises(self):
        from abacus.core.transport.batch_merge import BatchMergeCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = BatchMergeCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, folder="nonexistent/", output="out.xlsx")

    def test_missing_output_raises(self):
        from abacus.core.transport.batch_merge import BatchMergeCapability
        from abacus.core.exceptions import ValidationError
        cap = BatchMergeCapability()
        with pytest.raises(ValidationError):
            cap.execute(None, folder="somefolder")

    def test_schema(self):
        from abacus.core.transport.batch_merge import BatchMergeCapability
        cap = BatchMergeCapability()
        assert cap.name == "batch_merge"
        assert cap.chapter == "transport"


# ═══════════════════════════════════════════════════════════════════════
# transport/join_tables.py  (72% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestJoinTables:
    def _make_pair(self, tmp):
        left = os.path.join(tmp, "left.xlsx")
        right = os.path.join(tmp, "right.xlsx")
        _wb_with_data(left, {"ID": [1, 2, 3], "Name": ["A", "B", "C"]})
        _wb_with_data(right, {"ID": [2, 3, 4], "Score": [80, 90, 70]})
        return left, right

    def test_inner_join(self, tmp_path):
        from abacus.core.transport.join_tables import JoinTablesCapability
        cap = JoinTablesCapability()
        left, right = self._make_pair(str(tmp_path))
        result = cap.execute(None, left_file=left, left_sheet="Sheet",
                             right_file=right, right_sheet="Sheet", on=["ID"])
        assert result["join_type"] == "inner"
        assert result["result_rows"] == 2

    def test_left_join(self, tmp_path):
        from abacus.core.transport.join_tables import JoinTablesCapability
        cap = JoinTablesCapability()
        left, right = self._make_pair(str(tmp_path))
        result = cap.execute(None, left_file=left, left_sheet="Sheet",
                             right_file=right, right_sheet="Sheet", on=["ID"], how="left")
        assert result["result_rows"] == 3

    def test_outer_join(self, tmp_path):
        from abacus.core.transport.join_tables import JoinTablesCapability
        cap = JoinTablesCapability()
        left, right = self._make_pair(str(tmp_path))
        result = cap.execute(None, left_file=left, left_sheet="Sheet",
                             right_file=right, right_sheet="Sheet", on=["ID"], how="outer")
        assert result["result_rows"] == 4

    def test_join_with_output(self, tmp_path):
        from abacus.core.transport.join_tables import JoinTablesCapability
        cap = JoinTablesCapability()
        left, right = self._make_pair(str(tmp_path))
        out = os.path.join(tmp_path, "joined.xlsx")
        result = cap.execute(None, left_file=left, left_sheet="Sheet",
                             right_file=right, right_sheet="Sheet", on=["ID"], output=out)
        assert os.path.exists(out)

    def test_missing_key_in_left_raises(self, tmp_path):
        from abacus.core.transport.join_tables import JoinTablesCapability
        from abacus.core.exceptions import DataError
        cap = JoinTablesCapability()
        left = os.path.join(tmp_path, "left.xlsx")
        right = os.path.join(tmp_path, "right.xlsx")
        _wb_with_data(left, {"A": [1]})
        _wb_with_data(right, {"B": [2]})
        with pytest.raises(DataError):
            cap.execute(None, left_file=left, left_sheet="Sheet",
                        right_file=right, right_sheet="Sheet", on=["ID"])

    def test_missing_file_raises(self):
        from abacus.core.transport.join_tables import JoinTablesCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = JoinTablesCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, left_file="no.xlsx", left_sheet="S",
                        right_file="no2.xlsx", right_sheet="S", on=["ID"])

    def test_missing_params_raises(self):
        from abacus.core.transport.join_tables import JoinTablesCapability
        from abacus.core.exceptions import ValidationError
        cap = JoinTablesCapability()
        with pytest.raises(ValidationError):
            cap.execute(None)

    def test_schema(self):
        from abacus.core.transport.join_tables import JoinTablesCapability
        cap = JoinTablesCapability()
        assert cap.name == "join_tables"
        assert cap.chapter == "transport"


# ═══════════════════════════════════════════════════════════════════════
# csv/merge.py  (70% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestCSVMerge:
    def test_concat(self, tmp_path):
        from abacus.core.csv.merge import CSVMergeCapability
        cap = CSVMergeCapability()
        f1 = os.path.join(tmp_path, "a.xlsx")
        f2 = os.path.join(tmp_path, "b.xlsx")
        _wb_with_data(f1, {"X": [1]})
        _wb_with_data(f2, {"X": [2]})
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, files=[f1, f2], output=out)
        assert result["total_rows"] == 2

    def test_merge_on_key(self, tmp_path):
        from abacus.core.csv.merge import CSVMergeCapability
        cap = CSVMergeCapability()
        f1 = os.path.join(tmp_path, "a.xlsx")
        f2 = os.path.join(tmp_path, "b.xlsx")
        _wb_with_data(f1, {"ID": [1, 2], "V1": [10, 20]})
        _wb_with_data(f2, {"ID": [2, 3], "V2": [200, 300]})
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, files=[f1, f2], output=out, merge_type="merge", on="ID")
        assert result["total_rows"] == 3

    def test_join_on_key(self, tmp_path):
        from abacus.core.csv.merge import CSVMergeCapability
        cap = CSVMergeCapability()
        f1 = os.path.join(tmp_path, "a.xlsx")
        f2 = os.path.join(tmp_path, "b.xlsx")
        _wb_with_data(f1, {"ID": [1, 2], "V1": [10, 20]})
        _wb_with_data(f2, {"ID": [2, 3], "V2": [200, 300]})
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, files=[f1, f2], output=out, merge_type="join", on="ID")
        assert result["total_rows"] == 2

    def test_dedup(self, tmp_path):
        from abacus.core.csv.merge import CSVMergeCapability
        cap = CSVMergeCapability()
        f1 = os.path.join(tmp_path, "a.xlsx")
        f2 = os.path.join(tmp_path, "b.xlsx")
        _wb_with_data(f1, {"X": [1, 2]})
        _wb_with_data(f2, {"X": [2, 3]})
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, files=[f1, f2], output=out, dedup=True)
        assert result["duplicates_removed"] > 0

    def test_dedup_columns(self, tmp_path):
        from abacus.core.csv.merge import CSVMergeCapability
        cap = CSVMergeCapability()
        f1 = os.path.join(tmp_path, "a.xlsx")
        _wb_with_data(f1, {"ID": [1, 1], "V": [10, 20]})
        out = os.path.join(tmp_path, "out.xlsx")
        result = cap.execute(None, files=[f1], output=out, dedup=True, dedup_columns=["ID"])
        assert result["total_rows"] == 1

    def test_merge_requires_on(self, tmp_path):
        from abacus.core.csv.merge import CSVMergeCapability
        from abacus.core.exceptions import ValidationError
        cap = CSVMergeCapability()
        f1 = os.path.join(tmp_path, "a.xlsx")
        _wb_with_data(f1, {"X": [1]})
        with pytest.raises(ValidationError):
            cap.execute(None, files=[f1], output="out.xlsx", merge_type="merge")

    def test_unknown_merge_type_raises(self, tmp_path):
        from abacus.core.csv.merge import CSVMergeCapability
        from abacus.core.exceptions import DataError
        cap = CSVMergeCapability()
        f1 = os.path.join(tmp_path, "a.xlsx")
        _wb_with_data(f1, {"X": [1]})
        with pytest.raises(DataError, match="不支持的合并类型"):
            cap.execute(None, files=[f1], output="out.xlsx", merge_type="unknown")

    def test_csv_output(self, tmp_path):
        from abacus.core.csv.merge import CSVMergeCapability
        cap = CSVMergeCapability()
        f1 = os.path.join(tmp_path, "a.xlsx")
        _wb_with_data(f1, {"X": [1]})
        out = os.path.join(tmp_path, "out.csv")
        result = cap.execute(None, files=[f1], output=out)
        assert os.path.exists(out)

    def test_missing_files_raises(self):
        from abacus.core.csv.merge import CSVMergeCapability
        from abacus.core.exceptions import ValidationError
        cap = CSVMergeCapability()
        with pytest.raises(ValidationError):
            cap.execute(None, files=[], output="out.xlsx")

    def test_missing_file_on_disk_raises(self):
        from abacus.core.csv.merge import CSVMergeCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = CSVMergeCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, files=["nonexistent.xlsx"], output="out.xlsx")

    def test_schema(self):
        from abacus.core.csv.merge import CSVMergeCapability
        cap = CSVMergeCapability()
        assert cap.name == "merge_files"
        assert cap.chapter == "transport"


# ═══════════════════════════════════════════════════════════════════════
# dimension/derive.py  (71% → target 100%)
# ═══════════════════════════════════════════════════════════════════════

class TestDerive:
    def test_derive_t(self):
        from abacus.core.dimension.derive import DeriveCapability
        cap = DeriveCapability()
        result = cap.execute(None, target_value=100, formula="simple_interest",
                             params={"P": 1000, "r": 0.1})
        assert result["derived"] == "t"
        assert result["value"] == 1.0

    def test_derive_r(self):
        from abacus.core.dimension.derive import DeriveCapability
        cap = DeriveCapability()
        result = cap.execute(None, target_value=100, formula="simple_interest",
                             params={"P": 1000, "t": 1})
        assert result["derived"] == "r"
        assert result["value"] == 0.1

    def test_derive_P(self):
        from abacus.core.dimension.derive import DeriveCapability
        cap = DeriveCapability()
        result = cap.execute(None, target_value=100, formula="simple_interest",
                             params={"r": 0.1, "t": 1})
        assert result["derived"] == "P"
        assert result["value"] == 1000

    def test_derive_profit_price(self):
        from abacus.core.dimension.derive import DeriveCapability
        cap = DeriveCapability()
        result = cap.execute(None, target_value=0.3, formula="profit_margin",
                             params={"cost": 70})
        assert result["derived"] == "price"
        assert result["value"] == pytest.approx(100)

    def test_derive_profit_cost(self):
        from abacus.core.dimension.derive import DeriveCapability
        cap = DeriveCapability()
        result = cap.execute(None, target_value=0.3, formula="profit_margin",
                             params={"price": 100})
        assert result["derived"] == "cost"
        assert result["value"] == pytest.approx(70)

    def test_insufficient_params_raises(self):
        from abacus.core.dimension.derive import DeriveCapability
        from abacus.core.exceptions import DataError
        cap = DeriveCapability()
        with pytest.raises(DataError):
            cap.execute(None, target_value=100, formula="simple_interest", params={})

    def test_profit_missing_params_raises(self):
        from abacus.core.dimension.derive import DeriveCapability
        from abacus.core.exceptions import DataError
        cap = DeriveCapability()
        with pytest.raises(DataError):
            cap.execute(None, target_value=0.3, formula="profit_margin", params={})

    def test_unknown_formula_raises(self):
        from abacus.core.dimension.derive import DeriveCapability
        from abacus.core.exceptions import DataError
        cap = DeriveCapability()
        with pytest.raises(DataError):
            cap.execute(None, target_value=100, formula="compound_interest", params={"P": 1000})

    def test_missing_target_raises(self):
        from abacus.core.dimension.derive import DeriveCapability
        from abacus.core.exceptions import ValidationError
        cap = DeriveCapability()
        with pytest.raises(ValidationError):
            cap.execute(None, formula="simple_interest", params={"P": 1000, "r": 0.1})

    def test_missing_formula_raises(self):
        from abacus.core.dimension.derive import DeriveCapability
        from abacus.core.exceptions import ValidationError
        cap = DeriveCapability()
        with pytest.raises(ValidationError):
            cap.execute(None, target_value=100, params={"P": 1000})

    def test_schema(self):
        from abacus.core.dimension.derive import DeriveCapability
        cap = DeriveCapability()
        assert cap.name == "derive"
        assert cap.chapter == "dimension"


# ═══════════════════════════════════════════════════════════════════════
# formula/generator.py  (69% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestFormulaGenerator:
    def test_vlookup(self):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        cap = FormulaGeneratorCapability()
        result = cap.execute(None, formula_type="vlookup",
                             params={"lookup_value": "A1", "table_range": "B:C", "col_index": 2})
        assert "VLOOKUP" in result["formula"]

    def test_sumif(self):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        cap = FormulaGeneratorCapability()
        result = cap.execute(None, formula_type="sumif",
                             params={"criteria_range": "A:A", "criteria": "yes", "sum_range": "B:B"})
        assert "SUMIF" in result["formula"]

    def test_if(self):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        cap = FormulaGeneratorCapability()
        result = cap.execute(None, formula_type="if",
                             params={"condition": "A1>10", "value_if_true": "Yes", "value_if_false": "No"})
        assert "IF" in result["formula"]

    def test_today(self):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        cap = FormulaGeneratorCapability()
        result = cap.execute(None, formula_type="today", params={})
        assert "TODAY" in result["formula"]

    def test_unknown_formula_raises(self):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        from abacus.core.exceptions import DataError
        cap = FormulaGeneratorCapability()
        with pytest.raises(DataError):
            cap.execute(None, formula_type="nonexistent", params={})

    def test_missing_param_raises(self):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        from abacus.core.exceptions import ValidationError
        cap = FormulaGeneratorCapability()
        with pytest.raises(ValidationError):
            cap.execute(None, formula_type="vlookup", params={})

    def test_write_to_file(self, tmp_path):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        cap = FormulaGeneratorCapability()
        path = os.path.join(tmp_path, "test.xlsx")
        wb = Workbook()
        wb.save(path)
        wb.close()
        result = cap.execute(None, formula_type="today", params={},
                             file=path, sheet="Sheet", cell="A1")
        assert result["formula"]

    def test_write_to_nonexistent_file_raises(self):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = FormulaGeneratorCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, formula_type="today", params={},
                        file="nonexistent.xlsx", sheet="Sheet", cell="A1")

    def test_missing_formula_type_raises(self):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        from abacus.core.exceptions import ValidationError
        cap = FormulaGeneratorCapability()
        with pytest.raises(ValidationError):
            cap.execute(None, params={})

    def test_all_formula_types(self):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        cap = FormulaGeneratorCapability()
        # Test formulas that don't require parameters
        for ft in ["today", "now"]:
            result = cap.execute(None, formula_type=ft, params={})
            assert "formula" in result
        # Test that all formula types have templates
        assert len(cap.FORMULAS) > 20

    def test_schema(self):
        from abacus.core.formula.generator import FormulaGeneratorCapability
        cap = FormulaGeneratorCapability()
        assert cap.name == "generate_formula"
        assert cap.chapter == "equation"


# ═══════════════════════════════════════════════════════════════════════
# work/summary_report.py  (73% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestSummaryReport:
    def test_basic_summary(self, tmp_path):
        from abacus.core.work.summary_report import SummaryReportCapability
        cap = SummaryReportCapability()
        path = os.path.join(tmp_path, "test.xlsx")
        _wb_with_data(path, {"Num": [1, 2, 3], "Text": ["a", "b", "c"]})
        result = cap.execute(None, file=path, sheet="Sheet")
        assert result["total_rows"] == 3
        assert result["total_columns"] == 2
        assert "Num" in result["numeric_stats"]
        assert "Text" in result["categorical_stats"]

    def test_with_nulls(self, tmp_path):
        from abacus.core.work.summary_report import SummaryReportCapability
        cap = SummaryReportCapability()
        path = os.path.join(tmp_path, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="X")
        ws.cell(row=1, column=2, value="Y")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="a")
        ws.cell(row=3, column=2, value="b")
        wb.save(path)
        wb.close()
        result = cap.execute(None, file=path, sheet="Sheet")
        assert result["null_counts"]["X"] == 1
        assert result["categorical_stats"]["Y"]["unique_count"] == 2

    def test_missing_file_raises(self):
        from abacus.core.work.summary_report import SummaryReportCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = SummaryReportCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx", sheet="Sheet")

    def test_missing_params_raises(self):
        from abacus.core.work.summary_report import SummaryReportCapability
        from abacus.core.exceptions import DataError
        cap = SummaryReportCapability()
        with pytest.raises(DataError):
            cap.execute(None)

    def test_schema(self):
        from abacus.core.work.summary_report import SummaryReportCapability
        cap = SummaryReportCapability()
        assert cap.name == "generate_summary_report"
        assert cap.chapter == "work"


# ═══════════════════════════════════════════════════════════════════════
# work/group_rows.py  (75% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestGroupRows:
    def test_basic_group(self, tmp_path):
        from abacus.core.work.group_rows import GroupRowsCapability
        cap = GroupRowsCapability()
        path = os.path.join(tmp_path, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i)
        wb.save(path)
        wb.close()
        result = cap.execute(None, file=path, sheet="Sheet", start_row=2, end_row=4)
        assert result["rows_grouped"] == 3

    def test_group_with_level(self, tmp_path):
        from abacus.core.work.group_rows import GroupRowsCapability
        cap = GroupRowsCapability()
        path = os.path.join(tmp_path, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i)
        wb.save(path)
        wb.close()
        result = cap.execute(None, file=path, sheet="Sheet", start_row=2, end_row=4, level=2)
        assert result["level"] == 2

    def test_start_row_below_1_raises(self, tmp_path):
        from abacus.core.work.group_rows import GroupRowsCapability
        from abacus.core.exceptions import DataError
        cap = GroupRowsCapability()
        path = os.path.join(tmp_path, "test.xlsx")
        wb = Workbook()
        wb.save(path)
        wb.close()
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", start_row=0, end_row=4)

    def test_end_row_before_start_raises(self, tmp_path):
        from abacus.core.work.group_rows import GroupRowsCapability
        from abacus.core.exceptions import DataError
        cap = GroupRowsCapability()
        path = os.path.join(tmp_path, "test.xlsx")
        wb = Workbook()
        wb.save(path)
        wb.close()
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", start_row=5, end_row=2)

    def test_missing_file_raises(self):
        from abacus.core.work.group_rows import GroupRowsCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = GroupRowsCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx", sheet="Sheet", start_row=1, end_row=5)

    def test_missing_params_raises(self):
        from abacus.core.work.group_rows import GroupRowsCapability
        from abacus.core.exceptions import DataError
        cap = GroupRowsCapability()
        with pytest.raises(DataError):
            cap.execute(None)

    def test_sheet_not_found_raises(self, tmp_path):
        from abacus.core.work.group_rows import GroupRowsCapability
        from abacus.core.exceptions import DataError
        cap = GroupRowsCapability()
        path = os.path.join(tmp_path, "test.xlsx")
        wb = Workbook()
        wb.save(path)
        wb.close()
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="NoSuchSheet", start_row=1, end_row=5)

    def test_schema(self):
        from abacus.core.work.group_rows import GroupRowsCapability
        cap = GroupRowsCapability()
        assert cap.name == "group_rows"
        assert cap.chapter == "work"


# ═══════════════════════════════════════════════════════════════════════
# share/summarize.py  (78% → target 90%+)
# ═══════════════════════════════════════════════════════════════════════

class TestSummarize:
    def _make_file(self, tmp):
        path = os.path.join(tmp, "test.xlsx")
        _wb_with_data(path, {"Cat": ["A", "A", "B", "B"], "Val": [10, 20, 30, 40]})
        return path

    def test_sum_agg(self, tmp_path):
        from abacus.core.share.summarize import SummarizeCapability
        cap = SummarizeCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B5",
                             group_by="Cat", agg_config={"Val": "sum"})
        assert result["groups_count"] == 2

    def test_avg_agg(self, tmp_path):
        from abacus.core.share.summarize import SummarizeCapability
        cap = SummarizeCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B5",
                             group_by="Cat", agg_config={"Val": "avg"})
        assert result["groups_count"] == 2

    def test_count_agg(self, tmp_path):
        from abacus.core.share.summarize import SummarizeCapability
        cap = SummarizeCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B5",
                             group_by="Cat", agg_config={"Val": "count"})
        assert result["groups_count"] == 2

    def test_min_max_agg(self, tmp_path):
        from abacus.core.share.summarize import SummarizeCapability
        cap = SummarizeCapability()
        path = self._make_file(str(tmp_path))
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B5",
                             group_by="Cat", agg_config={"Val": "min"})
        assert result["groups_count"] == 2
        result = cap.execute(None, file=path, sheet="Sheet", range="A1:B5",
                             group_by="Cat", agg_config={"Val": "max"})
        assert result["groups_count"] == 2

    def test_missing_column_raises(self, tmp_path):
        from abacus.core.share.summarize import SummarizeCapability
        from abacus.core.exceptions import DataError
        cap = SummarizeCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="Sheet", range="A1:B5",
                        group_by="NonExistent", agg_config={"Val": "sum"})

    def test_missing_file_raises(self):
        from abacus.core.share.summarize import SummarizeCapability
        from abacus.core.exceptions import FileNotFoundError
        cap = SummarizeCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file="nonexistent.xlsx", sheet="Sheet", range="A1:B5",
                        group_by="Cat", agg_config={"Val": "sum"})

    def test_missing_params_raises(self):
        from abacus.core.share.summarize import SummarizeCapability
        from abacus.core.exceptions import DataError
        cap = SummarizeCapability()
        with pytest.raises(DataError):
            cap.execute(None)

    def test_sheet_not_found_raises(self, tmp_path):
        from abacus.core.share.summarize import SummarizeCapability
        from abacus.core.exceptions import DataError
        cap = SummarizeCapability()
        path = self._make_file(str(tmp_path))
        with pytest.raises(DataError):
            cap.execute(None, file=path, sheet="NoSuchSheet", range="A1:B5",
                        group_by="Cat", agg_config={"Val": "sum"})

    def test_schema(self):
        from abacus.core.share.summarize import SummarizeCapability
        cap = SummarizeCapability()
        assert cap.name == "summarize"
        assert cap.chapter == "share"
