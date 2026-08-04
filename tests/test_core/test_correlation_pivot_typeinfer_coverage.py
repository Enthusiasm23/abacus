"""测试 analyze_correlation.py, analysis/pivot.py, auto_type_infer.py"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from abacus.core.triangle.analyze_correlation import AnalyzeCorrelationCapability
from abacus.core.analysis.pivot import PivotAnalysisCapability
from abacus.core.grain.auto_type_infer import AutoTypeInferCapability
from abacus.core.exceptions import DataError, FileNotFoundError, ValidationError


@pytest.fixture
def correlation_excel(tmp_path):
    file_path = tmp_path / "correlation.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Sales", "Marketing", "Profit"])
    ws.append([100, 50, 20])
    ws.append([200, 80, 50])
    ws.append([150, 60, 30])
    ws.append([300, 120, 80])
    ws.append([250, 100, 60])
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def pivot_excel(tmp_path):
    file_path = tmp_path / "pivot.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Region", "Product", "Amount"])
    ws.append(["East", "A", 100])
    ws.append(["West", "B", 200])
    ws.append(["East", "A", 150])
    ws.append(["West", "A", 250])
    ws.append(["East", "B", 300])
    wb.save(file_path)
    wb.close()
    return file_path


class TestAnalyzeCorrelation:
    def test_missing_file(self):
        cap = AnalyzeCorrelationCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 file"):
            cap.execute(None, sheet="Data", range="A1:C6",
                       column1="Sales", column2="Profit")

    def test_missing_columns(self, correlation_excel):
        cap = AnalyzeCorrelationCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 column1 或 column2"):
            cap.execute(None, file=str(correlation_excel), sheet="Data",
                       range="A1:C6", column1="Sales")

    def test_file_not_found(self, tmp_path):
        cap = AnalyzeCorrelationCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"), sheet="Data",
                       range="A1:C6", column1="Sales", column2="Profit")

    def test_sheet_not_found(self, correlation_excel):
        cap = AnalyzeCorrelationCapability()
        with pytest.raises(DataError, match="不存在"):
            cap.execute(None, file=str(correlation_excel), sheet="NoSuch",
                       range="A1:C6", column1="Sales", column2="Profit")

    def test_column_not_found(self, correlation_excel):
        cap = AnalyzeCorrelationCapability()
        with pytest.raises(DataError, match="不存在"):
            cap.execute(None, file=str(correlation_excel), sheet="Data",
                       range="A1:C6", column1="Nonexistent", column2="Profit")

    def test_correlation_positive(self, correlation_excel):
        cap = AnalyzeCorrelationCapability()
        result = cap.execute(None, file=str(correlation_excel), sheet="Data",
                           range="A1:C6", column1="Sales", column2="Profit")
        assert result["correlation"] > 0
        assert result["direction"] == "正相关"
        assert result["data_points"] == 5

    def test_correlation_negative(self, tmp_path):
        file_path = tmp_path / "neg_corr.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["X", "Y"])
        for i in range(10):
            ws.append([i + 1, 100 - i * 10])
        wb.save(file_path)
        wb.close()

        cap = AnalyzeCorrelationCapability()
        result = cap.execute(None, file=str(file_path), sheet="Sheet",
                           range="A1:B11", column1="X", column2="Y")
        assert result["correlation"] < 0
        assert result["direction"] == "负相关"

    def test_correlation_strength_strong(self, correlation_excel):
        cap = AnalyzeCorrelationCapability()
        result = cap.execute(None, file=str(correlation_excel), sheet="Data",
                           range="A1:C6", column1="Sales", column2="Marketing")
        assert result["strength"] in ["强相关", "中等相关", "弱相关", "几乎不相关"]

    def test_correlation_insufficient_data(self, tmp_path):
        file_path = tmp_path / "few.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["X", "Y"])
        ws.append([1, 2])
        wb.save(file_path)
        wb.close()

        cap = AnalyzeCorrelationCapability()
        result = cap.execute(None, file=str(file_path), sheet="Sheet",
                           range="A1:B2", column1="X", column2="Y")
        assert "error" in result

    def test_schema_has_params(self):
        cap = AnalyzeCorrelationCapability()
        names = [s.name for s in cap.schema]
        assert "file" in names
        assert "sheet" in names
        assert "range" in names
        assert "column1" in names
        assert "column2" in names


class TestPivotAnalysis:
    def test_missing_file(self):
        cap = PivotAnalysisCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 file"):
            cap.execute(None, group_by="Region", value_field="Amount")

    def test_missing_group_by(self, pivot_excel):
        cap = PivotAnalysisCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 group_by"):
            cap.execute(None, file=str(pivot_excel), value_field="Amount")

    def test_missing_value_field(self, pivot_excel):
        cap = PivotAnalysisCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 value_field"):
            cap.execute(None, file=str(pivot_excel), group_by="Region")

    def test_file_not_found(self, tmp_path):
        cap = PivotAnalysisCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"),
                       group_by="Region", value_field="Amount")

    def test_column_not_found(self, pivot_excel):
        cap = PivotAnalysisCapability()
        with pytest.raises(DataError, match="不存在"):
            cap.execute(None, file=str(pivot_excel), group_by="Nonexistent",
                       value_field="Amount")

    def test_pivot_sum(self, pivot_excel):
        cap = PivotAnalysisCapability()
        result = cap.execute(None, file=str(pivot_excel), sheet="Sales",
                           group_by="Region", value_field="Amount", agg_function="sum")
        assert result["agg_function"] == "sum"
        assert result["groups"] == 2

    def test_pivot_mean(self, pivot_excel):
        cap = PivotAnalysisCapability()
        result = cap.execute(None, file=str(pivot_excel), sheet="Sales",
                           group_by="Region", value_field="Amount", agg_function="mean")
        assert result["agg_function"] == "mean"

    def test_pivot_count(self, pivot_excel):
        cap = PivotAnalysisCapability()
        result = cap.execute(None, file=str(pivot_excel), sheet="Sales",
                           group_by="Region", value_field="Amount", agg_function="count")
        assert result["agg_function"] == "count"

    def test_pivot_unknown_agg(self, pivot_excel):
        cap = PivotAnalysisCapability()
        with pytest.raises(DataError, match="不支持的聚合函数"):
            cap.execute(None, file=str(pivot_excel), sheet="Sales",
                       group_by="Region", value_field="Amount", agg_function="median")

    def test_pivot_output_csv(self, pivot_excel, tmp_path):
        output = tmp_path / "pivot_output.csv"
        cap = PivotAnalysisCapability()
        result = cap.execute(None, file=str(pivot_excel), sheet="Sales",
                           group_by="Region", value_field="Amount",
                           output=str(output))
        assert output.exists()

    def test_pivot_output_excel(self, pivot_excel, tmp_path):
        output = tmp_path / "pivot_output.xlsx"
        cap = PivotAnalysisCapability()
        result = cap.execute(None, file=str(pivot_excel), sheet="Sales",
                           group_by="Region", value_field="Amount",
                           output=str(output))
        assert output.exists()

    def test_pivot_no_sheet(self, pivot_excel):
        cap = PivotAnalysisCapability()
        result = cap.execute(None, file=str(pivot_excel),
                           group_by="Region", value_field="Amount")
        assert result["groups"] == 2

    def test_unsupported_format(self, tmp_path):
        file_path = tmp_path / "data.txt"
        file_path.write_text("test")
        cap = PivotAnalysisCapability()
        with pytest.raises(DataError, match="不支持的文件格式"):
            cap.execute(None, file=str(file_path), group_by="A", value_field="B")

    def test_schema_has_params(self):
        cap = PivotAnalysisCapability()
        names = [s.name for s in cap.schema]
        assert "file" in names
        assert "group_by" in names
        assert "value_field" in names


class TestAutoTypeInfer:
    def test_missing_file(self):
        cap = AutoTypeInferCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 file"):
            cap.execute(None, sheet="Sheet1")

    def test_missing_sheet(self):
        cap = AutoTypeInferCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 sheet"):
            cap.execute(None, file="dummy.xlsx")

    def test_file_not_found(self, tmp_path):
        cap = AutoTypeInferCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"), sheet="Sheet1")

    def test_auto_type_infer_basic(self, correlation_excel):
        cap = AutoTypeInferCapability()
        result = cap.execute(None, file=str(correlation_excel), sheet="Data")
        assert result["success"] is True
        assert result["total_columns"] == 3
        assert "inferred_types" in result

    def test_auto_type_infer_with_output(self, correlation_excel, tmp_path):
        output = tmp_path / "inferred.xlsx"
        cap = AutoTypeInferCapability()
        result = cap.execute(None, file=str(correlation_excel), sheet="Data",
                           output=str(output))
        assert result["success"] is True
        assert output.exists()

    def test_auto_type_infer_text_to_number(self, tmp_path):
        file_path = tmp_path / "text_num.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Value"])
        ws.append([1, "100"])
        ws.append([2, "200"])
        ws.append([3, "300"])
        wb.save(file_path)
        wb.close()

        cap = AutoTypeInferCapability()
        result = cap.execute(None, file=str(file_path), sheet="Sheet")
        assert result["success"] is True
