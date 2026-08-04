"""测试方程章扩展能力"""

import pytest

from abacus.core.equation import CreateFormulaCapability
from abacus.core.dimension import SolveEquationCapability


class TestCreateFormula:
    def test_create_sum_formula(self, tmp_path):
        """创建求和公式"""
        from openpyxl import Workbook
        file_path = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = 30
        wb.save(file_path)
        wb.close()
        
        cap = CreateFormulaCapability()
        result = cap.execute(None, file=str(file_path), sheet="Sheet",
                           cell="A4", formula="SUM(A1:A3)")
        assert result["created"] is True


class TestSolveEquation:
    def test_linear_equation(self):
        """一元一次方程"""
        cap = SolveEquationCapability()
        result = cap.execute(None, equation="2x + 3 = 7")
        assert result["type"] == "linear"
        assert result["solution"] == 2.0
    
    def test_quadratic_equation(self):
        """一元二次方程"""
        cap = SolveEquationCapability()
        result = cap.execute(None, equation="x^2-4=0")
        assert result["type"] == "quadratic"
        assert len(result["solutions"]) == 2
