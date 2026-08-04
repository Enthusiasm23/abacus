"""测试报表生成能力"""

import pytest
import pandas as pd
from pathlib import Path

from abacus.core.report import BasicReportCapability, AdvancedReportCapability, TemplateReportCapability


@pytest.fixture
def sample_csv(tmp_path):
    """创建测试用 CSV 文件"""
    file_path = tmp_path / "test.csv"
    df = pd.DataFrame({
        "Name": ["Alice", "Bob", "Charlie"],
        "Sales": [100, 200, 150],
        "Region": ["North", "South", "North"]
    })
    df.to_csv(file_path, index=False)
    return file_path


class TestBasicReport:
    def test_create_basic_report(self, sample_csv, tmp_path):
        """创建基础报表"""
        cap = BasicReportCapability()
        output = tmp_path / "basic.xlsx"
        result = cap.execute(None, data_source=str(sample_csv), output=str(output),
                           title="Sales Report")
        assert result["rows"] == 3
        assert result["columns"] == 3
        assert output.exists()
    
    def test_create_report_with_sheet_name(self, sample_csv, tmp_path):
        """指定工作表名称"""
        cap = BasicReportCapability()
        output = tmp_path / "basic.xlsx"
        result = cap.execute(None, data_source=str(sample_csv), output=str(output),
                           sheet_name="Data")
        assert result["sheet_name"] == "Data"


class TestAdvancedReport:
    def test_create_advanced_report(self, sample_csv, tmp_path):
        """创建高级报表"""
        cap = AdvancedReportCapability()
        output = tmp_path / "advanced.xlsx"
        result = cap.execute(None, data_source=str(sample_csv), output=str(output),
                           chart_type="bar", include_dashboard=True)
        assert result["sheets"] >= 3
        assert output.exists()
    
    def test_create_report_with_line_chart(self, sample_csv, tmp_path):
        """创建折线图报表"""
        cap = AdvancedReportCapability()
        output = tmp_path / "advanced.xlsx"
        result = cap.execute(None, data_source=str(sample_csv), output=str(output),
                           chart_type="line")
        assert output.exists()


class TestTemplateReport:
    def test_fill_template_with_dict(self, tmp_path):
        """字典数据填充模板"""
        import openpyxl
        
        # 创建模板
        template = tmp_path / "template.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "{{name}}"
        ws["B2"] = "{{value}}"
        wb.save(template)
        
        # 填充数据
        cap = TemplateReportCapability()
        output = tmp_path / "filled.xlsx"
        result = cap.execute(None, template=str(template), output=str(output),
                           data={"A2": "Alice", "B2": 100})
        assert result["filled"] is True
        assert output.exists()
