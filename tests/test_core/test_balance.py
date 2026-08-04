"""测试盈不足章（数据验证）能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.balance import ValidateRangeCapability, ValidateTypeCapability, ValidateFormulaCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["C1"] = "Formula"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["C2"] = "=SUM(B2:B4)"
    ws["A3"] = "Bob"
    ws["B3"] = 200
    ws["C3"] = None
    ws["A4"] = "Charlie"
    ws["B4"] = 150
    ws["C4"] = None
    wb.save(file_path)
    wb.close()
    return file_path


class TestValidateRange:
    def test_validate_no_empty(self, sample_excel):
        """验证无空值"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B4", rules={"no_empty": True})
        assert result["valid"] is True
    
    def test_validate_with_empty(self, sample_excel):
        """验证有空值"""
        cap = ValidateRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:C4", rules={"no_empty": True})
        assert result["issues_count"] > 0


class TestValidateType:
    def test_validate_number_type(self, sample_excel):
        """验证数字类型"""
        cap = ValidateTypeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B2:B4", expected_type="number")
        assert result["valid"] is True
    
    def test_validate_text_type(self, sample_excel):
        """验证文本类型"""
        cap = ValidateTypeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A2:A4", expected_type="text")
        assert result["valid"] is True


class TestValidateFormula:
    def test_validate_formulas(self, sample_excel):
        """验证公式"""
        cap = ValidateFormulaCapability()
        result = cap.execute(None, file=str(sample_excel))
        assert result["total_formulas"] == 1
        assert result["valid"] == 1
