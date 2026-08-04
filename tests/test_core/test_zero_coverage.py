"""Tests for 5 zero-coverage capabilities: Variance, DataValidation, FileValidate, Style, Subtotal"""

import os
import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook

from abacus.core.finance.variance import VarianceCapability
from abacus.core.balance.data_validation import DataValidationCapability
from abacus.core.balance.file_validate import FileValidateCapability
from abacus.core.style.styling import StyleCapability
from abacus.core.share.subtotal import SubtotalCapability
from abacus.core.exceptions import DataError, FileNotFoundError, ValidationError


# ============================================================
# VarianceCapability
# ============================================================

class TestVarianceCapability:
    def setup_method(self):
        self.cap = VarianceCapability()
        self.tmpdir = tempfile.mkdtemp()

    def _create_budget_actual(self, budget_data, actual_data):
        path = os.path.join(self.tmpdir, "variance.xlsx")
        wb = Workbook()
        ws_b = wb.active
        ws_b.title = "Budget"
        for row in budget_data:
            ws_b.append(row)
        ws_a = wb.create_sheet("Actual")
        for row in actual_data:
            ws_a.append(row)
        wb.save(path)
        wb.close()
        return path

    def test_positive_basic(self):
        path = self._create_budget_actual(
            [["Item", "Amount"], ["Sales", 100], ["Cost", 50]],
            [["Item", "Amount"], ["Sales", 120], ["Cost", 45]],
        )
        result = self.cap.execute(None, file=path, budget_sheet="Budget", actual_sheet="Actual")
        assert result["total_variance"] == 15  # (120-100)+(45-50)=15
        assert result["favorable_count"] == 1
        assert result["unfavorable_count"] == 1

    def test_negative_missing_file(self):
        with pytest.raises(ValidationError):
            self.cap.execute(None)

    def test_negative_nonexistent_file(self):
        with pytest.raises(DataError):
            self.cap.execute(None, file="nonexistent.xlsx", budget_sheet="Budget", actual_sheet="Actual")

    def test_edge_zero_budget(self):
        path = self._create_budget_actual(
            [["Item", "Amount"], ["Sales", 0]],
            [["Item", "Amount"], ["Sales", 50]],
        )
        result = self.cap.execute(None, file=path, budget_sheet="Budget", actual_sheet="Actual")
        assert result["total_variance"] == 50

    def test_edge_with_output(self):
        path = self._create_budget_actual(
            [["Item", "Amount"], ["Sales", 100]],
            [["Item", "Amount"], ["Sales", 110]],
        )
        output = os.path.join(self.tmpdir, "output.xlsx")
        result = self.cap.execute(None, file=path, budget_sheet="Budget", actual_sheet="Actual", output=output)
        assert os.path.exists(output)
        assert result["total_variance"] == 10

    def test_edge_threshold(self):
        path = self._create_budget_actual(
            [["Item", "Amount"], ["Sales", 100], ["Cost", 100]],
            [["Item", "Amount"], ["Sales", 105], ["Cost", 95]],
        )
        result = self.cap.execute(None, file=path, budget_sheet="Budget", actual_sheet="Actual", threshold=0.1)
        assert result["material_variances"] == 0  # 5% < 10%

    def test_properties(self):
        assert self.cap.name == "variance_analysis"
        assert self.cap.chapter == "triangle"


# ============================================================
# DataValidationCapability
# ============================================================

class TestDataValidationCapability:
    def setup_method(self):
        self.cap = DataValidationCapability()
        self.tmpdir = tempfile.mkdtemp()

    def _create_file(self):
        path = os.path.join(self.tmpdir, "validation.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Status", "Score"])
        for i in range(5):
            ws.append([f"User{i+1}", "", 0])
        wb.save(path)
        wb.close()
        return path

    def test_positive_list_validation(self):
        path = self._create_file()
        result = self.cap.execute(None, file=path, sheet="Data", range="B2:B6",
                                  validation_type="list", formula1="Active,Inactive,Pending")
        assert result["applied"] is True
        assert result["validation_type"] == "list"

    def test_positive_number_validation(self):
        path = self._create_file()
        result = self.cap.execute(None, file=path, sheet="Data", range="C2:C6",
                                  validation_type="number", operator="between",
                                  formula1="0", formula2="100")
        assert result["applied"] is True

    def test_positive_date_validation(self):
        path = self._create_file()
        result = self.cap.execute(None, file=path, sheet="Data", range="A2:A6",
                                  validation_type="date")
        assert result["applied"] is True

    def test_positive_text_length(self):
        path = self._create_file()
        result = self.cap.execute(None, file=path, sheet="Data", range="A2:A6",
                                  validation_type="text_length", formula1="1", formula2="50")
        assert result["applied"] is True

    def test_negative_missing_file(self):
        with pytest.raises(ValidationError):
            self.cap.execute(None)

    def test_negative_missing_sheet(self):
        path = self._create_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="NoSuchSheet", range="A1:A5",
                             validation_type="list", formula1="a,b")

    def test_negative_unknown_type(self):
        path = self._create_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="Data", range="A1:A5",
                             validation_type="unknown_type")

    def test_negative_list_no_formula1(self):
        path = self._create_file()
        with pytest.raises(ValidationError):
            self.cap.execute(None, file=path, sheet="Data", range="A1:A5",
                             validation_type="list")

    def test_positive_with_error_message(self):
        path = self._create_file()
        result = self.cap.execute(None, file=path, sheet="Data", range="C2:C6",
                                  validation_type="number", formula1="0", formula2="100",
                                  error_message="Please enter 0-100")
        assert result["applied"] is True

    def test_properties(self):
        assert self.cap.name == "set_data_validation"
        assert self.cap.chapter == "balance"


# ============================================================
# FileValidateCapability
# ============================================================

class TestFileValidateCapability:
    def setup_method(self):
        self.cap = FileValidateCapability()
        self.tmpdir = tempfile.mkdtemp()

    def _create_valid_file(self):
        path = os.path.join(self.tmpdir, "valid.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["A", 1])
        ws.append(["B", 2])
        wb.save(path)
        wb.close()
        return path

    def _create_error_file(self):
        path = os.path.join(self.tmpdir, "errors.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "#REF!"
        ws["A2"] = "=1/0"
        wb.save(path)
        wb.close()
        return path

    def test_positive_valid_file(self):
        path = self._create_valid_file()
        result = self.cap.execute(None, file=path)
        assert result["valid"] is True
        assert len(result["errors"]) == 0

    def test_positive_with_formula_errors(self):
        path = self._create_error_file()
        result = self.cap.execute(None, file=path)
        assert result["valid"] is True  # warnings, not errors
        assert len(result["warnings"]) > 0

    def test_negative_missing_file(self):
        with pytest.raises(ValidationError):
            self.cap.execute(None)

    def test_negative_nonexistent_file(self):
        with pytest.raises(FileNotFoundError):
            self.cap.execute(None, file="nonexistent.xlsx")

    def test_negative_not_excel(self):
        bad_path = os.path.join(self.tmpdir, "bad.xlsx")
        with open(bad_path, "w") as f:
            f.write("not an excel file")
        result = self.cap.execute(None, file=bad_path)
        assert result["valid"] is False
        assert any("ZIP" in e or "Invalid" in e for e in result["errors"])

    def test_edge_empty_workbook(self):
        path = os.path.join(self.tmpdir, "empty.xlsx")
        wb = Workbook()
        wb.save(path)
        wb.close()
        result = self.cap.execute(None, file=path)
        assert result["valid"] is True

    def test_properties(self):
        assert self.cap.name == "validate_file"
        assert self.cap.chapter == "balance"


# ============================================================
# StyleCapability
# ============================================================

class TestStyleCapability:
    def setup_method(self):
        self.cap = StyleCapability()
        self.tmpdir = tempfile.mkdtemp()

    def _create_file(self):
        path = os.path.join(self.tmpdir, "style.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Value", "Status"])
        ws.append(["A", 100, "Good"])
        ws.append(["B", -50, "Bad"])
        wb.save(path)
        wb.close()
        return path

    def test_positive_apply_header(self):
        path = self._create_file()
        result = self.cap.execute(None, file=path, sheet="Data", action="apply_header",
                                  range="A1:C1", industry="finance")
        assert result["action"] == "apply_header"

    def test_positive_apply_kpi(self):
        path = self._create_file()
        result = self.cap.execute(None, file=path, sheet="Data", action="apply_kpi",
                                  range="B2:B3")
        assert result["action"] == "apply_kpi"

    def test_positive_auto_width(self):
        path = self._create_file()
        result = self.cap.execute(None, file=path, sheet="Data", action="auto_width")
        assert result["action"] == "auto_width"

    def test_positive_get_styles(self):
        path = self._create_file()
        result = self.cap.execute(None, file=path, sheet="Data", action="get", range="A1:C1")
        assert result["action"] == "get"
        assert len(result["styles"]) == 3

    def test_positive_all_industries(self):
        path = self._create_file()
        for ind in ["finance", "ecommerce", "saas", "internet"]:
            result = self.cap.execute(None, file=path, sheet="Data", action="apply_header",
                                      range="A1:C1", industry=ind)
            assert result["industry"] == ind

    def test_negative_missing_file(self):
        with pytest.raises(DataError):
            self.cap.execute(None)

    def test_negative_nonexistent_file(self):
        with pytest.raises(FileNotFoundError):
            self.cap.execute(None, file="nonexistent.xlsx", sheet="Data", action="auto_width")

    def test_negative_unknown_action(self):
        path = self._create_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="Data", action="unknown")

    def test_negative_header_no_range(self):
        path = self._create_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="Data", action="apply_header")

    def test_negative_kpi_no_range(self):
        path = self._create_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="Data", action="apply_kpi")

    def test_negative_get_no_range(self):
        path = self._create_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="Data", action="get")

    def test_properties(self):
        assert self.cap.name == "manage_style"
        assert self.cap.chapter == "work"


# ============================================================
# SubtotalCapability
# ============================================================

class TestSubtotalCapability:
    def setup_method(self):
        self.cap = SubtotalCapability()
        self.tmpdir = tempfile.mkdtemp()

    def _create_sales_file(self):
        path = os.path.join(self.tmpdir, "sales.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append(["Region", "Product", "Sales"])
        ws.append(["East", "A", 100])
        ws.append(["East", "B", 200])
        ws.append(["West", "A", 150])
        ws.append(["West", "B", 250])
        ws.append(["East", "A", 300])
        wb.save(path)
        wb.close()
        return path

    def test_positive_sum(self):
        path = self._create_sales_file()
        result = self.cap.execute(None, file=path, sheet="Sales", range="A1:C6",
                                  group_column="Region", function="sum")
        assert result["success"] is True
        assert result["groups_count"] == 2
        for grp in result["summary"]:
            if grp["Region"] == "East":
                assert grp["Sales"] == 600  # 100+200+300
            elif grp["Region"] == "West":
                assert grp["Sales"] == 400  # 150+250

    def test_positive_mean(self):
        path = self._create_sales_file()
        result = self.cap.execute(None, file=path, sheet="Sales", range="A1:C6",
                                  group_column="Region", function="mean")
        assert result["success"] is True
        assert result["function"] == "mean"

    def test_positive_count(self):
        path = self._create_sales_file()
        result = self.cap.execute(None, file=path, sheet="Sales", range="A1:C6",
                                  group_column="Region", function="count")
        assert result["success"] is True

    def test_positive_min_max(self):
        path = self._create_sales_file()
        r_min = self.cap.execute(None, file=path, sheet="Sales", range="A1:C6",
                                  group_column="Region", function="min")
        r_max = self.cap.execute(None, file=path, sheet="Sales", range="A1:C6",
                                  group_column="Region", function="max")
        assert r_min["success"] is True
        assert r_max["success"] is True

    def test_negative_missing_file(self):
        with pytest.raises(DataError):
            self.cap.execute(None)

    def test_negative_missing_sheet(self):
        path = self._create_sales_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="NoSuchSheet", range="A1:C6",
                             group_column="Region")

    def test_negative_invalid_function(self):
        path = self._create_sales_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="Sales", range="A1:C6",
                             group_column="Region", function="median")

    def test_negative_column_not_found(self):
        path = self._create_sales_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="Sales", range="A1:C6",
                             group_column="Nonexistent")

    def test_negative_missing_range(self):
        path = self._create_sales_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="Sales",
                             group_column="Region")

    def test_negative_missing_group_column(self):
        path = self._create_sales_file()
        with pytest.raises(DataError):
            self.cap.execute(None, file=path, sheet="Sales", range="A1:C6")

    def test_properties(self):
        assert self.cap.name == "subtotal"
        assert self.cap.chapter == "share"
