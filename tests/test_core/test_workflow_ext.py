"""测试工作流扩展能力（边界情况、错误处理、格式化扩展）"""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from abacus.core.workflow import SpreadsheetWorkflowCapability, FormattingWorkflowCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["A3"] = "Bob"
    ws["B3"] = 200
    wb.save(file_path)
    wb.close()
    return file_path


class TestSpreadsheetWorkflowExtended:
    def test_missing_action(self, sample_excel):
        """缺少 action 参数"""
        cap = SpreadsheetWorkflowCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=str(sample_excel))

    def test_unknown_action(self, sample_excel):
        """未知 action"""
        cap = SpreadsheetWorkflowCapability()
        with pytest.raises(Exception):
            cap.execute(None, action="unknown", file=str(sample_excel))

    def test_create_missing_output(self):
        """创建时缺少 output"""
        cap = SpreadsheetWorkflowCapability()
        with pytest.raises(Exception):
            cap.execute(None, action="create")

    def test_read_missing_file(self):
        """读取时文件不存在"""
        cap = SpreadsheetWorkflowCapability()
        with pytest.raises(Exception):
            cap.execute(None, action="read", file="nonexistent.xlsx")

    def test_read_missing_sheet(self, sample_excel):
        """读取时工作表不存在"""
        cap = SpreadsheetWorkflowCapability()
        with pytest.raises(Exception):
            cap.execute(None, action="read", file=str(sample_excel), sheet="Nonexistent")

    def test_read_default_sheet(self, sample_excel):
        """读取默认工作表"""
        cap = SpreadsheetWorkflowCapability()
        result = cap.execute(None, action="read", file=str(sample_excel))
        assert result["action"] == "read"
        assert result["rows"] >= 2

    def test_edit_missing_file(self):
        """编辑时文件不存在"""
        cap = SpreadsheetWorkflowCapability()
        with pytest.raises(Exception):
            cap.execute(None, action="edit", file="nonexistent.xlsx",
                       data={"rows": [["X", 1]]})

    def test_edit_with_range(self, sample_excel):
        """编辑 - 指定范围"""
        cap = SpreadsheetWorkflowCapability()
        result = cap.execute(None, action="edit", file=str(sample_excel), sheet="Data",
                           data={"rows": [["Dave", 250]]}, range="A5")
        assert result["action"] == "edit"
        assert result["rows_written"] == 1

    def test_format_missing_file(self):
        """格式化时文件不存在"""
        cap = SpreadsheetWorkflowCapability()
        with pytest.raises(Exception):
            cap.execute(None, action="format", file="nonexistent.xlsx",
                       range="A1:B1", bold=True)

    def test_format_missing_sheet(self, sample_excel):
        """格式化时工作表不存在"""
        cap = SpreadsheetWorkflowCapability()
        with pytest.raises(Exception):
            cap.execute(None, action="format", file=str(sample_excel),
                       sheet="Nonexistent", range="A1:B1", bold=True)

    def test_create_with_empty_data(self, tmp_path):
        """创建 - 空数据"""
        cap = SpreadsheetWorkflowCapability()
        output = tmp_path / "empty.xlsx"
        result = cap.execute(None, action="create", output=str(output), sheet="Empty")
        assert result["action"] == "create"
        assert result["rows"] == 0
        assert output.exists()

    def test_format_bold_and_font_size(self, sample_excel):
        """格式化 - 粗体和字号"""
        cap = SpreadsheetWorkflowCapability()
        result = cap.execute(None, action="format", file=str(sample_excel), sheet="Data",
                           range="A1:B1", bold=True, font_size=14)
        assert result["action"] == "format"
        wb = load_workbook(sample_excel)
        cell = wb["Data"]["A1"]
        assert cell.font.bold is True
        assert cell.font.size == 14
        wb.close()

    def test_format_bg_color(self, sample_excel):
        """格式化 - 背景颜色"""
        cap = SpreadsheetWorkflowCapability()
        result = cap.execute(None, action="format", file=str(sample_excel), sheet="Data",
                           range="A1:B1", bg_color="4472C4")
        assert result["action"] == "format"
        wb = load_workbook(sample_excel)
        cell = wb["Data"]["A1"]
        assert cell.fill.fgColor.rgb is not None
        wb.close()


class TestFormattingWorkflowExtended:
    def test_border_formatting(self, sample_excel):
        """边框格式化"""
        cap = FormattingWorkflowCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B1",
                           border={"left": {"style": "thin", "color": "000000"},
                                   "right": {"style": "thin", "color": "000000"},
                                   "top": {"style": "thin", "color": "000000"},
                                   "bottom": {"style": "thin", "color": "000000"}})
        assert result["cells_formatted"] == 2

    def test_alignment_formatting(self, sample_excel):
        """对齐格式化"""
        cap = FormattingWorkflowCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B1",
                           alignment={"horizontal": "center", "vertical": "center", "wrap_text": True})
        assert result["cells_formatted"] == 2
        wb = load_workbook(sample_excel)
        cell = wb["Data"]["A1"]
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"
        assert cell.alignment.wrap_text is True
        wb.close()

    def test_font_with_all_properties(self, sample_excel):
        """字体 - 所有属性"""
        cap = FormattingWorkflowCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:A1",
                           font={"name": "Arial", "size": 14, "bold": True, "italic": True, "color": "FF0000"})
        assert result["cells_formatted"] == 1
        wb = load_workbook(sample_excel)
        cell = wb["Data"]["A1"]
        assert cell.font.name == "Arial"
        assert cell.font.size == 14
        assert cell.font.bold is True
        assert cell.font.italic is True
        wb.close()

    def test_fill_with_pattern_type(self, sample_excel):
        """填充 - 指定 pattern_type"""
        cap = FormattingWorkflowCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:A1",
                           fill={"color": "FFFF00", "pattern_type": "gray0625"})
        assert result["cells_formatted"] == 1

    def test_missing_file(self):
        """文件不存在"""
        cap = FormattingWorkflowCapability()
        with pytest.raises(Exception):
            cap.execute(None, file="nonexistent.xlsx", sheet="Data",
                       range="A1:B1", font={"bold": True})

    def test_missing_sheet(self, sample_excel):
        """工作表不存在"""
        cap = FormattingWorkflowCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=str(sample_excel), sheet="Nonexistent",
                       range="A1:B1", font={"bold": True})

    def test_combined_formatting(self, sample_excel):
        """组合格式化"""
        cap = FormattingWorkflowCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B1",
                           font={"bold": True, "color": "FFFFFF"},
                           fill={"color": "4472C4"},
                           border={"left": {"style": "thin"}},
                           alignment={"horizontal": "center"})
        assert result["cells_formatted"] == 2
