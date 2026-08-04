"""测试勾股章（数据分析）能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.triangle import AnalyzeStatsCapability, AnalyzeTrendCapability, AnalyzeCorrelationCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Month"
    ws["B1"] = "Sales"
    ws["C1"] = "Profit"
    ws["A2"] = "Jan"
    ws["B2"] = 100
    ws["C2"] = 20
    ws["A3"] = "Feb"
    ws["B3"] = 150
    ws["C3"] = 30
    ws["A4"] = "Mar"
    ws["B4"] = 200
    ws["C4"] = 40
    ws["A5"] = "Apr"
    ws["B5"] = 250
    ws["C5"] = 50
    ws["A6"] = "May"
    ws["B6"] = 300
    ws["C6"] = 60
    wb.save(file_path)
    wb.close()
    return file_path


class TestAnalyzeStats:
    def test_analyze_sales(self, sample_excel):
        """分析销售统计"""
        cap = AnalyzeStatsCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B1:B6")
        assert "Sales" in result["statistics"]
        stats = result["statistics"]["Sales"]
        assert stats["count"] == 5
        assert stats["mean"] == 200
        assert stats["min"] == 100
        assert stats["max"] == 300


class TestAnalyzeTrend:
    def test_analyze_sales_trend(self, sample_excel):
        """分析销售趋势"""
        cap = AnalyzeTrendCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B1:B6", value_column="Sales")
        assert result["data_points"] == 5
        assert result["trend_direction"] == "上升"


class TestAnalyzeCorrelation:
    def test_analyze_sales_profit_correlation(self, sample_excel):
        """分析销售利润相关性"""
        cap = AnalyzeCorrelationCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B1:C6", column1="Sales", column2="Profit")
        assert result["data_points"] == 5
        assert abs(result["correlation"]) > 0.9  # 强相关
