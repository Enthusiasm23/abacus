"""测试公式重算"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from abacus.core.equation.recalc import FormulaRecalcCapability


class TestFormulaRecalc:
    """测试公式重算"""
    
    def test_capability_properties(self):
        """测试能力属性"""
        cap = FormulaRecalcCapability()
        assert cap.name == "recalc_formulas"
        assert cap.chapter == "equation"
        assert "LibreOffice" in cap.description
    
    def test_missing_file(self):
        """测试缺少文件参数"""
        cap = FormulaRecalcCapability()
        with pytest.raises(Exception):
            cap.execute(None)
    
    def test_file_not_found(self):
        """测试文件不存在"""
        cap = FormulaRecalcCapability()
        with pytest.raises(Exception):
            cap.execute(None, file="nonexistent.xlsx")
    
    def test_scan_errors_with_errors(self):
        """测试扫描错误 - 有错误的情况"""
        wb = Workbook()
        ws = wb.active
        
        # 写入带错误的值
        ws["A1"] = "#REF!"
        ws["A2"] = "#N/A"
        ws["A3"] = "#VALUE!"
        ws["A4"] = "正常值"
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            file_path = f.name
        
        cap = FormulaRecalcCapability()
        errors = cap._scan_errors(file_path)
        
        assert len(errors) == 3
        assert errors[0]["error"] == "#REF!"
        assert errors[1]["error"] == "#N/A"
        assert errors[2]["error"] == "#VALUE!"
        
        os.unlink(file_path)
    
    def test_scan_errors_no_errors(self):
        """测试扫描错误 - 无错误的情况"""
        wb = Workbook()
        ws = wb.active
        
        ws["A1"] = "正常值"
        ws["A2"] = 123
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            file_path = f.name
        
        cap = FormulaRecalcCapability()
        errors = cap._scan_errors(file_path)
        
        assert len(errors) == 0
        
        os.unlink(file_path)
    
    def test_scan_errors_mixed_cells(self):
        """测试扫描错误 - 混合单元格"""
        wb = Workbook()
        ws = wb.active
        
        ws["A1"] = "#DIV/0!"
        ws["B1"] = 100
        ws["C1"] = "#NAME?"
        ws["D1"] = "文本"
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            file_path = f.name
        
        cap = FormulaRecalcCapability()
        errors = cap._scan_errors(file_path)
        
        assert len(errors) == 2
        error_types = [e["error"] for e in errors]
        assert "#DIV/0!" in error_types
        assert "#NAME?" in error_types
        
        os.unlink(file_path)
    
    def test_scan_errors_multiple_sheets(self):
        """测试扫描错误 - 多个工作表"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        
        ws1["A1"] = "#NULL!"
        ws2["A1"] = "#NUM!"
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            file_path = f.name
        
        cap = FormulaRecalcCapability()
        errors = cap._scan_errors(file_path)
        
        assert len(errors) == 2
        sheets = [e["sheet"] for e in errors]
        assert "Sheet1" in sheets
        assert "Sheet2" in sheets
        
        os.unlink(file_path)
    
    def test_scan_errors_cell_coordinates(self):
        """测试扫描错误 - 单元格坐标"""
        wb = Workbook()
        ws = wb.active
        
        ws["A1"] = "#REF!"
        ws["B2"] = "#N/A"
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            file_path = f.name
        
        cap = FormulaRecalcCapability()
        errors = cap._scan_errors(file_path)
        
        assert errors[0]["cell"] == "A1"
        assert errors[1]["cell"] == "B2"
        
        os.unlink(file_path)
    
    def test_scan_errors_inexistent_file(self):
        """测试扫描错误 - 不存在的文件"""
        cap = FormulaRecalcCapability()
        errors = cap._scan_errors("nonexistent.xlsx")
        assert errors == []
