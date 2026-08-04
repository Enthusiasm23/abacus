"""测试 transform_pipeline.py 未覆盖步骤：convert_unit, standardize"""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from abacus.core.grain.transform_pipeline import TransformPipelineCapability
from abacus.core.exceptions import DataError


@pytest.fixture
def pipeline_data(tmp_path):
    file_path = tmp_path / "pipeline_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Amount", "Status"])
    ws.append(["Alice", 100, "active"])
    ws.append(["Bob", 200, "inactive"])
    ws.append(["Charlie", 300, "active"])
    wb.save(file_path)
    wb.close()
    return file_path


class TestTransformPipelineConvertUnit:
    def test_convert_unit_step(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "convert_unit", "range": "B2:B4", "factor": 2}])
        assert result["steps_succeeded"] == 1
        r = result["results"][0]["result"]
        assert r["converted"] == 3
        assert r["factor"] == 2

        wb = load_workbook(pipeline_data)
        assert wb["Data"]["B2"].value == 200.0
        assert wb["Data"]["B3"].value == 400.0
        assert wb["Data"]["B4"].value == 600.0
        wb.close()

    def test_convert_unit_missing_range(self, pipeline_data):
        cap = TransformPipelineCapability()
        with pytest.raises(DataError, match="range is required"):
            cap.execute(None, file=str(pipeline_data), sheet="Data",
                       steps=[{"type": "convert_unit", "factor": 2}])

    def test_convert_unit_with_non_numeric(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "convert_unit", "range": "A2:A4", "factor": 10}])
        r = result["results"][0]["result"]
        assert r["converted"] == 0


class TestTransformPipelineStandardize:
    def test_standardize_upper(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "standardize", "range": "C2:C4", "text_case": "upper"}])
        assert result["steps_succeeded"] == 1
        r = result["results"][0]["result"]
        assert r["standardized"] == 3

        wb = load_workbook(pipeline_data)
        assert wb["Data"]["C2"].value == "ACTIVE"
        wb.close()

    def test_standardize_lower(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "standardize", "range": "C2:C4", "text_case": "lower"}])
        r = result["results"][0]["result"]
        assert r["standardized"] == 3

    def test_standardize_title(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "standardize", "range": "C2:C4", "text_case": "title"}])
        r = result["results"][0]["result"]
        assert r["standardized"] == 3

    def test_standardize_missing_range(self, pipeline_data):
        cap = TransformPipelineCapability()
        with pytest.raises(DataError, match="range is required"):
            cap.execute(None, file=str(pipeline_data), sheet="Data",
                       steps=[{"type": "standardize", "text_case": "upper"}])

    def test_standardize_skips_non_strings(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "standardize", "range": "B2:B4", "text_case": "upper"}])
        r = result["results"][0]["result"]
        assert r["standardized"] == 0


class TestTransformPipelineConvertTypeEdgeCases:
    def test_convert_type_to_float(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "convert_type", "range": "B2:B4", "target_type": "float"}])
        assert result["steps_succeeded"] == 1
        r = result["results"][0]["result"]
        assert r["converted"] == 3
        assert r["target_type"] == "float"

    def test_convert_type_to_str(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "convert_type", "range": "B2:B4", "target_type": "str"}])
        r = result["results"][0]["result"]
        assert r["converted"] == 3

    def test_convert_type_invalid_value(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "convert_type", "range": "A2:A4", "target_type": "int"}])
        r = result["results"][0]["result"]
        assert r["converted"] == 0

    def test_convert_type_missing_range(self, pipeline_data):
        cap = TransformPipelineCapability()
        with pytest.raises(DataError, match="range is required"):
            cap.execute(None, file=str(pipeline_data), sheet="Data",
                       steps=[{"type": "convert_type", "target_type": "int"}])


class TestTransformPipelineConvertFormatEdgeCases:
    def test_convert_format_currency(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "convert_format", "range": "B2:B4", "format_type": "currency"}])
        r = result["results"][0]["result"]
        assert r["formatted"] == 3
        assert r["format"] == "$#,##0.00"

    def test_convert_format_percentage(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "convert_format", "range": "B2:B4", "format_type": "percentage"}])
        r = result["results"][0]["result"]
        assert r["format"] == "0.00%"

    def test_convert_format_date(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "convert_format", "range": "B2:B4", "format_type": "date"}])
        r = result["results"][0]["result"]
        assert r["format"] == "yyyy-mm-dd"

    def test_convert_format_text(self, pipeline_data):
        cap = TransformPipelineCapability()
        result = cap.execute(None, file=str(pipeline_data), sheet="Data",
                           steps=[{"type": "convert_format", "range": "B2:B4", "format_type": "text"}])
        r = result["results"][0]["result"]
        assert r["format"] == "@"

    def test_convert_format_missing_range(self, pipeline_data):
        cap = TransformPipelineCapability()
        with pytest.raises(DataError, match="range is required"):
            cap.execute(None, file=str(pipeline_data), sheet="Data",
                       steps=[{"type": "convert_format", "format_type": "number"}])


class TestTransformPipelineReplaceValueEdgeCases:
    def test_replace_value_missing_old(self, pipeline_data):
        cap = TransformPipelineCapability()
        with pytest.raises(DataError, match="range and old_value are required"):
            cap.execute(None, file=str(pipeline_data), sheet="Data",
                       steps=[{"type": "replace_value", "range": "C2:C4"}])
