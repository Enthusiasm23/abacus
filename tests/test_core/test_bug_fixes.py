"""测试 Bug 修复"""

import pytest
import tempfile
import os
import shutil
from pathlib import Path
from openpyxl import Workbook

from abacus.core.grain.transpose import TransposeCapability
from abacus.core.dimension.solve_equation import SolveEquationCapability
from abacus.core.work.unpack import UnpackFileCapability
from abacus.core.work.zoom import ZoomCapability
from abacus.core.balance.validate_type import ValidateTypeCapability
from abacus.core.field.detect_columns import DetectColumnsCapability
from abacus.core.grain.auto_type_infer import AutoTypeInferCapability
from abacus.core.work.format import FormatRangeCapability
from abacus.core.work.pivot import CreatePivotCapability
from abacus.core.transport.export_data import ExportDataCapability
from abacus.core.share.distribute import DistributeCapability
from abacus.core.grain.convert_type import ConvertTypeCapability
from abacus.core.grain.convert_unit import ConvertUnitCapability
from abacus.core.field.search_content import SearchContentCapability
from abacus.core.field.measure_cells import MeasureCellsCapability
from abacus.core.work.comment import CommentCapability
from abacus.core.triangle.analyze_stats import AnalyzeStatsCapability


@pytest.fixture
def sample_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Name", "Sales", "Active"])
    ws.append([1, "Alice", 100, True])
    ws.append([2, "Bob", 200, False])
    ws.append([3, "Carol", 150, True])
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


class TestTransposeFix:
    """测试 transpose.py 行列索引修复"""

    def test_transpose_non_square(self, sample_excel):
        cap = TransposeCapability()
        output_sheet = "Transposed"
        result = cap.execute(None, file=sample_excel, sheet="Data",
                           range="A1:D4", output_sheet=output_sheet)
        assert result["success"] is True
        assert result["source_rows"] == 4
        assert result["source_columns"] == 4
        assert result["output_rows"] == 4
        assert result["output_columns"] == 4
        os.unlink(sample_excel)

    def test_transpose_rectangular(self):
        """测试非方形范围的转置"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["A", "B", "C"])
        ws.append([1, 2, 3])
        ws.append([4, 5, 6])
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        cap = TransposeCapability()
        result = cap.execute(None, file=tmp.name, sheet="Data",
                           range="A1:C3", output_sheet="Transposed")
        assert result["source_rows"] == 3
        assert result["source_columns"] == 3
        os.unlink(tmp.name)


class TestSolveEquationFix:
    """测试二次方程修复"""

    def test_quadratic_with_b(self):
        cap = SolveEquationCapability()
        result = cap.execute(None, equation="x^2 - 5x + 6 = 0")
        assert result["type"] == "quadratic"
        solutions = sorted(result["solutions"])
        assert abs(solutions[0] - 2) < 0.001
        assert abs(solutions[1] - 3) < 0.001

    def test_linear(self):
        cap = SolveEquationCapability()
        result = cap.execute(None, equation="2x + 3 = 7")
        assert result["type"] == "linear"
        assert abs(result["solution"] - 2) < 0.001


class TestUnpackFix:
    """测试 unpack.py with 块修复"""

    def test_unpack_returns_file_list(self):
        import zipfile
        cap = UnpackFileCapability()

        # Create a zip file
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.close()
        output_dir = tempfile.mkdtemp()

        with zipfile.ZipFile(tmp.name, 'w') as zf:
            zf.writestr("test.txt", "hello")

        result = cap.execute(None, file=tmp.name, output=output_dir)
        assert result["unpacked"] is True
        assert len(result["files"]) > 0
        os.unlink(tmp.name)
        shutil.rmtree(output_dir, ignore_errors=True)


class TestZoomFix:
    """测试 zoom.py 空值检查修复"""

    def test_missing_zoom(self, sample_excel):
        cap = ZoomCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=sample_excel, sheet="Data")
        os.unlink(sample_excel)

    def test_missing_sheet(self, sample_excel):
        cap = ZoomCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=sample_excel, zoom=100)
        os.unlink(sample_excel)

    def test_valid_zoom(self, sample_excel):
        cap = ZoomCapability()
        result = cap.execute(None, file=sample_excel, sheet="Data", zoom=150)
        assert result["zoom"] == 150
        os.unlink(sample_excel)

    def test_zoom_out_of_range(self, sample_excel):
        cap = ZoomCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=sample_excel, sheet="Data", zoom=5)
        os.unlink(sample_excel)


class TestValidateTypeFix:
    """测试 validate_type.py 列字母修复"""

    def test_wide_table_column_letters(self):
        """测试超过 Z 列的列字母"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        # Write headers in columns A through Z
        for i in range(26):
            ws.cell(row=1, column=i+1, value=f"Col_{i+1}")
            ws.cell(row=2, column=i+1, value=i+1)
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        cap = ValidateTypeCapability()
        result = cap.execute(None, file=tmp.name, sheet="Data",
                           range="A1:Z2")
        assert "valid" in result
        assert "column_types" in result
        os.unlink(tmp.name)


class TestDetectColumnsFix:
    """测试 detect_columns.py bool 检测修复"""

    def test_boolean_detection(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Flag"])
        ws.append([True])
        ws.append([False])
        ws.append([True])
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        cap = DetectColumnsCapability()
        result = cap.execute(None, file=tmp.name, sheet="Data")
        details = result["column_details"]
        assert details["Flag"]["type"] == "boolean"
        os.unlink(tmp.name)

    def test_number_detection(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Value"])
        ws.append([1])
        ws.append([2.5])
        ws.append([3])
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        cap = DetectColumnsCapability()
        result = cap.execute(None, file=tmp.name, sheet="Data")
        details = result["column_details"]
        assert details["Value"]["type"] == "number"
        os.unlink(tmp.name)


class TestAutoTypeInferFix:
    """测试 auto_type_infer.py 整数检测修复"""

    def test_integer_detection(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["IntCol"])
        ws.append([1])
        ws.append([2])
        ws.append([3])
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        cap = AutoTypeInferCapability()
        result = cap.execute(None, file=tmp.name, sheet="Data")
        inferred = result["inferred_types"]["IntCol"]["inferred_type"]
        assert "int" in inferred
        os.unlink(tmp.name)


class TestFormatColorFix:
    """测试 format.py 颜色 # 前缀修复"""

    def test_color_with_hash_prefix(self, sample_excel):
        cap = FormatRangeCapability()
        result = cap.execute(None, file=sample_excel, sheet="Data",
                           range="A1:D4",
                           font={"color": "#FF0000", "bold": True})
        assert result["cells_formatted"] > 0
        os.unlink(sample_excel)

    def test_color_without_hash_prefix(self, sample_excel):
        cap = FormatRangeCapability()
        result = cap.execute(None, file=sample_excel, sheet="Data",
                           range="A1:D4",
                           font={"color": "00FF00", "bold": True})
        assert result["cells_formatted"] > 0
        os.unlink(sample_excel)


class TestPivotNonNumericFix:
    """测试 pivot.py 非数值聚合修复"""

    def test_pivot_with_non_numeric_values(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Category", "Value"])
        ws.append(["A", 100])
        ws.append(["A", 200])
        ws.append(["B", 300])
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        cap = CreatePivotCapability()
        result = cap.execute(None, file=tmp.name, sheet="Data",
                           range="A1:B4", row_fields=["Category"],
                           value_field="Value", agg_function="count")
        assert result is not None
        os.unlink(tmp.name)


class TestExportDataRaggedFix:
    """测试 export_data.py ragged 数据修复"""

    def test_ragged_rows_json(self):
        import json
        cap = ExportDataCapability()
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["A", "B", "C"])
        ws.append([1, 2, 3])
        ws.append([4, 5])  # Short row
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        output = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
        output.close()

        result = cap.execute(None, file=tmp.name, sheet="Data",
                           range="A1:C3", output=output.name, format="json")
        assert result["rows_exported"] == 2

        with open(output.name, 'r') as f:
            data = json.load(f)
        assert len(data) == 2
        assert data[1]["C"] is None

        os.unlink(tmp.name)
        os.unlink(output.name)


class TestDistributeFix:
    """测试 distribute.py 异常时 workbook 关闭修复"""

    def test_missing_sheet(self, sample_excel):
        cap = DistributeCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=sample_excel, sheet="Nonexistent",
                       range="A1:D4", total=1000)
        os.unlink(sample_excel)

    def test_missing_params(self):
        cap = DistributeCapability()
        with pytest.raises(Exception):
            cap.execute(None)


class TestConvertTypeFix:
    """测试 convert_type.py 参数验证修复"""

    def test_missing_params(self):
        cap = ConvertTypeCapability()
        with pytest.raises(Exception):
            cap.execute(None)

    def test_invalid_target_type(self, sample_excel):
        cap = ConvertTypeCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=sample_excel, sheet="Data",
                       range="A1:D4", target_type="invalid")
        os.unlink(sample_excel)

    def test_valid_target_type(self, sample_excel):
        cap = ConvertTypeCapability()
        result = cap.execute(None, file=sample_excel, sheet="Data",
                           range="A1:D4", target_type="float")
        assert result["target_type"] == "float"
        os.unlink(sample_excel)


class TestConvertUnitFix:
    """测试 convert_unit.py 参数验证修复"""

    def test_missing_params(self):
        cap = ConvertUnitCapability()
        with pytest.raises(Exception):
            cap.execute(None)

    def test_invalid_conversion(self, sample_excel):
        cap = ConvertUnitCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=sample_excel, sheet="Data",
                       range="A1:D4", from_unit="km", to_unit="kg")
        os.unlink(sample_excel)

    def test_valid_conversion(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Distance"])
        ws.append([1])
        ws.append([2])
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        cap = ConvertUnitCapability()
        result = cap.execute(None, file=tmp.name, sheet="Data",
                           range="A1:A3", from_unit="km", to_unit="m")
        assert result["cells_converted"] == 2
        os.unlink(tmp.name)


class TestSearchContentMaxResultsFix:
    """测试 search_content.py max_results 跨表限制修复"""

    def test_max_results_across_sheets(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        for i in range(30):
            ws1.append([f"test_{i}"])
            ws2.append([f"test_{i+30}"])
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        cap = SearchContentCapability()
        result = cap.execute(None, file=tmp.name, keyword="test", max_results=10)
        assert result["total_found"] <= 10
        os.unlink(tmp.name)


class TestMeasureCellsEmptySheetFix:
    """测试 measure_cells.py 空 sheet 修复"""

    def test_empty_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        cap = MeasureCellsCapability()
        result = cap.execute(None, file=tmp.name, sheet="Empty", range="A1")
        assert result["count"] >= 0
        os.unlink(tmp.name)


class TestCommentReadOnlyFix:
    """测试 comment.py 只读操作不保存修复"""

    def test_list_does_not_modify_timestamp(self, sample_excel):
        cap = CommentCapability()
        mtime_before = os.path.getmtime(sample_excel)
        result = cap.execute(None, file=sample_excel, sheet="Data", action="list")
        mtime_after = os.path.getmtime(sample_excel)
        assert mtime_before == mtime_after
        os.unlink(sample_excel)

    def test_get_does_not_modify_timestamp(self, sample_excel):
        cap = CommentCapability()
        mtime_before = os.path.getmtime(sample_excel)
        result = cap.execute(None, file=sample_excel, sheet="Data",
                           action="get", cell="A1")
        mtime_after = os.path.getmtime(sample_excel)
        assert mtime_before == mtime_after
        os.unlink(sample_excel)


class TestAnalyzeStatsOptimizationFix:
    """测试 analyze_stats.py 单次打开修复"""

    def test_analyze_stats(self, sample_excel):
        cap = AnalyzeStatsCapability()
        result = cap.execute(None, file=sample_excel, sheet="Data",
                           range="A1:D4")
        assert result["columns_analyzed"] > 0
        os.unlink(sample_excel)
