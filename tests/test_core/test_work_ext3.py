"""测试商功章扩展能力（批量执行、批量转换、批量验证、图表、格式化、数组公式）"""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from abacus.core.work import (
    BatchExecuteCapability,
    BatchTransformCapability,
    BatchValidateCapability,
    CreatePivotCapability,
    FormatRangeCapability,
    CreateChartCapability,
    UpdateChartCapability,
    DeleteChartCapability,
    ListChartsCapability,
    SetArrayFormulaCapability,
)


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Sales"
    ws["C1"] = "Region"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["C2"] = "North"
    ws["A3"] = "Bob"
    ws["B3"] = 200
    ws["C3"] = "South"
    ws["A4"] = "Charlie"
    ws["B4"] = 150
    ws["C4"] = "North"
    wb.save(file_path)
    wb.close()
    return file_path


class TestBatchExecute:
    def test_merge_cells(self, sample_excel):
        """合并单元格"""
        cap = BatchExecuteCapability()
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "merge", "sheet": "Data", "range": "A1:C1"}
        ])
        assert result["executed"] == 1
        assert result["results"][0]["status"] == "success"

    def test_unmerge_cells(self, sample_excel):
        """取消合并单元格"""
        cap = BatchExecuteCapability()
        cap.execute(None, file=str(sample_excel), operations=[
            {"type": "merge", "sheet": "Data", "range": "A1:C1"}
        ])
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "unmerge", "sheet": "Data", "range": "A1:C1"}
        ])
        assert result["executed"] == 1

    def test_write_cell(self, sample_excel):
        """写入单元格"""
        cap = BatchExecuteCapability()
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "write", "sheet": "Data", "cell": "D1", "value": "New"}
        ])
        assert result["executed"] == 1
        wb = load_workbook(sample_excel)
        assert wb["Data"]["D1"].value == "New"
        wb.close()

    def test_apply_style(self, sample_excel):
        """应用样式"""
        cap = BatchExecuteCapability()
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "style", "sheet": "Data", "cell": "A1",
             "font": {"bold": True}, "fill": {"patternType": "solid", "fgColor": "4472C4"}}
        ])
        assert result["executed"] == 1

    def test_unknown_operation(self, sample_excel):
        """未知操作类型"""
        cap = BatchExecuteCapability()
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "unknown_op", "sheet": "Data"}
        ])
        assert result["results"][0]["status"] == "skipped"

    def test_multiple_operations(self, sample_excel):
        """多个操作批量执行"""
        cap = BatchExecuteCapability()
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "write", "sheet": "Data", "cell": "D1", "value": "X"},
            {"type": "merge", "sheet": "Data", "range": "D2:D3"},
            {"type": "style", "sheet": "Data", "cell": "A1", "font": {"bold": True}},
        ])
        assert result["total"] == 3
        assert result["executed"] == 3

    def test_missing_file(self):
        """文件不存在"""
        cap = BatchExecuteCapability()
        with pytest.raises(Exception):
            cap.execute(None, file="nonexistent.xlsx", operations=[])


class TestBatchTransformExtended:
    def test_fill_formula(self, sample_excel):
        """填充公式"""
        cap = BatchTransformCapability()
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "fill_formula", "sheet": "Data", "cell": "D2", "formula": "=B2*2"}
        ])
        assert result["operations"] == 1
        wb = load_workbook(sample_excel)
        assert wb["Data"]["D2"].value == "=B2*2"
        wb.close()

    def test_fill_value(self, sample_excel):
        """填充值"""
        cap = BatchTransformCapability()
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "fill_value", "sheet": "Data", "cell": "D1", "value": "Test"}
        ])
        assert result["operations"] == 1
        wb = load_workbook(sample_excel)
        assert wb["Data"]["D1"].value == "Test"
        wb.close()

    def test_copy_format(self, sample_excel):
        """复制格式"""
        cap = BatchTransformCapability()
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "copy_format", "sheet": "Data", "source": "A1", "target": "B1"}
        ])
        assert result["operations"] == 1

    def test_clear_content(self, sample_excel):
        """清除内容"""
        cap = BatchTransformCapability()
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "clear_content", "sheet": "Data", "range": "A2:B2"}
        ])
        assert result["operations"] == 1
        wb = load_workbook(sample_excel, data_only=True)
        assert wb["Data"]["A2"].value is None
        assert wb["Data"]["B2"].value is None
        wb.close()

    def test_clear_format(self, sample_excel):
        """清除格式"""
        cap = BatchTransformCapability()
        cap.execute(None, file=str(sample_excel), operations=[
            {"type": "fill_value", "sheet": "Data", "cell": "A1", "value": "X"}
        ])
        result = cap.execute(None, file=str(sample_excel), operations=[
            {"type": "clear_format", "sheet": "Data", "range": "A1:A1"}
        ])
        assert result["operations"] == 1


class TestBatchValidateExtended:
    def test_validate_unique(self, sample_excel):
        """验证唯一性"""
        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(sample_excel), validations=[
            {"sheet": "Data", "range": "A2:A4", "rule": "unique"}
        ])
        assert result["validations"] == 1
        assert result["results"][0]["valid"] is True

    def test_validate_unique_with_duplicates(self, tmp_path):
        """验证唯一性 - 有重复值"""
        file_path = tmp_path / "dup.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Alice"
        ws["A2"] = "Alice"
        ws["A3"] = "Bob"
        wb.save(file_path)
        wb.close()

        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(file_path), validations=[
            {"sheet": ws.title, "range": "A1:A3", "rule": "unique"}
        ])
        assert result["results"][0]["valid"] is False

    def test_validate_range(self, sample_excel):
        """验证数值范围"""
        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(sample_excel), validations=[
            {"sheet": "Data", "range": "B2:B4", "rule": "range", "min": 50, "max": 250}
        ])
        assert result["validations"] == 1
        assert result["results"][0]["valid"] is True

    def test_validate_range_out_of_bounds(self, tmp_path):
        """验证数值范围 - 超出范围"""
        file_path = tmp_path / "range.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = 200
        ws["A3"] = 300
        wb.save(file_path)
        wb.close()

        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(file_path), validations=[
            {"sheet": ws.title, "range": "A1:A3", "rule": "range", "min": 0, "max": 100}
        ])
        assert result["results"][0]["valid"] is False

    def test_validate_pattern(self, tmp_path):
        """验证文本模式"""
        file_path = tmp_path / "pattern.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "ABC-123"
        ws["A2"] = "XYZ-456"
        ws["A3"] = "invalid"
        wb.save(file_path)
        wb.close()

        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(file_path), validations=[
            {"sheet": ws.title, "range": "A1:A3", "rule": "pattern", "pattern": r"^[A-Z]{3}-\d{3}$"}
        ])
        assert result["results"][0]["valid"] is False

    def test_validate_date(self, tmp_path):
        """验证日期格式"""
        file_path = tmp_path / "date.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "2024-01-15"
        ws["A2"] = "2024/02/20"
        wb.save(file_path)
        wb.close()

        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(file_path), validations=[
            {"sheet": ws.title, "range": "A1:A2", "rule": "date"}
        ])
        assert result["results"][0]["valid"] is False

    def test_validate_email(self, tmp_path):
        """验证邮箱格式"""
        file_path = tmp_path / "email.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "test@example.com"
        ws["A2"] = "not-an-email"
        wb.save(file_path)
        wb.close()

        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(file_path), validations=[
            {"sheet": ws.title, "range": "A1:A2", "rule": "email"}
        ])
        assert result["results"][0]["valid"] is False

    def test_validate_numeric(self, tmp_path):
        """验证数值类型"""
        file_path = tmp_path / "numeric.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 100
        ws["A2"] = "text"
        wb.save(file_path)
        wb.close()

        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(file_path), validations=[
            {"sheet": ws.title, "range": "A1:A2", "rule": "numeric"}
        ])
        assert result["results"][0]["valid"] is False

    def test_validate_no_empty_with_empty(self, tmp_path):
        """验证无空值 - 有空值"""
        file_path = tmp_path / "empty.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Value"
        ws["A2"] = None
        wb.save(file_path)
        wb.close()

        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(file_path), validations=[
            {"sheet": ws.title, "range": "A1:A2", "rule": "no_empty"}
        ])
        assert result["results"][0]["valid"] is False

    def test_validate_sheet_not_found(self, sample_excel):
        """验证时工作表不存在"""
        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(sample_excel), validations=[
            {"sheet": "Nonexistent", "range": "A1:A2", "rule": "no_empty"}
        ])
        assert result["results"][0]["valid"] is False


class TestCreatePivotExtended:
    def test_pivot_avg(self, sample_excel):
        """透视表 - 平均值"""
        cap = CreatePivotCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:C4", row_fields=["Region"],
                           value_field="Sales", agg_function="avg")
        assert result["rows"] == 2

    def test_pivot_count(self, sample_excel):
        """透视表 - 计数"""
        cap = CreatePivotCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:C4", row_fields=["Region"],
                           value_field="Sales", agg_function="count")
        assert result["rows"] == 2

    def test_pivot_min(self, sample_excel):
        """透视表 - 最小值"""
        cap = CreatePivotCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:C4", row_fields=["Region"],
                           value_field="Sales", agg_function="min")
        assert result["rows"] == 2

    def test_pivot_max(self, sample_excel):
        """透视表 - 最大值"""
        cap = CreatePivotCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:C4", row_fields=["Region"],
                           value_field="Sales", agg_function="max")
        assert result["rows"] == 2


class TestFormatRangeExtended:
    def test_format_border(self, sample_excel):
        """格式化 - 边框"""
        cap = FormatRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:C1",
                           border={"left": {"style": "thin", "color": "000000"},
                                   "right": {"style": "thin", "color": "000000"}})
        assert result["cells_formatted"] == 3

    def test_format_alignment(self, sample_excel):
        """格式化 - 对齐"""
        cap = FormatRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:C1",
                           alignment={"horizontal": "center", "vertical": "center", "wrap_text": True})
        assert result["cells_formatted"] == 3

    def test_format_number(self, sample_excel):
        """格式化 - 数字格式"""
        cap = FormatRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B2:B4", number_format="#,##0.00")
        assert result["cells_formatted"] == 3

    def test_format_conditional_color_scale(self, sample_excel):
        """格式化 - 条件格式（色阶）"""
        cap = FormatRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B2:B4",
                           conditional={"type": "color_scale"})
        assert result["has_conditional"] is True

    def test_format_conditional_data_bar(self, sample_excel):
        """格式化 - 条件格式（数据条）"""
        cap = FormatRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B2:B4",
                           conditional={"type": "data_bar"})
        assert result["has_conditional"] is True

    def test_format_conditional_cell_is(self, sample_excel):
        """格式化 - 条件格式（单元格值）"""
        cap = FormatRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="B2:B4",
                           conditional={"type": "cell_is", "operator": "greaterThan",
                                       "formula": ["100"], "fill_color": "FF0000"})
        assert result["has_conditional"] is True

    def test_format_font_full(self, sample_excel):
        """格式化 - 完整字体设置"""
        cap = FormatRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:A1",
                           font={"name": "Arial", "size": 14, "bold": True, "italic": True, "color": "FF0000"})
        assert result["cells_formatted"] == 1


class TestChartExtended:
    def test_create_line_chart(self, sample_excel):
        """创建折线图"""
        cap = CreateChartCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B4", chart_type="line", title="Line Chart")
        assert result["chart_type"] == "line"

    def test_create_pie_chart(self, sample_excel):
        """创建饼图"""
        cap = CreateChartCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B4", chart_type="pie", title="Pie Chart")
        assert result["chart_type"] == "pie"

    def test_create_area_chart(self, sample_excel):
        """创建面积图"""
        cap = CreateChartCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B4", chart_type="area", title="Area Chart")
        assert result["chart_type"] == "area"

    def test_create_scatter_chart(self, sample_excel):
        """创建散点图"""
        cap = CreateChartCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B4", chart_type="scatter", title="Scatter Chart")
        assert result["chart_type"] == "scatter"

    def test_create_chart_with_axes(self, sample_excel):
        """创建图表 - 带坐标轴标题"""
        cap = CreateChartCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B4", chart_type="bar",
                           x_axis="Name", y_axis="Sales")
        assert result["chart_type"] == "bar"

    def test_create_chart_output_sheet(self, sample_excel):
        """创建图表 - 输出到指定工作表"""
        cap = CreateChartCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B4", chart_type="bar", output_sheet="Charts")
        assert result["output_sheet"] == "Charts"
        wb = load_workbook(sample_excel)
        assert "Charts" in wb.sheetnames
        wb.close()

    def test_update_chart_title(self, sample_excel):
        """更新图表标题"""
        create_cap = CreateChartCapability()
        create_cap.execute(None, file=str(sample_excel), sheet="Data",
                          range="A1:B4", chart_type="bar")

        update_cap = UpdateChartCapability()
        result = update_cap.execute(None, file=str(sample_excel), sheet="Data",
                                   chart_index=0, title="Updated Title")
        assert result["updated"] is True

    def test_delete_chart(self, sample_excel):
        """删除图表"""
        create_cap = CreateChartCapability()
        create_cap.execute(None, file=str(sample_excel), sheet="Data",
                          range="A1:B4", chart_type="bar")

        delete_cap = DeleteChartCapability()
        result = delete_cap.execute(None, file=str(sample_excel), sheet="Data",
                                   chart_index=0)
        assert result["deleted"] is True

        list_cap = ListChartsCapability()
        result = list_cap.execute(None, file=str(sample_excel))
        assert result["total"] == 0

    def test_list_charts_after_create(self, sample_excel):
        """创建后列出图表"""
        create_cap = CreateChartCapability()
        create_cap.execute(None, file=str(sample_excel), sheet="Data",
                          range="A1:B4", chart_type="bar", title="Test")

        list_cap = ListChartsCapability()
        result = list_cap.execute(None, file=str(sample_excel), sheet="Data")
        assert result["total"] == 1
        assert result["charts"][0]["type"] == "BarChart"

    def test_chart_missing_file(self):
        """图表 - 文件不存在"""
        cap = CreateChartCapability()
        with pytest.raises(Exception):
            cap.execute(None, file="nonexistent.xlsx", sheet="Data",
                       range="A1:B4", chart_type="bar")

    def test_chart_invalid_type(self, sample_excel):
        """图表 - 无效类型"""
        cap = CreateChartCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=str(sample_excel), sheet="Data",
                       range="A1:B4", chart_type="invalid")


class TestSetArrayFormula:
    def test_set_array_formula(self, sample_excel):
        """设置数组公式"""
        cap = SetArrayFormulaCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="D2:D4", formula="=B2:B4*2")
        assert result["success"] is True
        assert result["formula"] == "=B2:B4*2"

    def test_array_formula_missing_params(self, sample_excel):
        """数组公式 - 缺少参数"""
        cap = SetArrayFormulaCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=str(sample_excel))

    def test_array_formula_file_not_found(self):
        """数组公式 - 文件不存在"""
        cap = SetArrayFormulaCapability()
        with pytest.raises(Exception):
            cap.execute(None, file="nonexistent.xlsx", sheet="Data",
                       range="A1:A1", formula="=1")
