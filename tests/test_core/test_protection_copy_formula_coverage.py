"""测试 protection.py, copy.py, create_formula.py, validate_formula.py, quality_check.py, lint.py, validate_type.py"""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from abacus.core.work.protection import (
    ProtectWorkbookCapability, ProtectSheetCapability,
    UnprotectSheetCapability, SetArrayFormulaCapability
)
from abacus.core.range.copy import CopyRangeCapability
from abacus.core.equation.create_formula import CreateFormulaCapability
from abacus.core.balance.validate_formula import ValidateFormulaCapability
from abacus.core.balance.quality_check import QualityCheckCapability
from abacus.core.balance.lint import ExcelLintCapability
from abacus.core.balance.validate_type import ValidateTypeCapability
from abacus.core.exceptions import DataError, FileNotFoundError, ValidationError


@pytest.fixture
def test_excel(tmp_path):
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Value", "Rate"])
    ws.append(["Alice", 100, 0.5])
    ws.append(["Bob", 200, 0.8])
    ws.append(["Charlie", 300, 0.3])
    wb.save(file_path)
    wb.close()
    return file_path


class TestProtectWorkbook:
    def test_missing_file(self):
        cap = ProtectWorkbookCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None)

    def test_file_not_found(self, tmp_path):
        cap = ProtectWorkbookCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"))

    def test_protect_workbook(self, test_excel):
        cap = ProtectWorkbookCapability()
        result = cap.execute(None, file=str(test_excel))
        assert result["success"] is True
        assert result["action"] == "protect_workbook"

    def test_protect_workbook_with_password(self, test_excel):
        cap = ProtectWorkbookCapability()
        result = cap.execute(None, file=str(test_excel), password="test123")
        assert result["success"] is True


class TestProtectSheet:
    def test_missing_file(self):
        cap = ProtectSheetCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Sheet1")

    def test_missing_sheet(self, test_excel):
        cap = ProtectSheetCapability()
        with pytest.raises(DataError, match="sheet parameter is required"):
            cap.execute(None, file=str(test_excel))

    def test_file_not_found(self, tmp_path):
        cap = ProtectSheetCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"), sheet="Sheet1")

    def test_sheet_not_found(self, test_excel):
        cap = ProtectSheetCapability()
        with pytest.raises(DataError, match="not found"):
            cap.execute(None, file=str(test_excel), sheet="NoSuchSheet")

    def test_protect_sheet(self, test_excel):
        cap = ProtectSheetCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1")
        assert result["success"] is True
        assert result["action"] == "protect_sheet"

    def test_protect_sheet_with_password(self, test_excel):
        cap = ProtectSheetCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1", password="pwd")
        assert result["success"] is True


class TestUnprotectSheet:
    def test_missing_file(self):
        cap = UnprotectSheetCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Sheet1")

    def test_missing_sheet(self, test_excel):
        cap = UnprotectSheetCapability()
        with pytest.raises(DataError, match="sheet parameter is required"):
            cap.execute(None, file=str(test_excel))

    def test_file_not_found(self, tmp_path):
        cap = UnprotectSheetCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"), sheet="Sheet1")

    def test_sheet_not_found(self, test_excel):
        cap = UnprotectSheetCapability()
        with pytest.raises(DataError, match="not found"):
            cap.execute(None, file=str(test_excel), sheet="NoSuchSheet")

    def test_unprotect_sheet(self, test_excel):
        cap = UnprotectSheetCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1")
        assert result["success"] is True
        assert result["action"] == "unprotect_sheet"


class TestSetArrayFormula:
    def test_missing_file(self):
        cap = SetArrayFormulaCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Sheet1", range="A1:A3", formula="=SUM(B1:B3)")

    def test_missing_sheet(self, test_excel):
        cap = SetArrayFormulaCapability()
        with pytest.raises(DataError, match="sheet parameter is required"):
            cap.execute(None, file=str(test_excel), range="A1:A3", formula="=SUM(B1:B3)")

    def test_missing_range(self, test_excel):
        cap = SetArrayFormulaCapability()
        with pytest.raises(DataError, match="range parameter is required"):
            cap.execute(None, file=str(test_excel), sheet="Sheet1", formula="=SUM(B1:B3)")

    def test_missing_formula(self, test_excel):
        cap = SetArrayFormulaCapability()
        with pytest.raises(DataError, match="formula parameter is required"):
            cap.execute(None, file=str(test_excel), sheet="Sheet1", range="A1:A3")

    def test_file_not_found(self, tmp_path):
        cap = SetArrayFormulaCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"), sheet="Sheet1",
                       range="A1:A3", formula="=SUM(B1:B3)")

    def test_sheet_not_found(self, test_excel):
        cap = SetArrayFormulaCapability()
        with pytest.raises(DataError, match="not found"):
            cap.execute(None, file=str(test_excel), sheet="NoSuchSheet",
                       range="A1:A3", formula="=SUM(B1:B3)")

    def test_set_array_formula(self, test_excel):
        cap = SetArrayFormulaCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1",
                           range="A1:A3", formula="=SUM(B1:B3)")
        assert result["success"] is True
        assert result["action"] == "set_array_formula"


class TestCopyRange:
    def test_missing_file(self):
        cap = CopyRangeCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Sheet1", source="A1:B3", target="D1")

    def test_file_not_found(self, tmp_path):
        cap = CopyRangeCapability()
        with pytest.raises(Exception):
            cap.execute(None, file=str(tmp_path / "no.xlsx"), sheet="Sheet1",
                       source="A1:B3", target="D1")

    def test_copy_values(self, test_excel):
        cap = CopyRangeCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1",
                           source="A1:B2", target="E1", copy_type="values")
        assert result["copy_type"] == "values"
        assert result["rows"] == 2
        assert result["columns"] == 2

    def test_copy_formulas(self, test_excel):
        cap = CopyRangeCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1",
                           source="A1:B2", target="E1", copy_type="formulas")
        assert result["copy_type"] == "formulas"

    def test_copy_formats(self, test_excel):
        cap = CopyRangeCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1",
                           source="A1:B2", target="E1", copy_type="formats")
        assert result["copy_type"] == "formats"

    def test_copy_all(self, test_excel):
        cap = CopyRangeCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1",
                           source="A1:B2", target="E1", copy_type="all")
        assert result["copy_type"] == "all"

    def test_copy_with_font_formatting(self, tmp_path):
        file_path = tmp_path / "format_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Bold"
        ws["A1"].font = ws["A1"].font.copy(bold=True, color="FF0000")
        ws["A1"].alignment = ws["A1"].alignment.copy(horizontal="center", vertical="center")
        ws["A1"].fill = ws["A1"].fill.copy(patternType="solid", fgColor="FFFF00")
        wb.save(file_path)
        wb.close()

        cap = CopyRangeCapability()
        result = cap.execute(None, file=str(file_path), sheet="Sheet",
                           source="A1:A1", target="B1", copy_type="all")
        assert result["rows"] == 1
        assert result["columns"] == 1


class TestCreateFormula:
    def test_missing_file(self):
        cap = CreateFormulaCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 file"):
            cap.execute(None, sheet="Sheet1", cell="E1", formula="SUM(A1:D1)")

    def test_missing_cell(self, test_excel):
        cap = CreateFormulaCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 cell"):
            cap.execute(None, file=str(test_excel), sheet="Sheet1", formula="SUM(A1:D1)")

    def test_missing_formula(self, test_excel):
        cap = CreateFormulaCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 formula"):
            cap.execute(None, file=str(test_excel), sheet="Sheet1", cell="E1")

    def test_formula_without_equals(self, test_excel):
        cap = CreateFormulaCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1",
                           cell="E1", formula="SUM(A1:D1)")
        assert result["created"] is True
        assert result["formula"] == "=SUM(A1:D1)"

    def test_formula_with_equals(self, test_excel):
        cap = CreateFormulaCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1",
                           cell="E1", formula="=SUM(A1:D1)")
        assert result["created"] is True
        assert result["formula"] == "=SUM(A1:D1)"


class TestValidateFormula:
    def test_missing_file(self):
        cap = ValidateFormulaCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 file"):
            cap.execute(None)

    def test_file_not_found(self, tmp_path):
        cap = ValidateFormulaCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"))

    def test_valid_formulas(self, test_excel):
        wb = load_workbook(test_excel)
        ws = wb["Sheet1"]
        ws["D1"] = "=SUM(B2:B4)"
        ws["D2"] = "=AVERAGE(B2:B4)"
        wb.save(test_excel)
        wb.close()

        cap = ValidateFormulaCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1")
        assert result["total_formulas"] >= 2
        assert result["valid"] >= 2

    def test_invalid_formula(self, tmp_path):
        file_path = tmp_path / "formula_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=SUM(B1"  # unclosed parenthesis
        ws["A2"] = "=IF(A1=1，2，3)"  # Chinese comma
        wb.save(file_path)
        wb.close()

        cap = ValidateFormulaCapability()
        result = cap.execute(None, file=str(file_path))
        assert result["invalid"] >= 2

    def test_validate_specific_cell(self, test_excel):
        wb = load_workbook(test_excel)
        ws = wb["Sheet1"]
        ws["D1"] = "=SUM(B2:B4)"
        wb.save(test_excel)
        wb.close()

        cap = ValidateFormulaCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1", cell="D1")
        assert result["total_formulas"] == 1

    def test_unsafe_formula(self, tmp_path):
        file_path = tmp_path / "unsafe.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = '=INDIRECT("A2")'
        wb.save(file_path)
        wb.close()

        cap = ValidateFormulaCapability()
        result = cap.execute(None, file=str(file_path))
        assert result["invalid"] >= 1


class TestQualityCheck:
    def test_missing_file(self):
        cap = QualityCheckCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 file"):
            cap.execute(None, sheet="Sheet1")

    def test_missing_sheet(self, test_excel):
        cap = QualityCheckCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 sheet"):
            cap.execute(None, file=str(test_excel))

    def test_file_not_found(self, tmp_path):
        cap = QualityCheckCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"), sheet="Sheet1")

    def test_sheet_not_found(self, test_excel):
        cap = QualityCheckCapability()
        with pytest.raises(DataError, match="不存在"):
            cap.execute(None, file=str(test_excel), sheet="NoSuchSheet")

    def test_quality_check_clean_data(self, test_excel):
        cap = QualityCheckCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1")
        assert result["success"] is True
        assert result["quality_score"] == 100
        assert result["issue_count"] == 0

    def test_quality_check_with_nulls(self, tmp_path):
        file_path = tmp_path / "null_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Value"])
        ws.append(["Alice", 100])
        ws.append([None, None])
        ws.append(["Charlie", 300])
        wb.save(file_path)
        wb.close()

        cap = QualityCheckCapability()
        result = cap.execute(None, file=str(file_path), sheet="Sheet")
        assert result["quality_score"] < 100
        assert result["issue_count"] > 0

    def test_quality_check_with_duplicates(self, tmp_path):
        file_path = tmp_path / "dup_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Value"])
        ws.append(["Alice", 100])
        ws.append(["Alice", 100])
        ws.append(["Bob", 200])
        wb.save(file_path)
        wb.close()

        cap = QualityCheckCapability()
        result = cap.execute(None, file=str(file_path), sheet="Sheet")
        # The quality check detects duplicates based on row equality
        assert result["success"] is True


class TestExcelLint:
    def test_missing_params(self):
        cap = ExcelLintCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 code 或 file"):
            cap.execute(None)

    def test_file_not_found(self, tmp_path):
        cap = ExcelLintCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.py"))

    def test_lint_code_with_issues(self):
        cap = ExcelLintCapability()
        code = '''
import openpyxl
wb = openpyxl.load_workbook("test.xlsx", data_only=True)
wb.save("output.xlsx")
'''
        result = cap.execute(None, code=code)
        assert result["total_issues"] >= 1

    def test_lint_code_clean(self):
        cap = ExcelLintCapability()
        code = '''
import openpyxl
wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active
ws["A1"] = "Hello"
wb.save("output.xlsx")
'''
        result = cap.execute(None, code=code)
        assert result["total_issues"] == 0

    def test_lint_code_chinese_comma(self):
        cap = ExcelLintCapability()
        code = 'formula = "=SUM(1，2)"'
        result = cap.execute(None, code=code)
        assert result["total_issues"] >= 1

    def test_lint_code_sheet_forbidden_char(self):
        cap = ExcelLintCapability()
        code = 'ws.title = "My:Sheet"'
        result = cap.execute(None, code=code)
        assert result["total_issues"] >= 1

    def test_lint_code_pattern_fill_no_type(self):
        cap = ExcelLintCapability()
        code = 'fill = PatternFill()'
        result = cap.execute(None, code=code)
        assert result["total_issues"] >= 1

    def test_lint_code_read_only_save(self):
        cap = ExcelLintCapability()
        code = '''
wb = openpyxl.load_workbook("test.xlsx", read_only=True)
wb.save("output.xlsx")
'''
        result = cap.execute(None, code=code)
        assert result["total_issues"] >= 1

    def test_lint_from_file(self, tmp_path):
        code_file = tmp_path / "test_lint.py"
        code_file.write_text('import openpyxl\nwb = openpyxl.load_workbook("t.xlsx", data_only=True)\nwb.save("o.xlsx")\n')
        cap = ExcelLintCapability()
        result = cap.execute(None, file=str(code_file))
        assert result["total_issues"] >= 1


class TestValidateType:
    def test_missing_file(self):
        cap = ValidateTypeCapability()
        with pytest.raises(ValidationError, match="缺少必要参数 file"):
            cap.execute(None, sheet="Sheet1", range="A1:A10")

    def test_file_not_found(self, tmp_path):
        cap = ValidateTypeCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(None, file=str(tmp_path / "no.xlsx"), sheet="Sheet1", range="A1:A10")

    def test_sheet_not_found(self, test_excel):
        cap = ValidateTypeCapability()
        with pytest.raises(DataError, match="不存在"):
            cap.execute(None, file=str(test_excel), sheet="NoSuchSheet", range="A1:A10")

    def test_validate_type_number(self, test_excel):
        cap = ValidateTypeCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1",
                           range="A1:C4", expected_type="number")
        assert result["columns_analyzed"] == 3
        assert any(i["type"] == "unexpected_type" for i in result["issues"])

    def test_validate_type_text(self, test_excel):
        cap = ValidateTypeCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1",
                           range="A1:A4", expected_type="text")
        assert result["valid"] is True

    def test_validate_type_mixed(self, tmp_path):
        file_path = tmp_path / "mixed.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Mixed"])
        ws.append([1])
        ws.append(["two"])
        ws.append([3])
        wb.save(file_path)
        wb.close()

        cap = ValidateTypeCapability()
        result = cap.execute(None, file=str(file_path), sheet="Sheet",
                           range="A1:A4")
        assert any(i["type"] == "mixed_types" for i in result["issues"])

    def test_validate_type_no_expected(self, test_excel):
        cap = ValidateTypeCapability()
        result = cap.execute(None, file=str(test_excel), sheet="Sheet1", range="A1:C4")
        assert result["columns_analyzed"] == 3
