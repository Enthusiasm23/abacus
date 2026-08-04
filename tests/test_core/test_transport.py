"""测试均输章（导入导出）能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.transport import ImportDataCapability, ExportDataCapability, MigrateCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["A3"] = "Bob"
    ws["B3"] = 200
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def sample_csv(tmp_path):
    """创建测试用 CSV 文件"""
    file_path = tmp_path / "test.csv"
    file_path.write_text("Name,Value\nAlice,100\nBob,200\n", encoding="utf-8")
    return file_path


class TestExportData:
    def test_export_to_csv(self, sample_excel, tmp_path):
        """导出为 CSV"""
        cap = ExportDataCapability()
        output = tmp_path / "output.csv"
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B3", output=str(output), format="csv")
        assert result["rows_exported"] == 2
        assert output.exists()
    
    def test_export_to_json(self, sample_excel, tmp_path):
        """导出为 JSON"""
        cap = ExportDataCapability()
        output = tmp_path / "output.json"
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B3", output=str(output), format="json")
        assert result["rows_exported"] == 2
        assert output.exists()


class TestImportData:
    def test_import_csv(self, sample_csv, tmp_path):
        """导入 CSV"""
        cap = ImportDataCapability()
        output = tmp_path / "imported.xlsx"
        result = cap.execute(None, file=str(output), source=str(sample_csv),
                           source_type="csv", sheet="Imported")
        assert result["rows_imported"] == 3  # header + 2 data rows


class TestMigrate:
    def test_migrate_sheets(self, sample_excel, tmp_path):
        """迁移工作表"""
        cap = MigrateCapability()
        target = tmp_path / "migrated.xlsx"
        result = cap.execute(None, source=str(sample_excel), target=str(target))
        assert result["sheets_migrated"] == 1
