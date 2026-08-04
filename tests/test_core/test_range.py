"""测试 Range 扩展能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.range.clear import ClearRangeCapability
from abacus.core.range.copy import CopyRangeCapability
from abacus.core.range.find import FindReplaceCapability
from abacus.core.range.hyperlink import HyperlinkCapability
from abacus.core.range.lock import CellLockCapability
from abacus.core.range.size import ColumnRowSizeCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["A3"] = "Bob"
    ws["B3"] = 200
    ws["A4"] = "Charlie"
    ws["B4"] = 300
    wb.save(file_path)
    wb.close()
    return file_path


class TestClearRange:
    def test_clear_all(self, sample_excel):
        """清除所有内容"""
        cap = ClearRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1", 
                           range="A1:B2", clear_type="all")
        assert result["cells_cleared"] == 4
    
    def test_clear_contents(self, sample_excel):
        """仅清除内容"""
        cap = ClearRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           range="A1:B2", clear_type="contents")
        assert result["cells_cleared"] == 4
    
    def test_clear_formats(self, sample_excel):
        """仅清除格式"""
        cap = ClearRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           range="A1:B2", clear_type="formats")
        assert result["cells_cleared"] == 4


class TestCopyRange:
    def test_copy_all(self, sample_excel):
        """复制所有内容"""
        cap = CopyRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           source="A1:B2", target="D1", copy_type="all")
        assert result["rows"] == 2
        assert result["columns"] == 2
    
    def test_copy_values(self, sample_excel):
        """仅复制值"""
        cap = CopyRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           source="A1:B2", target="D1", copy_type="values")
        assert result["rows"] == 2


class TestFindReplace:
    def test_find(self, sample_excel):
        """查找文本"""
        cap = FindReplaceCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           find="Alice")
        assert result["found_count"] == 1
    
    def test_replace(self, sample_excel):
        """替换文本"""
        cap = FindReplaceCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           find="Alice", replace="Eve")
        assert result["replaced_count"] == 1
    
    def test_find_not_found(self, sample_excel):
        """查找不存在的文本"""
        cap = FindReplaceCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           find="NotExist")
        assert result["found_count"] == 0


class TestHyperlink:
    def test_add_hyperlink(self, sample_excel):
        """添加超链接"""
        cap = HyperlinkCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           action="add", cell="A1", url="https://example.com")
        assert result["action"] == "add"
    
    def test_list_hyperlinks(self, sample_excel):
        """列出超链接"""
        cap = HyperlinkCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           action="list")
        assert result["action"] == "list"
        assert isinstance(result["hyperlinks"], list)


class TestCellLock:
    def test_lock_cells(self, sample_excel):
        """锁定单元格"""
        cap = CellLockCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           range="A1:B2", locked=True)
        assert result["cells_updated"] == 4
    
    def test_unlock_cells(self, sample_excel):
        """解锁单元格"""
        cap = CellLockCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           range="A1:B2", locked=False)
        assert result["cells_updated"] == 4


class TestColumnRowSize:
    def test_set_column_width(self, sample_excel):
        """设置列宽"""
        cap = ColumnRowSizeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           action="set", dimension="column", index=1, size=20)
        assert result["action"] == "set"
    
    def test_get_column_width(self, sample_excel):
        """获取列宽"""
        cap = ColumnRowSizeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           action="get", dimension="column", index=1)
        assert result["action"] == "get"
    
    def test_set_row_height(self, sample_excel):
        """设置行高"""
        cap = ColumnRowSizeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           action="set", dimension="row", index=1, size=30)
        assert result["action"] == "set"
