"""测试新增能力 - 模糊匹配、质量检测、关联、批量合并、类型推断、标准化、摘要报告、变化检测、数据视图"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from abacus.core.grain import FuzzyMatchCapability, AutoTypeInferCapability, StandardizeCapability
from abacus.core.balance import QualityCheckCapability
from abacus.core.transport import JoinTablesCapability, BatchMergeCapability
from abacus.core.work import SummaryReportCapability, DiffReportCapability, DataViewCapability


@pytest.fixture
def sample_excel():
    """创建示例 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 添加表头
    ws.append(["ID", "Name", "Sales", "Profit", "Date"])
    
    # 添加数据
    ws.append([1, "Product A", 1000, 200, "2024-01-15"])
    ws.append([2, "Product B", 1500, 300, "2024-01-16"])
    ws.append([3, "Product A", 800, 160, "2024-01-17"])
    ws.append([4, "Product C", 2000, 400, "2024-01-18"])
    ws.append([5, "Product B", 1200, 240, "2024-01-19"])
    
    # 保存到临时文件
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        wb.save(f.name)
        return f.name


@pytest.fixture
def sample_excel_with_nulls():
    """创建包含空值的示例 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 添加表头
    ws.append(["ID", "Name", "Sales", "Profit"])
    
    # 添加数据（包含空值）
    ws.append([1, "Product A", 1000, 200])
    ws.append([2, None, 1500, None])
    ws.append([3, "Product A", None, 160])
    ws.append([4, "Product C", 2000, 400])
    ws.append([5, None, None, 240])
    
    # 保存到临时文件
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        wb.save(f.name)
        return f.name


@pytest.fixture
def sample_excel_for_join():
    """创建用于关联的示例 Excel 文件"""
    # 左表
    wb_left = Workbook()
    ws_left = wb_left.active
    ws_left.title = "Sheet1"
    ws_left.append(["ID", "Name", "Sales"])
    ws_left.append([1, "Product A", 1000])
    ws_left.append([2, "Product B", 1500])
    ws_left.append([3, "Product C", 2000])
    
    left_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    wb_left.save(left_file)
    
    # 右表
    wb_right = Workbook()
    ws_right = wb_right.active
    ws_right.title = "Sheet1"
    ws_right.append(["ID", "Region", "Manager"])
    ws_right.append([1, "East", "Alice"])
    ws_right.append([2, "West", "Bob"])
    ws_right.append([4, "North", "Charlie"])
    
    right_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    wb_right.save(right_file)
    
    return left_file, right_file


@pytest.fixture
def sample_excel_for_diff():
    """创建用于变化检测的示例 Excel 文件"""
    # 旧版本
    wb_old = Workbook()
    ws_old = wb_old.active
    ws_old.title = "Sheet1"
    ws_old.append(["ID", "Name", "Sales"])
    ws_old.append([1, "Product A", 1000])
    ws_old.append([2, "Product B", 1500])
    ws_old.append([3, "Product C", 2000])
    
    old_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    wb_old.save(old_file)
    
    # 新版本
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Sheet1"
    ws_new.append(["ID", "Name", "Sales"])
    ws_new.append([1, "Product A", 1200])
    ws_new.append([2, "Product B", 1500])
    ws_new.append([4, "Product D", 2500])
    
    new_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    wb_new.save(new_file)
    
    return old_file, new_file


class TestFuzzyMatch:
    """测试模糊匹配"""
    
    def test_fuzzy_match_basic(self, sample_excel):
        """测试基本模糊匹配"""
        cap = FuzzyMatchCapability()
        result = cap.execute(None, file=sample_excel, sheet="Sheet1", target_columns=["Sales", "Profit"])
        
        assert result["success"] is True
        assert "matches" in result
        assert result["match_count"] >= 1
        
        # 清理临时文件
        os.unlink(sample_excel)
    
    def test_fuzzy_match_with_threshold(self, sample_excel):
        """测试带阈值的模糊匹配"""
        cap = FuzzyMatchCapability()
        result = cap.execute(None, file=sample_excel, sheet="Sheet1", target_columns=["Sales"], threshold=0.8)
        
        assert result["success"] is True
        
        # 清理临时文件
        os.unlink(sample_excel)


class TestQualityCheck:
    """测试质量检测"""
    
    def test_quality_check_basic(self, sample_excel):
        """测试基本质量检测"""
        cap = QualityCheckCapability()
        result = cap.execute(None, file=sample_excel, sheet="Sheet1")
        
        assert result["success"] is True
        assert "quality_score" in result
        assert "issues" in result
        assert result["total_rows"] == 5
        assert result["total_columns"] == 5
        
        # 清理临时文件
        os.unlink(sample_excel)
    
    def test_quality_check_with_nulls(self, sample_excel_with_nulls):
        """测试包含空值的质量检测"""
        cap = QualityCheckCapability()
        result = cap.execute(None, file=sample_excel_with_nulls, sheet="Sheet1")
        
        assert result["success"] is True
        assert result["quality_score"] < 100  # 有空值，质量分数应该降低
        assert result["issue_count"] > 0
        
        # 清理临时文件
        os.unlink(sample_excel_with_nulls)


class TestJoinTables:
    """测试关联表"""
    
    def test_join_tables_inner(self, sample_excel_for_join):
        """测试 INNER JOIN"""
        left_file, right_file = sample_excel_for_join
        
        cap = JoinTablesCapability()
        result = cap.execute(None, left_file=left_file, left_sheet="Sheet1", 
                           right_file=right_file, right_sheet="Sheet1", 
                           on=["ID"], how="inner")
        
        assert result["success"] is True
        assert result["result_rows"] == 2  # 只有 ID 1 和 2 匹配
        
        # 清理临时文件
        os.unlink(left_file)
        os.unlink(right_file)
    
    def test_join_tables_left(self, sample_excel_for_join):
        """测试 LEFT JOIN"""
        left_file, right_file = sample_excel_for_join
        
        cap = JoinTablesCapability()
        result = cap.execute(None, left_file=left_file, left_sheet="Sheet1", 
                           right_file=right_file, right_sheet="Sheet1", 
                           on=["ID"], how="left")
        
        assert result["success"] is True
        assert result["result_rows"] == 3  # 左表所有行
        
        # 清理临时文件
        os.unlink(left_file)
        os.unlink(right_file)


class TestBatchMerge:
    """测试批量合并"""
    
    def test_batch_merge_basic(self, sample_excel):
        """测试基本批量合并"""
        # 创建临时目录
        with tempfile.TemporaryDirectory() as tmpdir:
            # 复制文件到临时目录
            import shutil
            for i in range(3):
                shutil.copy(sample_excel, os.path.join(tmpdir, f"file{i}.xlsx"))
            
            cap = BatchMergeCapability()
            output_file = os.path.join(tmpdir, "merged.xlsx")
            result = cap.execute(None, folder=tmpdir, pattern="*.xlsx", output=output_file)
            
            assert result["success"] is True
            assert result["file_count"] == 3
            assert result["total_rows"] == 15  # 3 文件 * 5 行
        
        # 清理临时文件
        os.unlink(sample_excel)


class TestAutoTypeInfer:
    """测试自动类型推断"""
    
    def test_auto_type_infer_basic(self, sample_excel):
        """测试基本类型推断"""
        cap = AutoTypeInferCapability()
        result = cap.execute(None, file=sample_excel, sheet="Sheet1")
        
        assert result["success"] is True
        assert "inferred_types" in result
        assert "conversions" in result
        
        # 清理临时文件
        os.unlink(sample_excel)


class TestStandardize:
    """测试数据标准化"""
    
    def test_standardize_basic(self, sample_excel):
        """测试基本标准化"""
        cap = StandardizeCapability()
        result = cap.execute(None, file=sample_excel, sheet="Sheet1", text_case="lower")
        
        assert result["success"] is True
        assert result["operation_count"] > 0
        
        # 清理临时文件
        os.unlink(sample_excel)


class TestSummaryReport:
    """测试摘要报告"""
    
    def test_summary_report_basic(self, sample_excel):
        """测试基本摘要报告"""
        cap = SummaryReportCapability()
        result = cap.execute(None, file=sample_excel, sheet="Sheet1")
        
        assert result["success"] is True
        assert "total_rows" in result
        assert "total_columns" in result
        assert result["total_rows"] == 5
        assert result["total_columns"] == 5
        
        # 清理临时文件
        os.unlink(sample_excel)


class TestDiffReport:
    """测试变化检测"""
    
    def test_diff_report_basic(self, sample_excel_for_diff):
        """测试基本变化检测"""
        old_file, new_file = sample_excel_for_diff
        
        cap = DiffReportCapability()
        result = cap.execute(None, old_file=old_file, old_sheet="Sheet1", 
                           new_file=new_file, new_sheet="Sheet1")
        
        assert result["success"] is True
        assert "old_rows" in result
        assert "new_rows" in result
        assert result["old_rows"] == 3
        assert result["new_rows"] == 3
        
        # 清理临时文件
        os.unlink(old_file)
        os.unlink(new_file)


class TestDataView:
    """测试数据视图"""
    
    def test_data_view_create(self, sample_excel):
        """测试创建数据视图"""
        cap = DataViewCapability()
        result = cap.execute(None, file=sample_excel, sheet="Sheet1", 
                           action="create", view_name="sales_view", 
                           columns=["ID", "Name", "Sales"])
        
        assert result["success"] is True
        assert result["view_name"] == "sales_view"
        
        # 清理临时文件
        os.unlink(sample_excel)
    
    def test_data_view_list(self, sample_excel):
        """测试列出数据视图"""
        # 先创建视图
        cap = DataViewCapability()
        cap.execute(None, file=sample_excel, sheet="Sheet1", 
                   action="create", view_name="sales_view", 
                   columns=["ID", "Name", "Sales"])
        
        # 列出视图
        result = cap.execute(None, file=sample_excel, sheet="Sheet1", action="list")
        
        assert result["success"] is True
        assert "sales_view" in result["views"]
        
        # 清理临时文件
        os.unlink(sample_excel)
