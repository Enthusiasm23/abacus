"""测试商功章扩展能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.work import (
    BatchTransformCapability, BatchValidateCapability,
    CreatePivotCapability, FormatRangeCapability,
    CreateChartCapability, UpdateChartCapability, 
    ListChartsCapability, DeleteChartCapability
)


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Sales"
    ws["C1"] = "Region"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["C2"] = "North"
    ws["A3"] = "Bob"
    ws["B3"] = 200
    ws["C3"] = "South"
    ws["A4"] = "Charlie"
    ws["B4"] = 150
    ws["C4"] = "North"
    wb.save(file_path)
    wb.close()
    return file_path


class TestBatchTransform:
    def test_replace(self, sample_excel):
        """批量替换"""
        cap = BatchTransformCapability()
        result = cap.execute(None, file=str(sample_excel),
                           operations=[{"type": "replace", "sheet": "Data", 
                                       "old": "Alice", "new": "Eve"}])
        assert result["operations"] == 1


class TestBatchValidate:
    def test_validate_no_empty(self, sample_excel):
        """验证无空值"""
        cap = BatchValidateCapability()
        result = cap.execute(None, file=str(sample_excel),
                           validations=[{"sheet": "Data", "range": "A1:C4", 
                                        "rule": "no_empty"}])
        assert result["validations"] == 1


class TestCreatePivot:
    def test_create_pivot(self, sample_excel):
        """创建透视表"""
        cap = CreatePivotCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:C4", row_fields=["Region"],
                           value_field="Sales", agg_function="sum")
        assert result["rows"] == 2


class TestFormatRange:
    def test_format_header(self, sample_excel):
        """格式化表头"""
        cap = FormatRangeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:C1", font={"bold": True, "color": "FFFFFF"},
                           fill={"color": "4472C4"})
        assert result["cells_formatted"] == 3


class TestChart:
    def test_create_chart(self, sample_excel):
        """创建图表"""
        cap = CreateChartCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Data",
                           range="A1:B4", chart_type="bar", title="Sales Chart")
        assert result["chart_type"] == "bar"
    
    def test_list_charts(self, sample_excel):
        """列出图表"""
        # 先创建图表
        create_cap = CreateChartCapability()
        create_cap.execute(None, file=str(sample_excel), sheet="Data",
                         range="A1:B4", chart_type="bar")
        
        # 列出图表
        list_cap = ListChartsCapability()
        result = list_cap.execute(None, file=str(sample_excel))
        assert result["total"] > 0
