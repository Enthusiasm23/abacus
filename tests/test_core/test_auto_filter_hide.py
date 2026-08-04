"""测试商功章 - 自动筛选和隐藏显示能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.work.auto_filter import AutoFilterCapability
from abacus.core.work.hide_show import HideShowCapability
from abacus.core.exceptions import DataError, FileNotFoundError


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Sales"
    ws["C1"] = "Region"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["C2"] = "North"
    ws["A3"] = "Bob"
    ws["B3"] = 200
    ws["C3"] = "South"
    ws["A4"] = "Charlie"
    ws["B4"] = 150
    ws["C4"] = "North"
    wb.save(file_path)
    wb.close()
    return file_path


class TestAutoFilter:
    def test_set_auto_filter(self, sample_excel):
        """设置自动筛选"""
        cap = AutoFilterCapability()
        result = cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="set",
            range="A1:C4",
        )
        assert result["action"] == "set"
        assert result["range"] == "A1:C4"

    def test_get_auto_filter(self, sample_excel):
        """查询自动筛选"""
        cap = AutoFilterCapability()
        # 先设置筛选
        cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="set",
            range="A1:C4",
        )
        # 查询筛选
        result = cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="get",
        )
        assert result["action"] == "get"
        assert result["has_filter"] is True
        assert result["range"] == "A1:C4"

    def test_get_auto_filter_none(self, sample_excel):
        """查询无筛选时返回 None"""
        cap = AutoFilterCapability()
        result = cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="get",
        )
        assert result["action"] == "get"
        assert result["has_filter"] is False
        assert result["range"] is None

    def test_remove_auto_filter(self, sample_excel):
        """删除自动筛选"""
        cap = AutoFilterCapability()
        # 先设置筛选
        cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="set",
            range="A1:C4",
        )
        # 删除筛选
        result = cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="remove",
        )
        assert result["action"] == "remove"
        # 验证已删除
        result = cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="get",
        )
        assert result["has_filter"] is False

    def test_set_auto_filter_missing_range(self, sample_excel):
        """设置筛选时缺少范围参数"""
        cap = AutoFilterCapability()
        with pytest.raises(DataError, match="range required"):
            cap.execute(
                None,
                file=str(sample_excel),
                sheet="Data",
                action="set",
            )

    def test_unknown_action(self, sample_excel):
        """未知操作"""
        cap = AutoFilterCapability()
        with pytest.raises(DataError, match="Unknown action"):
            cap.execute(
                None,
                file=str(sample_excel),
                sheet="Data",
                action="invalid",
            )

    def test_file_not_found(self):
        """文件不存在"""
        cap = AutoFilterCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(
                None,
                file="nonexistent.xlsx",
                sheet="Data",
                action="get",
            )

    def test_sheet_not_found(self, sample_excel):
        """工作表不存在"""
        cap = AutoFilterCapability()
        with pytest.raises(DataError, match="Sheet.*not found"):
            cap.execute(
                None,
                file=str(sample_excel),
                sheet="NonExistent",
                action="get",
            )

    def test_missing_file_param(self):
        """缺少 file 参数"""
        cap = AutoFilterCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Data", action="get")


class TestHideShow:
    def test_hide_row(self, sample_excel):
        """隐藏行"""
        cap = HideShowCapability()
        result = cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="hide",
            dimension="row",
            index=2,
        )
        assert result["action"] == "hide"
        assert result["dimension"] == "row"
        assert result["index"] == 2
        assert result["applied"] is True

    def test_show_row(self, sample_excel):
        """显示行"""
        cap = HideShowCapability()
        # 先隐藏
        cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="hide",
            dimension="row",
            index=2,
        )
        # 再显示
        result = cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="show",
            dimension="row",
            index=2,
        )
        assert result["action"] == "show"
        assert result["applied"] is True

    def test_hide_column(self, sample_excel):
        """隐藏列"""
        cap = HideShowCapability()
        result = cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="hide",
            dimension="column",
            index=2,
        )
        assert result["action"] == "hide"
        assert result["dimension"] == "column"
        assert result["index"] == 2
        assert result["applied"] is True

    def test_show_column(self, sample_excel):
        """显示列"""
        cap = HideShowCapability()
        # 先隐藏
        cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="hide",
            dimension="column",
            index=2,
        )
        # 再显示
        result = cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="show",
            dimension="column",
            index=2,
        )
        assert result["action"] == "show"
        assert result["applied"] is True

    def test_unknown_action(self, sample_excel):
        """未知操作"""
        cap = HideShowCapability()
        with pytest.raises(DataError, match="Unknown action"):
            cap.execute(
                None,
                file=str(sample_excel),
                sheet="Data",
                action="invalid",
                dimension="row",
                index=1,
            )

    def test_unknown_dimension(self, sample_excel):
        """未知维度"""
        cap = HideShowCapability()
        with pytest.raises(DataError, match="Unknown dimension"):
            cap.execute(
                None,
                file=str(sample_excel),
                sheet="Data",
                action="hide",
                dimension="invalid",
                index=1,
            )

    def test_file_not_found(self):
        """文件不存在"""
        cap = HideShowCapability()
        with pytest.raises(FileNotFoundError):
            cap.execute(
                None,
                file="nonexistent.xlsx",
                sheet="Data",
                action="hide",
                dimension="row",
                index=1,
            )

    def test_sheet_not_found(self, sample_excel):
        """工作表不存在"""
        cap = HideShowCapability()
        with pytest.raises(DataError, match="Sheet.*not found"):
            cap.execute(
                None,
                file=str(sample_excel),
                sheet="NonExistent",
                action="hide",
                dimension="row",
                index=1,
            )

    def test_missing_file_param(self):
        """缺少 file 参数"""
        cap = HideShowCapability()
        with pytest.raises(DataError, match="file parameter is required"):
            cap.execute(None, sheet="Data", action="hide", dimension="row", index=1)

    def test_hide_persists_after_reload(self, sample_excel):
        """隐藏设置在重新加载文件后仍然有效"""
        from openpyxl import load_workbook

        cap = HideShowCapability()
        cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="hide",
            dimension="row",
            index=2,
        )

        # 重新加载验证
        wb = load_workbook(sample_excel)
        ws = wb["Data"]
        assert ws.row_dimensions[2].hidden is True
        wb.close()

    def test_hide_column_persists_after_reload(self, sample_excel):
        """隐藏列设置在重新加载文件后仍然有效"""
        from openpyxl import load_workbook

        cap = HideShowCapability()
        cap.execute(
            None,
            file=str(sample_excel),
            sheet="Data",
            action="hide",
            dimension="column",
            index=2,
        )

        # 重新加载验证
        wb = load_workbook(sample_excel)
        ws = wb["Data"]
        assert ws.column_dimensions["B"].hidden is True
        wb.close()
