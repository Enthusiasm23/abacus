"""测试衰分章（分组汇总）能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.share import GroupByCapability, DistributeCapability, SummarizeCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Region"
    ws["B1"] = "Product"
    ws["C1"] = "Amount"
    ws["A2"] = "North"
    ws["B2"] = "A"
    ws["C2"] = 100
    ws["A3"] = "South"
    ws["B3"] = "B"
    ws["C3"] = 200
    ws["A4"] = "North"
    ws["B4"] = "B"
    ws["C4"] = 150
    ws["A5"] = "South"
    ws["B5"] = "A"
    ws["C5"] = 250
    wb.save(file_path)
    wb.close()
    return file_path


class TestGroupBy:
    def test_group_by_region(self, sample_excel):
        """按区域分组"""
        cap = GroupByCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sales",
                           range="A1:C5", group_columns=["Region"])
        assert result["groups_count"] == 2
    
    def test_group_by_product(self, sample_excel):
        """按产品分组"""
        cap = GroupByCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sales",
                           range="A1:C5", group_columns=["Product"])
        assert result["groups_count"] == 2


class TestDistribute:
    def test_equal_distribution(self, sample_excel):
        """等比分配"""
        cap = DistributeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sales",
                           range="A1:C5", total=1000, method="equal")
        assert result["distributed"] is True


class TestSummarize:
    def test_summarize_by_region(self, sample_excel):
        """按区域汇总"""
        cap = SummarizeCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sales",
                           range="A1:C5", group_by="Region",
                           agg_config={"Amount": "sum"})
        assert result["groups_count"] == 2
