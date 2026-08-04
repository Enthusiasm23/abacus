"""测试 Table 和 NamedRange 能力"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from abacus.core.table import TableCapability
from abacus.core.named_range import NamedRangeCapability


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用 Excel 文件"""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["A3"] = "Bob"
    ws["B3"] = 200
    wb.save(file_path)
    wb.close()
    return file_path


class TestTable:
    def test_create_table(self, sample_excel):
        """创建表格"""
        cap = TableCapability()
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           action="create", range="A1:B3", table_name="TestTable")
        assert result["action"] == "create"
        assert result["table_name"] == "TestTable"
    
    def test_list_tables(self, sample_excel):
        """列出表格"""
        cap = TableCapability()
        # 先创建一个表格
        cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                   action="create", range="A1:B3", table_name="TestTable")
        # 列出表格
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           action="list")
        assert result["action"] == "list"
        assert len(result["tables"]) > 0
    
    def test_delete_table(self, sample_excel):
        """删除表格"""
        cap = TableCapability()
        # 先创建一个表格
        cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                   action="create", range="A1:B3", table_name="TestTable")
        # 删除表格
        result = cap.execute(None, file=str(sample_excel), sheet="Sheet1",
                           action="delete", table_name="TestTable")
        assert result["action"] == "delete"


class TestNamedRange:
    def test_create_named_range(self, sample_excel):
        """创建命名范围"""
        cap = NamedRangeCapability()
        result = cap.execute(None, file=str(sample_excel),
                           action="create", name="TestRange", 
                           refers_to="Sheet1!$A$1:$B$3")
        assert result["action"] == "create"
        assert result["name"] == "TestRange"
    
    def test_list_named_ranges(self, sample_excel):
        """列出命名范围"""
        cap = NamedRangeCapability()
        # 先创建一个命名范围
        cap.execute(None, file=str(sample_excel),
                   action="create", name="TestRange",
                   refers_to="Sheet1!$A$1:$B$3")
        # 列出命名范围
        result = cap.execute(None, file=str(sample_excel),
                           action="list")
        assert result["action"] == "list"
        assert len(result["named_ranges"]) > 0
    
    def test_read_named_range(self, sample_excel):
        """读取命名范围"""
        cap = NamedRangeCapability()
        # 先创建一个命名范围
        cap.execute(None, file=str(sample_excel),
                   action="create", name="TestRange",
                   refers_to="Sheet1!$A$1:$B$3")
        # 读取命名范围
        result = cap.execute(None, file=str(sample_excel),
                           action="read", name="TestRange")
        assert result["action"] == "read"
        assert result["name"] == "TestRange"
    
    def test_delete_named_range(self, sample_excel):
        """删除命名范围"""
        cap = NamedRangeCapability()
        # 先创建一个命名范围
        cap.execute(None, file=str(sample_excel),
                   action="create", name="TestRange",
                   refers_to="Sheet1!$A$1:$B$3")
        # 删除命名范围
        result = cap.execute(None, file=str(sample_excel),
                           action="delete", name="TestRange")
        assert result["action"] == "delete"
