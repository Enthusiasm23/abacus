"""公式生成器 - 常用公式字符串生成"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class FormulaGeneratorCapability(Capability):
    """公式生成器 - 生成常用 Excel 公式"""

    # 公式模板
    FORMULAS = {
        # 查找函数
        "vlookup": "=VLOOKUP({lookup_value},{table_range},{col_index},FALSE)",
        "xlookup": "=XLOOKUP({lookup_value},{lookup_range},{return_range},{if_not_found})",
        "index_match": "=INDEX({return_range},MATCH({lookup_value},{lookup_range},0))",
        # 条件求和
        "sumif": "=SUMIF({criteria_range},{criteria},{sum_range})",
        "sumifs": "=SUMIFS({sum_range},{criteria_range1},{criteria1},{criteria_range2},{criteria2})",
        "countif": "=COUNTIF({criteria_range},{criteria})",
        "countifs": "=COUNTIFS({criteria_range1},{criteria1},{criteria_range2},{criteria2})",
        "averageif": "=AVERAGEIF({criteria_range},{criteria},{average_range})",
        "averageifs": "=AVERAGEIFS({average_range},{criteria_range1},{criteria1},{criteria_range2},{criteria2})",
        # 逻辑函数
        "if": "=IF({condition},{value_if_true},{value_if_false})",
        "ifs": "=IFS({condition1},{value1},{condition2},{value2},TRUE,{default})",
        "iferror": "=IFERROR({value},{value_if_error})",
        "ifna": "=IFNA({value},{value_if_na})",
        # 日期函数
        "today": "=TODAY()",
        "now": "=NOW()",
        "datedif": '=DATEDIF({start_date},{end_date},"{unit}")',
        "eomonth": "=EOMONTH({start_date},{months})",
        "networkdays": "=NETWORKDAYS({start_date},{end_date})",
        "workday": "=WORKDAY({start_date},{days})",
        # 文本函数
        "left": "=LEFT({text},{num_chars})",
        "right": "=RIGHT({text},{num_chars})",
        "mid": "=MID({text},{start_num},{num_chars})",
        "concatenate": "=CONCATENATE({text1},{text2})",
        "textjoin": "=TEXTJOIN({delimiter},TRUE,{text1},{text2})",
        "substitute": "=SUBSTITUTE({text},{old_text},{new_text})",
        # 数学函数
        "sum": "=SUM({range})",
        "average": "=AVERAGE({range})",
        "max": "=MAX({range})",
        "min": "=MIN({range})",
        "round": "=ROUND({number},{num_digits})",
        "abs": "=ABS({number})",
        # 财务函数
        "npv": "=NPV({rate},{values})",
        "irr": "=IRR({values})",
        "pmt": "=PMT({rate},{nper},{pv})",
        "fv": "=FV({rate},{nper},{pmt})",
        "pv": "=PV({rate},{nper},{pmt})",
    }

    @property
    def name(self) -> str:
        return "generate_formula"

    @property
    def chapter(self) -> str:
        return "equation"

    @property
    def description(self) -> str:
        return "生成常用 Excel 公式（VLOOKUP、SUMIFS、IF 等）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="formula_type", type="string", description="公式类型", required=True
            ),
            CapabilitySchema(name="params", type="object", description="公式参数", required=True),
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径（可选）", required=False
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=False),
            CapabilitySchema(name="cell", type="string", description="单元格位置", required=False),
        ]

    def execute(self, context: Any, **params) -> Any:
        formula_type = params.get("formula_type")
        formula_params = params.get("params", {})
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        cell = params.get("cell")

        if not formula_type:
            raise ValidationError("执行失败: 缺少必要参数 formula_type")

        # 生成公式
        formula = self._generate_formula(formula_type, formula_params)

        # 如果指定了文件，写入公式
        if file_path and sheet_name and cell:
            self._write_formula(file_path, sheet_name, cell, formula)

        return {"formula_type": formula_type, "formula": formula, "params": formula_params}

    def _generate_formula(self, formula_type: str, params: dict) -> str:
        """生成公式"""
        template = self.FORMULAS.get(formula_type)
        if not template:
            raise DataError(
                f"数据操作失败: 不支持的公式类型 {formula_type}，可用类型: {list(self.FORMULAS.keys())}"
            )

        try:
            return template.format(**params)
        except KeyError as e:
            raise ValidationError(f"执行失败: 缺少必要参数 {e}")

    def _write_formula(self, file_path: str, sheet_name: str, cell: str, formula: str):
        """写入公式到 Excel"""
        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]
            ws[cell] = formula
            wb.save(file_path)
            wb.close()
        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"公式写入失败: {e}")
            raise DataError(f"数据操作失败: {e}")
