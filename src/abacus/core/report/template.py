"""模板填充报表生成"""

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class TemplateReportCapability(Capability):
    """模板填充 - 基于模板生成报表"""

    @property
    def name(self) -> str:
        return "fill_template"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "基于模板填充数据生成报表（支持命名单元格、批量填充）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="template", type="string", description="模板文件路径", required=True
            ),
            CapabilitySchema(
                name="data_source",
                type="string",
                description="数据源（CSV 文件路径）",
                required=False,
            ),
            CapabilitySchema(
                name="output", type="string", description="输出文件路径", required=True
            ),
            CapabilitySchema(
                name="data", type="object", description="填充数据（字典格式）", required=False
            ),
            CapabilitySchema(
                name="sheet_name", type="string", description="工作表名称", required=False
            ),
            CapabilitySchema(
                name="start_cell", type="string", description="起始单元格", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        template = params.get("template")
        data_source = params.get("data_source")
        output = params.get("output")
        data = params.get("data")
        sheet_name = params.get("sheet_name")
        start_cell = params.get("start_cell", "A1")

        if not template:
            raise DataError("template parameter is required")
        if not output:
            raise DataError("output parameter is required")

        # 加载模板
        template_path = Path(template)
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template}")

        wb = load_workbook(template_path)

        # 加载数据
        if data_source:
            data_path = Path(data_source)
            if not data_path.exists():
                raise FileNotFoundError(f"Data source not found: {data_source}")

            suffix = data_path.suffix.lower()
            if suffix == ".csv":
                df = pd.read_csv(data_path)
            elif suffix in [".xlsx", ".xls"]:
                df = pd.read_excel(data_path)
            else:
                raise DataError(f"Unsupported file format: {suffix}")

            # 填充 DataFrame 数据
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            self._fill_dataframe(ws, df, start_cell)

        elif data:
            # 填充字典数据
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            self._fill_dict(ws, data)

        # 保存
        wb.save(output)

        return {"template": template, "output": output, "filled": True}

    def _fill_dataframe(self, ws, df, start_cell):
        """填充 DataFrame 数据"""
        # 解析起始单元格
        from ..cell_utils import parse_cell_reference

        start_row, start_col = parse_cell_reference(start_cell)

        # 写入表头
        for col_idx, col_name in enumerate(df.columns, start_col):
            ws.cell(row=start_row, column=col_idx, value=col_name)

        # 写入数据
        for row_idx, row in enumerate(df.iterrows(), start_row + 1):
            _, row_data = row
            for col_idx, value in enumerate(row_data, start_col):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _fill_dict(self, ws, data):
        """填充字典数据"""
        for key, value in data.items():
            # 尝试查找命名单元格
            if key in ws.defined_names:
                named_range = ws.defined_names[key]
                # 获取单元格引用
                cell_ref = str(named_range.attr_text).split("!")[-1]
                ws[cell_ref] = value
            else:
                # 直接使用单元格引用
                try:
                    ws[key] = value
                except:
                    logger.warning(f"Could not fill cell {key}")
