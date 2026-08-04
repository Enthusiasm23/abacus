"""基础报表生成"""

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class BasicReportCapability(Capability):
    """基础报表生成 - 从数据生成 Excel 报表"""

    @property
    def name(self) -> str:
        return "create_basic_report"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "从数据生成基础 Excel 报表（自动格式化、列宽调整、冻结首行）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="data_source",
                type="string",
                description="数据源（CSV 文件路径）",
                required=True,
            ),
            CapabilitySchema(
                name="output", type="string", description="输出文件路径", required=True
            ),
            CapabilitySchema(
                name="sheet_name", type="string", description="工作表名称", required=False
            ),
            CapabilitySchema(name="title", type="string", description="报表标题", required=False),
        ]

    def execute(self, context: Any, **params) -> Any:
        data_source = params.get("data_source")
        output = params.get("output")
        sheet_name = params.get("sheet_name", "Data")
        title = params.get("title")

        if not data_source:
            raise DataError("data_source parameter is required")
        if not output:
            raise DataError("output parameter is required")

        # 加载数据
        path = Path(data_source)
        if not path.exists():
            raise FileNotFoundError(f"Data source not found: {data_source}")

        suffix = path.suffix.lower()
        if suffix == ".csv":
            df = pd.read_csv(path)
        elif suffix in [".xlsx", ".xls"]:
            df = pd.read_excel(path)
        else:
            raise DataError(f"Unsupported file format: {suffix}")

        # 创建报表
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 添加标题
        start_row = 1
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            ws["A1"] = title
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center")
            start_row = 3

        # 写入表头
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(patternType="solid", fgColor="2F5496")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for row_idx, row in enumerate(df.iterrows(), start_row + 1):
            _, row_data = row
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # 数字格式
                if isinstance(value, (int, float)):
                    if isinstance(value, float):
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "#,##0"

        # 自动调整列宽
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # 冻结首行
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

        # 添加自动筛选
        ws.auto_filter.ref = (
            f"A{start_row}:{get_column_letter(len(df.columns))}{start_row + len(df)}"
        )

        # 保存
        wb.save(output)

        return {
            "output": output,
            "rows": len(df),
            "columns": len(df.columns),
            "sheet_name": sheet_name,
        }
