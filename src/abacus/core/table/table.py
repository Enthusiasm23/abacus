"""表格操作"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class TableCapability(Capability):
    """表格管理"""

    @property
    def name(self) -> str:
        return "manage_table"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "管理 Excel 表格（创建/列出/删除/追加）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="action",
                type="string",
                description="操作（create//delete/append）",
                required=True,
            ),
            CapabilitySchema(
                name="table_name", type="string", description="表格名称", required=False
            ),
            CapabilitySchema(name="range", type="string", description="数据范围", required=False),
            CapabilitySchema(name="style", type="string", description="表格样式", required=False),
            CapabilitySchema(name="data", type="array", description="追加数据", required=False),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        action = params.get("action")
        table_name = params.get("table_name")
        range_str = params.get("range")
        style_name = params.get("style", "TableStyleMedium9")
        data = params.get("data")

        if not file_path:
            raise DataError("file parameter is required")
        if not action:
            raise DataError("action parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            if action == "create":
                if not range_str:
                    raise DataError("range required for create action")

                start_row, start_col, end_row, end_col = parse_range(range_str)
                if end_row is None:
                    end_row = ws.max_row
                if end_col is None:
                    end_col = ws.max_column

                ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

                if not table_name:
                    table_name = f"Table_{len(ws.tables) + 1}"

                tab = Table(displayName=table_name, ref=ref)
                style = TableStyleInfo(
                    name=style_name,
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                tab.tableStyleInfo = style
                ws.add_table(tab)

                result = {"action": "create", "table_name": table_name, "range": ref}

            elif action == "list":
                tables = []
                for name, tab in ws.tables.items():
                    # tab 可能是 Table 对象或字符串
                    if hasattr(tab, "ref"):
                        tables.append(
                            {
                                "name": name,
                                "range": tab.ref,
                                "style": tab.tableStyleInfo.name if tab.tableStyleInfo else None,
                            }
                        )
                    else:
                        tables.append({"name": name, "range": str(tab), "style": None})
                result = {"action": "list", "tables": tables}

            elif action == "delete":
                if not table_name:
                    raise DataError("table_name required for delete action")

                if table_name in ws.tables:
                    del ws.tables[table_name]
                    result = {"action": "delete", "table_name": table_name}
                else:
                    raise DataError(f"Table '{table_name}' not found")

            elif action == "append":
                if not table_name:
                    raise DataError("table_name required for append action")
                if not data:
                    raise DataError("data required for append action")

                if table_name not in ws.tables:
                    raise DataError(f"Table '{table_name}' not found")

                tab = ws.tables[table_name]
                # 解析表格范围获取列数
                ref = tab.ref
                start, end = ref.split(":")
                start_col = start[0]
                end_col = end[0]

                # 追加数据
                for row_data in data:
                    ws.append(row_data)

                # 更新表格范围
                end_row = ws.max_row
                tab.ref = f"{start}:{end_col}{end_row}"

                result = {"action": "append", "table_name": table_name, "rows_added": len(data)}

            else:
                raise DataError(f"Unknown action: {action}")

            wb.save(file_path)
            wb.close()

            return result

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to manage table: {e}")
            raise DataError(str(e))
