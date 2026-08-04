"""透视表向导 - 基于 jst-well-dan 的参考文档实现"""

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class PivotWizardCapability(Capability):
    """透视表向导 - 创建高级透视表"""

    @property
    def name(self) -> str:
        return "pivot_wizard"

    @property
    def chapter(self) -> str:
        return "share"

    @property
    def description(self) -> str:
        return "透视表向导（创建高级透视表、计算字段、分组汇总）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(name="file", type="string", description="文件路径", required=True),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(name="rows", type="array", description="行字段", required=True),
            CapabilitySchema(name="values", type="array", description="值字段", required=True),
            CapabilitySchema(
                name="agg_function", type="string", description="聚合函数", required=False
            ),
            CapabilitySchema(
                name="output", type="string", description="输出文件路径", required=False
            ),
            CapabilitySchema(
                name="output_sheet", type="string", description="输出工作表名称", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        row_fields = params.get("rows", [])
        value_fields = params.get("values", [])
        agg_function = params.get("agg_function", "sum")
        output = params.get("output")
        output_sheet = params.get("output_sheet", "Pivot")

        if not file_path:
            raise DataError("file parameter is required")
        if not row_fields:
            raise DataError("rows parameter is required")
        if not value_fields:
            raise DataError("values parameter is required")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # 加载数据
        try:
            df = pd.read_excel(path, sheet_name=sheet_name)
        except Exception as e:
            raise DataError(f"Failed to read Excel file: {e}")

        # 验证字段
        for field in row_fields + value_fields:
            if field not in df.columns:
                raise DataError(f"Field '{field}' not found in data")

        # 创建透视表
        pivot_df = self._create_pivot(df, row_fields, value_fields, agg_function)

        # 保存结果
        if not output:
            output = file_path

        self._save_pivot(pivot_df, output, output_sheet, row_fields, value_fields, agg_function)

        return {
            "file": file_path,
            "output": output,
            "output_sheet": output_sheet,
            "rows": len(pivot_df),
            "columns": len(pivot_df.columns),
            "row_fields": row_fields,
            "value_fields": value_fields,
            "agg_function": agg_function,
        }

    def _create_pivot(
        self, df: pd.DataFrame, row_fields: list[str], value_fields: list[str], agg_function: str
    ) -> pd.DataFrame:
        """创建透视表"""
        # 执行聚合
        agg_dict = {}
        for field in value_fields:
            agg_dict[field] = agg_function

        pivot_df = df.groupby(row_fields, dropna=False).agg(agg_dict).reset_index()

        # 添加合计行
        total_row = {}
        for field in row_fields:
            total_row[field] = "合计" if field == row_fields[0] else ""

        for field in value_fields:
            if agg_function == "sum":
                total_row[field] = pivot_df[field].sum()
            elif agg_function == "mean":
                total_row[field] = pivot_df[field].mean()
            elif agg_function == "count":
                total_row[field] = pivot_df[field].count()
            elif agg_function == "min":
                total_row[field] = pivot_df[field].min()
            elif agg_function == "max":
                total_row[field] = pivot_df[field].max()

        pivot_df = pd.concat([pivot_df, pd.DataFrame([total_row])], ignore_index=True)

        return pivot_df

    def _save_pivot(
        self,
        df: pd.DataFrame,
        output: str,
        sheet_name: str,
        row_fields: list[str],
        value_fields: list[str],
        agg_function: str,
    ):
        """保存透视表"""
        from openpyxl import Workbook

        wb = load_workbook(output) if Path(output).exists() else Workbook()

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(sheet_name)

        # 定义样式
        header_font = Font(name="微软雅黑", size=10, bold=True)
        header_fill = PatternFill(patternType="solid", fgColor="4472C4")
        header_alignment = Alignment(horizontal="center", vertical="center")

        data_font = Font(name="微软雅黑", size=10)

        # 写入表头
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for row_idx, row in enumerate(df.iterrows(), 2):
            _, row_data = row
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font

        # 合计行样式
        last_row = len(df) + 1
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=last_row, column=col_idx)
            cell.font = Font(name="微软雅黑", size=10, bold=True)
            cell.fill = PatternFill(patternType="solid", fgColor="D9D9D9")

        # 自适应列宽
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        max_length = max(max_length, cell_len)
                except:
                    pass

            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        wb.save(output)
