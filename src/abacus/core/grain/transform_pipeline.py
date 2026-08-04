"""商功章 - 数据转换管道：支持链式数据转换"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class TransformPipelineCapability(Capability):
    """数据转换管道：支持链式数据转换"""

    @property
    def name(self) -> str:
        return "transform_pipeline"

    @property
    def chapter(self) -> str:
        return "grain"

    @property
    def description(self) -> str:
        return "数据转换管道，支持链式执行多个转换步骤"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称"),
            CapabilitySchema(name="steps", type="array", description="转换步骤列表", required=True),
            CapabilitySchema(
                name="stop_on_error", type="boolean", description="遇到错误是否停止（默认 True）"
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行转换管道"""
        file_path = params.get("file")
        sheet = params.get("sheet")
        steps = params.get("steps", [])
        stop_on_error = params.get("stop_on_error", True)

        if not file_path:
            raise DataError("file parameter is required")
        if not steps:
            raise DataError("steps parameter is required")

        return self._execute_pipeline(file_path, sheet, steps, stop_on_error)

    def _execute_pipeline(
        self, filepath: str, sheet: str, steps: list, stop_on_error: bool
    ) -> dict[str, Any]:
        """执行转换管道"""
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {filepath}")

        wb = None
        try:
            wb = load_workbook(path)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

            results = []
            for i, step in enumerate(steps):
                step_type = step.get("type")
                try:
                    result = self._execute_step(ws, step)
                    results.append(
                        {"step": i + 1, "type": step_type, "status": "success", "result": result}
                    )
                except Exception as e:
                    logger.error(f"Step {i + 1} ({step_type}) failed: {e}")
                    results.append(
                        {"step": i + 1, "type": step_type, "status": "failed", "error": str(e)}
                    )
                    if stop_on_error:
                        raise DataError(f"Pipeline failed at step {i + 1}: {e}")

            wb.save(path)

            return {
                "file": filepath,
                "sheet": ws.title,
                "steps_executed": len(results),
                "steps_succeeded": sum(1 for r in results if r["status"] == "success"),
                "results": results,
            }
        finally:
            if wb:
                wb.close()

    def _execute_step(self, ws, step: dict) -> dict[str, Any]:
        """执行单个转换步骤"""
        step_type = step.get("type")

        if step_type == "convert_type":
            return self._step_convert_type(ws, step)
        elif step_type == "convert_format":
            return self._step_convert_format(ws, step)
        elif step_type == "convert_unit":
            return self._step_convert_unit(ws, step)
        elif step_type == "standardize":
            return self._step_standardize(ws, step)
        elif step_type == "fill_value":
            return self._step_fill_value(ws, step)
        elif step_type == "replace_value":
            return self._step_replace_value(ws, step)
        else:
            raise DataError(f"Unknown step type: {step_type}")

    def _step_convert_type(self, ws, step: dict) -> dict:
        """类型转换"""
        range_str = step.get("range")
        target_type = step.get("target_type", "str")

        if not range_str:
            raise DataError("range is required for convert_type")

        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        converted = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    try:
                        if target_type == "int":
                            cell.value = int(cell.value)
                        elif target_type == "float":
                            cell.value = float(cell.value)
                        elif target_type == "str":
                            cell.value = str(cell.value)
                        converted += 1
                    except (ValueError, TypeError):
                        pass

        return {"converted": converted, "target_type": target_type}

    def _step_convert_format(self, ws, step: dict) -> dict:
        """格式转换"""
        range_str = step.get("range")
        format_type = step.get("format_type", "general")

        if not range_str:
            raise DataError("range is required for convert_format")

        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        format_map = {
            "number": "#,##0.00",
            "currency": "$#,##0.00",
            "percentage": "0.00%",
            "date": "yyyy-mm-dd",
            "text": "@",
        }

        number_format = format_map.get(format_type, format_type)
        formatted = 0

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.number_format = number_format
                    formatted += 1

        return {"formatted": formatted, "format": number_format}

    def _step_convert_unit(self, ws, step: dict) -> dict:
        """单位转换"""
        range_str = step.get("range")
        factor = step.get("factor", 1)

        if not range_str:
            raise DataError("range is required for convert_unit")

        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        converted = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    try:
                        cell.value = float(cell.value) * factor
                        converted += 1
                    except (ValueError, TypeError):
                        pass

        return {"converted": converted, "factor": factor}

    def _step_standardize(self, ws, step: dict) -> dict:
        """标准化"""
        range_str = step.get("range")
        text_case = step.get("text_case")

        if not range_str:
            raise DataError("range is required for standardize")

        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        standardized = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, str):
                    if text_case == "upper":
                        cell.value = cell.value.upper()
                    elif text_case == "lower":
                        cell.value = cell.value.lower()
                    elif text_case == "title":
                        cell.value = cell.value.title()
                    standardized += 1

        return {"standardized": standardized, "text_case": text_case}

    def _step_fill_value(self, ws, step: dict) -> dict:
        """填充值"""
        range_str = step.get("range")
        value = step.get("value")

        if not range_str:
            raise DataError("range is required for fill_value")

        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        filled = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    cell.value = value
                    filled += 1

        return {"filled": filled, "value": value}

    def _step_replace_value(self, ws, step: dict) -> dict:
        """替换值"""
        range_str = step.get("range")
        old_value = step.get("old_value")
        new_value = step.get("new_value")

        if not range_str or old_value is None:
            raise DataError("range and old_value are required for replace_value")

        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        replaced = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value == old_value:
                    cell.value = new_value
                    replaced += 1

        return {"replaced": replaced, "old_value": old_value, "new_value": new_value}
