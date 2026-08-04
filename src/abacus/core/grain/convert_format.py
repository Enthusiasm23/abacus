"""粟米章 - 转格式：转换数据格式"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class ConvertFormatCapability(Capability):
    """转格式：转换数据格式"""

    @property
    def name(self) -> str:
        return "convert_format"

    @property
    def chapter(self) -> str:
        return "grain"

    @property
    def description(self) -> str:
        return "转换数据格式（日期、数字、文本等）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(
                name="format_type",
                type="string",
                description="目标格式（date/number/text/percentage）",
                required=True,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        format_type = params.get("format_type")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if not range_str:
            raise DataError("range parameter is required")
        if not format_type:
            raise DataError("format_type parameter is required")

        return self._convert_format(file_path, sheet_name, range_str, format_type)

    def _convert_format(
        self, filepath: str, sheet_name: str, range_str: str, format_type: str
    ) -> dict[str, Any]:
        """转换数据格式"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            format_map = {
                "date": "YYYY-MM-DD",
                "number": "#,##0.00",
                "text": "@",
                "percentage": "0.0%",
                "currency": "$#,##0.00",
            }

            if format_type not in format_map:
                raise DataError(
                    f"Unsupported format type: {format_type}. Supported: {(format_map.keys())}"
                )

            count = 0
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.number_format = format_map[format_type]
                    count += 1

            wb.save(filepath)
            wb.close()

            return {"file": filepath, "format_type": format_type, "cells_formatted": count}

        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Failed to convert format: {e}")
            raise DataError(str(e))
