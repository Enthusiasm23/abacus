"""粟米章 - 文本分列"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class TextToColumnsCapability(Capability):
    """文本分列"""

    @property
    def name(self) -> str:
        return "text_to_columns"

    @property
    def chapter(self) -> str:
        return "grain"

    @property
    def description(self) -> str:
        return "文本分列（按分隔符拆分）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="column", type="string", description="列标识（如 A）", required=True
            ),
            CapabilitySchema(
                name="delimiter", type="string", description="分隔符", required=False, default=","
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        column = params.get("column")
        delimiter = params.get("delimiter", ",")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if not column:
            raise DataError("column parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            from openpyxl.utils import column_index_from_string

            col_idx = column_index_from_string(column.upper())

            max_row = ws.max_row
            rows_split = 0
            max_parts = 0

            all_values = []
            for row in range(1, max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value is not None and delimiter in str(cell_value):
                    parts = str(cell_value).split(delimiter)
                    all_values.append(parts)
                    max_parts = max(max_parts, len(parts))
                    rows_split += 1
                else:
                    all_values.append([cell_value])

            for row_idx, parts in enumerate(all_values, 1):
                for part_idx, value in enumerate(parts):
                    ws.cell(row=row_idx, column=col_idx + part_idx, value=value)

            wb.save(file_path)
            wb.close()

            return {
                "success": True,
                "column": column,
                "delimiter": delimiter,
                "rows_split": rows_split,
                "columns_created": max_parts,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to split text to columns: {e}")
            raise DataError(str(e))
