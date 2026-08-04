"""粟米章 - 模糊匹配：自动识别相似列名"""

import logging
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class FuzzyMatchCapability(Capability):
    """模糊匹配：自动识别相似列名"""

    @property
    def name(self) -> str:
        return "fuzzy_match_columns"

    @property
    def chapter(self) -> str:
        return "grain"

    @property
    def description(self) -> str:
        return "自动识别相似列名（如'销售额'和'销售金额'）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="target_columns", type="array", description="目标列名列表", required=True
            ),
            CapabilitySchema(
                name="threshold",
                type="number",
                description="相似度阈值（0-1，默认 0.6）",
                required=False,
                default=0.6,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行模糊匹配"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        target_columns = params.get("target_columns", [])
        threshold = params.get("threshold", 0.6)

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if not target_columns:
            raise DataError("target_columns parameter is required")

        return self._fuzzy_match(file_path, sheet_name, target_columns, threshold)

    def _fuzzy_match(
        self, filepath: str, sheet_name: str, target_columns: list[str], threshold: float
    ) -> dict[str, Any]:
        """模糊匹配列名"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, read_only=True, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            # 获取表头
            source_columns = []
            for row in ws.iter_rows(max_row=1, values_only=True):
                source_columns = [
                    str(cell) if cell is not None else f"Column_{i + 1}"
                    for i, cell in enumerate(row)
                ]
                break

            wb.close()

            # 模糊匹配
            matches = []
            for target in target_columns:
                best_match = None
                best_score = 0

                for source in source_columns:
                    score = self._similarity(target, source)
                    if score > best_score:
                        best_score = score
                        best_match = source

                matches.append(
                    {
                        "target": target,
                        "best_match": best_match,
                        "score": round(best_score, 3),
                        "matched": best_score >= threshold,
                    }
                )

            return {
                "success": True,
                "file": filepath,
                "sheet": sheet_name,
                "source_columns": source_columns,
                "target_columns": target_columns,
                "matches": matches,
                "match_count": sum(1 for m in matches if m["matched"]),
            }

        except FileNotFoundError:
            raise
        except DataError:
            raise
        except Exception as e:
            logger.error(f"Failed to fuzzy match columns: {e}")
            raise DataError(str(e))

    def _similarity(self, a: str, b: str) -> float:
        """计算两个字符串的相似度"""
        return SequenceMatcher(None, a, b).ratio()
