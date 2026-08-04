"""粟米章 - 转类型：深度实现数据类型转换"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class ConvertTypeCapability(Capability):
    @property
    def name(self) -> str:
        return "convert_type"

    @property
    def chapter(self) -> str:
        return "grain"

    @property
    def description(self) -> str:
        return "深度转换数据类型（文本转数字、数字转文本等）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(
                name="target_type",
                type="string",
                description="目标类型（number/text/date）",
                required=True,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        target_type = params.get("target_type")

        if not file_path:
            raise DataError(key="param_required", param="file")
        if not target_type:
            raise DataError(key="param_required", param="target_type")

        valid_types = ["number", "text", "date", "int", "float", "str"]
        if target_type not in valid_types:
            raise DataError(key="invalid_type", type=target_type, valid_types=valid_types)

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(file=file_path)

            wb = load_workbook(path)
            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            count = 0
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        if target_type == "number":
                            try:
                                cell.value = float(cell.value)
                                count += 1
                            except (ValueError, TypeError):
                                pass
                        elif target_type == "text":
                            cell.value = str(cell.value)
                            count += 1

            wb.save(file_path)
            wb.close()

            return {"file": file_path, "target_type": target_type, "cells_converted": count}
        except Exception as e:
            raise DataError(str(e))
