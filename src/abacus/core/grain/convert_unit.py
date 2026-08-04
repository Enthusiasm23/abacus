"""粟米章 - 转单位：深度实现单位转换"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)

# 单位转换表
UNIT_CONVERSIONS = {
    ("km", "m"): 1000,
    ("m", "cm"): 100,
    ("cm", "mm"): 10,
    ("kg", "g"): 1000,
    ("g", "mg"): 1000,
    ("lb", "kg"): 0.453592,
    ("oz", "g"): 28.3495,
    ("mile", "km"): 1.60934,
    ("inch", "cm"): 2.54,
    ("ft", "m"): 0.3048,
    ("celsius", "fahrenheit"): lambda c: c * 9 / 5 + 32,
    ("fahrenheit", "celsius"): lambda f: (f - 32) * 5 / 9,
}


class ConvertUnitCapability(Capability):
    @property
    def name(self) -> str:
        return "convert_unit"

    @property
    def chapter(self) -> str:
        return "grain"

    @property
    def description(self) -> str:
        return "深度转换单位（长度、重量、温度等）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(name="from_unit", type="string", description="源单位", required=True),
            CapabilitySchema(name="to_unit", type="string", description="目标单位", required=True),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        from_unit = params.get("from_unit")
        to_unit = params.get("to_unit")

        if not file_path:
            raise DataError("file parameter is required")

        try:
            key = (from_unit.lower(), to_unit.lower())
            if key not in UNIT_CONVERSIONS:
                raise DataError(f"Unsupported conversion: {from_unit} -> {to_unit}")

            converter = UNIT_CONVERSIONS[key]

            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            count = 0
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        if callable(converter):
                            cell.value = converter(cell.value)
                        else:
                            cell.value = cell.value * converter
                        count += 1

            wb.save(file_path)
            wb.close()

            return {
                "file": file_path,
                "from_unit": from_unit,
                "to_unit": to_unit,
                "cells_converted": count,
            }
        except Exception as e:
            raise DataError(str(e))
