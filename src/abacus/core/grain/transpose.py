"""粟米章 - 转置数据"""

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class TransposeCapability(Capability):
    """转置数据"""

    @property
    def name(self) -> str:
        return "transpose"

    @property
    def chapter(self) -> str:
        return "grain"

    @property
    def description(self) -> str:
        return "转置数据（行列互换）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="range", type="string", description="数据范围（A1 表示法）", required=True
            ),
            CapabilitySchema(
                name="output_sheet", type="string", description="输出工作表名称", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        data_range = params.get("range")
        output_sheet = params.get("output_sheet")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if not data_range:
            raise DataError("range parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            from ..cell_utils import parse_range

            min_col, min_row, max_col, max_row = parse_range(data_range)

            data = []
            for row in range(min_row, max_row + 1):
                row_data = []
                for col in range(min_col, max_col + 1):
                    row_data.append(ws.cell(row=row, column=col).value)
                data.append(row_data)

            df = pd.DataFrame(data)
            transposed = df.T

            target_sheet = output_sheet or sheet_name
            if target_sheet not in wb.sheetnames:
                wb.create_sheet(target_sheet)
            ws_out = wb[target_sheet]

            for r_idx, row in enumerate(transposed.values.tolist(), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_out.cell(row=r_idx, column=c_idx, value=value)

            wb.save(file_path)
            wb.close()

            return {
                "success": True,
                "source_range": data_range,
                "source_rows": max_row - min_row + 1,
                "source_columns": max_col - min_col + 1,
                "output_sheet": target_sheet,
                "output_rows": transposed.shape[0],
                "output_columns": transposed.shape[1],
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to transpose data: {e}")
            raise DataError(str(e))
