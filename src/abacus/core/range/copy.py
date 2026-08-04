"""范围操作 - 复制"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class CopyRangeCapability(Capability):
    """复制范围"""

    @property
    def name(self) -> str:
        return "copy_range"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "复制范围（值/公式/格式）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="源工作表", required=True),
            CapabilitySchema(name="source", type="string", description="源范围", required=True),
            CapabilitySchema(
                name="target", type="string", description="目标位置（如 F1）", required=True
            ),
            CapabilitySchema(
                name="copy_type",
                type="string",
                description="复制类型（all/values/formulas/formats）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        source = params.get("source")
        target = params.get("target")
        copy_type = params.get("copy_type", "all")

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            src_start, src_col, src_end_row, src_end_col = parse_range(source)
            tgt_start, tgt_col, _, _ = parse_range(target)

            if src_end_row is None:
                src_end_row = ws.max_row
            if src_end_col is None:
                src_end_col = ws.max_column

            rows = src_end_row - src_start + 1
            cols = src_end_col - src_col + 1

            for r in range(rows):
                for c in range(cols):
                    src_cell = ws.cell(row=src_start + r, column=src_col + c)
                    tgt_cell = ws.cell(row=tgt_start + r, column=tgt_col + c)

                    if copy_type in ["all", "values"]:
                        tgt_cell.value = src_cell.value

                    if copy_type in ["all", "formats"]:
                        # 复制字体
                        try:
                            font_args = {}
                            if src_cell.font.name:
                                font_args["name"] = src_cell.font.name
                            if src_cell.font.size:
                                font_args["size"] = src_cell.font.size
                            if src_cell.font.bold:
                                font_args["bold"] = src_cell.font.bold
                            if src_cell.font.italic:
                                font_args["italic"] = src_cell.font.italic
                            if src_cell.font.color and src_cell.font.color.rgb:
                                font_args["color"] = src_cell.font.color.rgb
                            if font_args:
                                tgt_cell.font = Font(**font_args)
                        except Exception:
                            pass

                        # 复制填充
                        try:
                            if src_cell.fill.patternType:
                                fill_args = {"patternType": src_cell.fill.patternType}
                                if src_cell.fill.fgColor and src_cell.fill.fgColor.rgb:
                                    fill_args["fgColor"] = src_cell.fill.fgColor.rgb
                                if src_cell.fill.bgColor and src_cell.fill.bgColor.rgb:
                                    fill_args["bgColor"] = src_cell.fill.bgColor.rgb
                                tgt_cell.fill = PatternFill(**fill_args)
                        except Exception:
                            pass

                        # 复制边框
                        try:
                            border_args = {}
                            for side_name in ["left", "right", "top", "bottom"]:
                                side = getattr(src_cell.border, side_name, None)
                                if side and side.style:
                                    border_args[side_name] = side
                            if border_args:
                                tgt_cell.border = Border(**border_args)
                        except Exception:
                            pass

                        # 复制对齐
                        try:
                            align_args = {}
                            if src_cell.alignment.horizontal:
                                align_args["horizontal"] = src_cell.alignment.horizontal
                            if src_cell.alignment.vertical:
                                align_args["vertical"] = src_cell.alignment.vertical
                            if src_cell.alignment.wrap_text:
                                align_args["wrap_text"] = src_cell.alignment.wrap_text
                            if align_args:
                                tgt_cell.alignment = Alignment(**align_args)
                        except Exception:
                            pass

                        tgt_cell.number_format = src_cell.number_format

            wb.save(file_path)
            wb.close()

            return {
                "file": file_path,
                "source": source,
                "target": target,
                "copy_type": copy_type,
                "rows": rows,
                "columns": cols,
            }

        except Exception as e:
            logger.error(f"Failed to copy range: {e}")
            raise DataError(str(e))
