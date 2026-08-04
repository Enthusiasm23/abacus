"""范围操作 - 清除"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class ClearRangeCapability(Capability):
    """清除范围内容"""

    @property
    def name(self) -> str:
        return "clear_range"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "清除范围内容（值/公式/格式/全部）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(
                name="clear_type",
                type="string",
                description="清除类型（all/contents/formats）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        clear_type = params.get("clear_type", "all")

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            count = 0
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)

                    if clear_type == "all":
                        cell.value = None
                        cell.font = Font()
                        cell.fill = PatternFill()
                        cell.border = Border()
                        cell.alignment = Alignment()
                        cell.number_format = "General"
                    elif clear_type == "contents":
                        cell.value = None
                    elif clear_type == "formats":
                        cell.font = Font()
                        cell.fill = PatternFill()
                        cell.border = Border()
                        cell.alignment = Alignment()
                        cell.number_format = "General"

                    count += 1

            wb.save(file_path)
            wb.close()

            return {
                "file": file_path,
                "sheet": sheet_name,
                "range": range_str,
                "clear_type": clear_type,
                "cells_cleared": count,
            }

        except Exception as e:
            logger.error(f"Failed to clear range: {e}")
            raise DataError(str(e))
