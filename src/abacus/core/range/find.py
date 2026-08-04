"""范围操作 - 查找替换"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class FindReplaceCapability(Capability):
    """查找替换"""

    @property
    def name(self) -> str:
        return "find_replace"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "查找替换文本"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="find", type="string", description="查找内容", required=True),
            CapabilitySchema(name="replace", type="string", description="替换内容", required=False),
            CapabilitySchema(
                name="range", type="string", description="查找范围（可选）", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        find_text = params.get("find")
        replace_text = params.get("replace")
        range_str = params.get("range")

        if not file_path:
            raise DataError("file parameter is required")
        if not find_text:
            raise DataError("find parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            if range_str:
                start_row, start_col, end_row, end_col = parse_range(range_str)
                if end_row is None:
                    end_row = ws.max_row
                if end_col is None:
                    end_col = ws.max_column
            else:
                start_row, start_col = 1, 1
                end_row, end_col = ws.max_row, ws.max_column

            found = []
            replaced = 0

            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        if find_text in cell.value:
                            found.append({"cell": cell.coordinate, "value": cell.value})

                            if replace_text is not None:
                                cell.value = cell.value.replace(find_text, replace_text)
                                replaced += 1

            wb.save(file_path)
            wb.close()

            return {
                "file": file_path,
                "sheet": sheet_name,
                "find": find_text,
                "replace": replace_text,
                "found_count": len(found),
                "replaced_count": replaced,
                "found_cells": found[:20],
            }

        except Exception as e:
            logger.error(f"Failed to find/replace: {e}")
            raise DataError(str(e))
