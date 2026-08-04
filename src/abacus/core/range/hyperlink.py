"""范围操作 - 超链接"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class HyperlinkCapability(Capability):
    """超链接管理"""

    @property
    def name(self) -> str:
        return "manage_hyperlink"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "管理超链接（添加/删除/列出）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="action", type="string", description="操作（add/remove/）", required=True
            ),
            CapabilitySchema(name="cell", type="string", description="单元格位置", required=False),
            CapabilitySchema(name="url", type="string", description="链接地址", required=False),
            CapabilitySchema(name="text", type="string", description="显示文本", required=False),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        action = params.get("action")
        cell = params.get("cell")
        url = params.get("url")
        text = params.get("text")

        if not file_path:
            raise DataError("file parameter is required")
        if not action:
            raise DataError("action parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            if action == "add":
                if not cell or not url:
                    raise DataError("cell and url required for add action")
                ws[cell].hyperlink = url
                ws[cell].style = "Hyperlink"
                if text:
                    ws[cell].value = text
                result = {"action": "add", "cell": cell, "url": url, "text": text}

            elif action == "remove":
                if not cell:
                    raise DataError("cell required for remove action")
                ws[cell].hyperlink = None
                result = {"action": "remove", "cell": cell}

            elif action == "list":
                hyperlinks = []
                for row in ws.iter_rows():
                    for c in row:
                        if c.hyperlink:
                            hyperlinks.append(
                                {
                                    "cell": c.coordinate,
                                    "url": c.hyperlink.target if c.hyperlink else None,
                                }
                            )
                result = {"action": "list", "hyperlinks": hyperlinks}

            else:
                raise DataError(f"Unknown action: {action}")

            wb.save(file_path)
            wb.close()

            return result

        except Exception as e:
            logger.error(f"Failed to manage hyperlink: {e}")
            raise DataError(str(e))
