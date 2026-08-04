"""范围操作 - 行列大小"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class ColumnRowSizeCapability(Capability):
    """行列大小管理"""

    @property
    def name(self) -> str:
        return "manage_size"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "管理行列大小（列宽/行高）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="action", type="string", description="操作（set/get/auto）", required=True
            ),
            CapabilitySchema(
                name="dimension", type="string", description="维度（row/column）", required=True
            ),
            CapabilitySchema(name="index", type="number", description="行号或列号", required=True),
            CapabilitySchema(
                name="size", type="number", description="大小（可选）", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        action = params.get("action")
        dimension = params.get("dimension")
        index = params.get("index")
        size = params.get("size")

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            if action == "set":
                if size is None:
                    raise DataError("size required for set action")

                if dimension == "column":
                    ws.column_dimensions[get_column_letter(index)].width = size
                elif dimension == "row":
                    ws.row_dimensions[index].height = size
                result = {"action": "set", "dimension": dimension, "index": index, "size": size}

            elif action == "get":
                if dimension == "column":
                    current = ws.column_dimensions[get_column_letter(index)].width
                elif dimension == "row":
                    current = ws.row_dimensions[index].height
                result = {"action": "get", "dimension": dimension, "index": index, "size": current}

            elif action == "auto":
                if dimension == "column":
                    max_len = 0
                    for row in ws.iter_rows(min_col=index, max_col=index):
                        for cell in row:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[get_column_letter(index)].width = max_len + 2
                    result = {
                        "action": "auto",
                        "dimension": dimension,
                        "index": index,
                        "size": max_len + 2,
                    }
                else:
                    raise DataError("auto only supported for columns")

            else:
                raise DataError(f"Unknown action: {action}")

            wb.save(file_path)
            wb.close()

            return result

        except Exception as e:
            logger.error(f"Failed to manage size: {e}")
            raise DataError(str(e))
