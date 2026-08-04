"""范围操作 - 单元格锁定"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class CellLockCapability(Capability):
    """单元格锁定"""

    @property
    def name(self) -> str:
        return "manage_cell_lock"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "管理单元格锁定（锁定/解锁）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(name="locked", type="boolean", description="是否锁定", required=True),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        locked = params.get("locked", True)

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            count = 0
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.protection = cell.protection.copy(locked=locked)
                    count += 1

            wb.save(file_path)
            wb.close()

            return {
                "file": file_path,
                "sheet": sheet_name,
                "range": range_str,
                "locked": locked,
                "cells_updated": count,
            }

        except Exception as e:
            logger.error(f"Failed to manage cell lock: {e}")
            raise DataError(str(e))
