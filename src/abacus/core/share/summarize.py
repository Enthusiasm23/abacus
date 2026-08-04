"""衰分章 - 汇总：深度实现分组汇总"""

import logging
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class SummarizeCapability(Capability):
    """汇总：深度实现分组汇总"""

    @property
    def name(self) -> str:
        return "summarize"

    @property
    def chapter(self) -> str:
        return "share"

    @property
    def description(self) -> str:
        return "分组汇总（支持多列、多聚合函数）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(name="group_by", type="string", description="分组列名", required=True),
            CapabilitySchema(
                name="agg_config", type="object", description="聚合配置 {列名: 函数}", required=True
            ),
            CapabilitySchema(
                name="output_sheet", type="string", description="输出工作表", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        group_by = params.get("group_by")
        agg_config = params.get("agg_config", {})
        output_sheet = params.get("output_sheet")

        if not file_path:
            raise DataError("file parameter is required")
        if not group_by:
            raise DataError("group_by parameter is required")

        return self._summarize(file_path, sheet_name, range_str, group_by, agg_config, output_sheet)

    def _summarize(
        self,
        filepath: str,
        sheet_name: str,
        range_str: str,
        group_by: str,
        agg_config: dict,
        output_sheet: str = None,
    ) -> dict:
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 读取表头
            headers = []
            for col in range(start_col, end_col + 1):
                headers.append(ws.cell(row=start_row, column=col).value)

            # 验证列名
            if group_by not in headers:
                raise DataError(f"Column '{group_by}' not found")

            # 读取数据
            data = []
            for row in range(start_row + 1, end_row + 1):
                row_data = {}
                for col in range(start_col, end_col + 1):
                    header = headers[col - start_col]
                    row_data[header] = ws.cell(row=row, column=col).value
                data.append(row_data)

            wb.close()

            # 执行分组
            groups = defaultdict(list)
            for row in data:
                key = row.get(group_by)
                groups[key].append(row)

            # 执行聚合
            result = []
            for key, rows in groups.items():
                summary = {group_by: key}

                for col, func in agg_config.items():
                    values = [row.get(col) for row in rows if row.get(col) is not None]

                    if func == "sum":
                        summary[f"sum({col})"] = sum(float(v) for v in values) if values else 0
                    elif func == "avg":
                        summary[f"avg({col})"] = (
                            sum(float(v) for v in values) / len(values) if values else 0
                        )
                    elif func == "count":
                        summary[f"count({col})"] = len(values)
                    elif func == "min":
                        summary[f"min({col})"] = min(float(v) for v in values) if values else 0
                    elif func == "max":
                        summary[f"max({col})"] = max(float(v) for v in values) if values else 0

                summary["count"] = len(rows)
                result.append(summary)

            # 排序
            result.sort(key=lambda x: str(x.get(group_by, "")))

            return {
                "file": filepath,
                "sheet": sheet_name,
                "range": range_str,
                "group_by": group_by,
                "agg_config": agg_config,
                "groups_count": len(result),
                "result": result,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to summarize: {e}")
            raise DataError(str(e))
