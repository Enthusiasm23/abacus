"""衰分章 - 分配：深度实现按比例分配"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class DistributeCapability(Capability):
    @property
    def name(self) -> str:
        return "distribute"

    @property
    def chapter(self) -> str:
        return "share"

    @property
    def description(self) -> str:
        return "深度按比例分配（等比、加权、自定义）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(name="total", type="number", description="分配总额", required=True),
            CapabilitySchema(
                name="method",
                type="string",
                description="分配方法（equal/weighted）",
                required=False,
            ),
            CapabilitySchema(
                name="weight_column",
                type="string",
                description="权重列名（加权分配时）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        total = params.get("total")
        method = params.get("method", "equal")
        weight_column = params.get("weight_column")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if not range_str:
            raise DataError("range parameter is required")
        if total is None:
            raise DataError("total parameter is required")

        valid_methods = ["equal", "weighted"]
        if method not in valid_methods:
            raise DataError(f"Invalid method: {method}. Must be one of {valid_methods}")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 读取表头
            headers = []
            for col in range(start_col, end_col + 1):
                headers.append(ws.cell(row=start_row, column=col).value)

            # 计算分配
            if method == "equal":
                count = end_row - start_row
                if count <= 0:
                    raise DataError("No data rows to distribute")
                per_row = total / count

                for row in range(start_row + 1, end_row + 1):
                    ws.cell(row=row, column=end_col + 1, value=per_row)

            elif method == "weighted":
                if not weight_column:
                    raise DataError("weight_column required for weighted distribution")

                weight_idx = headers.index(weight_column)
                weights = []
                for row in range(start_row + 1, end_row + 1):
                    weights.append(ws.cell(row=row, column=start_col + weight_idx).value or 0)

                total_weight = sum(weights)
                if total_weight == 0:
                    raise DataError("Total weight is 0")

                for i, row in enumerate(range(start_row + 1, end_row + 1)):
                    allocated = (weights[i] / total_weight) * total
                    ws.cell(row=row, column=end_col + 1, value=allocated)

            wb.save(file_path)
            wb.close()

            return {"file": file_path, "total": total, "method": method, "distributed": True}
        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            raise DataError(str(e))
