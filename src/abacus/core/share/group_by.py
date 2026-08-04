"""衰分章 - 分组：深度实现按字段分组"""

import logging
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class GroupByCapability(Capability):
    """分组：深度实现按字段分组"""

    @property
    def name(self) -> str:
        return "group_by"

    @property
    def chapter(self) -> str:
        return "share"

    @property
    def description(self) -> str:
        return "按字段分组（支持多列分组、统计）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(
                name="group_columns", type="array", description="分组列名列表", required=True
            ),
            CapabilitySchema(
                name="agg_column", type="string", description="聚合列名", required=False
            ),
            CapabilitySchema(
                name="agg_function",
                type="string",
                description="聚合函数（sum/avg/count/min/max）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        group_columns = params.get("group_columns", [])
        agg_column = params.get("agg_column")
        agg_function = params.get("agg_function", "count")

        if not file_path:
            raise DataError("file parameter is required")
        if not group_columns:
            raise DataError("group_columns parameter is required")

        return self._group_by(
            file_path, sheet_name, range_str, group_columns, agg_column, agg_function
        )

    def _group_by(
        self,
        filepath: str,
        sheet_name: str,
        range_str: str,
        group_columns: list[str],
        agg_column: str,
        agg_function: str,
    ) -> dict[str, Any]:
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 读取表头
            headers = []
            for col in range(start_col, end_col + 1):
                headers.append(ws.cell(row=start_row, column=col).value)

            # 验证列名
            for col_name in group_columns + ([agg_column] if agg_column else []):
                if col_name not in headers:
                    raise DataError(f"Column '{col_name}' not found in headers")

            # 读取数据
            data = []
            for row in range(start_row + 1, end_row + 1):
                row_data = {}
                for col in range(start_col, end_col + 1):
                    header = headers[col - start_col]
                    row_data[header] = ws.cell(row=row, column=col).value
                data.append(row_data)

            wb.close()

            # 执行分组
            groups = defaultdict(list)
            for row in data:
                key = tuple(row.get(col) for col in group_columns)
                groups[key].append(row)

            # 执行聚合
            result = []
            for key, rows in groups.items():
                group_data = {col: val for col, val in zip(group_columns, key)}

                if agg_column:
                    values = [
                        row.get(agg_column) for row in rows if row.get(agg_column) is not None
                    ]

                    if agg_function == "sum":
                        agg_value = sum(float(v) for v in values) if values else 0
                    elif agg_function == "avg":
                        agg_value = sum(float(v) for v in values) / len(values) if values else 0
                    elif agg_function == "count":
                        agg_value = len(values)
                    elif agg_function == "min":
                        agg_value = min(float(v) for v in values) if values else 0
                    elif agg_function == "max":
                        agg_value = max(float(v) for v in values) if values else 0
                    else:
                        raise DataError(f"Unknown aggregation function: {agg_function}")

                    group_data[f"{agg_function}({agg_column})"] = agg_value

                group_data["count"] = len(rows)
                result.append(group_data)

            # 排序
            result.sort(key=lambda x: tuple(str(x.get(col, "")) for col in group_columns))

            return {
                "file": filepath,
                "sheet": sheet_name,
                "range": range_str,
                "group_columns": group_columns,
                "groups_count": len(result),
                "result": result,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to group by: {e}")
            raise DataError(str(e))
