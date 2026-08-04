"""衰分章 - 分类汇总"""

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class SubtotalCapability(Capability):
    """分类汇总"""

    @property
    def name(self) -> str:
        return "subtotal"

    @property
    def chapter(self) -> str:
        return "share"

    @property
    def description(self) -> str:
        return "分类汇总（按字段分组聚合）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="range", type="string", description="数据范围（A1 表示法）", required=True
            ),
            CapabilitySchema(
                name="group_column", type="string", description="分组列名", required=True
            ),
            CapabilitySchema(
                name="function",
                type="string",
                description="聚合函数（sum/mean/count/min/max）",
                required=False,
                default="sum",
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        data_range = params.get("range")
        group_column = params.get("group_column")
        function = params.get("function", "sum")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if not data_range:
            raise DataError("range parameter is required")
        if not group_column:
            raise DataError("group_column parameter is required")

        valid_functions = ["sum", "mean", "count", "min", "max"]
        if function not in valid_functions:
            raise DataError(f"Invalid function: {function}. Must be one of {valid_functions}")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            from ..cell_utils import parse_range

            min_row, min_col, max_row, max_col = parse_range(data_range)

            headers = []
            for col in range(min_col, max_col + 1):
                cell_value = ws.cell(row=min_row, column=col).value
                headers.append(str(cell_value) if cell_value else f"Col{col}")

            data = []
            for row in range(min_row + 1, max_row + 1):
                row_data = {}
                for idx, col in enumerate(range(min_col, max_col + 1)):
                    row_data[headers[idx]] = ws.cell(row=row, column=col).value
                data.append(row_data)

            wb.close()

            if group_column not in headers:
                raise DataError(f"Column '{group_column}' not found in range. Available: {headers}")

            df = pd.DataFrame(data)

            numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
            numeric_cols = [c for c in numeric_cols if c != group_column]

            if not numeric_cols:
                numeric_cols = [c for c in headers if c != group_column]

            agg_dict = {col: function for col in numeric_cols}
            result = df.groupby(group_column).agg(agg_dict).reset_index()

            summary_data = result.to_dict(orient="records")

            return {
                "success": True,
                "group_column": group_column,
                "function": function,
                "groups_count": len(summary_data),
                "summary": summary_data,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to compute subtotal: {e}")
            raise DataError(str(e))
