"""勾股章 - 统计分析：深度实现统计分析"""

import logging
import statistics
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class AnalyzeStatsCapability(Capability):
    """统计分析：深度实现统计分析"""

    @property
    def name(self) -> str:
        return "analyze_stats"

    @property
    def chapter(self) -> str:
        return "triangle"

    @property
    def description(self) -> str:
        return "深度统计分析（均值、中位数、标准差、分位数等）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(
                name="columns",
                type="array",
                description="分析列名列表（可选，默认全部数值列）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        columns = params.get("columns")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")

        return self._analyze_stats(file_path, sheet_name, range_str, columns)

    def _analyze_stats(
        self, filepath: str, sheet_name: str, range_str: str, columns: list[str] = None
    ) -> dict[str, Any]:
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 读取表头
            headers = []
            for col in range(start_col, end_col + 1):
                headers.append(ws.cell(row=start_row, column=col).value)

            # 确定要分析的列
            if columns:
                analyze_cols = [headers.index(col) for col in columns if col in headers]
            else:
                # 自动检测数值列
                analyze_cols = []
                for col in range(start_col, end_col + 1):
                    for row in range(start_row + 1, min(start_row + 10, end_row + 1)):
                        val = ws.cell(row=row, column=col).value
                        if isinstance(val, (int, float)):
                            analyze_cols.append(col - start_col)
                            break

            wb.close()

            # 读取数据并计算统计
            results = {}
            for col_idx in analyze_cols:
                col_name = headers[col_idx]
                values = []

                wb2 = load_workbook(path, data_only=True)
                ws2 = wb2[sheet_name]
                for row in range(start_row + 1, end_row + 1):
                    val = ws2.cell(row=row, column=start_col + col_idx).value
                    if isinstance(val, (int, float)):
                        values.append(float(val))
                wb2.close()

                if not values:
                    continue

                stats = {
                    "count": len(values),
                    "sum": sum(values),
                    "mean": statistics.mean(values),
                    "median": statistics.median(values),
                    "min": min(values),
                    "max": max(values),
                    "range": max(values) - min(values),
                }

                if len(values) >= 2:
                    stats["stdev"] = statistics.stdev(values)
                    stats["variance"] = statistics.variance(values)

                # 分位数
                sorted_values = sorted(values)
                n = len(sorted_values)
                stats["q1"] = sorted_values[n // 4] if n >= 4 else None
                stats["q3"] = sorted_values[3 * n // 4] if n >= 4 else None

                results[col_name] = stats

            return {
                "file": filepath,
                "sheet": sheet_name,
                "range": range_str,
                "columns_analyzed": len(results),
                "statistics": results,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"统计分析失败: {e}")
            raise DataError(f"数据操作失败: {e}")
