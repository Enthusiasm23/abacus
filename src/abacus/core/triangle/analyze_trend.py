"""勾股章 - 趋势分析：深度实现趋势分析"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class AnalyzeTrendCapability(Capability):
    """趋势分析：深度实现趋势分析"""

    @property
    def name(self) -> str:
        return "analyze_trend"

    @property
    def chapter(self) -> str:
        return "triangle"

    @property
    def description(self) -> str:
        return "深度趋势分析（增长、环比、同比）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(
                name="value_column", type="string", description="值列名", required=True
            ),
            CapabilitySchema(
                name="time_column", type="string", description="时间列名", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        value_column = params.get("value_column")
        time_column = params.get("time_column")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")
        if not value_column:
            raise ValidationError("执行失败: 缺少必要参数 value_column")

        return self._analyze_trend(file_path, sheet_name, range_str, value_column, time_column)

    def _analyze_trend(
        self,
        filepath: str,
        sheet_name: str,
        range_str: str,
        value_column: str,
        time_column: str = None,
    ) -> dict[str, Any]:
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 读取表头
            headers = []
            for col in range(start_col, end_col + 1):
                headers.append(ws.cell(row=start_row, column=col).value)

            # 验证列名
            if value_column not in headers:
                raise DataError(f"数据操作失败: 列 '{value_column}' 不存在")

            value_idx = headers.index(value_column)

            # 读取数据
            values = []
            for row in range(start_row + 1, end_row + 1):
                val = ws.cell(row=row, column=start_col + value_idx).value
                if isinstance(val, (int, float)):
                    values.append(float(val))

            wb.close()

            if len(values) < 2:
                return {
                    "file": filepath,
                    "sheet": sheet_name,
                    "error": "Need at least 2 data points for trend analysis",
                }

            # 计算趋势
            total_change = values[-1] - values[0]
            total_change_pct = (total_change / values[0] * 100) if values[0] != 0 else None

            # 环比变化
            period_changes = []
            for i in range(1, len(values)):
                change = values[i] - values[i - 1]
                change_pct = (change / values[i - 1] * 100) if values[i - 1] != 0 else None
                period_changes.append(
                    {"period": i, "value": values[i], "change": change, "change_pct": change_pct}
                )

            # 简单线性趋势
            n = len(values)
            x_mean = (n - 1) / 2
            y_mean = sum(values) / n
            numerator = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(values))
            denominator = sum((i - x_mean) ** 2 for i in range(n))
            slope = numerator / denominator if denominator != 0 else 0
            intercept = y_mean - slope * x_mean

            trend_direction = "上升" if slope > 0 else "下降" if slope < 0 else "平稳"

            return {
                "file": filepath,
                "sheet": sheet_name,
                "data_points": len(values),
                "first_value": values[0],
                "last_value": values[-1],
                "total_change": total_change,
                "total_change_pct": total_change_pct,
                "trend_direction": trend_direction,
                "slope": slope,
                "intercept": intercept,
                "period_changes": period_changes[:10],  # 只返回前10个
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"趋势分析失败: {e}")
            raise DataError(f"数据操作失败: {e}")
