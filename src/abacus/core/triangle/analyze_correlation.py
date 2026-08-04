"""勾股章 - 相关性分析：深度实现相关性分析"""

import logging
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class AnalyzeCorrelationCapability(Capability):
    """相关性分析：深度实现相关性分析"""

    @property
    def name(self) -> str:
        return "analyze_correlation"

    @property
    def chapter(self) -> str:
        return "triangle"

    @property
    def description(self) -> str:
        return "深度相关性分析（皮尔逊相关系数、散点图数据）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(name="column1", type="string", description="第一列名", required=True),
            CapabilitySchema(name="column2", type="string", description="第二列名", required=True),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        column1 = params.get("column1")
        column2 = params.get("column2")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")
        if not column1 or not column2:
            raise ValidationError("执行失败: 缺少必要参数 column1 或 column2")

        return self._analyze_correlation(file_path, sheet_name, range_str, column1, column2)

    def _analyze_correlation(
        self, filepath: str, sheet_name: str, range_str: str, column1: str, column2: str
    ) -> dict[str, Any]:
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 读取表头
            headers = []
            for col in range(start_col, end_col + 1):
                headers.append(ws.cell(row=start_row, column=col).value)

            # 验证列名
            if column1 not in headers:
                raise DataError(f"数据操作失败: 列 '{column1}' 不存在")
            if column2 not in headers:
                raise DataError(f"数据操作失败: 列 '{column2}' 不存在")

            col1_idx = headers.index(column1)
            col2_idx = headers.index(column2)

            # 读取配对数据
            pairs = []
            for row in range(start_row + 1, end_row + 1):
                val1 = ws.cell(row=row, column=start_col + col1_idx).value
                val2 = ws.cell(row=row, column=start_col + col2_idx).value
                if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
                    pairs.append((float(val1), float(val2)))

            wb.close()

            if len(pairs) < 2:
                return {"file": filepath, "error": "Need at least 2 paired data points"}

            # 计算皮尔逊相关系数
            n = len(pairs)
            sum_x = sum(p[0] for p in pairs)
            sum_y = sum(p[1] for p in pairs)
            sum_xy = sum(p[0] * p[1] for p in pairs)
            sum_x2 = sum(p[0] ** 2 for p in pairs)
            sum_y2 = sum(p[1] ** 2 for p in pairs)

            numerator = n * sum_xy - sum_x * sum_y
            denominator = math.sqrt((n * sum_x2 - sum_x**2) * (n * sum_y2 - sum_y**2))

            correlation = numerator / denominator if denominator != 0 else 0

            # 解释相关性
            if abs(correlation) >= 0.8:
                strength = "强相关"
            elif abs(correlation) >= 0.5:
                strength = "中等相关"
            elif abs(correlation) >= 0.3:
                strength = "弱相关"
            else:
                strength = "几乎不相关"

            direction = "正相关" if correlation > 0 else "负相关" if correlation < 0 else "无相关"

            return {
                "file": filepath,
                "sheet": sheet_name,
                "column1": column1,
                "column2": column2,
                "data_points": n,
                "correlation": round(correlation, 4),
                "strength": strength,
                "direction": direction,
                "r_squared": round(correlation**2, 4),
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"相关性分析失败: {e}")
            raise DataError(f"数据操作失败: {e}")
