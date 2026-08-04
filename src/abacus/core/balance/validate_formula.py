"""盈不足章 - 验公式：深度实现公式验证"""

import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class ValidateFormulaCapability(Capability):
    """验公式：深度实现公式验证"""

    # 不安全函数列表
    UNSAFE_FUNCTIONS = {"INDIRECT", "HYPERLINK", "WEBSERVICE", "DGET", "RTD"}

    @property
    def name(self) -> str:
        return "validate_formula"

    @property
    def chapter(self) -> str:
        return "balance"

    @property
    def description(self) -> str:
        return "深度验证公式正确性（语法、安全、引用）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=False),
            CapabilitySchema(name="cell", type="string", description="单元格位置", required=False),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行公式验证"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        cell = params.get("cell")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")

        return self._validate_formulas(file_path, sheet_name, cell)

    def _validate_formulas(
        self, filepath: str, sheet_name: str = None, cell: str = None
    ) -> dict[str, Any]:
        """验证所有公式"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path, data_only=False)
            results = []

            for name in wb.sheetnames:
                if sheet_name and name != sheet_name:
                    continue

                ws = wb[name]

                for row in ws.iter_rows():
                    for c in row:
                        if cell and c.coordinate != cell:
                            continue

                        if isinstance(c.value, str) and c.value.startswith("="):
                            is_valid, message = self._validate_single_formula(c.value)
                            results.append(
                                {
                                    "sheet": name,
                                    "cell": c.coordinate,
                                    "formula": c.value,
                                    "valid": is_valid,
                                    "message": message,
                                }
                            )

            wb.close()

            total = len(results)
            valid = sum(1 for r in results if r["valid"])
            invalid = total - valid

            return {
                "file": filepath,
                "total_formulas": total,
                "valid": valid,
                "invalid": invalid,
                "results": results,
            }

        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"公式验证失败: {e}")
            raise DataError(f"数据操作失败: {e}")

    def _validate_single_formula(self, formula: str) -> [bool, str]:
        """验证单个公式"""
        # 1. 检查是否以 = 开头
        if not formula.startswith("="):
            return False, "Formula must start with '='"

        # 2. 检查括号平衡
        parens = 0
        for i, c in enumerate(formula):
            if c == "(":
                parens += 1
            elif c == ")":
                parens -= 1
            if parens < 0:
                return False, f"Unmatched closing parenthesis at position {i}"

        if parens > 0:
            return False, "Unclosed parenthesis"

        # 3. 检查不安全函数
        func_pattern = r"([A-Z]+)\("
        funcs = re.findall(func_pattern, formula)

        for func in funcs:
            if func in self.UNSAFE_FUNCTIONS:
                return False, f"Unsafe function: {func}"

        # 4. 检查单元格引用格式
        cell_refs = re.findall(r"[A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?", formula)
        for ref in cell_refs:
            if ":" in ref:
                start, end = ref.split(":")
                if not self._validate_cell_ref(start) or not self._validate_cell_ref(end):
                    return False, f"Invalid cell reference: {ref}"
            else:
                if not self._validate_cell_ref(ref):
                    return False, f"Invalid cell reference: {ref}"

        # 5. 检查中文逗号
        if "，" in formula:
            return False, "Formula contains Chinese comma (，) instead of English comma (,)"

        return True, "Formula is valid"

    def _validate_cell_ref(self, ref: str) -> bool:
        """验证单元格引用格式"""
        match = re.match(r"([A-Z]+)([0-9]+)", ref)
        if not match:
            return False
        col, row = match.groups()
        return bool(col and row)
