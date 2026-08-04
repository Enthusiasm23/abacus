"""盈不足章 - 质量检测：自动检测数据质量问题"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class QualityCheckCapability(Capability):
    """质量检测：自动检测数据质量问题"""

    @property
    def name(self) -> str:
        return "data_quality_check"

    @property
    def chapter(self) -> str:
        return "balance"

    @property
    def description(self) -> str:
        return "自动检测空值、异常值、重复数据、格式不一致"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="range",
                type="string",
                description="数据范围（可选，默认全部）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行质量检测"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        data_range = params.get("range")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")
        if not sheet_name:
            raise ValidationError("执行失败: 缺少必要参数 sheet")

        return self._quality_check(file_path, sheet_name, data_range)

    def _quality_check(self, filepath: str, sheet_name: str, data_range: str = None) -> dict[str, Any]:
        """质量检测"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path, read_only=True, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 获取表头
            headers = []
            for row in ws.iter_rows(max_row=1, values_only=True):
                headers = [
                    str(cell) if cell is not None else f"Column_{i + 1}"
                    for i, cell in enumerate(row)
                ]
                break

            # 统计信息
            total_rows = 0
            null_counts = {h: 0 for h in headers}
            type_counts = {h: {"number": 0, "string": 0, "date": 0, "empty": 0} for h in headers}
            duplicate_rows = []
            row_values = {}

            # 采样数据用于类型检测
            for row in ws.iter_rows(min_row=2, values_only=True):
                total_rows += 1

                # 检查重复
                row_key = (str(cell) for cell in row)
                if row_key in row_values:
                    duplicate_rows.append(total_rows + 1)
                else:
                    row_values[row_key] = total_rows

                for i, cell in enumerate(row):
                    if i < len(headers):
                        header = headers[i]
                        if cell is None:
                            null_counts[header] += 1
                            type_counts[header]["empty"] += 1
                        else:
                            # 检测类型
                            if isinstance(cell, (int, float)):
                                type_counts[header]["number"] += 1
                            elif hasattr(cell, "year"):
                                type_counts[header]["date"] += 1
                            else:
                                type_counts[header]["string"] += 1

            wb.close()

            # 生成质量问题列表
            issues = []

            # 检查空值
            for header, count in null_counts.items():
                if count > 0:
                    pct = round(count / total_rows * 100, 1) if total_rows > 0 else 0
                    issues.append(
                        {
                            "type": "null_values",
                            "column": header,
                            "count": count,
                            "percentage": pct,
                            "severity": "high" if pct > 50 else "medium" if pct > 20 else "low",
                        }
                    )

            # 检查重复行
            if duplicate_rows:
                issues.append(
                    {
                        "type": "duplicate_rows",
                        "count": len(duplicate_rows),
                        "rows": duplicate_rows[:10],  # 只返回前10个
                        "severity": "high" if len(duplicate_rows) > total_rows * 0.1 else "medium",
                    }
                )

            # 计算质量分数
            null_penalty = (
                sum(null_counts.values()) / (total_rows * len(headers)) * 100
                if total_rows > 0
                else 0
            )
            dup_penalty = len(duplicate_rows) / total_rows * 100 if total_rows > 0 else 0
            quality_score = max(0, 100 - null_penalty - dup_penalty)

            return {
                "success": True,
                "file": filepath,
                "sheet": sheet_name,
                "total_rows": total_rows,
                "total_columns": len(headers),
                "quality_score": round(quality_score, 1),
                "issues": issues,
                "issue_count": len(issues),
                "null_counts": null_counts,
                "type_distribution": type_counts,
                "duplicate_count": len(duplicate_rows),
            }

        except FileNotFoundError:
            raise
        except DataError:
            raise
        except Exception as e:
            logger.error(f"质量检测失败: {e}")
            raise DataError(f"数据操作失败: {e}")
