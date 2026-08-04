"""盈不足章 - 文件验证：验证 Excel 文件结构"""

import logging
import zipfile
from pathlib import Path
from typing import Any

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class FileValidateCapability(Capability):
    """文件验证：验证 Excel 文件结构和内容"""

    @property
    def name(self) -> str:
        return "validate_file"

    @property
    def chapter(self) -> str:
        return "balance"

    @property
    def description(self) -> str:
        return "验证 Excel 文件结构（ZIP 格式、XML 结构、公式错误）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件操作失败: 文件不存在 {file_path}")

        return self._validate_file(path)

    def _validate_file(self, path: Path) -> dict:
        """验证文件"""
        results = {"file": str(path), "valid": True, "checks": [], "errors": [], "warnings": []}

        try:
            with zipfile.ZipFile(path, "r") as zf:
                results["checks"].append("ZIP format: OK")
        except zipfile.BadZipFile:
            results["valid"] = False
            results["errors"].append("Invalid ZIP format")
            return results

        try:
            with zipfile.ZipFile(path, "r") as zf:
                required_files = ["[Content_Types].xml", "xl/workbook.xml", "xl/styles.xml"]
                for f in required_files:
                    if f in zf.namelist():
                        results["checks"].append(f"{f}: OK")
                    else:
                        results["valid"] = False
                        results["errors"].append(f"Missing required file: {f}")
        except Exception as e:
            results["errors"].append(f"Failed to check required files: {e}")

        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, data_only=True)

            error_types = ["#REF!", "#N/A", "#VALUE!", "#NAME?", "#DIV/0!"]
            formula_errors = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            for error in error_types:
                                if error in cell.value:
                                    formula_errors.append(
                                        {
                                            "sheet": sheet_name,
                                            "cell": cell.coordinate,
                                            "error": error,
                                        }
                                    )

            wb.close()

            if formula_errors:
                results["warnings"].append(f"Found {len(formula_errors)} formula errors")
                results["formula_errors"] = formula_errors
            else:
                results["checks"].append("Formula errors: None")

        except Exception as e:
            results["warnings"].append(f"Failed to check formulas: {e}")

        return results
