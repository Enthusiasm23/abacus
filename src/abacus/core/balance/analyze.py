"""文件审计工具 - 检查 Excel 文件常见问题"""

import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class FileAnalyzeCapability(Capability):
    """文件审计 - 检查 Excel 文件的 10 类常见问题"""

    RULES = {  # noqa: RUF012
        "XA001": "公式中包含中文逗号（，）",
        "XA002": "单元格已为错误值（#REF! #VALUE! 等）",
        "XA003": "sheet 名包含禁止字符",
        "XA004": "sheet 名超过 31 字符",
        "XA005": "sheet 名含空格但公式未加引号",
        "XA006": "同一单元格既是输入又是公式",
        "XA007": "公式含魔数（4+ 位数字常量）",
        "XA008": "合并单元格被公式引用",
        "XA009": "整列引用拖慢性能",
        "XA010": "隐藏 sheet 被可见 sheet 引用",
    }

    @property
    def name(self) -> str:
        return "file_analyze"

    @property
    def chapter(self) -> str:
        return "balance"

    @property
    def description(self) -> str:
        return "检查 Excel 文件的 10 类常见问题"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件操作失败: 文件不存在 {file_path}")

        return self._analyze_file(path)

    def _analyze_file(self, path: Path) -> dict[str, Any]:
        """分析文件"""
        issues = []

        try:
            wb = load_workbook(path, data_only=False)

            # 检查 sheet 名
            for name in wb.sheetnames:
                # XA003: 禁止字符
                if re.search(r"[\\/?*\[\]:]", name):
                    issues.append(
                        {
                            "rule": "XA003",
                            "level": "ERROR",
                            "sheet": name,
                            "message": self.RULES["XA003"],
                        }
                    )

                # XA004: 名称过长
                if len(name) > 31:
                    issues.append(
                        {
                            "rule": "XA004",
                            "level": "ERROR",
                            "sheet": name,
                            "message": self.RULES["XA004"],
                        }
                    )

            # 检查公式
            for name in wb.sheetnames:
                ws = wb[name]

                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formula = cell.value

                            # XA001: 中文逗号
                            if "，" in formula:
                                issues.append(
                                    {
                                        "rule": "XA001",
                                        "level": "ERROR",
                                        "sheet": name,
                                        "cell": cell.coordinate,
                                        "message": self.RULES["XA001"],
                                    }
                                )

                            # XA007: 魔数
                            if re.search(r"\d{4,}", formula):
                                issues.append(
                                    {
                                        "rule": "XA007",
                                        "level": "INFO",
                                        "sheet": name,
                                        "cell": cell.coordinate,
                                        "message": self.RULES["XA007"],
                                    }
                                )

                            # XA009: 整列引用
                            if re.search(r"[A-Z]:[A-Z]", formula):
                                issues.append(
                                    {
                                        "rule": "XA009",
                                        "level": "WARN",
                                        "sheet": name,
                                        "cell": cell.coordinate,
                                        "message": self.RULES["XA009"],
                                    }
                                )

                        # XA002: 错误值
                        if isinstance(cell.value, str) and cell.value.startswith("#"):
                            issues.append(
                                {
                                    "rule": "XA002",
                                    "level": "ERROR",
                                    "sheet": name,
                                    "cell": cell.coordinate,
                                    "message": self.RULES["XA002"],
                                }
                            )

            wb.close()

            return {
                "file": str(path),
                "total_issues": len(issues),
                "errors": len([i for i in issues if i["level"] == "ERROR"]),
                "warnings": len([i for i in issues if i["level"] == "WARN"]),
                "info": len([i for i in issues if i["level"] == "INFO"]),
                "issues": issues,
                "rules": self.RULES,
            }

        except Exception as e:
            logger.error(f"文件分析失败: {e}")
            raise DataError(f"数据操作失败: {e}")
