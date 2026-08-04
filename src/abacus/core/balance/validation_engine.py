"""盈不足章 - 数据验证规则引擎：支持自定义规则、规则组合、规则链"""

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


@dataclass
class RuleResult:
    """单条规则验证结果"""
    valid: bool
    cell: str = ""
    rule_name: str = ""
    errors: list = field(default_factory=list)


class ValidationRule:
    """验证规则基类"""

    @property
    def name(self) -> str:
        return "base"

    def validate(self, value: Any, cell: str = "") -> RuleResult:
        raise NotImplementedError


class TypeRule(ValidationRule):
    """类型验证规则"""

    def __init__(self, expected_type: str):
        self.expected_type = expected_type

    @property
    def name(self) -> str:
        return "type"

    def validate(self, value: Any, cell: str = "") -> RuleResult:
        if value is None:
            return RuleResult(valid=False, cell=cell, rule_name=self.name,
                              errors=[f"值为空，期望类型 {self.expected_type}"])

        type_map = {
            "int": (int,),
            "float": (int, float),
            "str": (str,),
        }

        if self.expected_type == "date":
            from datetime import date, datetime
            valid = isinstance(value, (date, datetime))
        elif self.expected_type in type_map:
            valid = isinstance(value, type_map[self.expected_type])
        else:
            valid = True

        if valid:
            return RuleResult(valid=True, cell=cell, rule_name=self.name)
        return RuleResult(
            valid=False, cell=cell, rule_name=self.name,
            errors=[f"类型错误：期望 {self.expected_type}，实际 {type(value).__name__}"]
        )


class RangeRule(ValidationRule):
    """范围验证规则"""

    def __init__(self, min_val: float = None, max_val: float = None):
        self.min_val = min_val
        self.max_val = max_val

    @property
    def name(self) -> str:
        return "range"

    def validate(self, value: Any, cell: str = "") -> RuleResult:
        if value is None:
            return RuleResult(valid=True, cell=cell, rule_name=self.name)

        try:
            num_val = float(value)
        except (TypeError, ValueError):
            return RuleResult(
                valid=False, cell=cell, rule_name=self.name,
                errors=[f"无法转换为数值: {value}"]
            )

        errors = []
        if self.min_val is not None and num_val < self.min_val:
            errors.append(f"值 {num_val} 小于最小值 {self.min_val}")
        if self.max_val is not None and num_val > self.max_val:
            errors.append(f"值 {num_val} 大于最大值 {self.max_val}")

        if errors:
            return RuleResult(valid=False, cell=cell, rule_name=self.name, errors=errors)
        return RuleResult(valid=True, cell=cell, rule_name=self.name)


class NotEmptyRule(ValidationRule):
    """非空验证规则"""

    @property
    def name(self) -> str:
        return "not_empty"

    def validate(self, value: Any, cell: str = "") -> RuleResult:
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return RuleResult(
                valid=False, cell=cell, rule_name=self.name,
                errors=["值为空"]
            )
        return RuleResult(valid=True, cell=cell, rule_name=self.name)


class RegexRule(ValidationRule):
    """正则表达式验证规则"""

    def __init__(self, pattern: str):
        self.pattern = pattern

    @property
    def name(self) -> str:
        return "regex"

    def validate(self, value: Any, cell: str = "") -> RuleResult:
        if value is None:
            return RuleResult(valid=True, cell=cell, rule_name=self.name)

        str_val = str(value)
        if re.match(self.pattern, str_val):
            return RuleResult(valid=True, cell=cell, rule_name=self.name)
        return RuleResult(
            valid=False, cell=cell, rule_name=self.name,
            errors=[f"值 '{str_val}' 不匹配模式 '{self.pattern}'"]
        )


class CustomRule(ValidationRule):
    """自定义验证规则"""

    def __init__(self, func: Callable, name: str = "custom", error_message: str = None):
        self._func = func
        self._name = name
        self._error_message = error_message

    @property
    def name(self) -> str:
        return self._name

    def validate(self, value: Any, cell: str = "") -> RuleResult:
        try:
            if self._func(value):
                return RuleResult(valid=True, cell=cell, rule_name=self.name)
            msg = self._error_message or f"自定义规则 '{self.name}' 验证失败"
            return RuleResult(valid=False, cell=cell, rule_name=self.name, errors=[msg])
        except Exception as e:
            return RuleResult(
                valid=False, cell=cell, rule_name=self.name,
                errors=[f"规则执行错误: {e}"]
            )


class AndRule(ValidationRule):
    """AND 组合规则 - 所有子规则必须通过"""

    def __init__(self, rules: list):
        self.rules = rules

    @property
    def name(self) -> str:
        return "and"

    def validate(self, value: Any, cell: str = "") -> RuleResult:
        all_errors = []
        for rule in self.rules:
            result = rule.validate(value, cell)
            if not result.valid:
                all_errors.extend(result.errors)

        if all_errors:
            return RuleResult(valid=False, cell=cell, rule_name=self.name, errors=all_errors)
        return RuleResult(valid=True, cell=cell, rule_name=self.name)


class OrRule(ValidationRule):
    """OR 组合规则 - 任一子规则通过即可"""

    def __init__(self, rules: list):
        self.rules = rules

    @property
    def name(self) -> str:
        return "or"

    def validate(self, value: Any, cell: str = "") -> RuleResult:
        all_errors = []
        for rule in self.rules:
            result = rule.validate(value, cell)
            if result.valid:
                return RuleResult(valid=True, cell=cell, rule_name=self.name)
            all_errors.extend(result.errors)

        return RuleResult(valid=False, cell=cell, rule_name=self.name, errors=all_errors)


class ValidationChain:
    """规则链 - 按顺序执行多个规则"""

    def __init__(self, rules: list):
        self.rules = rules

    def validate(self, value: Any, cell: str = "") -> RuleResult:
        all_errors = []
        for rule in self.rules:
            result = rule.validate(value, cell)
            if not result.valid:
                all_errors.extend(result.errors)
                break
        if all_errors:
            return RuleResult(valid=False, cell=cell, rule_name="chain", errors=all_errors)
        return RuleResult(valid=True, cell=cell, rule_name="chain")


def _build_rule(rule_config: dict) -> ValidationRule:
    """从配置构建规则对象"""
    rule_type = rule_config.get("type", "")
    params = rule_config.get("params", {})
    sub_rules = rule_config.get("rules", [])

    if rule_type == "type":
        return TypeRule(expected_type=params.get("expected_type", "str"))
    elif rule_type == "range":
        return RangeRule(
            min_val=params.get("min_val"),
            max_val=params.get("max_val"),
        )
    elif rule_type == "not_empty":
        return NotEmptyRule()
    elif rule_type == "regex":
        return RegexRule(pattern=params.get("pattern", ".*"))
    elif rule_type == "custom":
        expression = params.get("expression", "True")
        func = lambda v, _expr=expression: eval(_expr, {"value": v})
        return CustomRule(func=func, name="custom", error_message=params.get("error_message"))
    elif rule_type == "and":
        return AndRule([_build_rule(r) for r in sub_rules])
    elif rule_type == "or":
        return OrRule([_build_rule(r) for r in sub_rules])
    else:
        raise DataError(f"不支持的规则类型: {rule_type}")


class ValidationEngineCapability(Capability):
    """数据验证规则引擎：支持自定义规则、规则组合（AND/OR）、规则链"""

    @property
    def name(self) -> str:
        return "validation_engine"

    @property
    def chapter(self) -> str:
        return "balance"

    @property
    def description(self) -> str:
        return "数据验证规则引擎（自定义规则、AND/OR 组合、规则链）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="sheet", type="string", description="工作表名称", required=True
            ),
            CapabilitySchema(
                name="range", type="string", description="数据范围，使用 A1 表示法", required=True
            ),
            CapabilitySchema(
                name="rules", type="array", description="验证规则列表", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        rules_config = params.get("rules", [])

        if not file_path:
            raise DataError("执行失败: 缺少必要参数 file")

        return self._validate(file_path, sheet_name, range_str, rules_config)

    def _validate(
        self, filepath: str, sheet_name: str, range_str: str, rules_config: list
    ) -> dict[str, Any]:
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            from ..cell_utils import parse_range
            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            rules = [_build_rule(rc) for rc in rules_config]

            total_cells = 0
            passed_cells = 0
            failed_cells = 0
            cell_results = []

            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    total_cells += 1

                    if not rules:
                        passed_cells += 1
                        cell_results.append({
                            "cell": cell.coordinate,
                            "valid": True,
                            "errors": [],
                        })
                        continue

                    all_errors = []
                    for rule in rules:
                        result = rule.validate(cell.value, cell.coordinate)
                        if not result.valid:
                            all_errors.extend(result.errors)

                    valid = len(all_errors) == 0
                    if valid:
                        passed_cells += 1
                    else:
                        failed_cells += 1

                    cell_results.append({
                        "cell": cell.coordinate,
                        "valid": valid,
                        "errors": all_errors,
                    })

            wb.close()

            return {
                "file": filepath,
                "sheet": sheet_name,
                "range": range_str,
                "valid": failed_cells == 0,
                "total_cells": total_cells,
                "passed_cells": passed_cells,
                "failed_cells": failed_cells,
                "results": cell_results,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"验证引擎执行失败: {e}")
            raise DataError(f"数据操作失败: {e}")
