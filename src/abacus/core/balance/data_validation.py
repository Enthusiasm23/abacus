"""盈不足章 - 数据验证：设置单元格数据验证规则"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class DataValidationCapability(Capability):
    """数据验证：设置单元格数据验证规则"""

    @property
    def name(self) -> str:
        return "set_data_validation"

    @property
    def chapter(self) -> str:
        return "balance"

    @property
    def description(self) -> str:
        return "设置单元格数据验证（下拉列表、数值范围、日期范围等）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(
                name="validation_type",
                type="string",
                description="验证类型（/number/date/text_length）",
                required=True,
            ),
            CapabilitySchema(
                name="operator",
                type="string",
                description="运算符（between/notBetween/equal/notEqual等）",
                required=False,
            ),
            CapabilitySchema(
                name="formula1", type="string", description="验证公式1", required=False
            ),
            CapabilitySchema(
                name="formula2", type="string", description="验证公式2（between时）", required=False
            ),
            CapabilitySchema(
                name="error_message", type="string", description="错误提示消息", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        validation_type = params.get("validation_type")
        operator = params.get("operator")
        formula1 = params.get("formula1")
        formula2 = params.get("formula2")
        error_message = params.get("error_message")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")

        return self._set_validation(
            file_path,
            sheet_name,
            range_str,
            validation_type,
            operator,
            formula1,
            formula2,
            error_message,
        )

    def _set_validation(
        self,
        filepath: str,
        sheet_name: str,
        range_str: str,
        validation_type: str,
        operator: str = None,
        formula1: str = None,
        formula2: str = None,
        error_message: str = None,
    ) -> dict[str, Any]:
        """设置数据验证"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 创建数据验证规则
            if validation_type == "list":
                # 下拉列表
                if not formula1:
                    raise ValidationError("执行失败: 下拉列表验证需要 formula1 参数（逗号分隔的值）")
                dv = DataValidation(type="list", formula1=f'"{formula1}"', allow_blank=True)
            elif validation_type == "number":
                # 数值范围
                dv = DataValidation(
                    type="whole",
                    operator=operator or "between",
                    formula1=formula1 or "0",
                    formula2=formula2 or "100",
                    allow_blank=True,
                )
            elif validation_type == "date":
                # 日期范围
                dv = DataValidation(
                    type="date",
                    operator=operator or "between",
                    formula1=formula1 or "2020-01-01",
                    formula2=formula2 or "2030-12-31",
                    allow_blank=True,
                )
            elif validation_type == "text_length":
                # 文本长度
                dv = DataValidation(
                    type="textLength",
                    operator=operator or "between",
                    formula1=formula1 or "0",
                    formula2=formula2 or "100",
                    allow_blank=True,
                )
            else:
                raise DataError(f"数据操作失败: 不支持的验证类型 {validation_type}")

            # 设置错误消息
            if error_message:
                dv.error = error_message
                dv.errorTitle = "Validation Error"

            # 应用到范围
            dv.add(range_str)
            ws.add_data_validation(dv)

            wb.save(filepath)
            wb.close()

            return {
                "file": filepath,
                "sheet": sheet_name,
                "range": range_str,
                "validation_type": validation_type,
                "applied": True,
            }

        except (FileNotFoundError, DataError, ValidationError):
            raise
        except Exception as e:
            logger.error(f"数据验证设置失败: {e}")
            raise DataError(f"数据操作失败: {e}")
