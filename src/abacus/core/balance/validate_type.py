"""盈不足章 - 验类型：深度实现数据类型验证"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class ValidateTypeCapability(Capability):
    """验类型：深度实现数据类型验证"""

    TYPE_MAP = {
        "number": (int, float),
        "text": (str,),
        "date": (str,),  # Excel 日期存储为数字
        "boolean": (bool,),
    }

    @property
    def name(self) -> str:
        return "validate_type"

    @property
    def chapter(self) -> str:
        return "balance"

    @property
    def description(self) -> str:
        return "深度验证数据类型（列类型一致性、混合类型检测）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(
                name="expected_type",
                type="string",
                description="期望类型（number/text/date/boolean）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        expected_type = params.get("expected_type")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")

        return self._validate_type(file_path, sheet_name, range_str, expected_type)

    def _validate_type(
        self, filepath: str, sheet_name: str, range_str: str, expected_type: str = None
    ) -> dict[str, Any]:
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 按列分析类型
            column_types = {}
            issues = []

            for col in range(start_col, end_col + 1):
                col_letter = ws.cell(row=start_row, column=col).coordinate[0]
                type_counts = {}

                for row in range(start_row, end_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        type_name = type(cell.value).__name__
                        type_counts[type_name] = type_counts.get(type_name, 0) + 1

                if type_counts:
                    dominant_type = max(type_counts.items(), key=lambda x: x[1])
                    column_types[col_letter] = {
                        "dominant_type": dominant_type[0],
                        "counts": type_counts,
                        "is_mixed": len(type_counts) > 1,
                    }

                    # 检查混合类型
                    if len(type_counts) > 1:
                        issues.append(
                            {"type": "mixed_types", "column": col_letter, "types": type_counts}
                        )

                    # 检查期望类型
                    if expected_type and expected_type in self.TYPE_MAP:
                        expected_types = self.TYPE_MAP[expected_type]
                        for type_name, count in type_counts.items():
                            if type_name not in [t.__name__ for t in expected_types]:
                                issues.append(
                                    {
                                        "type": "unexpected_type",
                                        "column": col_letter,
                                        "expected": expected_type,
                                        "actual": type_name,
                                        "count": count,
                                    }
                                )

            wb.close()

            return {
                "file": filepath,
                "sheet": sheet_name,
                "range": range_str,
                "columns_analyzed": len(column_types),
                "issues_count": len(issues),
                "valid": len(issues) == 0,
                "column_types": column_types,
                "issues": issues,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"类型验证失败: {e}")
            raise DataError(f"数据操作失败: {e}")
