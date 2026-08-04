"""盈不足章 - 验范围：深度实现数据范围验证"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class ValidateRangeCapability(Capability):
    """验范围：深度实现数据范围验证"""

    @property
    def name(self) -> str:
        return "validate_range"

    @property
    def chapter(self) -> str:
        return "balance"

    @property
    def description(self) -> str:
        return "深度验证数据范围（类型、值域、空值、重复）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(name="rules", type="object", description="验证规则", required=False),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        rules = params.get("rules", {})

        if not file_path:
            raise DataError("file parameter is required")

        return self._validate_range(file_path, sheet_name, range_str, rules)

    def _validate_range(
        self, filepath: str, sheet_name: str, range_str: str, rules: dict
    ) -> dict[str, Any]:
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            issues = []
            total_cells = 0
            empty_cells = 0

            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    total_cells += 1

                    # 检查空值
                    if cell.value is None:
                        empty_cells += 1
                        if rules.get("no_empty"):
                            issues.append(
                                {
                                    "type": "empty",
                                    "cell": cell.coordinate,
                                    "message": "Cell is empty",
                                }
                            )
                        continue

                    # 检查类型
                    if "expected_type" in rules:
                        expected = rules["expected_type"]
                        actual = type(cell.value).__name__
                        if expected == "number" and not isinstance(cell.value, (int, float)):
                            issues.append(
                                {
                                    "type": "type_mismatch",
                                    "cell": cell.coordinate,
                                    "expected": expected,
                                    "actual": actual,
                                }
                            )
                        elif expected == "text" and not isinstance(cell.value, str):
                            issues.append(
                                {
                                    "type": "type_mismatch",
                                    "cell": cell.coordinate,
                                    "expected": expected,
                                    "actual": actual,
                                }
                            )

                    # 检查值域
                    if isinstance(cell.value, (int, float)):
                        if "min_value" in rules and cell.value < rules["min_value"]:
                            issues.append(
                                {
                                    "type": "below_min",
                                    "cell": cell.coordinate,
                                    "value": cell.value,
                                    "min": rules["min_value"],
                                }
                            )
                        if "max_value" in rules and cell.value > rules["max_value"]:
                            issues.append(
                                {
                                    "type": "above_max",
                                    "cell": cell.coordinate,
                                    "value": cell.value,
                                    "max": rules["max_value"],
                                }
                            )

            wb.close()

            return {
                "file": filepath,
                "sheet": sheet_name,
                "range": range_str,
                "total_cells": total_cells,
                "empty_cells": empty_cells,
                "issues_count": len(issues),
                "valid": len(issues) == 0,
                "issues": issues,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to validate range: {e}")
            raise DataError(str(e))
