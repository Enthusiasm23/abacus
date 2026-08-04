"""CSV 数据可视化 - 自动生成图表"""

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class CSVVisualizeCapability(Capability):
    """CSV 数据可视化 - 自动生成图表和仪表板"""

    @property
    def name(self) -> str:
        return "visualize_data"

    @property
    def chapter(self) -> str:
        return "triangle"

    @property
    def description(self) -> str:
        return "CSV 数据可视化（自动生成图表、仪表板、统计摘要）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="CSV/Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="output", type="string", description="输出文件路径", required=True
            ),
            CapabilitySchema(
                name="chart_type",
                type="string",
                description="图表类型（bar/line/pie/auto）",
                required=False,
            ),
            CapabilitySchema(
                name="include_dashboard",
                type="boolean",
                description="是否包含仪表板",
                required=False,
            ),
            CapabilitySchema(
                name="include_stats", type="boolean", description="是否包含统计摘要", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        output = params.get("output")
        chart_type = params.get("chart_type", "auto")
        include_dashboard = params.get("include_dashboard", True)
        include_stats = params.get("include_stats", True)

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")
        if not output:
            raise ValidationError("执行失败: 缺少必要参数 output")

        # 加载数据
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件操作失败: 文件不存在 {file_path}")

        suffix = path.suffix.lower()
        if suffix == ".csv":
            df = pd.read_csv(path)
        elif suffix in [".xlsx", ".xls"]:
            df = pd.read_excel(path)
        else:
            raise DataError(f"数据操作失败: 不支持的文件格式 {suffix}")

        # 创建可视化
        wb = Workbook()

        # Sheet 1: 原始数据
        ws_data = wb.active
        ws_data.title = "Data"
        self._write_data_sheet(ws_data, df)

        # Sheet 2: 图表
        ws_chart = wb.create_sheet("Charts")
        self._create_charts(ws_chart, df, chart_type)

        # Sheet 3: 统计摘要（可选）
        if include_stats:
            ws_stats = wb.create_sheet("Statistics")
            self._create_statistics(ws_stats, df)

        # Sheet 4: 仪表板（可选）
        if include_dashboard:
            ws_dashboard = wb.create_sheet("Dashboard")
            self._create_dashboard(ws_dashboard, df)

        # 保存
        wb.save(output)

        return {
            "file": file_path,
            "output": output,
            "sheets": len(wb.sheetnames),
            "rows": len(df),
            "columns": len(df.columns),
            "chart_type": chart_type,
        }

    def _write_data_sheet(self, ws, df):
        """写入数据表"""
        # 表头
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(patternType="solid", fgColor="2F5496")

        # 数据
        for row_idx, row in enumerate(df.iterrows(), 2):
            _, row_data = row
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _create_charts(self, ws, df, chart_type):
        """创建图表"""
        numeric_cols = df.select_dtypes(include=["number"]).columns

        if len(numeric_cols) == 0:
            ws["A1"] = "没有数值列，无法创建图表"
            return

        # 写入数据
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, row in enumerate(df.head(20).iterrows(), 2):
            _, row_data = row
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 创建图表
        if chart_type == "auto":
            # 自动选择图表类型
            if len(numeric_cols) == 1:
                chart_type = "bar"
            elif len(df) > 10:
                chart_type = "line"
            else:
                chart_type = "bar"

        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()

        chart.title = "数据图表"
        chart.style = 10
        chart.y_axis.title = "数值"
        chart.x_axis.title = "类别"

        # 添加数据系列
        data = Reference(
            ws,
            min_col=2,
            min_row=1,
            max_col=min(len(numeric_cols) + 1, len(df.columns)),
            max_row=min(21, len(df) + 1),
        )
        cats = Reference(ws, min_col=1, min_row=2, max_row=min(21, len(df) + 1))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "A" + str(min(25, len(df) + 3)))

    def _create_statistics(self, ws, df):
        """创建统计摘要"""
        ws["A1"] = "统计摘要"
        ws["A1"].font = Font(bold=True, size=14)

        # 基础统计
        ws["A3"] = "基础统计"
        ws["A3"].font = Font(bold=True)

        stats = [
            ("总行数", len(df)),
            ("总列数", len(df.columns)),
            ("数值列数", len(df.select_dtypes(include=["number"]).columns)),
            ("分类列数", len(df.select_dtypes(include=["object"]).columns)),
            ("缺失值总数", int(df.isnull().sum().sum())),
            ("重复行数", int(df.duplicated().sum())),
        ]

        for i, (name, value) in enumerate(stats, 4):
            ws[f"A{i}"] = name
            ws[f"B{i}"] = value

        # 数值列统计
        numeric_cols = df.select_dtypes(include=["number"]).columns
        if len(numeric_cols) > 0:
            ws[f"A{len(stats) + 5}"] = "数值列统计"
            ws[f"A{len(stats) + 5}"].font = Font(bold=True)

            for i, col in enumerate(numeric_cols, len(stats) + 6):
                ws[f"A{i}"] = col
                ws[f"B{i}"] = f"均值: {df[col].mean():.2f}"
                ws[f"C{i}"] = f"标准差: {df[col].std():.2f}"
                ws[f"D{i}"] = f"最小值: {df[col].min()}"
                ws[f"E{i}"] = f"最大值: {df[col].max()}"

    def _create_dashboard(self, ws, df):
        """创建仪表板"""
        ws["A1"] = "数据仪表板"
        ws["A1"].font = Font(bold=True, size=16)

        # KPI 指标
        numeric_cols = df.select_dtypes(include=["number"]).columns

        if len(numeric_cols) > 0:
            ws["A3"] = "关键指标"
            ws["A3"].font = Font(bold=True, size=12)

            for i, col in enumerate(numeric_cols[:5], 4):
                ws[f"A{i}"] = col
                ws[f"B{i}"] = df[col].sum()
                ws[f"B{i}"].number_format = "#,##0"
                ws[f"C{i}"] = df[col].mean()
                ws[f"C{i}"].number_format = "#,##0.00"
