"""商功章 - 打印区域：设置打印区域"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class PrintAreaCapability(Capability):
    """打印区域：设置打印区域"""

    @property
    def name(self) -> str:
        return "set_print_area"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "设置打印区域"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="range", type="string", description="打印区域（如 A1:D10）", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")

        if not file_path:
            raise DataError("file parameter is required")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(path)

        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        ws.print_area = range_str

        wb.save(file_path)
        wb.close()

        return {"file": file_path, "sheet": sheet_name, "print_area": range_str, "set": True}
