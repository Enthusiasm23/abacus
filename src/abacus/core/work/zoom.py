"""商功章 - 缩放：控制工作表缩放"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class ZoomCapability(Capability):
    """缩放：控制工作表缩放"""

    @property
    def name(self) -> str:
        return "set_zoom"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "控制工作表缩放（10-400%）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="zoom", type="number", description="缩放比例（10-400）", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        zoom = params.get("zoom")

        if not file_path:
            raise DataError("file parameter is required")

        if zoom < 10 or zoom > 400:
            raise DataError("Zoom must be between 10 and 400")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(path)

        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        ws.sheet_view.zoomScale = zoom

        wb.save(file_path)
        wb.close()

        return {"file": file_path, "sheet": sheet_name, "zoom": zoom, "set": True}
