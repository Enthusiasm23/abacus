"""商功章 - 工作簿和工作表保护"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.protection import SheetProtection

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class ProtectWorkbookCapability(Capability):
    """保护工作簿"""

    @property
    def name(self) -> str:
        return "protect_workbook"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "保护工作簿"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="password", type="string", description="保护密码", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        password = params.get("password")

        if not file_path:
            raise DataError("file parameter is required")

        return self._protect_workbook(file_path, password)

    def _protect_workbook(self, filepath: str, password: str = None) -> dict[str, Any]:
        """保护工作簿"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            # 设置工作簿保护
            wb.security = WorkbookProtection(workbookPassword=password, lockStructure=True)

            wb.save(filepath)
            wb.close()

            return {"success": True, "action": "protect_workbook", "file": filepath}

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to protect workbook: {e}")
            raise DataError(str(e))


class ProtectSheetCapability(Capability):
    """保护工作表"""

    @property
    def name(self) -> str:
        return "protect_sheet"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "保护工作表"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="password", type="string", description="保护密码", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        password = params.get("password")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")

        return self._protect_sheet(file_path, sheet_name, password)

    def _protect_sheet(
        self, filepath: str, sheet_name: str, password: str = None
    ) -> dict[str, Any]:
        """保护工作表"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            # 设置工作表保护
            ws.protection = SheetProtection(password=password, sheet=True)

            wb.save(filepath)
            wb.close()

            return {"success": True, "action": "protect_sheet", "sheet": sheet_name}

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to protect sheet: {e}")
            raise DataError(str(e))


class UnprotectSheetCapability(Capability):
    """解除工作表保护"""

    @property
    def name(self) -> str:
        return "unprotect_sheet"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "解除工作表保护"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="password", type="string", description="保护密码", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        password = params.get("password")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")

        return self._unprotect_sheet(file_path, sheet_name, password)

    def _unprotect_sheet(
        self, filepath: str, sheet_name: str, password: str = None
    ) -> dict[str, Any]:
        """解除工作表保护"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            # 解除工作表保护
            ws.protection = SheetProtection(sheet=False)

            wb.save(filepath)
            wb.close()

            return {"success": True, "action": "unprotect_sheet", "sheet": sheet_name}

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to unprotect sheet: {e}")
            raise DataError(str(e))


class SetArrayFormulaCapability(Capability):
    """设置数组公式"""

    @property
    def name(self) -> str:
        return "set_array_formula"

    @property
    def chapter(self) -> str:
        return "equation"

    @property
    def description(self) -> str:
        return "设置数组公式"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="range", type="string", description="数组公式范围", required=True
            ),
            CapabilitySchema(
                name="formula", type="string", description="数组公式内容", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        formula = params.get("formula")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if not range_str:
            raise DataError("range parameter is required")
        if not formula:
            raise DataError("formula parameter is required")

        return self._set_array_formula(file_path, sheet_name, range_str, formula)

    def _set_array_formula(
        self, filepath: str, sheet_name: str, range_str: str, formula: str
    ) -> dict[str, Any]:
        """设置数组公式"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            # 设置数组公式 - 解析范围，在第一个单元格设置公式
            from ..cell_utils import parse_range

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 将公式设置在范围的第一个单元格
            ws.cell(row=start_row, column=start_col).value = formula

            wb.save(filepath)
            wb.close()

            return {
                "success": True,
                "action": "set_array_formula",
                "sheet": sheet_name,
                "range": range_str,
                "formula": formula,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to set array formula: {e}")
            raise DataError(str(e))
