"""商功章 - 分组行"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class GroupRowsCapability(Capability):
    """分组行"""

    @property
    def name(self) -> str:
        return "group_rows"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "分组行（折叠/展开）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="start_row", type="integer", description="起始行号", required=True
            ),
            CapabilitySchema(name="end_row", type="integer", description="结束行号", required=True),
            CapabilitySchema(
                name="level", type="integer", description="分组层级", required=False, default=1
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        start_row = params.get("start_row")
        end_row = params.get("end_row")
        level = params.get("level", 1)

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if start_row is None:
            raise DataError("start_row parameter is required")
        if end_row is None:
            raise DataError("end_row parameter is required")
        if start_row < 1:
            raise DataError("start_row must be >= 1")
        if end_row < start_row:
            raise DataError("end_row must be >= start_row")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            for row_idx in range(start_row, end_row + 1):
                ws.row_dimensions[row_idx].outlineLevel = level

            wb.save(file_path)
            wb.close()

            return {
                "success": True,
                "start_row": start_row,
                "end_row": end_row,
                "level": level,
                "rows_grouped": end_row - start_row + 1,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to group rows: {e}")
            raise DataError(str(e))
