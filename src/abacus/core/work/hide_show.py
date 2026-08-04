"""商功章 - 隐藏显示：隐藏和显示行列"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class HideShowCapability(Capability):
    """隐藏显示：隐藏和显示行列"""

    @property
    def name(self) -> str:
        return "manage_visibility"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "管理行列可见性（隐藏/显示行和列）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="action", type="string", description="操作（hide/show）", required=True
            ),
            CapabilitySchema(
                name="dimension", type="string", description="维度（row/column）", required=True
            ),
            CapabilitySchema(name="index", type="number", description="行号或列号", required=True),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        action = params.get("action")
        dimension = params.get("dimension")
        index = params.get("index")

        if not file_path:
            raise DataError("file parameter is required")

        return self._manage_visibility(file_path, sheet_name, action, dimension, index)

    def _manage_visibility(
        self, filepath: str, sheet_name: str, action: str, dimension: str, index: int
    ) -> dict[str, Any]:
        """管理行列可见性"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            if dimension == "row":
                if action == "hide":
                    ws.row_dimensions[index].hidden = True
                elif action == "show":
                    ws.row_dimensions[index].hidden = False
                else:
                    raise DataError(f"Unknown action: {action}")
            elif dimension == "column":
                col_letter = get_column_letter(index)
                if action == "hide":
                    ws.column_dimensions[col_letter].hidden = True
                elif action == "show":
                    ws.column_dimensions[col_letter].hidden = False
                else:
                    raise DataError(f"Unknown action: {action}")
            else:
                raise DataError(f"Unknown dimension: {dimension}")

            wb.save(filepath)
            wb.close()

            return {"action": action, "dimension": dimension, "index": index, "applied": True}

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to manage visibility: {e}")
            raise DataError(str(e))
