"""商功章 - 高级筛选：支持复杂条件的数据筛选"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class AdvancedFilterCapability(Capability):
    """高级筛选：支持复杂条件的数据筛选"""

    @property
    def name(self) -> str:
        return "advanced_filter"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "高级筛选（支持 AND/OR/NOT 逻辑、比较运算符、文本匹配、数值和日期范围筛选）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="range",
                type="string",
                description="数据范围（如 A1:D100），可选，默认整个工作表",
                required=False,
            ),
            CapabilitySchema(
                name="conditions", type="object", description="筛选条件（JSON 对象）", required=True
            ),
            CapabilitySchema(
                name="return_type",
                type="string",
                description="返回类型：'data'（数据）或 'rows'（行号），默认 'data'",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行高级筛选"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        conditions = params.get("conditions")
        return_type = params.get("return_type", "data")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if not conditions:
            raise DataError("conditions parameter is required")

        return self._advanced_filter(file_path, sheet_name, range_str, conditions, return_type)

    def _advanced_filter(
        self,
        file_path: str,
        sheet_name: str,
        range_str: str,
        conditions: dict[str, Any],
        return_type: str,
    ) -> dict[str, Any]:
        """执行高级筛选"""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(path)
        try:
            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            if return_type not in ("data", "rows"):
                raise DataError("return_type must be 'data' or 'rows'")

            # 获取数据范围
            if range_str:
                from ..cell_utils import parse_range

                min_row, min_col, max_row, max_col = parse_range(range_str)
            else:
                min_row = 1
                max_row = ws.max_row
                min_col = 1
                max_col = ws.max_column

            # 读取表头（第一行）
            headers = []
            for col in range(min_col, max_col + 1):
                cell_value = ws.cell(row=min_row, column=col).value
                headers.append(cell_value if cell_value is not None else f"Col{col}")

            # 读取数据行
            data_rows = []
            row_numbers = []

            for row in range(min_row + 1, max_row + 1):
                row_data = {}
                for col_idx, col in enumerate(range(min_col, max_col + 1)):
                    row_data[headers[col_idx]] = ws.cell(row=row, column=col).value

                # 评估条件
                if self._evaluate_condition(conditions, row_data):
                    data_rows.append(row_data)
                    row_numbers.append(row)

            # 返回结果
            if return_type == "data":
                result = {
                    "sheet": sheet_name,
                    "range": range_str or f"A1:{ws.max_column}{ws.max_row}",
                    "headers": headers,
                    "rows": data_rows,
                    "total_matched": len(data_rows),
                }
            else:
                result = {
                    "sheet": sheet_name,
                    "range": range_str or f"A1:{ws.max_column}{ws.max_row}",
                    "headers": headers,
                    "row_numbers": row_numbers,
                    "total_matched": len(row_numbers),
                }

            return result

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Advanced filter failed: {e}")
            raise DataError(str(e))
        finally:
            wb.close()

    def _evaluate_condition(self, condition: dict[str, Any], row_data: dict[str, Any]) -> bool:
        """评估单个条件"""
        condition_type = condition.get("type")

        if condition_type == "group":
            logic = condition.get("logic", "AND").upper()
            sub_conditions = condition.get("conditions", [])

            if logic == "AND":
                return all(self._evaluate_condition(sub, row_data) for sub in sub_conditions)
            elif logic == "OR":
                return any(self._evaluate_condition(sub, row_data) for sub in sub_conditions)
            elif logic == "NOT":
                if len(sub_conditions) != 1:
                    raise DataError("NOT condition must have exactly one sub-condition")
                return not self._evaluate_condition(sub_conditions[0], row_data)
            else:
                raise DataError(f"Unsupported logic: {logic}")

        elif condition_type == "condition":
            field = condition.get("field")
            operator = condition.get("operator")
            value = condition.get("value")

            if field not in row_data:
                raise DataError(f"Field '{field}' not found in data")

            cell_value = row_data[field]
            return self._compare(cell_value, operator, value)

        else:
            raise DataError(f"Unknown condition type: {condition_type}")

    def _compare(self, cell_value: Any, operator: str, target_value: Any) -> bool:
        """比较两个值"""
        # 处理空值
        if cell_value is None:
            return operator in ("==", "is") and target_value is None

        # 类型转换
        try:
            if isinstance(target_value, (int, float)) and isinstance(cell_value, str):
                cell_value = float(cell_value)
            elif isinstance(target_value, str) and isinstance(cell_value, (int, float)):
                target_value = float(target_value)
        except (ValueError, TypeError):
            pass

        # 数值比较
        if operator in (">", "<", ">=", "<="):
            try:
                if isinstance(cell_value, (int, float)) and isinstance(target_value, (int, float)):
                    if operator == ">":
                        return cell_value > target_value
                    elif operator == "<":
                        return cell_value < target_value
                    elif operator == ">=":
                        return cell_value >= target_value
                    elif operator == "<=":
                        return cell_value <= target_value
                else:
                    raise DataError(f"Cannot compare non-numeric values with {operator}")
            except TypeError:
                return False

        # 相等比较
        elif operator in ("==", "!="):
            if operator == "==":
                return cell_value == target_value
            else:
                return cell_value != target_value

        # 文本比较
        elif operator in ("contains", "starts_with", "ends_with"):
            if not isinstance(cell_value, str):
                cell_value = str(cell_value) if cell_value is not None else ""
            if not isinstance(target_value, str):
                target_value = str(target_value) if target_value is not None else ""

            if operator == "contains":
                return target_value in cell_value
            elif operator == "starts_with":
                return cell_value.startswith(target_value)
            elif operator == "ends_with":
                return cell_value.endswith(target_value)

        # 范围比较（需要特殊处理）
        elif operator == "between":
            if not isinstance(target_value, (list, tuple)) or len(target_value) != 2:
                raise DataError(
                    "'between' operator requires a list or tuple of two values [min, max]"
                )

            min_val, max_val = target_value
            try:
                if isinstance(cell_value, (int, float)):
                    return min_val <= cell_value <= max_val
                else:
                    return False
            except TypeError:
                return False

        else:
            raise DataError(f"Unsupported operator: {operator}")

        return False
