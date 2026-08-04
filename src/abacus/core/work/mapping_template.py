"""商功章 - 模板生成：创建数据映射模板"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..base import Capability, CapabilitySchema

logger = logging.getLogger(__name__)


class CreateMappingTemplateCapability(Capability):
    """创建数据映射模板"""

    @property
    def name(self) -> str:
        return "create_mapping_template"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "创建数据映射模板，定义目标表与源表的映射关系"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="output", type="string", description="输出文件路径（可选，默认带时间戳）"
            ),
            CapabilitySchema(name="source_count", type="integer", description="源表数量（默认4）"),
            CapabilitySchema(name="quiet", type="boolean", description="静默模式，不输出日志"),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行创建模板"""
        output = params.get("output")
        source_count = params.get("source_count", 4)
        quiet = params.get("quiet", False)

        return self._create_template(source_count, output, quiet)

    def _create_template(self, source_count: int, output: str, quiet: bool) -> dict[str, Any]:
        """创建模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "{目标表中文名}"

        # 样式定义
        title_font = Font(bold=True, size=12, name="微软雅黑", color="FFFFFF")
        label_font = Font(bold=True, size=12, name="微软雅黑")
        placeholder_font = Font(size=12, name="微软雅黑", color="FF0000")
        data_font = Font(size=12, name="微软雅黑")

        target_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        source_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        field_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        data_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        def set_cell(cell, value, font=None, fill=None, align=None):
            cell.value = value
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if align:
                cell.alignment = align
            cell.border = thin_border

        # 动态计算
        base_rows = 5
        extra_rows = max(0, source_count - 4)
        data_last_row = base_rows + extra_rows
        field_row = data_last_row + 1

        # 标题行
        ws.merge_cells(f"A1:A{data_last_row}")
        set_cell(ws["A1"], "目标表", title_font, target_fill, center_align)
        set_cell(ws["B1"], "目标表英文名", label_font, data_fill, center_align)
        ws.merge_cells("C1:F1")
        set_cell(ws["C1"], "{目标表英文名}", placeholder_font, data_fill, center_align)
        ws.merge_cells(f"G1:G{data_last_row}")
        set_cell(ws["G1"], "源表", title_font, source_fill, center_align)

        for col, name in enumerate(
            ["源表中文名", "源表英文名", "关联关系", "过滤条件", "源表别名", "SCHEMA", "源表备注"],
            8,
        ):
            set_cell(ws.cell(row=1, column=col), name, label_font, data_fill, center_align)

        # 目标表属性
        set_cell(ws["B2"], "目标表中文名", label_font, data_fill, center_align)
        ws.merge_cells("C2:F2")
        set_cell(ws["C2"], "{目标表中文名}", placeholder_font, data_fill, center_align)

        set_cell(ws["B3"], "主键/分布键/粒度", label_font, data_fill, center_align)
        ws.merge_cells("C3:F3")
        set_cell(ws["C3"], "{主键}", placeholder_font, data_fill, center_align)

        set_cell(ws["B4"], "频率/调度/时间", label_font, data_fill, center_align)
        ws.merge_cells("C4:F4")
        set_cell(ws["C4"], "{调度策略}", placeholder_font, data_fill, center_align)

        set_cell(ws["B5"], "目标表业务定义", label_font, data_fill, center_align)
        if data_last_row > 5:
            ws.merge_cells(f"B5:B{data_last_row}")
            ws.merge_cells(f"C5:F{data_last_row}")
        else:
            ws.merge_cells("C5:F5")
        set_cell(ws["C5"], "{业务逻辑实体定义}", placeholder_font, data_fill, center_align)

        # 源表区域
        for row in range(2, data_last_row + 1):
            if row - 1 <= source_count:
                suffix = str(row - 1) if row > 2 else ""
                set_cell(
                    ws[f"H{row}"],
                    f"{{源表中文名{suffix}}}",
                    placeholder_font,
                    data_fill,
                    center_align,
                )
                set_cell(
                    ws[f"I{row}"],
                    f"{{源表英文名{suffix}}}",
                    placeholder_font,
                    data_fill,
                    center_align,
                )
                set_cell(ws[f"J{row}"], "{关联关系}", placeholder_font, data_fill, center_align)
                set_cell(ws[f"K{row}"], "{过滤条件}", placeholder_font, data_fill, center_align)
                set_cell(
                    ws[f"L{row}"], f"{{别名{suffix}}}", placeholder_font, data_fill, center_align
                )
                set_cell(
                    ws[f"M{row}"], f"{{SCHEMA{suffix}}}", placeholder_font, data_fill, center_align
                )
                set_cell(ws[f"N{row}"], "{源表备注}", placeholder_font, data_fill, center_align)
            else:
                for col in range(8, 15):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = data_fill
                    cell.border = thin_border
                    cell.alignment = center_align

        for row in range(1, data_last_row + 1):
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                if not cell.fill or cell.fill.patternType is None:
                    cell.fill = data_fill
                if not cell.border or cell.border.left.style is None:
                    cell.border = thin_border
                if not cell.alignment or cell.alignment.horizontal is None:
                    cell.alignment = center_align

        # 字段映射表
        headers = [
            "字段序号",
            "字段英文名",
            "字段中文名",
            "字段属性说明",
            "类型及长度(位数)",
            "是否可为空",
            "来源别名",
            "来源表名",
            "来源字段",
            "加工逻辑",
            "加工说明",
            "基线版本",
            "版本日期",
            "目标表备注",
        ]
        for col, h in enumerate(headers, 1):
            set_cell(ws.cell(row=field_row, column=col), h, title_font, field_fill, center_align)

        set_cell(ws[f"A{field_row + 1}"], 1, data_font, data_fill, center_align)
        set_cell(ws[f"G{field_row + 1}"], "{别名}", placeholder_font, data_fill, center_align)
        set_cell(ws[f"A{field_row + 2}"], 2, data_font, data_fill, center_align)
        set_cell(ws[f"G{field_row + 2}"], "{别名}", placeholder_font, data_fill, center_align)

        for row in [field_row + 1, field_row + 2]:
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                if not cell.fill or cell.fill.patternType is None:
                    cell.fill = data_fill
                if not cell.border or cell.border.left.style is None:
                    cell.border = thin_border
                if not cell.alignment or cell.alignment.horizontal is None:
                    cell.alignment = center_align

        # 样式设置
        for col, w in {
            "A": 9.75,
            "B": 18.625,
            "C": 11.875,
            "D": 14.125,
            "E": 16.375,
            "F": 11.875,
            "G": 9.75,
            "H": 14.5,
            "I": 14.5,
            "J": 11.125,
            "K": 11.125,
            "L": 9.75,
            "M": 12.875,
            "N": 11.875,
        }.items():
            ws.column_dimensions[col].width = w

        for row in range(1, 5):
            ws.row_dimensions[row].height = 18.0
        for row in range(5, data_last_row + 1):
            ws.row_dimensions[row].height = 17.25
        ws.row_dimensions[field_row].height = 18.0
        for row in [field_row + 1, field_row + 2]:
            ws.row_dimensions[row].height = 17.25

        # 填写说明Sheet
        ws2 = wb.create_sheet(title="填写说明")
        self._create_instructions_sheet(ws2, source_count)

        # 示例Sheet
        ws3 = wb.create_sheet(title="填写示例")
        self._create_example_sheet(ws3)

        # 保存
        if output is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output = Path(f"mapping_template_{timestamp}.xlsx")
        else:
            output = Path(output)

        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)

        if not quiet:
            logger.info(f"Created template: {output}")

        return {
            "output": str(output),
            "source_count": source_count,
            "sheets": wb.sheetnames,
        }

    def _create_instructions_sheet(self, ws, source_count: int):
        """创建填写说明Sheet"""
        title_font_doc = Font(bold=True, size=16, name="微软雅黑", color="4472C4")
        h1_font = Font(bold=True, size=14, name="微软雅黑", color="333333")
        h2_font = Font(bold=True, size=12, name="微软雅黑", color="4472C4")
        h3_font = Font(bold=True, size=11, name="微软雅黑", color="333333")
        normal_font = Font(size=11, name="微软雅黑")

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        docs = [
            ("A1", "数据映射模板 - 填写说明", title_font_doc),
            ("A2", "本文档详细说明如何填写数据映射模板。", normal_font),
            ("A4", "一、模板概述", h1_font),
            ("A5", "本模板用于定义数据仓库中【目标表】与【源表】之间的映射关系。", normal_font),
            ("A6", "模板包含表级映射信息和字段级映射信息。", normal_font),
            ("A7", "所有需要填写的地方都用 {占位符} 格式标记，红色字体显示。", normal_font),
            ("A9", "二、表格结构详解", h1_font),
            ("A11", "【Row 1: 标题行】", h2_font),
            ("A12", "  A1: 目标表标题 | G1: 源表标题", normal_font),
            ("A13", "  C1:F1: 目标表英文名 | H1-N1: 源表属性名", normal_font),
            ("A15", "【Row 2-5: 数据区域】", h2_font),
            ("A16", "  Row 2: 目标表中文名 + 源表1", normal_font),
            ("A17", "  Row 3: 主键 + 源表2", normal_font),
            ("A18", "  Row 4: 调度策略 + 源表3", normal_font),
            ("A19", "  Row 5: 业务定义 + 源表4+", normal_font),
            ("A21", "【字段映射表】", h2_font),
            ("A22", "  定义字段级别的映射关系", normal_font),
            ("A24", "三、占位符说明", h1_font),
            (
                "A25",
                "  {目标表中文名}, {目标表英文名}, {主键}, {调度策略}, {业务逻辑实体定义}",
                normal_font,
            ),
            (
                "A26",
                "  {源表中文名}, {源表英文名}, {关联关系}, {过滤条件}, {别名}, {SCHEMA}",
                normal_font,
            ),
            ("A28", "四、注意事项", h1_font),
            ("A29", "  1. Sheet名称最长31个字符", normal_font),
            ("A30", "  2. 每个源表的别名必须唯一", normal_font),
        ]

        for ref, text, font in docs:
            ws[ref] = text
            ws[ref].font = font

        for row in range(1, 31):
            if ws[f"A{row}"].value:
                ws.row_dimensions[row].height = 20
        ws.row_dimensions[1].height = 30
        ws.column_dimensions["A"].width = 80

    def _create_example_sheet(self, ws):
        """创建填写示例Sheet"""
        title_font = Font(bold=True, size=12, name="微软雅黑", color="FFFFFF")
        label_font = Font(bold=True, size=12, name="微软雅黑")
        data_font = Font(size=12, name="微软雅黑")
        field_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        data_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        target_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        source_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        no_wrap_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        def set_cell(cell, value, font=None, fill=None, align=None):
            cell.value = value
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if align:
                cell.alignment = align
            cell.border = thin_border

        # Row 1
        ws.merge_cells("A1:A5")
        set_cell(ws["A1"], "目标表", title_font, target_fill, center_align)
        set_cell(ws["B1"], "目标表英文名", label_font, data_fill, center_align)
        ws.merge_cells("C1:F1")
        set_cell(ws["C1"], "dwd_user_order_detail", data_font, data_fill, no_wrap_center)
        ws.merge_cells("G1:G5")
        set_cell(ws["G1"], "源表", title_font, source_fill, center_align)
        for col, name in enumerate(
            ["源表中文名", "源表英文名", "关联关系", "过滤条件", "源表别名", "SCHEMA", "源表备注"],
            8,
        ):
            set_cell(ws.cell(row=1, column=col), name, label_font, data_fill, no_wrap_center)

        # Row 2-5
        for row in range(2, 6):
            for col in range(2, 7):
                cell = ws.cell(row=row, column=col)
                cell.fill = data_fill
                cell.border = thin_border
                cell.alignment = no_wrap_center

        set_cell(ws["B2"], "目标表中文名", label_font, data_fill, no_wrap_center)
        ws.merge_cells("C2:F2")
        set_cell(ws["C2"], "用户订单明细表", data_font, data_fill, no_wrap_center)
        set_cell(ws["B3"], "主键/分布键/粒度", label_font, data_fill, no_wrap_center)
        ws.merge_cells("C3:F3")
        set_cell(ws["C3"], "order_detail_id", data_font, data_fill, no_wrap_center)
        set_cell(ws["B4"], "频率/调度/时间", label_font, data_fill, no_wrap_center)
        ws.merge_cells("C4:F4")
        set_cell(ws["C4"], "T+1 | 0点 | 每日", data_font, data_fill, no_wrap_center)
        set_cell(ws["B5"], "目标表业务定义", label_font, data_fill, no_wrap_center)
        ws.merge_cells("C5:F5")
        set_cell(ws["C5"], "用户订单明细宽表", data_font, data_fill, no_wrap_center)

        # 源表
        sources = [
            (
                "订单主表",
                "ods_order_master",
                "form t1",
                "where t1.del_flag='N'",
                "t1",
                "sdiipd",
                "订单主表",
            ),
            (
                "订单明细表",
                "ods_order_detail",
                "ON t1.order_id=t2.order_id",
                "where t2.del_flag='N'",
                "t2",
                "sdiipd",
                "订单明细",
            ),
            (
                "商品信息表",
                "ods_product",
                "ON t2.product_id=t3.product_id",
                "where t3.del_flag='N'",
                "t3",
                "sdiipd",
                "商品主数据",
            ),
            (
                "用户信息表",
                "ods_user",
                "ON t1.user_id=t4.user_id",
                "where t4.del_flag='N'",
                "t4",
                "sdiipd",
                "用户主数据",
            ),
        ]
        for i, src in enumerate(sources, 2):
            set_cell(ws.cell(row=i, column=8), src[0], data_font, data_fill, no_wrap_center)
            set_cell(ws.cell(row=i, column=9), src[1], data_font, data_fill, no_wrap_center)
            set_cell(ws.cell(row=i, column=10), src[2], data_font, data_fill, no_wrap_center)
            set_cell(ws.cell(row=i, column=11), src[3], data_font, data_fill, no_wrap_center)
            set_cell(ws.cell(row=i, column=12), src[4], data_font, data_fill, no_wrap_center)
            set_cell(ws.cell(row=i, column=13), src[5], data_font, data_fill, no_wrap_center)
            set_cell(ws.cell(row=i, column=14), src[6], data_font, data_fill, no_wrap_center)

        # 字段映射表头
        for col, h in enumerate(
            [
                "字段序号",
                "字段英文名",
                "字段中文名",
                "字段属性说明",
                "类型及长度(位数)",
                "是否可为空",
                "来源别名",
                "来源表名",
                "来源字段",
                "加工逻辑",
                "加工说明",
                "基线版本",
                "版本日期",
                "目标表备注",
            ],
            1,
        ):
            set_cell(ws.cell(row=6, column=col), h, title_font, field_fill, center_align)

        # 示例字段
        fields = [
            [
                1,
                "order_detail_id",
                "订单明细ID",
                "唯一标识",
                "varchar(64)",
                "NO",
                "t2",
                "ods_order_detail",
                "detail_id",
                "t2.detail_id",
                "直接映射",
                "v1.0",
                "2024-01-15",
                "主键",
            ],
            [
                2,
                "order_id",
                "订单号",
                "订单标识",
                "varchar(64)",
                "NO",
                "t1",
                "ods_order_master",
                "order_id",
                "t1.order_id",
                "直接映射",
                "v1.0",
                "2024-01-15",
                "关联键",
            ],
            [
                3,
                "user_id",
                "用户ID",
                "用户标识",
                "varchar(64)",
                "NO",
                "t1",
                "ods_order_master",
                "user_id",
                "t1.user_id",
                "直接映射",
                "v1.0",
                "2024-01-15",
                "关联键",
            ],
            [
                4,
                "product_name",
                "商品名称",
                "商品名",
                "varchar(200)",
                "NO",
                "t3",
                "ods_product",
                "product_name",
                "t3.product_name",
                "直接映射",
                "v1.0",
                "2024-01-15",
                "",
            ],
            [
                5,
                "quantity",
                "购买数量",
                "数量",
                "int",
                "NO",
                "t2",
                "ods_order_detail",
                "quantity",
                "t2.quantity",
                "直接映射",
                "v1.0",
                "2024-01-15",
                "",
            ],
            [
                6,
                "total_amount",
                "订单金额",
                "金额",
                "decimal(12,2)",
                "NO",
                "t1",
                "ods_order_master",
                "total_amount",
                "t1.total_amount",
                "直接映射",
                "v1.0",
                "2024-01-15",
                "",
            ],
            [
                7,
                "order_status",
                "订单状态",
                "状态",
                "varchar(20)",
                "NO",
                "t1",
                "ods_order_master",
                "status",
                "CASE t1.status WHEN 0 THEN '待付款' WHEN 1 THEN '已付款' ELSE '未知' END",
                "状态码转中文",
                "v1.0",
                "2024-01-15",
                "",
            ],
            [
                8,
                "user_name",
                "用户姓名",
                "姓名",
                "varchar(50)",
                "YES",
                "t4",
                "ods_user",
                "user_name",
                "t4.user_name",
                "直接映射",
                "v1.0",
                "2024-01-15",
                "",
            ],
        ]
        for row_idx, field in enumerate(fields, 7):
            for col_idx, value in enumerate(field, 1):
                set_cell(
                    ws.cell(row=row_idx, column=col_idx),
                    value,
                    data_font,
                    data_fill,
                    no_wrap_center
                    if col_idx in [1, 6, 7, 12, 13]
                    else Alignment(horizontal="left", vertical="center"),
                )

        # 列宽和行高
        for col, w in {
            "A": 9.75,
            "B": 18.625,
            "C": 11.875,
            "D": 14.125,
            "E": 16.375,
            "F": 11.875,
            "G": 9.75,
            "H": 14.5,
            "I": 14.5,
            "J": 11.125,
            "K": 11.125,
            "L": 9.75,
            "M": 12.875,
            "N": 11.875,
        }.items():
            ws.column_dimensions[col].width = w

        for row in range(1, 6):
            ws.row_dimensions[row].height = 18.0
        ws.row_dimensions[6].height = 18.0
        for row in range(7, 15):
            ws.row_dimensions[row].height = 17.25
