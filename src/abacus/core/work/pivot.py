"""商功章 - 透视表：深度实现透视表功能"""

import logging
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class CreatePivotCapability(Capability):
    """建透视表：创建数据透视表"""

    @property
    def name(self) -> str:
        return "create_pivot"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "创建数据透视表"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="sheet", type="string", description="源数据工作表", required=True
            ),
            CapabilitySchema(name="range", type="string", description="源数据范围", required=True),
            CapabilitySchema(
                name="row_fields", type="array", description="行字段列表", required=True
            ),
            CapabilitySchema(
                name="value_field", type="string", description="值字段", required=True
            ),
            CapabilitySchema(
                name="agg_function",
                type="string",
                description="聚合函数（sum/avg/count/min/max）",
                required=False,
            ),
            CapabilitySchema(
                name="output_sheet", type="string", description="输出工作表名称", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行透视表创建"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        row_fields = params.get("row_fields", [])
        value_field = params.get("value_field")
        agg_function = params.get("agg_function", "sum")
        output_sheet = params.get("output_sheet")

        if not file_path:
            raise DataError("file parameter is required")
        if not row_fields:
            raise DataError("row_fields parameter is required")
        if not value_field:
            raise DataError("value_field parameter is required")

        return self._create_pivot(
            file_path, sheet_name, range_str, row_fields, value_field, agg_function, output_sheet
        )

    def _create_pivot(
        self,
        filepath: str,
        sheet_name: str,
        range_str: str,
        row_fields: list[str],
        value_field: str,
        agg_function: str,
        output_sheet: str = None,
    ) -> dict[str, Any]:
        """创建透视表"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            # 解析范围获取数据
            from ..cell_utils import parse_range

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 读取表头
            headers = []
            for col in range(start_col, end_col + 1):
                headers.append(ws.cell(row=start_row, column=col).value)

            # 读取数据
            data = []
            for row in range(start_row + 1, end_row + 1):
                row_data = {}
                for col in range(start_col, end_col + 1):
                    header = headers[col - start_col]
                    cell_value = ws.cell(row=row, column=col).value
                    # 跳过公式单元格
                    if isinstance(cell_value, str) and cell_value.startswith("="):
                        row_data[header] = None
                    else:
                        row_data[header] = cell_value
                data.append(row_data)

            # 执行透视
            pivot_result = self._pivot_data(data, headers, row_fields, value_field, agg_function)

            # 创建输出工作表
            if not output_sheet:
                output_sheet = f"Pivot_{uuid.uuid4().hex[:8]}"

            if output_sheet in wb.sheetnames:
                del wb[output_sheet]

            pivot_ws = wb.create_sheet(output_sheet)

            # 写入透视表
            # 表头
            for i, field in enumerate(row_fields + [f"{agg_function}({value_field})"]):
                pivot_ws.cell(row=1, column=i + 1, value=field)

            # 数据
            for row_idx, row_data in enumerate(pivot_result, 2):
                for col_idx, value in enumerate(row_data, 1):
                    pivot_ws.cell(row=row_idx, column=col_idx, value=value)

            # 创建表格样式
            pivot_range = f"A1:{get_column_letter(len(row_fields) + 1)}{len(pivot_result) + 1}"
            pivot_table = Table(displayName=f"PivotTable_{uuid.uuid4().hex[:8]}", ref=pivot_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            pivot_table.tableStyleInfo = style
            pivot_ws.add_table(pivot_table)

            wb.save(filepath)
            wb.close()

            return {
                "file": filepath,
                "output_sheet": output_sheet,
                "rows": len(pivot_result),
                "columns": len(row_fields) + 1,
                "agg_function": agg_function,
                "pivot_range": pivot_range,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to create pivot table: {e}")
            raise DataError(str(e))

    def _pivot_data(
        self,
        data: list,
        headers: list[str],
        row_fields: list[str],
        value_field: str,
        agg_function: str,
    ) -> []:
        """执行透视计算"""
        from collections import defaultdict

        # 按行字段分组
        groups = defaultdict(list)
        for row in data:
            key = tuple(row.get(f) for f in row_fields)
            groups[key].append(row)

        # 聚合计算
        result = []
        for key, rows in groups.items():
            values = [row.get(value_field) for row in rows if row.get(value_field) is not None]

            if not values:
                agg_value = 0
            elif agg_function == "sum":
                agg_value = sum(float(v) for v in values)
            elif agg_function == "avg":
                agg_value = sum(float(v) for v in values) / len(values)
            elif agg_function == "count":
                agg_value = len(values)
            elif agg_function == "min":
                agg_value = min(float(v) for v in values)
            elif agg_function == "max":
                agg_value = max(float(v) for v in values)
            else:
                raise DataError(f"Unknown aggregation function: {agg_function}")

            result.append(list(key) + [agg_value])

        # 排序
        result.sort(key=lambda x: x[:-1])

        return result
