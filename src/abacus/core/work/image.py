"""商功章 - 插入图片到单元格"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class InsertImageCapability(Capability):
    """插入图片到单元格"""

    @property
    def name(self) -> str:
        return "insert_excel_image"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "插入图片到单元格"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="cell", type="string", description="单元格位置", required=True),
            CapabilitySchema(
                name="image_path", type="string", description="图片文件路径", required=True
            ),
            CapabilitySchema(
                name="width", type="integer", description="图片宽度（像素）", required=False
            ),
            CapabilitySchema(
                name="height", type="integer", description="图片高度（像素）", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        cell = params.get("cell")
        image_path = params.get("image_path")
        width = params.get("width")
        height = params.get("height")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if not cell:
            raise DataError("cell parameter is required")
        if not image_path:
            raise DataError("image_path parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            img_path = Path(image_path)
            if not img_path.exists():
                raise FileNotFoundError(f"Image file not found: {image_path}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            img = XlImage(str(img_path))

            if width:
                img.width = width
            if height:
                img.height = height

            ws.add_image(img, cell)

            wb.save(file_path)
            wb.close()

            return {
                "success": True,
                "cell": cell,
                "image_path": str(image_path),
                "width": img.width,
                "height": img.height,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to insert image: {e}")
            raise DataError(str(e))
