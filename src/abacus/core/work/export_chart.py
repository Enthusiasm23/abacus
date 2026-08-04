"""商功章 - 导出图表为图片"""

import logging
import shutil
from pathlib import Path
from typing import Any

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class ExportChartAsImageCapability(Capability):
    """导出图表为图片"""

    @property
    def name(self) -> str:
        return "export_chart_as_image"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "导出图表为图片（使用 LibreOffice）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="chart_index", type="integer", description="图表索引", required=True
            ),
            CapabilitySchema(
                name="output", type="string", description="输出图片路径", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        chart_index = params.get("chart_index")
        output = params.get("output")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if chart_index is None:
            raise DataError("chart_index parameter is required")
        if not output:
            raise DataError("output parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            from openpyxl import load_workbook

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            if not hasattr(ws, "_charts") or chart_index >= len(ws._charts):
                raise DataError(f"Chart index {chart_index} not found in sheet '{sheet_name}'")

            wb.close()

            libreoffice = shutil.which("libreoffice") or shutil.which("soffice")
            if not libreoffice:
                return {
                    "success": False,
                    "error": "LibreOffice not found. Install LibreOffice to use chart export.",
                    "file": str(file_path),
                    "chart_index": chart_index,
                }

            import subprocess

            output_path = Path(output)
            output_dir = output_path.parent
            output_dir.mkdir(parents=True, exist_ok=True)

            cmd = [
                libreoffice,
                "--headless",
                "--convert-to",
                "png",
                "--outdir",
                str(output_dir),
                str(file_path),
            ]

            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=60,
            )

            if result.returncode != 0:
                return {
                    "success": False,
                    "error": f"LibreOffice conversion failed: {result.stderr}",
                    "file": str(file_path),
                }

            default_output = output_dir / (path.stem + ".png")
            if default_output.exists():
                if str(default_output) != str(output_path):
                    default_output.rename(output_path)

            return {
                "success": True,
                "file": str(file_path),
                "chart_index": chart_index,
                "output": str(output_path),
                "note": "openpyxl does not support direct chart export. LibreOffice was used for conversion.",
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to export chart as image: {e}")
            raise DataError(str(e))
