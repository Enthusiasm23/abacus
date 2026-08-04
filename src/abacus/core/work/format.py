"""商功章 - 格式化：深度实现单元格格式化"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class FormatRangeCapability(Capability):
    """格式化：深度实现单元格格式化"""

    @property
    def name(self) -> str:
        return "format_range"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "深度格式化单元格（字体、颜色、边框、条件格式等）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(
                name="font",
                type="object",
                description="字体设置（name, size, bold, italic, color）",
                required=False,
            ),
            CapabilitySchema(
                name="fill",
                type="object",
                description="填充设置（color, pattern_type）",
                required=False,
            ),
            CapabilitySchema(
                name="border", type="object", description="边框设置（style, color）", required=False
            ),
            CapabilitySchema(
                name="alignment",
                type="object",
                description="对齐设置（horizontal, vertical, wrap_text）",
                required=False,
            ),
            CapabilitySchema(
                name="number_format", type="string", description="数字格式", required=False
            ),
            CapabilitySchema(
                name="conditional", type="object", description="条件格式设置", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行格式化"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        font_params = params.get("font")
        fill_params = params.get("fill")
        border_params = params.get("border")
        alignment_params = params.get("alignment")
        number_format = params.get("number_format")
        conditional = params.get("conditional")

        if not file_path:
            raise DataError("file parameter is required")

        return self._format_range(
            file_path,
            sheet_name,
            range_str,
            font_params,
            fill_params,
            border_params,
            alignment_params,
            number_format,
            conditional,
        )

    def _format_range(
        self,
        filepath: str,
        sheet_name: str,
        range_str: str,
        font_params: dict = None,
        fill_params: dict = None,
        border_params: dict = None,
        alignment_params: dict = None,
        number_format: str = None,
        conditional: dict = None,
    ) -> dict:
        """执行格式化"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            from ..cell_utils import parse_range

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            count = 0
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)

                    # 应用字体
                    if font_params:
                        font_args = {}
                        if "name" in font_params:
                            font_args["name"] = font_params["name"]
                        if "size" in font_params:
                            font_args["size"] = font_params["size"]
                        if "bold" in font_params:
                            font_args["bold"] = font_params["bold"]
                        if "italic" in font_params:
                            font_args["italic"] = font_params["italic"]
                        if "color" in font_params:
                            color = font_params["color"]
                            if color.startswith("#"):
                                color = color[1:]
                            if not color.startswith("FF"):
                                color = f"FF{color}"
                            font_args["color"] = Color(rgb=color)
                        cell.font = Font(**font_args)

                    # 应用填充
                    if fill_params:
                        fill_args = {}
                        if "color" in fill_params:
                            color = fill_params["color"]
                            if color.startswith("#"):
                                color = color[1:]
                            if not color.startswith("FF"):
                                color = f"FF{color}"
                            fill_args["fgColor"] = Color(rgb=color)
                        if "pattern_type" in fill_params:
                            fill_args["patternType"] = fill_params["pattern_type"]
                        else:
                            fill_args["patternType"] = "solid"
                        cell.fill = PatternFill(**fill_args)

                    # 应用边框
                    if border_params:
                        border_args = {}
                        for side_name in ["left", "right", "top", "bottom"]:
                            if side_name in border_params:
                                side_params = border_params[side_name]
                                side_args = {}
                                if "style" in side_params:
                                    side_args["style"] = side_params["style"]
                                if "color" in side_params:
                                    color = side_params["color"]
                                    if color.startswith("#"):
                                        color = color[1:]
                                    if not color.startswith("FF"):
                                        color = f"FF{color}"
                                    side_args["color"] = Color(rgb=color)
                                border_args[side_name] = Side(**side_args)
                        cell.border = Border(**border_args)

                    # 应用对齐
                    if alignment_params:
                        align_args = {}
                        if "horizontal" in alignment_params:
                            align_args["horizontal"] = alignment_params["horizontal"]
                        if "vertical" in alignment_params:
                            align_args["vertical"] = alignment_params["vertical"]
                        if "wrap_text" in alignment_params:
                            align_args["wrap_text"] = alignment_params["wrap_text"]
                        cell.alignment = Alignment(**align_args)

                    # 应用数字格式
                    if number_format:
                        cell.number_format = number_format

                    count += 1

            # 应用条件格式
            if conditional:
                self._apply_conditional_formatting(ws, range_str, conditional)

            wb.save(filepath)
            wb.close()

            return {
                "file": filepath,
                "sheet": sheet_name,
                "range": range_str,
                "cells_formatted": count,
                "has_conditional": conditional is not None,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to format range: {e}")
            raise DataError(str(e))

    def _apply_conditional_formatting(self, ws, range_str: str, conditional: dict):
        """应用条件格式"""
        rule_type = conditional.get("type")

        if rule_type == "color_scale":
            colors = conditional.get("colors", ["F8696B", "FFEB84", "63BE7B"])
            rule = ColorScaleRule(
                start_type="min",
                start_color=colors[0],
                mid_type="percentile",
                mid_value=50,
                mid_color=colors[1],
                end_type="max",
                end_color=colors[2],
            )
            ws.conditional_formatting.add(range_str, rule)

        elif rule_type == "data_bar":
            color = conditional.get("color", "638EC6")
            rule = DataBarRule(start_type="min", end_type="max", color=color)
            ws.conditional_formatting.add(range_str, rule)

        elif rule_type == "cell_is":
            operator = conditional.get("operator", "greaterThan")
            formula = conditional.get("formula", ["0"])
            fill_color = conditional.get("fill_color", "FF0000")
            if not fill_color.startswith("FF"):
                fill_color = f"FF{fill_color}"
            rule = CellIsRule(
                operator=operator, formula=formula, fill=PatternFill(bgColor=Color(rgb=fill_color))
            )
            ws.conditional_formatting.add(range_str, rule)
