"""商功章 - 图表：深度实现图表功能"""

import logging
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.chart.label import DataLabelList

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)

CHART_TYPES = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "area": AreaChart,
    "scatter": ScatterChart,
}


class CreateChartCapability(Capability):
    """建图表：深度实现图表创建"""

    @property
    def name(self) -> str:
        return "create_chart"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "深度创建图表（柱形图、折线图、饼图、面积图、散点图）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="sheet", type="string", description="数据工作表名称", required=True
            ),
            CapabilitySchema(
                name="range", type="string", description="数据范围（含表头）", required=True
            ),
            CapabilitySchema(
                name="chart_type",
                type="string",
                description="图表类型（bar/line/pie/area/scatter）",
                required=True,
            ),
            CapabilitySchema(name="title", type="string", description="图表标题", required=False),
            CapabilitySchema(name="x_axis", type="string", description="X轴标题", required=False),
            CapabilitySchema(name="y_axis", type="string", description="Y轴标题", required=False),
            CapabilitySchema(
                name="output_sheet",
                type="string",
                description="输出工作表（可选，默认新建）",
                required=False,
            ),
            CapabilitySchema(
                name="position", type="string", description="图表位置（如 A1）", required=False
            ),
            CapabilitySchema(
                name="width", type="number", description="图表宽度（厘米）", required=False
            ),
            CapabilitySchema(
                name="height", type="number", description="图表高度（厘米）", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        chart_type = params.get("chart_type")
        title = params.get("title")
        x_axis = params.get("x_axis")
        y_axis = params.get("y_axis")
        output_sheet = params.get("output_sheet")
        position = params.get("position", "A1")
        width = params.get("width", 15)
        height = params.get("height", 10)

        if not file_path:
            raise DataError("file parameter is required")
        if not chart_type:
            raise DataError("chart_type parameter is required")
        if chart_type not in CHART_TYPES:
            raise DataError(
                f"Unsupported chart type: {chart_type}. Supported: {(CHART_TYPES.keys())}"
            )

        return self._create_chart(
            file_path,
            sheet_name,
            range_str,
            chart_type,
            title,
            x_axis,
            y_axis,
            output_sheet,
            position,
            width,
            height,
        )

    def _create_chart(
        self,
        filepath: str,
        sheet_name: str,
        range_str: str,
        chart_type: str,
        title: str = None,
        x_axis: str = None,
        y_axis: str = None,
        output_sheet: str = None,
        position: str = "A1",
        width: float = 15,
        height: float = 10,
    ) -> dict[str, Any]:
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            num_cols = end_col - start_col + 1
            num_rows = end_row - start_row + 1

            if num_cols < 2 or num_rows < 2:
                raise DataError("Data range must have at least 2 columns and 2 rows")

            chart_class = CHART_TYPES[chart_type]
            chart = chart_class()

            if title:
                chart.title = title
            if x_axis:
                chart.x_axis.title = x_axis
            if y_axis:
                chart.y_axis.title = y_axis

            chart.width = width * 0.5
            chart.height = height * 0.5

            cats = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=end_row)

            for col in range(start_col + 1, end_col + 1):
                data = Reference(ws, min_col=col, min_row=start_row, max_row=end_row)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)

            if chart_type == "pie":
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showPercent = True
                chart.dataLabels.showCatName = True

            if output_sheet:
                if output_sheet in wb.sheetnames:
                    ws_output = wb[output_sheet]
                else:
                    ws_output = wb.create_sheet(output_sheet)
            else:
                ws_output = ws

            chart_id = uuid.uuid4().hex[:8]
            ws_output.add_chart(chart, position)

            wb.save(filepath)
            wb.close()

            return {
                "file": filepath,
                "chart_type": chart_type,
                "title": title,
                "data_range": range_str,
                "output_sheet": output_sheet or sheet_name,
                "position": position,
                "size": f"{width}x{height} cm",
                "chart_id": chart_id,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to create chart: {e}")
            raise DataError(str(e))


class UpdateChartCapability(Capability):
    @property
    def name(self) -> str:
        return "update_chart"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "深度更新图表"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="sheet", type="string", description="图表所在工作表", required=True
            ),
            CapabilitySchema(
                name="chart_index", type="number", description="图表索引（从0开始）", required=True
            ),
            CapabilitySchema(name="title", type="string", description="新标题", required=False),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        chart_index = params.get("chart_index", 0)
        title = params.get("title")

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]
            charts = ws._charts

            if chart_index >= len(charts):
                raise DataError(f"Chart index {chart_index} out of range")

            chart = charts[chart_index]

            if title:
                chart.title = title

            wb.save(file_path)
            wb.close()

            return {
                "file": file_path,
                "sheet": sheet_name,
                "chart_index": chart_index,
                "updated": True,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to update chart: {e}")
            raise DataError(str(e))


class ListChartsCapability(Capability):
    @property
    def name(self) -> str:
        return "list_charts"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "深度列出工作表中的所有图表"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=False),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path, data_only=True)

            result = {"file": file_path, "charts": []}

            for name in wb.sheetnames:
                if sheet_name and name != sheet_name:
                    continue

                ws = wb[name]
                charts = ws._charts

                for i, chart in enumerate(charts):
                    chart_info = {
                        "index": i,
                        "sheet": name,
                        "type": type(chart).__name__,
                        "title": chart.title,
                    }
                    result["charts"].append(chart_info)

            wb.close()
            result["total"] = len(result["charts"])

            return result

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to  charts: {e}")
            raise DataError(str(e))


class DeleteChartCapability(Capability):
    @property
    def name(self) -> str:
        return "delete_chart"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "深度删除图表"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="chart_index", type="number", description="图表索引（从0开始）", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        chart_index = params.get("chart_index", 0)

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]
            charts = ws._charts

            if chart_index >= len(charts):
                raise DataError(f"Chart index {chart_index} out of range")

            deleted_chart = charts.pop(chart_index)

            wb.save(file_path)
            wb.close()

            return {
                "file": file_path,
                "sheet": sheet_name,
                "chart_index": chart_index,
                "deleted": True,
                "chart_type": type(deleted_chart).__name__,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to delete chart: {e}")
            raise DataError(str(e))
