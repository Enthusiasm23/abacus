"""商功章 - 批量执行：执行多个操作"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class BatchExecuteCapability(Capability):
    """批量执行：执行多个操作"""

    MAX_OPERATIONS = 1000  # 最大操作数量限制

    @property
    def name(self) -> str:
        return "batch_execute"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "批量执行多个操作（合并、样式、写入等）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="operations", type="array", description="操作列表", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行批量操作"""
        file_path = params.get("file")
        operations = params.get("operations", [])

        if not file_path:
            raise DataError("file parameter is required")

        if len(operations) > self.MAX_OPERATIONS:
            raise ValidationError(
                f"操作数量超出限制: {len(operations)} > {self.MAX_OPERATIONS}"
            )

        return self._batch_execute(file_path, operations)

    def _batch_execute(self, filepath: str, operations: list) -> dict[str, Any]:
        """执行批量操作"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)
            results = []

            for i, op in enumerate(operations):
                try:
                    op_type = op.get("type")
                    sheet_name = op.get("sheet")

                    if sheet_name and sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.active

                    if op_type == "merge":
                        result = self._merge_cells(ws, op)
                    elif op_type == "unmerge":
                        result = self._unmerge_cells(ws, op)
                    elif op_type == "write":
                        result = self._write_cell(ws, op)
                    elif op_type == "style":
                        result = self._apply_style(ws, op)
                    else:
                        result = {
                            "status": "skipped",
                            "reason": f"Unknown operation type: {op_type}",
                        }

                    results.append({"index": i, "type": op_type, **result})

                except Exception as e:
                    results.append(
                        {"index": i, "type": op.get("type"), "status": "error", "error": str(e)}
                    )

            wb.save(filepath)
            wb.close()

            return {
                "file": filepath,
                "total": len(operations),
                "executed": len([r for r in results if r.get("status") != "skipped"]),
                "results": results,
            }

        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Batch execution failed: {e}")
            raise DataError(str(e))

    def _merge_cells(self, ws, op: dict) -> dict:
        """合并单元格"""
        range_str = op.get("range")
        if not range_str:
            return {"status": "error", "error": "range is required for merge"}

        ws.merge_cells(range_str)
        return {"status": "success", "merged": range_str}

    def _unmerge_cells(self, ws, op: dict) -> dict:
        """取消合并单元格"""
        range_str = op.get("range")
        if not range_str:
            return {"status": "error", "error": "range is required for unmerge"}

        ws.unmerge_cells(range_str)
        return {"status": "success", "unmerged": range_str}

    def _write_cell(self, ws, op: dict) -> dict:
        """写入单元格"""
        cell = op.get("cell")
        value = op.get("value")

        if not cell:
            return {"status": "error", "error": "cell is required for write"}

        ws[cell] = value
        return {"status": "success", "cell": cell, "value": value}

    def _apply_style(self, ws, op: dict) -> dict:
        """应用样式"""
        from openpyxl.styles import Alignment, Font, PatternFill

        cell = op.get("cell")
        if not cell:
            return {"status": "error", "error": "cell is required for style"}

        cell_obj = ws[cell]

        if "font" in op:
            font_params = op["font"]
            cell_obj.font = Font(**font_params)

        if "fill" in op:
            fill_params = op["fill"]
            cell_obj.fill = PatternFill(**fill_params)

        if "alignment" in op:
            align_params = op["alignment"]
            cell_obj.alignment = Alignment(**align_params)

        return {"status": "success", "cell": cell}
