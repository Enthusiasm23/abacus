"""商功章 - 批量转换：深度实现批量转换"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class BatchTransformCapability(Capability):
    @property
    def name(self) -> str:
        return "batch_transform"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "深度批量转换（查找替换、公式填充、数据清洗）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="operations", type="array", description="转换操作列表", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        operations = params.get("operations", [])

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            results = []

            for op in operations:
                op_type = op.get("type")

                if op_type == "replace":
                    sheet = op.get("sheet", wb.sheetnames[0])
                    ws = wb[sheet]
                    old_value = op.get("old")
                    new_value = op.get("new")

                    count = 0
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value == old_value:
                                cell.value = new_value
                                count += 1

                    results.append({"type": "replace", "count": count})

                elif op_type == "fill_formula":
                    sheet = op.get("sheet", wb.sheetnames[0])
                    ws = wb[sheet]
                    cell = op.get("cell")
                    formula = op.get("formula")

                    if cell and formula:
                        ws[cell] = formula
                        results.append(
                            {
                                "type": "fill_formula",
                                "cell": cell,
                                "formula": formula,
                                "status": "success",
                            }
                        )
                    else:
                        results.append(
                            {
                                "type": "fill_formula",
                                "status": "error",
                                "error": "cell and formula required",
                            }
                        )

                elif op_type == "fill_value":
                    sheet = op.get("sheet", wb.sheetnames[0])
                    ws = wb[sheet]
                    cell = op.get("cell")
                    value = op.get("value")

                    if cell:
                        ws[cell] = value
                        results.append(
                            {
                                "type": "fill_value",
                                "cell": cell,
                                "value": value,
                                "status": "success",
                            }
                        )
                    else:
                        results.append(
                            {"type": "fill_value", "status": "error", "error": "cell required"}
                        )

                elif op_type == "copy_format":
                    sheet = op.get("sheet", wb.sheetnames[0])
                    ws = wb[sheet]
                    source = op.get("source")
                    target = op.get("target")

                    if source and target:
                        src_cell = ws[source]
                        tgt_cell = ws[target]

                        # 复制格式（使用 copy() 避免 StyleProxy 不可哈希问题）
                        from copy import copy

                        tgt_cell.font = copy(src_cell.font)
                        tgt_cell.fill = copy(src_cell.fill)
                        tgt_cell.border = copy(src_cell.border)
                        tgt_cell.alignment = copy(src_cell.alignment)
                        tgt_cell.number_format = src_cell.number_format

                        results.append(
                            {
                                "type": "copy_format",
                                "source": source,
                                "target": target,
                                "status": "success",
                            }
                        )
                    else:
                        results.append(
                            {
                                "type": "copy_format",
                                "status": "error",
                                "error": "source and target required",
                            }
                        )

                elif op_type == "clear_content":
                    sheet = op.get("sheet", wb.sheetnames[0])
                    ws = wb[sheet]
                    range_str = op.get("range")

                    if range_str:
                        from ..cell_utils import parse_range

                        start_row, start_col, end_row, end_col = parse_range(range_str)
                        if end_row is None:
                            end_row = ws.max_row
                        if end_col is None:
                            end_col = ws.max_column

                        count = 0
                        for row in range(start_row, end_row + 1):
                            for col in range(start_col, end_col + 1):
                                ws.cell(row=row, column=col).value = None
                                count += 1

                        results.append(
                            {
                                "type": "clear_content",
                                "range": range_str,
                                "cleared": count,
                                "status": "success",
                            }
                        )
                    else:
                        results.append(
                            {"type": "clear_content", "status": "error", "error": "range required"}
                        )

                elif op_type == "clear_format":
                    sheet = op.get("sheet", wb.sheetnames[0])
                    ws = wb[sheet]
                    range_str = op.get("range")

                    if range_str:
                        from openpyxl.styles import Alignment, Border, Font, PatternFill

                        from ..cell_utils import parse_range

                        start_row, start_col, end_row, end_col = parse_range(range_str)
                        if end_row is None:
                            end_row = ws.max_row
                        if end_col is None:
                            end_col = ws.max_column

                        count = 0
                        for row in range(start_row, end_row + 1):
                            for col in range(start_col, end_col + 1):
                                cell = ws.cell(row=row, column=col)
                                cell.font = Font()
                                cell.fill = PatternFill()
                                cell.border = Border()
                                cell.alignment = Alignment()
                                cell.number_format = "General"
                                count += 1

                        results.append(
                            {
                                "type": "clear_format",
                                "range": range_str,
                                "cleared": count,
                                "status": "success",
                            }
                        )
                    else:
                        results.append(
                            {"type": "clear_format", "status": "error", "error": "range required"}
                        )

            wb.save(file_path)
            wb.close()

            return {"file": file_path, "operations": len(operations), "results": results}
        except Exception as e:
            raise DataError(str(e))
