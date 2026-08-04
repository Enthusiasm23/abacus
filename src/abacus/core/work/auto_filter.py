"""商功章 - 自动筛选：设置和管理自动筛选"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class AutoFilterCapability(Capability):
    """自动筛选：设置和管理自动筛选"""

    @property
    def name(self) -> str:
        return "set_auto_filter"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "设置自动筛选（添加、删除、查询筛选）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="action", type="string", description="操作（set/remove/get）", required=True
            ),
            CapabilitySchema(name="range", type="string", description="筛选范围", required=False),
            CapabilitySchema(name="column", type="string", description="筛选列", required=False),
            CapabilitySchema(
                name="criteria", type="string", description="筛选条件", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        action = params.get("action")
        range_str = params.get("range")
        column = params.get("column")
        criteria = params.get("criteria")

        if not file_path:
            raise DataError("file parameter is required")

        return self._manage_filter(file_path, sheet_name, action, range_str, column, criteria)

    def _manage_filter(
        self,
        filepath: str,
        sheet_name: str,
        action: str,
        range_str: str = None,
        column: str = None,
        criteria: str = None,
    ) -> dict[str, Any]:
        """管理自动筛选"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            if action == "set":
                if not range_str:
                    raise DataError("range required for set action")
                ws.auto_filter.ref = range_str
                result = {"action": "set", "range": range_str}

            elif action == "remove":
                ws.auto_filter.ref = None
                result = {"action": "remove"}

            elif action == "get":
                ref = ws.auto_filter.ref
                has_filter = ref is not None
                filters = []
                if has_filter and ws.auto_filter.filterColumn:
                    for fc in ws.auto_filter.filterColumn:
                        col_filters = {
                            "column": fc.colId,
                            "filters": [f.val for f in fc.filters] if fc.filters else [],
                        }
                        if fc.customFilters:
                            col_filters["custom_filters"] = [
                                {"operator": cf.operator, "val": cf.val} for cf in fc.customFilters
                            ]
                        filters.append(col_filters)
                result = {
                    "action": "get",
                    "range": str(ref) if ref else None,
                    "has_filter": has_filter,
                    "filters": filters,
                }

            else:
                raise DataError(f"Unknown action: {action}")

            wb.save(filepath)
            wb.close()

            return result

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to manage auto filter: {e}")
            raise DataError(str(e))
