"""商功章 - 批注管理：添加、删除、获取批注"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class CommentCapability(Capability):
    """批注管理：添加、删除、获取批注"""

    @property
    def name(self) -> str:
        return "manage_comment"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "管理批注（添加、删除、获取批注）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="action", type="string", description="操作（add/delete/get/）", required=True
            ),
            CapabilitySchema(name="cell", type="string", description="单元格位置", required=False),
            CapabilitySchema(name="text", type="string", description="批注内容", required=False),
            CapabilitySchema(name="author", type="string", description="作者", required=False),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        action = params.get("action")
        cell = params.get("cell")
        text = params.get("text")
        author = params.get("author", "Abacus")

        if not file_path:
            raise DataError("file parameter is required")

        return self._manage_comment(file_path, sheet_name, action, cell, text, author)

    def _manage_comment(
        self,
        filepath: str,
        sheet_name: str,
        action: str,
        cell: str = None,
        text: str = None,
        author: str = "Abacus",
    ) -> dict[str, Any]:
        """管理批注"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            if action == "add":
                if not cell or not text:
                    raise DataError("cell and text required for add action")

                ws[cell].comment = Comment(text, author)
                result = {"action": "add", "cell": cell, "text": text, "author": author}

            elif action == "delete":
                if not cell:
                    raise DataError("cell required for delete action")

                ws[cell].comment = None
                result = {"action": "delete", "cell": cell}

            elif action == "get":
                if not cell:
                    raise DataError("cell required for get action")

                comment = ws[cell].comment
                if comment:
                    result = {
                        "action": "get",
                        "cell": cell,
                        "text": comment.text,
                        "author": comment.author,
                    }
                else:
                    result = {"action": "get", "cell": cell, "comment": None}

            elif action == "list":
                comments = []
                for row in ws.iter_rows():
                    for c in row:
                        if c.comment:
                            comments.append(
                                {
                                    "cell": c.coordinate,
                                    "text": c.comment.text,
                                    "author": c.comment.author,
                                }
                            )
                result = {"action": "list", "comments": comments}

            else:
                raise DataError(f"Unknown action: {action}")

            # 只有写操作才保存文件
            if action not in ("list", "get"):
                wb.save(filepath)
            wb.close()

            return result

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to manage comment: {e}")
            raise DataError(str(e))
