"""商功章 - 冻结窗格：冻结行和列"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class FreezePaneCapability(Capability):
    """冻结窗格：冻结行和列"""

    @property
    def name(self) -> str:
        return "freeze_panes"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "冻结窗格（冻结行、列或行列）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="rows", type="number", description="冻结行数", required=False),
            CapabilitySchema(name="columns", type="number", description="冻结列数", required=False),
            CapabilitySchema(
                name="cell",
                type="string",
                description="冻结位置（如 B2 表示冻结第一行和第一列）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        rows = params.get("rows")
        columns = params.get("columns")
        cell = params.get("cell")

        if not file_path:
            raise DataError("file parameter is required")

        return self._freeze_panes(file_path, sheet_name, rows, columns, cell)

    def _freeze_panes(
        self,
        filepath: str,
        sheet_name: str,
        rows: int = None,
        columns: int = None,
        cell: str = None,
    ) -> dict[str, Any]:
        """冻结窗格"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            if cell:
                # 使用单元格位置
                ws.freeze_panes = cell
                result = {"action": "freeze", "cell": cell}
            elif rows is not None or columns is not None:
                # 使用行列数
                freeze_row = (rows + 1) if rows else 1
                freeze_col = (columns + 1) if columns else 1
                freeze_cell = f"{get_column_letter(freeze_col)}{freeze_row}"
                ws.freeze_panes = freeze_cell
                result = {"action": "freeze", "rows": rows, "columns": columns}
            else:
                # 解除冻结
                ws.freeze_panes = None
                result = {"action": "unfreeze"}

            wb.save(filepath)
            wb.close()

            return result

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"Failed to freeze panes: {e}")
            raise DataError(str(e))
