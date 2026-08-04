"""商功章 - 批量验证：深度实现批量验证"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class BatchValidateCapability(Capability):
    @property
    def name(self) -> str:
        return "batch_validate"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "深度批量验证（多范围、多规则）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="validations", type="array", description="验证规则列表", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        validations = params.get("validations", [])

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path, data_only=True)
            results = []

            for validation in validations:
                sheet = validation.get("sheet", wb.sheetnames[0])
                range_str = validation.get("range")
                rule = validation.get("rule")

                if sheet not in wb.sheetnames:
                    results.append({"sheet": sheet, "valid": False, "error": "Sheet not found"})
                    continue

                ws = wb[sheet]

                if range_str:
                    start_row, start_col, end_row, end_col = parse_range(range_str)
                    if end_row is None:
                        end_row = ws.max_row
                    if end_col is None:
                        end_col = ws.max_column

                    issues = []

                    if rule == "unique":
                        values = []
                        for row in range(start_row, end_row + 1):
                            for col in range(start_col, end_col + 1):
                                cell_val = ws.cell(row=row, column=col).value
                                if cell_val is not None:
                                    values.append(cell_val)

                        duplicates = [v for v in values if values.count(v) > 1]
                        if duplicates:
                            issues.append({"rule": "unique", "duplicates": (set(duplicates))})

                    elif rule == "range":
                        min_value = validation.get("min")
                        max_value = validation.get("max")

                        for row in range(start_row, end_row + 1):
                            for col in range(start_col, end_col + 1):
                                cell_val = ws.cell(row=row, column=col).value
                                if isinstance(cell_val, (int, float)):
                                    if min_value is not None and cell_val < min_value:
                                        issues.append(
                                            {
                                                "rule": "range",
                                                "cell": f"{get_column_letter(col)}{row}",
                                                "value": cell_val,
                                                "min": min_value,
                                            }
                                        )
                                    if max_value is not None and cell_val > max_value:
                                        issues.append(
                                            {
                                                "rule": "range",
                                                "cell": f"{get_column_letter(col)}{row}",
                                                "value": cell_val,
                                                "max": max_value,
                                            }
                                        )

                    elif rule == "pattern":
                        import re

                        pattern = validation.get("pattern")

                        if pattern:
                            for row in range(start_row, end_row + 1):
                                for col in range(start_col, end_col + 1):
                                    cell_val = ws.cell(row=row, column=col).value
                                    if isinstance(cell_val, str) and not re.match(
                                        pattern, cell_val
                                    ):
                                        issues.append(
                                            {
                                                "rule": "pattern",
                                                "cell": f"{get_column_letter(col)}{row}",
                                                "value": cell_val,
                                                "pattern": pattern,
                                            }
                                        )

                    elif rule == "date":
                        from datetime import datetime

                        for row in range(start_row, end_row + 1):
                            for col in range(start_col, end_col + 1):
                                cell_val = ws.cell(row=row, column=col).value
                                if cell_val is not None:
                                    if isinstance(cell_val, str):
                                        try:
                                            datetime.strptime(cell_val, "%Y-%m-%d")
                                        except ValueError:
                                            issues.append(
                                                {
                                                    "rule": "date",
                                                    "cell": f"{get_column_letter(col)}{row}",
                                                    "value": cell_val,
                                                }
                                            )

                    elif rule == "email":
                        import re

                        email_pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"

                        for row in range(start_row, end_row + 1):
                            for col in range(start_col, end_col + 1):
                                cell_val = ws.cell(row=row, column=col).value
                                if isinstance(cell_val, str) and not re.match(
                                    email_pattern, cell_val
                                ):
                                    issues.append(
                                        {
                                            "rule": "email",
                                            "cell": f"{get_column_letter(col)}{row}",
                                            "value": cell_val,
                                        }
                                    )

                    else:
                        for row in range(start_row, end_row + 1):
                            for col in range(start_col, end_col + 1):
                                cell = ws.cell(row=row, column=col)

                                if rule == "no_empty" and cell.value is None:
                                    issues.append({"cell": cell.coordinate, "issue": "empty"})
                                elif rule == "numeric" and not isinstance(cell.value, (int, float)):
                                    issues.append(
                                        {
                                            "cell": cell.coordinate,
                                            "issue": "not_numeric",
                                            "value": cell.value,
                                        }
                                    )

                    results.append(
                        {
                            "sheet": sheet,
                            "range": range_str,
                            "rule": rule,
                            "valid": len(issues) == 0,
                            "issues": issues[:10],
                        }
                    )

            wb.close()

            return {"file": file_path, "validations": len(validations), "results": results}
        except Exception as e:
            raise DataError(str(e))
