"""样式工具 - 行业品牌色和格式"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)

# 行业品牌色
INDUSTRY_COLORS = {
    "finance": {
        "primary": "1F4E79",  # 深蓝
        "secondary": "2E75B6",  # 中蓝
        "accent": "FFC000",  # 金色
        "header_bg": "1F4E79",
        "header_font": "FFFFFF",
        "kpi_good": "00B050",
        "kpi_bad": "FF0000",
    },
    "ecommerce": {
        "primary": "FF6B35",  # 橙色
        "secondary": "004E89",  # 深蓝
        "accent": "FFA500",  # 亮橙
        "header_bg": "FF6B35",
        "header_font": "FFFFFF",
        "kpi_good": "00B050",
        "kpi_bad": "FF0000",
    },
    "saas": {
        "primary": "6C5CE7",  # 紫色
        "secondary": "A29BFE",  # 浅紫
        "accent": "00CEC9",  # 青色
        "header_bg": "6C5CE7",
        "header_font": "FFFFFF",
        "kpi_good": "00B050",
        "kpi_bad": "FF0000",
    },
    "internet": {
        "primary": "00B894",  # 绿色
        "secondary": "00CEC9",  # 青色
        "accent": "FDCB6E",  # 黄色
        "header_bg": "00B894",
        "header_font": "FFFFFF",
        "kpi_good": "00B050",
        "kpi_bad": "FF0000",
    },
}


class StyleCapability(Capability):
    """样式管理 - 行业品牌色和格式"""

    @property
    def name(self) -> str:
        return "manage_style"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "管理样式（行业品牌色、表头、KPI 格式）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="action",
                type="string",
                description="操作（apply_header/apply_kpi/auto_width/get）",
                required=True,
            ),
            CapabilitySchema(name="range", type="string", description="数据范围", required=False),
            CapabilitySchema(
                name="industry",
                type="string",
                description="行业（finance/ecommerce/saas/internet）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        action = params.get("action")
        range_str = params.get("range")
        industry = params.get("industry", "finance")

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)
            ws = wb[sheet_name]

            colors = INDUSTRY_COLORS.get(industry, INDUSTRY_COLORS["finance"])

            if action == "apply_header":
                if not range_str:
                    raise DataError("range required for apply_header")

                start_row, start_col, end_row, end_col = parse_range(range_str)

                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=start_row, column=col)
                    cell.font = Font(bold=True, color=colors["header_font"])
                    cell.fill = PatternFill(fill_type="solid", fgColor=colors["header_bg"])
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                result = {"action": "apply_header", "range": range_str, "industry": industry}

            elif action == "apply_kpi":
                if not range_str:
                    raise DataError("range required for apply_kpi")

                start_row, start_col, end_row, end_col = parse_range(range_str)

                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cell = ws.cell(row=row, column=col)
                        if isinstance(cell.value, (int, float)):
                            if cell.value >= 0:
                                cell.font = Font(color=colors["kpi_good"])
                            else:
                                cell.font = Font(color=colors["kpi_bad"])

                result = {"action": "apply_kpi", "range": range_str}

            elif action == "auto_width":
                for col_idx, col in enumerate(ws.columns, 1):
                    max_len = 0
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

                result = {"action": "auto_width"}

            elif action == "get":
                if not range_str:
                    raise DataError("range required for get action")

                start_row, start_col, end_row, end_col = parse_range(range_str)
                if end_row is None:
                    end_row = ws.max_row
                if end_col is None:
                    end_col = ws.max_column

                styles = []
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cell = ws.cell(row=row, column=col)
                        style_info = {
                            "cell": cell.coordinate,
                            "font": {
                                "name": cell.font.name,
                                "size": cell.font.size,
                                "bold": cell.font.bold,
                                "italic": cell.font.italic,
                            }
                            if cell.font
                            else None,
                            "fill": {
                                "color": cell.fill.fgColor.rgb if cell.fill.fgColor else None,
                            }
                            if cell.fill
                            else None,
                            "alignment": {
                                "horizontal": cell.alignment.horizontal,
                                "vertical": cell.alignment.vertical,
                            }
                            if cell.alignment
                            else None,
                        }
                        styles.append(style_info)

                result = {"action": "get", "range": range_str, "styles": styles}

            else:
                raise DataError(f"Unknown action: {action}")

            wb.save(file_path)
            wb.close()

            return result

        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Failed to manage style: {e}")
            raise DataError(str(e))
