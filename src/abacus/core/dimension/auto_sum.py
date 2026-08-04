"""方程章 - 自动求和"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class AutoSumCapability(Capability):
    """自动求和"""

    @property
    def name(self) -> str:
        return "auto_sum"

    @property
    def chapter(self) -> str:
        return "equation"

    @property
    def description(self) -> str:
        return "自动求和（在范围内设置 SUM 公式）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="range", type="string", description="数据范围（A1 表示法）", required=True
            ),
            CapabilitySchema(
                name="direction",
                type="string",
                description="求和方向（down/right）",
                required=False,
                default="down",
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        data_range = params.get("range")
        direction = params.get("direction", "down")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")
        if not sheet_name:
            raise ValidationError("执行失败: 缺少必要参数 sheet")
        if not data_range:
            raise ValidationError("执行失败: 缺少必要参数 range")

        if direction not in ("down", "right"):
            raise ValidationError("执行失败: direction 必须为 'down' 或 'right'")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {file_path}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            from ..cell_utils import parse_range

            min_col, min_row, max_col, max_row = parse_range(data_range)

            formulas_set = 0

            if direction == "down":
                for col in range(min_col, max_col + 1):
                    col_letter = get_column_letter(col)
                    sum_row = max_row + 1
                    formula = f"=SUM({col_letter}{min_row}:{col_letter}{max_row})"
                    ws.cell(row=sum_row, column=col, value=formula)
                    formulas_set += 1
            else:
                for row in range(min_row, max_row + 1):
                    sum_col = max_col + 1
                    start_letter = get_column_letter(min_col)
                    end_letter = get_column_letter(max_col)
                    formula = f"=SUM({start_letter}{row}:{end_letter}{row})"
                    ws.cell(row=row, column=sum_col, value=formula)
                    formulas_set += 1

            wb.save(file_path)
            wb.close()

            return {
                "success": True,
                "range": data_range,
                "direction": direction,
                "formulas_set": formulas_set,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"自动求和失败: {e}")
            raise DataError(f"数据操作失败: {e}")
