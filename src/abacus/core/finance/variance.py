"""差异分析 - 预算与实际对比"""

import logging
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, ValidationError

logger = logging.getLogger(__name__)


class VarianceCapability(Capability):
    """差异分析 - 预算与实际对比"""

    @property
    def name(self) -> str:
        return "variance_analysis"

    @property
    def chapter(self) -> str:
        return "triangle"

    @property
    def description(self) -> str:
        return "预算与实际差异分析（差异计算、重要性标记、瀑布图）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="budget_sheet", type="string", description="预算数据工作表", required=True
            ),
            CapabilitySchema(
                name="actual_sheet", type="string", description="实际数据工作表", required=True
            ),
            CapabilitySchema(
                name="output", type="string", description="输出文件路径", required=False
            ),
            CapabilitySchema(
                name="threshold",
                type="number",
                description="重要性阈值（如 0.1 表示 10%）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        budget_sheet = params.get("budget_sheet")
        actual_sheet = params.get("actual_sheet")
        output = params.get("output")
        threshold = params.get("threshold", 0.1)

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")

        # 读取数据
        try:
            budget_df = pd.read_excel(file_path, sheet_name=budget_sheet)
            actual_df = pd.read_excel(file_path, sheet_name=actual_sheet)
        except Exception as e:
            raise DataError(f"数据操作失败: 读取 Excel 文件失败 {e}")

        # 计算差异
        variance_results = self._calculate_variance(budget_df, actual_df, threshold)

        # 保存结果
        if output:
            self._save_results(variance_results, output)

        return {
            "file": file_path,
            "budget_sheet": budget_sheet,
            "actual_sheet": actual_sheet,
            "total_variance": float(variance_results["total_variance"]),
            "favorable_count": int(variance_results["favorable_count"]),
            "unfavorable_count": int(variance_results["unfavorable_count"]),
            "material_variances": int(variance_results["material_count"]),
        }

    def _calculate_variance(self, budget_df, actual_df, threshold):
        """计算差异"""
        results = {
            "items": [],
            "total_variance": 0,
            "favorable_count": 0,
            "unfavorable_count": 0,
            "material_count": 0,
        }

        # 假设第一列是项目名称
        for idx in range(min(len(budget_df), len(actual_df))):
            item_name = budget_df.iloc[idx, 0]
            budget_value = budget_df.iloc[idx, 1] if len(budget_df.columns) > 1 else 0
            actual_value = actual_df.iloc[idx, 1] if len(actual_df.columns) > 1 else 0

            # 计算差异
            variance = actual_value - budget_value
            variance_pct = variance / abs(budget_value) if budget_value != 0 else 0

            # 判断有利/不利
            is_favorable = variance > 0  # 简化：收入增加为有利
            is_material = abs(variance_pct) > threshold

            results["items"].append(
                {
                    "item": item_name,
                    "budget": budget_value,
                    "actual": actual_value,
                    "variance": variance,
                    "variance_pct": variance_pct,
                    "is_favorable": is_favorable,
                    "is_material": is_material,
                }
            )

            results["total_variance"] += variance
            if is_favorable:
                results["favorable_count"] += 1
            else:
                results["unfavorable_count"] += 1
            if is_material:
                results["material_count"] += 1

        return results

    def _save_results(self, results, output):
        """保存结果到 Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "差异分析"

        # 表头
        headers = ["项目", "预算", "实际", "差异", "差异%", "状态"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)

        # 数据
        for i, item in enumerate(results["items"], 2):
            ws[f"A{i}"] = item["item"]
            ws[f"B{i}"] = item["budget"]
            ws[f"C{i}"] = item["actual"]
            ws[f"D{i}"] = item["variance"]
            ws[f"E{i}"] = item["variance_pct"]
            ws[f"F{i}"] = "有利" if item["is_favorable"] else "不利"

            # 材料差异标记
            if item["is_material"]:
                ws[f"F{i}"].fill = PatternFill(patternType="solid", fgColor="FF0000")

        wb.save(output)
