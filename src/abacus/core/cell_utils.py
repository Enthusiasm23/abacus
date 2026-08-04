"""单元格工具函数"""

import re

from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter as _get_column_letter

from .exceptions import RangeError


def parse_cell_reference(cell_ref: str) -> tuple[int, int]:
    """解析单元格引用为 (行, 列) 索引

    Args:
        cell_ref: 单元格引用，如 'A1', 'BC123'

    Returns:
        (row, col) 元组，1-based

    Raises:
        ValueError: 格式无效
    """
    match = re.match(r"([A-Z]+)(\d+)", cell_ref.upper())
    if not match:
        raise RangeError(f"范围错误: 无效的单元格引用 {cell_ref}")
    col_str, row_str = match.groups()
    row = int(row_str)
    col = column_index_from_string(col_str)
    return row, col


def parse_range(
    range_str: str,
) -> (
    tuple[int, int, int, int]
    | tuple[int, int, None, int]
    | tuple[int, int, int, None]
    | tuple[int, int, None, None]
):
    """解析范围字符串为 (起始行, 起始列, 结束行, 结束列)

    Args:
        range_str: 范围字符串，如 'A1:D10', 'A1', 'A:D'

    Returns:
        (start_row, start_col, end_row, end_col) 元组
        结束行/列可能为 None（表示整行/整列）
    """
    if ":" not in range_str:
        row, col = parse_cell_reference(range_str)
        return row, col, None, None

    start, end = range_str.split(":")

    # 处理整列引用如 'A:D'（起始部分只有字母）
    if start and not any(c.isdigit() for c in start) and end and not any(c.isdigit() for c in end):
        start_col = column_index_from_string(start.upper())
        end_col = column_index_from_string(end.upper())
        return 1, start_col, None, end_col

    start_row, start_col = parse_cell_reference(start)

    # 处理整列引用如 'A1:D'（起始有行号，结束只有字母）
    if end and not any(c.isdigit() for c in end):
        end_col = column_index_from_string(end.upper())
        return start_row, start_col, None, end_col

    # 处理整行引用如 '1:10'
    if end and not any(c.isalpha() for c in end):
        end_row = int(end)
        return start_row, start_col, end_row, None

    # 处理标准范围如 'A1:D10'
    end_row, end_col = parse_cell_reference(end)
    return start_row, start_col, end_row, end_col


def validate_cell_reference(cell_ref: str) -> bool:
    """验证单元格引用格式

    Args:
        cell_ref: 单元格引用

    Returns:
        是否有效
    """
    if not cell_ref:
        return False

    col = row = ""
    for c in cell_ref:
        if c.isalpha():
            if row:
                return False
            col += c
        elif c.isdigit():
            row += c
        else:
            return False

    return bool(col and row)


def get_column_letter(col_index: int) -> str:
    """获取列字母

    Args:
        col_index: 列索引（1-based）

    Returns:
        列字母，如 'A', 'B', 'AA'
    """
    return _get_column_letter(col_index)
