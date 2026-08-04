"""Excel 转 Markdown - 将 Excel 表格转换为 Markdown 格式"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class ExcelToMarkdownCapability(Capability):
    """Excel 转 Markdown - 将 Excel 表格转换为 Markdown 格式"""

    @property
    def name(self) -> str:
        return "excel_to_markdown"

    @property
    def chapter(self) -> str:
        return "transport"

    @property
    def description(self) -> str:
        return "将 Excel 表格转换为 Markdown 格式（支持合并单元格、多工作表、样式）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="sheet",
                type="string",
                description="工作表名称（可选，默认全部）",
                required=False,
            ),
            CapabilitySchema(
                name="output", type="string", description="输出文件路径", required=False
            ),
            CapabilitySchema(
                name="merge_mode",
                type="string",
                description="合并单元格处理模式（tl/fill）",
                required=False,
            ),
            CapabilitySchema(
                name="include_styles",
                type="boolean",
                description="是否包含样式（粗体/斜体）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        output = params.get("output")
        merge_mode = params.get("merge_mode", "tl")
        include_styles = params.get("include_styles", True)

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件操作失败: 文件不存在 {file_path}")

        # 加载工作簿
        wb = load_workbook(path, data_only=True)

        # 确定要转换的工作表
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")
            sheets = [sheet_name]
        else:
            sheets = wb.sheetnames

        # 转换每个工作表
        result_parts = []
        for name in sheets:
            ws = wb[name]
            md = self._convert_sheet(ws, merge_mode, include_styles)
            result_parts.append(f"## {name}\n\n{md}")

        wb.close()

        # 合并结果
        markdown = "\n\n".join(result_parts)

        # 保存到文件
        if output:
            output_path = Path(output)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(markdown, encoding="utf-8")

        return {
            "file": file_path,
            "sheets_converted": len(sheets),
            "output": output,
            "markdown_length": len(markdown),
        }

    def _convert_sheet(self, ws, merge_mode: str, include_styles: bool) -> str:
        """转换单个工作表"""
        # 获取合并单元格信息
        merged_cells = {}
        for merge_range in ws.merged_cells.ranges:
            min_row, min_col = merge_range.min_row, merge_range.min_col
            max_row, max_col = merge_range.max_row, merge_range.max_col

            # 记录合并区域
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if row == min_row and col == min_col:
                        merged_cells[(row, col)] = {
                            "is_top_left": True,
                            "rowspan": max_row - min_row + 1,
                            "colspan": max_col - min_col + 1,
                        }
                    else:
                        merged_cells[(row, col)] = {"is_top_left": False}

        # 获取数据范围
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row == 0 or max_col == 0:
            return "*空工作表*"

        # 构建 Markdown 表格
        rows = []
        for row in range(1, max_row + 1):
            cells = []
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                value = cell.value if cell.value is not None else ""

                # 处理合并单元格
                if (row, col) in merged_cells:
                    merge_info = merged_cells[(row, col)]
                    if not merge_info["is_top_left"] and merge_mode == "tl":
                        value = ""
                        # fill 模式保持原值

                # 处理样式
                if include_styles and isinstance(value, str):
                    if cell.font and cell.font.bold:
                        value = f"**{value}**"
                    if cell.font and cell.font.italic:
                        value = f"*{value}*"

                # 转义 Markdown 特殊字符
                value = str(value).replace("|", "\\|").replace("\n", "<br>")
                cells.append(value)

            rows.append("| " + " | ".join(cells) + " |")

        # 添加表头分隔线
        if len(rows) > 0:
            separator = "| " + " | ".join(["---"] * max_col) + " |"
            rows.insert(1, separator)

        return "\n".join(rows)
