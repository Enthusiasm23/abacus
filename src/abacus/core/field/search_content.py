"""方田章 - 搜索内容：在 Excel 文件中搜索关键词"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class SearchContentCapability(Capability):
    """搜索内容：在 Excel 文件中搜索关键词"""

    @property
    def name(self) -> str:
        return "search_content"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "在 Excel 文件中搜索关键词"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="keyword", type="string", description="搜索关键词", required=True
            ),
            CapabilitySchema(
                name="sheet",
                type="string",
                description="工作表名称（可选，默认搜索所有）",
                required=False,
            ),
            CapabilitySchema(
                name="max_results",
                type="number",
                description="最大结果数（默认 50）",
                required=False,
                default=50,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行搜索"""
        file_path = params.get("file")
        keyword = params.get("keyword")
        sheet_name = params.get("sheet")
        max_results = params.get("max_results", 50)

        if not file_path:
            raise DataError("file parameter is required")
        if not keyword:
            raise DataError("keyword parameter is required")

        return self._search_content(file_path, keyword, sheet_name, max_results)

    def _search_content(
        self, filepath: str, keyword: str, sheet_name: str = None, max_results: int = 50
    ) -> dict[str, Any]:
        """搜索内容"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, read_only=True, data_only=True)

            results = []
            search_count = 0

            for name in wb.sheetnames:
                if sheet_name and name != sheet_name:
                    continue

                ws = wb[name]

                for row in ws.iter_rows(values_only=False):
                    for cell in row:
                        if len(results) >= max_results:
                            break

                        if cell.value is not None:
                            cell_value = str(cell.value)
                            if keyword.lower() in cell_value.lower():
                                results.append(
                                    {"sheet": name, "cell": cell.coordinate, "value": cell.value}
                                )
                                search_count += 1

                    if len(results) >= max_results:
                        break

            wb.close()

            return {
                "file": filepath,
                "keyword": keyword,
                "results": results,
                "total_found": len(results),
                "truncated": len(results) >= max_results,
            }

        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Failed to search content: {e}")
            raise DataError(str(e))
