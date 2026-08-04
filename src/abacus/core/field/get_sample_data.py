"""方田章 - 样本数据：获取指定工作表的样本数据"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class GetSampleDataCapability(Capability):
    """样本数据：获取指定工作表的样本数据"""

    @property
    def name(self) -> str:
        return "get_sample_data"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "获取指定工作表的样本数据"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="rows",
                type="number",
                description="样本行数（默认 10）",
                required=False,
                default=10,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行获取样本数据"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        rows = params.get("rows", 10)

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")

        return self._get_sample_data(file_path, sheet_name, rows)

    def _get_sample_data(self, filepath: str, sheet_name: str, rows: int = 10) -> dict[str, Any]:
        """获取样本数据"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, read_only=True, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            # 获取表头
            headers = []
            for row in ws.iter_rows(max_row=1, values_only=True):
                headers = [
                    str(cell) if cell is not None else f"Column_{i + 1}"
                    for i, cell in enumerate(row)
                ]
                break

            # 获取样本数据
            data_rows = []
            row_count = 0

            for row in ws.iter_rows(min_row=2, max_row=rows + 1, values_only=True):
                row_count += 1
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = cell
                data_rows.append(row_data)

            wb.close()

            return {
                "file": filepath,
                "sheet": sheet_name,
                "columns": headers,
                "rows_returned": row_count,
                "data": data_rows,
            }

        except FileNotFoundError:
            raise
        except DataError:
            raise
        except Exception as e:
            logger.error(f"Failed to get sample data: {e}")
            raise DataError(str(e))
