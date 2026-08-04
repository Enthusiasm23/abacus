"""方田章 - 量结构：读取工作表结构"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class MeasureStructureCapability(Capability):
    """量结构：读取工作表结构"""

    @property
    def name(self) -> str:
        return "measure_structure"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "读取工作表结构（行数、列数、合并单元格等）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="sheet",
                type="string",
                description="工作表名称（可选，不填返回所有工作表）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行读取"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")

        return self._read_structure(file_path, sheet_name)

    def _read_structure(self, filepath: str, sheet_name: str = None) -> dict:
        """读取工作表结构"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, read_only=False)

            result = {"file": filepath, "sheets": []}

            for name in wb.sheetnames:
                if sheet_name and name != sheet_name:
                    continue

                ws = wb[name]
                sheet_info = {
                    "name": name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "merged_cells": [str(mc) for mc in ws.merged_cells.ranges],
                }
                result["sheets"].append(sheet_info)

            wb.close()

            return result

        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Failed to read structure: {e}")
            raise DataError(str(e))
