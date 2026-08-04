"""方田章 - 检测列：检测列名和数据类型"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class DetectColumnsCapability(Capability):
    """检测列：检测列名和数据类型"""

    @property
    def name(self) -> str:
        return "detect_columns"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "检测列名和数据类型"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="sheet", type="string", description="工作表名称（必填）", required=True
            ),
            CapabilitySchema(
                name="sample_rows",
                type="number",
                description="采样行数（默认 100）",
                required=False,
                default=100,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行检测"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        sample_rows = params.get("sample_rows", 100)

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")
        if not sheet_name:
            raise ValidationError("执行失败: 缺少必要参数 sheet")

        return self._detect_columns(file_path, sheet_name, sample_rows)

    def _detect_columns(self, filepath: str, sheet_name: str, sample_rows: int = 100) -> dict[str, Any]:
        """检测列信息"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path, read_only=True, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 获取表头（第一行）
            headers = []
            for row in ws.iter_rows(max_row=1, values_only=True):
                headers = [
                    str(cell) if cell is not None else f"Column_{i + 1}"
                    for i, cell in enumerate(row)
                ]
                break

            # 采样数据检测类型
            column_types = {
                header: {"type": "unknown", "samples": 0, "nulls": 0} for header in headers
            }

            row_count = 0
            for row in ws.iter_rows(min_row=2, max_row=sample_rows + 1, values_only=True):
                row_count += 1
                for i, cell in enumerate(row):
                    if i < len(headers):
                        header = headers[i]
                        if cell is None:
                            column_types[header]["nulls"] += 1
                        else:
                            column_types[header]["samples"] += 1
                            # 检测类型（bool 必须在 int 之前，因为 bool 是 int 子类）
                            if isinstance(cell, bool):
                                column_types[header]["type"] = "boolean"
                            elif isinstance(cell, (int, float)):
                                column_types[header]["type"] = "number"
                            elif hasattr(cell, "year"):  # datetime
                                column_types[header]["type"] = "date"
                            else:
                                column_types[header]["type"] = "string"

            wb.close()

            return {
                "file": filepath,
                "sheet": sheet_name,
                "columns": headers,
                "column_count": len(headers),
                "sample_rows": row_count,
                "column_details": column_types,
            }

        except FileNotFoundError:
            raise
        except DataError:
            raise
        except Exception as e:
            logger.error(f"列检测失败: {e}")
            raise DataError(f"数据操作失败: {e}")
