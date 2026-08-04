"""方田章 - 量单元格：读取单元格详细信息"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError, SheetNotFoundError

logger = logging.getLogger(__name__)


class MeasureCellsCapability(Capability):
    """量单元格：读取单元格详细信息（值、公式、样式）"""

    @property
    def name(self) -> str:
        return "measure_cells"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "读取单元格详细信息"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行读取"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")

        return self._read_cells(file_path, sheet_name, range_str)

    def _read_cells(self, filepath: str, sheet_name: str, range_str: str) -> dict[str, Any]:
        """读取单元格详细信息"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(file=filepath)

            # data_only=False 以获取公式
            wb = load_workbook(path, read_only=False, data_only=False)

            if sheet_name not in wb.sheetnames:
                raise SheetNotFoundError(sheet=sheet_name)

            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)

            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            cells = []
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell_address = f"{get_column_letter(col)}{row}"

                    cell_data = {
                        "address": cell_address,
                        "value": cell.value,
                        "row": row,
                        "column": col,
                        "data_type": cell.data_type,
                    }

                    # 检查是否有公式
                    if cell.data_type == "f" or (
                        isinstance(cell.value, str) and cell.value.startswith("=")
                    ):
                        cell_data["formula"] = cell.value

                    # 样式信息
                    if cell.font:
                        cell_data["font"] = {
                            "bold": cell.font.bold,
                            "italic": cell.font.italic,
                            "size": cell.font.size,
                        }

                    cells.append(cell_data)

            wb.close()

            return {"range": range_str, "sheet": sheet_name, "cells": cells, "count": len(cells)}

        except (FileNotFoundError, SheetNotFoundError):
            raise
        except Exception as e:
            logger.error(f"Failed to read cells: {e}")
            raise DataError(str(e))
