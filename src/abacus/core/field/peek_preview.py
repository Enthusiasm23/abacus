"""方田章 - 预览：快速预览每个工作表的前几行数据"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class PeekPreviewCapability(Capability):
    """预览：快速预览每个工作表的前几行数据"""

    @property
    def name(self) -> str:
        return "peek_preview"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "快速预览每个工作表的前几行数据"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="rows",
                type="number",
                description="预览行数（默认 5）",
                required=False,
                default=5,
            ),
            CapabilitySchema(
                name="sheet",
                type="string",
                description="工作表名称（可选，默认预览所有）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行预览"""
        file_path = params.get("file")
        rows = params.get("rows", 5)
        sheet_name = params.get("sheet")

        if not file_path:
            raise DataError("file parameter is required")

        return self._peek_preview(file_path, rows, sheet_name)

    def _peek_preview(self, filepath: str, rows: int = 5, sheet_name: str = None) -> dict[str, Any]:
        """预览数据"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, read_only=True, data_only=True)

            result = {"file": filepath, "preview": []}

            for name in wb.sheetnames:
                if sheet_name and name != sheet_name:
                    continue

                ws = wb[name]
                preview_rows = []
                row_count = 0

                for row in ws.iter_rows(max_row=rows, values_only=True):
                    preview_rows.append((row))
                    row_count += 1

                result["preview"].append(
                    {"sheet": name, "rows_previewed": row_count, "data": preview_rows}
                )

            wb.close()

            return result

        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Failed to peek preview: {e}")
            raise DataError(str(e))
