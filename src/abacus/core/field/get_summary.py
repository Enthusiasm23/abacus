"""方田章 - 摘要：获取 Excel 文件摘要信息"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class GetSummaryCapability(Capability):
    """摘要：获取 Excel 文件摘要信息"""

    @property
    def name(self) -> str:
        return "get_summary"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "获取 Excel 文件摘要信息"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行获取摘要"""
        file_path = params.get("file")

        if not file_path:
            raise DataError("file parameter is required")

        return self._get_summary(file_path)

    def _get_summary(self, filepath: str) -> dict[str, Any]:
        """获取摘要信息"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, read_only=True, data_only=True)

            sheets_summary = []
            total_rows = 0
            total_cells = 0

            for name in wb.sheetnames:
                ws = wb[name]
                row_count = 0
                col_count = 0
                non_empty = 0

                for row in ws.iter_rows(values_only=True):
                    row_count += 1
                    row_cols = len(row)
                    if row_cols > col_count:
                        col_count = row_cols
                    for cell in row:
                        if cell is not None:
                            non_empty += 1

                sheets_summary.append(
                    {
                        "name": name,
                        "rows": row_count,
                        "columns": col_count,
                        "non_empty_cells": non_empty,
                    }
                )
                total_rows += row_count
                total_cells += non_empty

            wb.close()

            return {
                "file": filepath,
                "sheet_count": len(sheets_summary),
                "total_rows": total_rows,
                "total_non_empty_cells": total_cells,
                "sheets": sheets_summary,
            }

        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Failed to get summary: {e}")
            raise DataError(str(e))
