"""方田章 - 列工作表：返回 Excel 文件中所有工作表名称"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class ListSheetsCapability(Capability):
    """列工作表：返回 Excel 文件中所有工作表名称"""

    @property
    def name(self) -> str:
        return "list_sheets"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "返回 Excel 文件中所有工作表名称列表"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行列出工作表"""
        file_path = params.get("file")

        if not file_path:
            raise DataError("file parameter is required")

        return self._list_sheets(file_path)

    def _list_sheets(self, filepath: str) -> dict:
        """列出所有工作表名称"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, read_only=True)

            sheets: list[str] = wb.sheetnames

            wb.close()

            return {"file": filepath, "sheets": sheets, "count": len(sheets)}

        except FileNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Failed to  sheets: {e}")
            raise DataError(str(e))
