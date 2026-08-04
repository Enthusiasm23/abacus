"""方田章 - 量范围：读取指定范围数据"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError, SheetNotFoundError

logger = logging.getLogger(__name__)


class MeasureRangeCapability(Capability):
    """量范围：读取指定范围数据"""

    @property
    def name(self) -> str:
        return "measure_range"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "读取指定范围数据"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="range", type="string", description="数据范围，如 A1:D10", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行读取"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")

        if not file_path:
            raise DataError("file parameter is required")
        if not sheet_name:
            raise DataError("sheet parameter is required")
        if not range_str:
            raise DataError("range parameter is required")

        return self._read_range(file_path, sheet_name, range_str)

    def _read_range(self, filepath: str, sheet_name: str, range_str: str) -> dict[str, Any]:
        """读取指定范围数据"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            wb = load_workbook(path, read_only=True, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            # 解析范围
            start_row, start_col, end_row, end_col = parse_range(range_str)

            # 如果没有指定结束位置，使用工作表的最大范围
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 读取数据
            data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    row_data.append(cell.value)
                if any(v is not None for v in row_data):
                    data.append(row_data)

            wb.close()

            # 构建范围字符串
            range_result = (
                f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
            )

            return {
                "range": range_result,
                "sheet": sheet_name,
                "data": data,
                "rows": len(data),
                "columns": end_col - start_col + 1 if data else 0,
            }

        except SheetNotFoundError:
            raise
        except FileNotFoundError as e:
            raise DataError(str(e))
        except Exception as e:
            logger.error(f"Failed to read range: {e}")
            raise DataError(str(e))
