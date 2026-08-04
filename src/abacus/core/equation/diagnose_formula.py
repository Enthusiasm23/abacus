"""方程章 - 诊断公式：分析公式错误原因"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class DiagnoseFormulaCapability(Capability):
    """诊断公式：分析 Excel 公式错误"""

    ERROR_TYPES = {
        "#REF!": "引用错误 - 引用的单元格不存在",
        "#N/A": "值不可用 - 查找值未找到",
        "#VALUE!": "值错误 - 参数类型不匹配",
        "#NAME?": "名称错误 - 函数名或范围名不存在",
        "#DIV/0!": "除零错误 - 除数为零",
        "#NUM!": "数值错误 - 数值超出范围",
        "#NULL!": "空值错误 - 区域引用不相交",
    }

    @property
    def name(self) -> str:
        return "diagnose_formula"

    @property
    def chapter(self) -> str:
        return "equation"

    @property
    def description(self) -> str:
        return "诊断公式错误（分析 #REF!, #N/A, #VALUE!, #NAME?, #DIV/0! 等错误）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="sheet", type="string", description="工作表名称（可选）", required=False
            ),
            CapabilitySchema(
                name="cell", type="string", description="单元格位置（可选）", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        cell = params.get("cell")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")

        return self._diagnose(file_path, sheet_name, cell)

    def _diagnose(self, filepath: str, sheet_name: str = None, cell: str = None) -> dict[str, Any]:
        """诊断公式错误"""
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path, data_only=False)

            errors = []
            formulas_checked = 0

            for name in wb.sheetnames:
                if sheet_name and name != sheet_name:
                    continue

                ws = wb[name]

                for row in ws.iter_rows():
                    for c in row:
                        if cell and c.coordinate != cell:
                            continue

                        if isinstance(c.value, str) and c.value.startswith("="):
                            formulas_checked += 1

                            # 检查公式中的错误
                            for error_type, description in self.ERROR_TYPES.items():
                                if error_type in str(c.value):
                                    errors.append(
                                        {
                                            "sheet": name,
                                            "cell": c.coordinate,
                                            "formula": c.value,
                                            "error": error_type,
                                            "description": description,
                                        }
                                    )

                            # 检查公式语法问题
                            issues = self._check_formula_syntax(c.value)
                            if issues:
                                errors.append(
                                    {
                                        "sheet": name,
                                        "cell": c.coordinate,
                                        "formula": c.value,
                                        "issues": issues,
                                    }
                                )

            wb.close()

            return {
                "file": filepath,
                "formulas_checked": formulas_checked,
                "errors_found": len(errors),
                "errors": errors,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"公式诊断失败: {e}")
            raise DataError(f"数据操作失败: {e}")

    def _check_formula_syntax(self, formula: str) -> [str]:
        """检查公式语法"""
        issues = []

        # 检查括号平衡
        parens = 0
        for i, c in enumerate(formula):
            if c == "(":
                parens += 1
            elif c == ")":
                parens -= 1
            if parens < 0:
                issues.append(f"Unmatched closing parenthesis at position {i}")

        if parens > 0:
            issues.append("Unclosed parenthesis")

        # 检查中文逗号
        if "，" in formula:
            issues.append("Contains Chinese comma (，) instead of English comma (,)")

        # 检查不安全函数
        unsafe_funcs = ["INDIRECT", "HYPERLINK", "WEBSERVICE"]
        for func in unsafe_funcs:
            if func in formula:
                issues.append(f"Contains unsafe function: {func}")

        return issues
