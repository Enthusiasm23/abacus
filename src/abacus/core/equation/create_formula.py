"""方程章 - 建公式：创建公式"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import validate_cell_reference
from ..exceptions import DataError, FileNotFoundError, FormulaError, ValidationError

logger = logging.getLogger(__name__)


def _create_formula(
        filepath: str, sheet_name: str, cell: str, formula: str
) -> dict[str, Any]:
    """创建公式"""
    try:
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

        if not validate_cell_reference(cell):
            raise FormulaError(f"公式错误: 无效的单元格引用 {cell}")

        wb = load_workbook(path)

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        # 设置公式
        ws[cell] = formula

        wb.save(filepath)
        wb.close()

        return {
            "file": filepath,
            "sheet": sheet_name,
            "cell": cell,
            "formula": formula,
            "created": True,
        }

    except (FileNotFoundError, FormulaError):
        raise
    except Exception as e:
        logger.error(f"公式创建失败: {e}")
        raise DataError(f"数据操作失败: {e}")


class CreateFormulaCapability(Capability):
    """建公式：创建公式"""

    @property
    def name(self) -> str:
        return "create_formula"

    @property
    def chapter(self) -> str:
        return "equation"

    @property
    def description(self) -> str:
        return "在指定单元格创建公式"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="cell", type="string", description="单元格位置，如 E1", required=True
            ),
            CapabilitySchema(
                name="formula", type="string", description="公式内容，如 SUM(A1:D1)", required=True
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        """执行公式创建"""
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        cell = params.get("cell")
        formula = params.get("formula")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")
        if not cell:
            raise ValidationError("执行失败: 缺少必要参数 cell")
        if not formula:
            raise ValidationError("执行失败: 缺少必要参数 formula")

        # 确保公式以 = 开头
        if not formula.startswith("="):
            formula = "=" + formula

        return _create_formula(file_path, sheet_name, cell, formula)
