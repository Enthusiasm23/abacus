"""方程章 - 公式重算：使用 LibreOffice 重算 Excel 公式"""

import logging
import subprocess
from pathlib import Path
from typing import Any

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class FormulaRecalcCapability(Capability):
    """公式重算：使用 LibreOffice 重算 Excel 公式"""

    @property
    def name(self) -> str:
        return "recalc_formulas"

    @property
    def chapter(self) -> str:
        return "equation"

    @property
    def description(self) -> str:
        return "使用 LibreOffice 重算 Excel 公式（扫描所有错误）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(
                name="output", type="string", description="输出文件路径", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        output = params.get("output")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件操作失败: 文件不存在 {file_path}")

        try:
            result = subprocess.run(
                ["soffice", "--version"], capture_output=True, text=True, timeout=10
            )
            if result.returncode != 0:
                raise DataError("数据操作失败: LibreOffice 未安装，请先安装 LibreOffice")
        except FileNotFoundError:
            raise DataError("数据操作失败: LibreOffice 未安装，请先安装 LibreOffice")

        try:
            output_path = output or str(path.with_suffix(".recalc.xlsx"))

            result = subprocess.run(
                [
                    "soffice",
                    "--headless",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(path.parent),
                    str(path),
                ],
                capture_output=True,
                text=True,
                timeout=60,
            )

            if result.returncode != 0:
                raise DataError(f"数据操作失败: LibreOffice 转换失败 {result.stderr}")

            errors = self._scan_errors(output_path)

            return {
                "file": file_path,
                "output": output_path,
                "errors_found": len(errors),
                "errors": errors,
                "recalculated": True,
            }

        except subprocess.TimeoutExpired:
            raise DataError("数据操作失败: LibreOffice 转换超时")
        except Exception as e:
            logger.error(f"公式重算失败: {e}")
            raise DataError(f"数据操作失败: {e}")

    def _scan_errors(self, file_path: str) -> dict:
        """扫描 Excel 错误"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)
            errors = []

            error_types = ["#REF!", "#N/A", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#NUM!"]

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            for error in error_types:
                                if error in cell.value:
                                    errors.append(
                                        {
                                            "sheet": sheet_name,
                                            "cell": cell.coordinate,
                                            "error": error,
                                            "value": cell.value,
                                        }
                                    )

            wb.close()
            return errors

        except Exception as e:
            logger.warning(f"Failed to scan errors: {e}")
            return []
