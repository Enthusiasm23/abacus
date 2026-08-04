"""Excel 格式化工作流 - 基于 igorwarzocha 的参考文档实现"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class FormattingWorkflowCapability(Capability):
    """Excel 格式化工作流 - 批量格式化操作"""

    @property
    def name(self) -> str:
        return "formatting_workflow"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "Excel 格式化工作流（批量设置样式、边框、填充、对齐）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(name="file", type="string", description="文件路径", required=True),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="范围", required=True),
            CapabilitySchema(name="font", type="object", description="字体设置", required=False),
            CapabilitySchema(name="fill", type="object", description="填充设置", required=False),
            CapabilitySchema(name="border", type="object", description="边框设置", required=False),
            CapabilitySchema(
                name="alignment", type="object", description="对齐设置", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        font_params = params.get("font")
        fill_params = params.get("fill")
        border_params = params.get("border")
        alignment_params = params.get("alignment")

        if not file_path:
            raise DataError("file parameter is required")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(path)

        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        from ..cell_utils import parse_range

        start_row, start_col, end_row, end_col = parse_range(range_str)
        if end_row is None:
            end_row = ws.max_row
        if end_col is None:
            end_col = ws.max_column

        count = 0
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)

                # 应用字体
                if font_params:
                    font_args = {}
                    if "name" in font_params:
                        font_args["name"] = font_params["name"]
                    if "size" in font_params:
                        font_args["size"] = font_params["size"]
                    if "bold" in font_params:
                        font_args["bold"] = font_params["bold"]
                    if "italic" in font_params:
                        font_args["italic"] = font_params["italic"]
                    if "color" in font_params:
                        color = font_params["color"]
                        if not color.startswith("FF"):
                            color = f"FF{color}"
                        font_args["color"] = Color(rgb=color)
                    if font_args:
                        cell.font = Font(**font_args)

                # 应用填充
                if fill_params:
                    fill_args = {}
                    if "color" in fill_params:
                        color = fill_params["color"]
                        if not color.startswith("FF"):
                            color = f"FF{color}"
                        fill_args["fgColor"] = Color(rgb=color)
                    if "pattern_type" in fill_params:
                        fill_args["patternType"] = fill_params["pattern_type"]
                    else:
                        fill_args["patternType"] = "solid"
                    cell.fill = PatternFill(**fill_args)

                # 应用边框
                if border_params:
                    border_args = {}
                    for side_name in ["left", "right", "top", "bottom"]:
                        if side_name in border_params:
                            side_params = border_params[side_name]
                            side_args = {}
                            if "style" in side_params:
                                side_args["style"] = side_params["style"]
                            if "color" in side_params:
                                color = side_params["color"]
                                if not color.startswith("FF"):
                                    color = f"FF{color}"
                                side_args["color"] = Color(rgb=color)
                            border_args[side_name] = Side(**side_args)
                    cell.border = Border(**border_args)

                # 应用对齐
                if alignment_params:
                    align_args = {}
                    if "horizontal" in alignment_params:
                        align_args["horizontal"] = alignment_params["horizontal"]
                    if "vertical" in alignment_params:
                        align_args["vertical"] = alignment_params["vertical"]
                    if "wrap_text" in alignment_params:
                        align_args["wrap_text"] = alignment_params["wrap_text"]
                    cell.alignment = Alignment(**align_args)

                count += 1

        wb.save(file_path)
        wb.close()

        return {
            "file": file_path,
            "sheet": sheet_name,
            "range": range_str,
            "cells_formatted": count,
        }
