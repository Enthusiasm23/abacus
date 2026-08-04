"""Excel 工作流 - 基于 igorwarzocha 的参考文档实现"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class SpreadsheetWorkflowCapability(Capability):
    """Excel 工作流 - 创建、读取、编辑 Excel 文件"""

    @property
    def name(self) -> str:
        return "spreadsheet_workflow"

    @property
    def chapter(self) -> str:
        return "work"

    @property
    def description(self) -> str:
        return "Excel 工作流（创建、读取、编辑、格式化）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="action",
                type="string",
                description="操作（create/read/edit/format）",
                required=True,
            ),
            CapabilitySchema(name="file", type="string", description="文件路径", required=False),
            CapabilitySchema(
                name="output", type="string", description="输出文件路径", required=False
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=False),
            CapabilitySchema(
                name="data", type="object", description="数据（create/edit 时）", required=False
            ),
            CapabilitySchema(
                name="range", type="string", description="范围（read/edit 时）", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        action = params.get("action")
        file_path = params.get("file")
        output = params.get("output")
        sheet_name = params.get("sheet")
        data = params.get("data")
        range_str = params.get("range")

        if not action:
            raise DataError("action parameter is required")

        if action == "create":
            return self._create_spreadsheet(output, sheet_name, data)
        elif action == "read":
            return self._read_spreadsheet(file_path, sheet_name, range_str)
        elif action == "edit":
            return self._edit_spreadsheet(file_path, sheet_name, data, range_str)
        elif action == "format":
            return self._format_spreadsheet(file_path, sheet_name, range_str, params)
        else:
            raise DataError(f"Unknown action: {action}")

    def _create_spreadsheet(self, output: str, sheet_name: str, data: dict) -> dict:
        """创建新电子表格"""
        if not output:
            raise DataError("output parameter is required for create")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name or "Sheet1"

        if data:
            # 写入数据
            for row_idx, row_data in enumerate(data.get("rows", []), 1):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(output)
        wb.close()

        return {
            "action": "create",
            "output": output,
            "sheet": ws.title,
            "rows": len(data.get("rows", [])) if data else 0,
        }

    def _read_spreadsheet(self, file_path: str, sheet_name: str, range_str: str) -> dict:
        """读取电子表格"""
        if not file_path:
            raise DataError("file parameter is required for read")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(path, data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        # 读取数据
        data = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data.append((row))

        wb.close()

        return {
            "action": "read",
            "file": file_path,
            "sheet": sheet_name,
            "rows": len(data),
            "columns": len(data[0]) if data else 0,
            "data": data[:100],  # 限制返回前100行
        }

    def _edit_spreadsheet(
        self, file_path: str, sheet_name: str, data: dict, range_str: str
    ) -> dict:
        """编辑电子表格"""
        if not file_path:
            raise DataError("file parameter is required for edit")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(path)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 写入数据
        if data and "rows" in data:
            start_row = 1
            start_col = 1

            if range_str:
                from ..cell_utils import parse_range

                start_row, start_col, _, _ = parse_range(range_str)

            for row_idx, row_data in enumerate(data["rows"], start_row):
                for col_idx, value in enumerate(row_data, start_col):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(file_path)
        wb.close()

        return {
            "action": "edit",
            "file": file_path,
            "sheet": sheet_name,
            "rows_written": len(data.get("rows", [])) if data else 0,
        }

    def _format_spreadsheet(
        self, file_path: str, sheet_name: str, range_str: str, params: dict
    ) -> dict:
        """格式化电子表格"""
        if not file_path:
            raise DataError("file parameter is required for format")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(path)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 应用格式
        if range_str:
            from ..cell_utils import parse_range

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)

                    if "bold" in params or "font_size" in params:
                        font_kwargs = {}
                        if "bold" in params:
                            font_kwargs["bold"] = params["bold"]
                        if "font_size" in params:
                            font_kwargs["size"] = params["font_size"]
                        cell.font = Font(**font_kwargs)
                    if "bg_color" in params:
                        cell.fill = PatternFill(patternType="solid", fgColor=params["bg_color"])

        wb.save(file_path)
        wb.close()

        return {"action": "format", "file": file_path, "sheet": sheet_name, "range": range_str}
