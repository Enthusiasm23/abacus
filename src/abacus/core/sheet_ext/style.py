"""工作表样式操作"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class SheetStyleCapability(Capability):
    @property
    def name(self) -> str:
        return "manage_sheet_style"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "管理工作表样式（标签颜色）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="action", type="string", description="操作（set/get/clear）", required=True
            ),
            CapabilitySchema(
                name="color", type="string", description="颜色（如 FF0000）", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        action = params.get("action")
        color = params.get("color")

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            if action == "set":
                if not color:
                    raise DataError("color required for set action")
                ws.sheet_properties.tabColor = color
                result = {"action": "set", "sheet": sheet_name, "color": color}
            elif action == "get":
                current_color = ws.sheet_properties.tabColor
                result = {
                    "action": "get",
                    "sheet": sheet_name,
                    "color": str(current_color) if current_color else None,
                }
            elif action == "clear":
                ws.sheet_properties.tabColor = None
                result = {"action": "clear", "sheet": sheet_name}
            else:
                raise DataError(f"Unknown action: {action}")

            wb.save(file_path)
            wb.close()
            return result
        except Exception as e:
            logger.error(f"Failed to manage sheet style: {e}")
            raise DataError(str(e))
