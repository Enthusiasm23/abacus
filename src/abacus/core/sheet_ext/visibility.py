"""工作表可见性操作"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError

logger = logging.getLogger(__name__)


class SheetVisibilityCapability(Capability):
    @property
    def name(self) -> str:
        return "manage_sheet_visibility"

    @property
    def chapter(self) -> str:
        return "field"

    @property
    def description(self) -> str:
        return "管理工作表可见性（显示/隐藏/非常隐藏）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(
                name="action",
                type="string",
                description="操作（show/hide/very-hide/get）",
                required=True,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        action = params.get("action")

        if not file_path:
            raise DataError("file parameter is required")

        try:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(path)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            if action == "show":
                ws.sheet_state = "visible"
                result = {"action": "show", "sheet": sheet_name}
            elif action == "hide":
                ws.sheet_state = "hidden"
                result = {"action": "hide", "sheet": sheet_name}
            elif action == "very_hide":
                ws.sheet_state = "veryHidden"
                result = {"action": "very_hide", "sheet": sheet_name}
            elif action == "get":
                result = {"action": "get", "sheet": sheet_name, "state": ws.sheet_state}
            else:
                raise DataError(f"Unknown action: {action}")

            wb.save(file_path)
            wb.close()
            return result
        except Exception as e:
            logger.error(f"Failed to manage sheet visibility: {e}")
            raise DataError(str(e))
