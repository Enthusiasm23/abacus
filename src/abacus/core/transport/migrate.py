"""均输章 - 迁移：深度实现数据迁移"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class MigrateCapability(Capability):
    @property
    def name(self) -> str:
        return "migrate"

    @property
    def chapter(self) -> str:
        return "transport"

    @property
    def description(self) -> str:
        return "深度数据迁移（跨工作簿、跨工作表）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(name="source", type="string", description="源文件路径", required=True),
            CapabilitySchema(
                name="target", type="string", description="目标文件路径", required=True
            ),
            CapabilitySchema(
                name="sheets",
                type="array",
                description="要迁移的工作表列表（可选，默认全部）",
                required=False,
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        source = params.get("source")
        target = params.get("target")
        sheets = params.get("sheets")

        if not source or not target:
            raise ValidationError("执行失败: 缺少必要参数 source 或 target")

        try:
            source_path = Path(source)
            if not source_path.exists():
                raise FileNotFoundError(f"文件操作失败: 源文件不存在 {source}")

            wb = load_workbook(source_path)

            if sheets:
                sheets_to_migrate = [s for s in sheets if s in wb.sheetnames]
            else:
                sheets_to_migrate = wb.sheetnames

            # 保存到目标
            target_path = Path(target)
            target_path.parent.mkdir(parents=True, exist_ok=True)

            # 只保留要迁移的工作表
            for sheet in wb.sheetnames:
                if sheet not in sheets_to_migrate:
                    del wb[sheet]

            wb.save(target_path)
            wb.close()

            return {
                "source": source,
                "target": target,
                "sheets_migrated": len(sheets_to_migrate),
                "sheet_names": sheets_to_migrate,
            }
        except Exception as e:
            raise DataError(f"数据操作失败: {e}")
