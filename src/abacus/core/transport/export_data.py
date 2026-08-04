"""均输章 - 导出：深度实现数据导出"""

import csv
import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..base import Capability, CapabilitySchema
from ..cell_utils import parse_range
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class ExportDataCapability(Capability):
    """导出：深度实现数据导出"""

    @property
    def name(self) -> str:
        return "export_data"

    @property
    def chapter(self) -> str:
        return "transport"

    @property
    def description(self) -> str:
        return "深度导出数据（CSV、JSON）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径", required=True
            ),
            CapabilitySchema(name="sheet", type="string", description="工作表名称", required=True),
            CapabilitySchema(name="range", type="string", description="数据范围", required=True),
            CapabilitySchema(
                name="output", type="string", description="输出文件路径", required=True
            ),
            CapabilitySchema(
                name="format", type="string", description="输出格式（csv/json）", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        sheet_name = params.get("sheet")
        range_str = params.get("range")
        output = params.get("output")
        format_type = params.get("format", "csv")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")
        if not output:
            raise ValidationError("执行失败: 缺少必要参数 output")

        return self._export_data(file_path, sheet_name, range_str, output, format_type)

    def _export_data(
        self, filepath: str, sheet_name: str, range_str: str, output: str, format_type: str
    ) -> dict[str, Any]:
        try:
            path = Path(filepath)
            if not path.exists():
                raise FileNotFoundError(f"文件操作失败: 文件不存在 {filepath}")

            wb = load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise DataError(f"数据操作失败: 工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            start_row, start_col, end_row, end_col = parse_range(range_str)
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column

            # 读取数据
            headers = []
            for col in range(start_col, end_col + 1):
                headers.append(ws.cell(row=start_row, column=col).value)

            rows = []
            for row in range(start_row + 1, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    row_data.append(ws.cell(row=row, column=col).value)
                rows.append(row_data)

            wb.close()

            # 导出
            output_path = Path(output)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            if format_type == "csv":
                with open(output_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(rows)
            elif format_type == "json":
                data = []
                for row in rows:
                    row_dict = {}
                    for i, header in enumerate(headers):
                        row_dict[header] = row[i]
                    data.append(row_dict)

                with open(output_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            else:
                raise DataError(f"数据操作失败: 不支持的导出格式 {format_type}")

            return {
                "file": filepath,
                "output": str(output_path),
                "format": format_type,
                "rows_exported": len(rows),
                "columns_exported": len(headers),
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"数据导出失败: {e}")
            raise DataError(f"数据操作失败: {e}")
