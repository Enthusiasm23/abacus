"""均输章 - 导入：深度实现数据导入"""

import csv
import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from ..base import Capability, CapabilitySchema
from ..exceptions import DataError, FileNotFoundError, ValidationError

logger = logging.getLogger(__name__)


class ImportDataCapability(Capability):
    """导入：深度实现数据导入"""

    @property
    def name(self) -> str:
        return "import_data"

    @property
    def chapter(self) -> str:
        return "transport"

    @property
    def description(self) -> str:
        return "深度导入数据（CSV、JSON）"

    @property
    def schema(self) -> list[CapabilitySchema]:
        return [
            CapabilitySchema(
                name="file", type="string", description="Excel 文件路径（目标）", required=True
            ),
            CapabilitySchema(name="source", type="string", description="源文件路径", required=True),
            CapabilitySchema(
                name="source_type",
                type="string",
                description="源文件类型（csv/json）",
                required=False,
            ),
            CapabilitySchema(
                name="sheet", type="string", description="目标工作表名称", required=False
            ),
        ]

    def execute(self, context: Any, **params) -> Any:
        file_path = params.get("file")
        source = params.get("source")
        source_type = params.get("source_type", "csv")
        sheet_name = params.get("sheet", "Sheet1")

        if not file_path:
            raise ValidationError("执行失败: 缺少必要参数 file")
        if not source:
            raise ValidationError("执行失败: 缺少必要参数 source")

        return self._import_data(file_path, source, source_type, sheet_name)

    def _import_data(self, filepath: str, source: str, source_type: str, sheet_name: str) -> dict:
        try:
            source_path = Path(source)
            if not source_path.exists():
                raise FileNotFoundError(f"文件操作失败: 源文件不存在 {source}")

            # 读取源数据
            if source_type == "csv":
                with open(source_path, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    data = list(reader)
            elif source_type == "json":
                with open(source_path, "r", encoding="utf-8") as f:
                    json_data = json.load(f)
                    if isinstance(json_data, list) and json_data:
                        headers = list(json_data[0].keys())
                        data = [headers] + [[row.get(h) for h in headers] for row in json_data]
                    else:
                        raise DataError("数据操作失败: 无效的 JSON 格式，期望对象列表")
            else:
                raise DataError(f"数据操作失败: 不支持的源文件类型 {source_type}")

            if not data:
                raise DataError("数据操作失败: 没有可导入的数据")

            # 创建或打开目标工作簿
            filepath_path = Path(filepath)
            if filepath_path.exists():
                wb = load_workbook(filepath_path)
            else:
                wb = Workbook()
                filepath_path.parent.mkdir(parents=True, exist_ok=True)

            # 创建或获取工作表
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            # 写入数据
            for row_idx, row in enumerate(data, 1):
                for col_idx, value in enumerate(row, 1):
                    # 尝试转换数字
                    try:
                        if value and value.replace(".", "").replace("-", "").isdigit():
                            value = float(value) if "." in value else int(value)
                    except (ValueError, AttributeError):
                        pass
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(filepath)
            wb.close()

            return {
                "file": filepath,
                "source": source,
                "source_type": source_type,
                "sheet": sheet_name,
                "rows_imported": len(data),
                "columns_imported": len(data[0]) if data else 0,
            }

        except (FileNotFoundError, DataError):
            raise
        except Exception as e:
            logger.error(f"数据导入失败: {e}")
            raise DataError(f"数据操作失败: {e}")
